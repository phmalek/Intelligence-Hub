from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict
import sys
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from scipy.optimize import curve_fit

BASE_DIR = Path(__file__).resolve().parent.parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))


def normalize_market(value: str) -> str:
    return str(value or "").strip().upper()


def extract_targets_from_workbook(workbook_path: Path) -> pd.DataFrame:
    wb = load_workbook(workbook_path, data_only=True)
    if "Share Of MAL-Budget" not in wb.sheetnames:
        raise ValueError("Sheet 'Share Of MAL-Budget' not found.")
    ws = wb["Share Of MAL-Budget"]

    # Find row with market headers (starts with 'Market').
    market_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        if str(ws.cell(r, 1).value or "").strip().lower() == "market":
            market_row = r
            break
    if market_row is None:
        raise ValueError("Could not find market header row in 'Share Of MAL-Budget'.")

    # Find target rows.
    sales_target_row = None
    sales_target_focus_row = None
    required_dcfs_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        label = str(ws.cell(r, 1).value or "").strip().lower()
        if label == "sales target (2026)":
            sales_target_row = r
        if "sales target (focus)" in label:
            sales_target_focus_row = r
        if "reqired dcfs" in label or "required dcfs" in label:
            required_dcfs_row = r
    if sales_target_row is None:
        raise ValueError("Could not find 'Sales Target (2026)' row.")
    if sales_target_focus_row is None:
        raise ValueError("Could not find 'Sales Target (Focus)' row.")
    if required_dcfs_row is None:
        raise ValueError("Could not find 'Reqired DCFS' row.")

    records = []
    def _to_float(val):
        if val in (None, ""):
            return None
        try:
            return float(val)
        except Exception:
            return None

    for c in range(2, ws.max_column + 1):
        market = ws.cell(market_row, c).value
        if market in (None, ""):
            continue
        t_all = ws.cell(sales_target_row, c).value
        t_focus = ws.cell(sales_target_focus_row, c).value
        t_dcfs = ws.cell(required_dcfs_row, c).value
        t_all_f = _to_float(t_all)
        t_focus_f = _to_float(t_focus)
        t_dcfs_f = _to_float(t_dcfs)
        if t_all_f is None and t_focus_f is None and t_dcfs_f is None:
            continue
        records.append(
            {
                "Market_Target": str(market).strip(),
                "sales_target_2026": t_all_f,
                "sales_target_focus_2026": t_focus_f,
                "target_dcfs_2026": t_dcfs_f,
            }
        )
    target_df = pd.DataFrame(records)
    target_df["Market_Target_Norm"] = target_df["Market_Target"].map(normalize_market)
    return target_df


def build_budget_table() -> pd.DataFrame:
    budget_2026_raw = [
        ("PCGB", 944, 275000),
        ("PD", 1427, 260000),
        ("PKO", 713, 270000),
        ("POF", 1406, 152000),
        ("PIT", 1024, 145000),
        ("TRK", 269, 225000),
        ("PCL", 3068, 120000),
        ("PIB SPA", 2570, 80000),
        ("PNO", 2643, 70000),
        ("PIB POR", 597, 80000),
        ("PCH", 3301, 80000),
        ("PTW", 181, 230000),
        ("PCA", 2481, 70000),
        ("PPL", 2860, 70000),
        ("PJ", 1606, 60000),
        ("PBR", 1606, 60000),
        ("PCNA", 611, 245000),
    ]
    df = pd.DataFrame(budget_2026_raw, columns=["Market", "CPL_Scenario_5_EUR", "Budget_2026_EUR"])
    df["Market_Norm"] = df["Market"].map(normalize_market)
    return df


def _saturation_curve(x, a, b):
    return a * x / (b + x)


def _fit_market_curve(x_vals: np.ndarray, y_vals: np.ndarray):
    mask = np.isfinite(x_vals) & np.isfinite(y_vals) & (x_vals >= 0) & (y_vals >= 0)
    x = x_vals[mask]
    y = y_vals[mask]
    if len(x) < 3:
        return None, None
    a0 = max(float(np.max(y)), 1.0)
    b0 = max(float(np.median(x)), 1.0)
    try:
        params, _ = curve_fit(
            _saturation_curve,
            x,
            y,
            p0=[a0, b0],
            bounds=([0.0, 0.0], [np.inf, np.inf]),
            maxfev=20000,
        )
        return float(params[0]), float(params[1])
    except Exception:
        return None, None


def compute_curve_metrics_by_market(source_csv: Path, budget_df: pd.DataFrame) -> pd.DataFrame:
    src = pd.read_csv(source_csv)
    for col in ["Market", "Media Spend", "DCFS"]:
        if col not in src.columns:
            raise ValueError(f"Missing required column in source csv: {col}")
    src["Media Spend"] = pd.to_numeric(src["Media Spend"], errors="coerce")
    src["DCFS"] = pd.to_numeric(src["DCFS"], errors="coerce")

    time_col = None
    for c in ["Date", "report_date", "calendar_week"]:
        if c in src.columns:
            time_col = c
            break
    if time_col is None:
        raise ValueError("Need one of Date/report_date/calendar_week for curve fitting.")

    agg = (
        src.groupby(["Market", time_col], dropna=False)
        .agg({"Media Spend": "sum", "DCFS": "sum"})
        .reset_index()
    )
    spend_2025 = (
        src.groupby("Market", dropna=False)
        .agg(spend_2025=("Media Spend", "sum"), dcfs_2025=("DCFS", "sum"))
        .reset_index()
    )

    records = []
    budget_map = {normalize_market(r["Market"]): float(r["Budget_2026_EUR"]) for _, r in budget_df.iterrows()}

    for market, g in agg.groupby("Market", dropna=False):
        market_norm = normalize_market(market)
        budget_2026 = budget_map.get(market_norm)
        if budget_2026 is None:
            continue
        x = g["Media Spend"].to_numpy(dtype=float)
        y = g["DCFS"].to_numpy(dtype=float)
        a, b = _fit_market_curve(x, y)
        rec = {
            "Market": market,
            "Market_Norm": market_norm,
            "curve_a": a,
            "curve_b": b,
        }
        if a is not None and b is not None and b > 0:
            spend_hist = float(
                spend_2025.loc[spend_2025["Market"] == market, "spend_2025"].iloc[0]
            ) if market in set(spend_2025["Market"]) else np.nan
            pred_2026 = float(_saturation_curve(budget_2026, a, b))
            pred_2025 = float(_saturation_curve(spend_hist, a, b)) if np.isfinite(spend_hist) else np.nan
            delta_ref = max(5000.0, 0.05 * budget_2026)
            marginal = float(_saturation_curve(budget_2026 + delta_ref, a, b) - _saturation_curve(budget_2026, a, b)) / delta_ref
            rec.update(
                {
                    "predicted_dcfs_2026_curve": pred_2026,
                    "predicted_dcfs_2025_curve": pred_2025,
                    "marginal_dcfs_per_eur": marginal,
                }
            )
        else:
            rec.update(
                {
                    "predicted_dcfs_2026_curve": np.nan,
                    "predicted_dcfs_2025_curve": np.nan,
                    "marginal_dcfs_per_eur": np.nan,
                }
            )
        records.append(rec)

    return pd.DataFrame(records)


def map_targets_to_budget(budget_df: pd.DataFrame, target_df: pd.DataFrame) -> pd.DataFrame:
    out = budget_df.copy()

    # Standard direct mapping.
    merged = out.merge(
        target_df[["Market_Target_Norm", "sales_target_2026", "sales_target_focus_2026", "target_dcfs_2026"]],
        left_on="Market_Norm",
        right_on="Market_Target_Norm",
        how="left",
    )

    # Handle PIB split in budget vs combined PIB target in workbook.
    pib_target = target_df.loc[target_df["Market_Target_Norm"] == "PIB", "sales_target_focus_2026"]
    if not pib_target.empty:
        pib_total = float(pib_target.iloc[0])
        pib_mask = merged["Market_Norm"].isin(["PIB SPA", "PIB POR"])
        pib_budget_sum = merged.loc[pib_mask, "Budget_2026_EUR"].sum()
        if pib_budget_sum > 0:
            merged.loc[pib_mask, "sales_target_focus_2026"] = (
                merged.loc[pib_mask, "Budget_2026_EUR"] / pib_budget_sum * pib_total
            )

    return merged


def assign_clusters(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["has_required_data"] = out["marginal_dcfs_per_eur"].notna() & out["target_dcfs_2026"].notna() & out["predicted_dcfs_2026_curve"].notna()

    valid = out["has_required_data"]
    out.loc[valid, "target_gap_dcfs"] = out.loc[valid, "target_dcfs_2026"] - out.loc[valid, "predicted_dcfs_2026_curve"]
    out.loc[valid, "target_gap_rate"] = out.loc[valid, "target_gap_dcfs"] / out.loc[valid, "target_dcfs_2026"]
    # Keep compatibility aliases with previous downstream scripts.
    out.loc[valid, "opportunity_score"] = out.loc[valid, "marginal_dcfs_per_eur"]
    out.loc[valid, "market_gap_raw"] = out.loc[valid, "target_gap_rate"]
    out.loc[valid, "opportunity_pct"] = out.loc[valid, "marginal_dcfs_per_eur"].rank(pct=True) * 100
    out.loc[valid, "market_gap_pct"] = out.loc[valid, "target_gap_rate"].rank(pct=True) * 100

    valid_idx = out.index[valid]
    out.loc[valid_idx, "opp_band"] = pd.qcut(
        out.loc[valid_idx, "opportunity_pct"], q=3, labels=["Low", "Mid", "High"]
    ).astype(str)
    out.loc[valid_idx, "gap_band"] = pd.qcut(
        out.loc[valid_idx, "market_gap_pct"], q=3, labels=["Low", "Mid", "High"]
    ).astype(str)

    matrix_to_uplift: Dict[tuple[str, str], str] = {
        ("High", "High"): "+60%",
        ("High", "Mid"): "+60%",
        ("Mid", "High"): "+60%",
        ("High", "Low"): "+50%",
        ("Mid", "Mid"): "+50%",
        ("Low", "High"): "+50%",
        ("Mid", "Low"): "+40%",
        ("Low", "Mid"): "+40%",
        ("Low", "Low"): "+40%",
    }

    out["aggressiveness_uplift"] = "Data Gap"
    for i in valid_idx:
        key = (out.at[i, "opp_band"], out.at[i, "gap_band"])
        out.at[i, "aggressiveness_uplift"] = matrix_to_uplift.get(key, "Data Gap")

    out["cluster_label"] = out["aggressiveness_uplift"].map(
        {
            "+60%": "Cluster A - High aggressiveness",
            "+50%": "Cluster B - Medium aggressiveness",
            "+40%": "Cluster C - Baseline aggressiveness",
            "Data Gap": "Cluster D - Data gap (manual review)",
        }
    )
    return out


def save_scatter_html(df: pd.DataFrame, out_html: Path) -> None:
    fig = px.scatter(
        df,
        x="market_gap_pct",
        y="opportunity_pct",
        color="aggressiveness_uplift",
        text="Market",
        size="Budget_2026_EUR",
        size_max=38,
        color_discrete_map={"+60%": "#1b9e77", "+50%": "#d95f02", "+40%": "#7570b3", "Data Gap": "#9CA3AF"},
        title="Market Clusters: Incremental Opportunity vs 2026 Target Gap",
    )
    fig.update_traces(textposition="top center")
    fig.update_layout(
        xaxis_title="Target Gap Percentile (higher = larger shortfall vs 2026 DCFS target)",
        yaxis_title="Incremental Opportunity Percentile (marginal DCFS per EUR)",
        legend_title="Recommended uplift band",
        template="plotly_white",
    )
    fig.write_html(out_html)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run market cluster analysis with targets from reverse-funnel workbook.")
    parser.add_argument(
        "--source-csv",
        type=Path,
        default=Path("pwc reports/outputs/python_output_all.csv"),
        help="Source performance CSV used for opportunity scoring.",
    )
    parser.add_argument(
        "--target-workbook",
        type=Path,
        default=Path("budget_setting/2026_03_03_RevFLight_Simplified.xlsx"),
        help="Workbook containing market-level 2026 targets.",
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=Path("budget_setting/market_cluster_aggressiveness_output.csv"),
        help="Latest output CSV path (also copied into timestamped run folder).",
    )
    parser.add_argument(
        "--output-html",
        type=Path,
        default=Path("budget_setting/market_clusters_scatter.html"),
        help="Latest scatter HTML path (also copied into timestamped run folder).",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=Path("budget_setting/runs"),
        help="Root folder where timestamped run outputs are created.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    budget_df = build_budget_table()
    target_df = extract_targets_from_workbook(args.target_workbook)
    curve_df = compute_curve_metrics_by_market(args.source_csv, budget_df)

    merged = map_targets_to_budget(budget_df, target_df)
    merged = merged.merge(
        curve_df[
            [
                "Market_Norm",
                "curve_a",
                "curve_b",
                "predicted_dcfs_2026_curve",
                "predicted_dcfs_2025_curve",
                "marginal_dcfs_per_eur",
            ]
        ],
        on="Market_Norm",
        how="left",
    )

    out = assign_clusters(merged)

    # Timestamped output folder per run.
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = args.output_root / run_ts
    run_dir.mkdir(parents=True, exist_ok=True)

    run_csv = run_dir / "market_cluster_aggressiveness_output.csv"
    run_html = run_dir / "market_clusters_scatter.html"

    out.to_csv(run_csv, index=False)
    save_scatter_html(out, run_html)

    # Keep latest compatibility files in budget_setting root.
    out.to_csv(args.output_csv, index=False)
    save_scatter_html(out, args.output_html)

    print(f"Saved run folder: {run_dir}")
    print(f"Saved: {run_csv}")
    print(f"Saved: {run_html}")
    print(f"Updated latest: {args.output_csv}")
    print(f"Updated latest: {args.output_html}")
    print("Columns available:", ", ".join(out.columns))


if __name__ == "__main__":
    main()

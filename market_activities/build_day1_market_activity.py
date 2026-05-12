from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / '2026-03-23_CTG_Kalkulation_final.xlsx'
LOCAL_OUT = BASE_DIR / 'local_activity_normalized.csv'
CTG_OUT = BASE_DIR / 'ctg_timeline_normalized.csv'
SIGNALS_OUT = BASE_DIR / 'market_activity_weekly_signals.csv'
REPORT_OUT = BASE_DIR / 'market_activity_day1_report.md'
CHANNEL_TAXONOMY_OUT = BASE_DIR / 'channel_taxonomy_day2.csv'
DAY2_REPORT_OUT = BASE_DIR / 'market_activity_day2_report.md'

FOCUS_MODELS = ['Macan BEV', 'Cayenne E3 II', 'Cayenne E4']
LOWER_FUNNEL_CHANNELS = {
    'Search',
    'Google Ads',
    'Google Inventory Ads',
    'META Inventory Ads',
    'Display',
    'Retail Marketplace',
}
UPPER_FUNNEL_CHANNELS = {
    'Social',
    'Paid Social',
    'Programmatic',
    'CTV',
    'DOOH',
    'YouTube',
    'Cinemas',
    'Streaming TV',
    'Branded Content',
    'Other Online Channel',
    'Brand Days',
}


def parse_date(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace('Z', '+00:00')
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError:
        return pd.to_datetime(text, errors='coerce').to_pydatetime()


def week_start(date_value):
    if date_value is None or pd.isna(date_value):
        return None
    return date_value - timedelta(days=date_value.weekday())


def normalize_model(value):
    text = str(value or '').strip()
    compact = re.sub(r'\s+', ' ', text).lower()
    if compact in {'macan h2', 'macan bev'}:
        return 'Macan BEV'
    if 'cayenne e3' in compact:
        return 'Cayenne E3 II'
    if 'cayenne e4' in compact:
        return 'Cayenne E4'
    return text


def normalize_channel(value):
    text = str(value or '').strip()
    lookup = text.lower()
    if 'google inventory' in lookup:
        return 'Google Inventory Ads'
    if 'meta inventory' in lookup:
        return 'META Inventory Ads'
    if lookup in {'google ads', 'search', 'sem', 'sea', 'always-on sea', 'always on sem'}:
        return 'Search' if lookup != 'google ads' else 'Google Ads'
    if 'paid social' in lookup or lookup == 'social':
        return 'Paid Social' if 'paid' in lookup else 'Social'
    if 'programmatic' in lookup:
        return 'Programmatic'
    if 'streaming tv' in lookup:
        return 'Streaming TV'
    if 'youtube' in lookup:
        return 'YouTube'
    if 'dooh' in lookup:
        return 'DOOH'
    if 'cinema' in lookup:
        return 'Cinemas'
    if 'branded content' in lookup:
        return 'Branded Content'
    if 'display' in lookup:
        return 'Display'
    if 'ctv' in lookup:
        return 'CTV'
    if 'other online' in lookup:
        return 'Other Online Channel'
    if 'webmotors' in lookup or 'marketplace' in lookup:
        return 'Retail Marketplace'
    if 'brand days' in lookup:
        return 'Brand Days'
    return text


def split_channels(value):
    text = str(value or '').strip()
    if not text:
        return []
    parts = [normalize_channel(part) for part in re.split(r',|&', text)]
    return sorted({part for part in parts if part})


def weeks_between(start, end):
    start_week = week_start(start)
    end_week = week_start(end)
    if start_week is None or end_week is None:
        return []
    weeks = []
    current = start_week
    while current <= end_week:
        weeks.append(current)
        current += timedelta(days=7)
    return weeks


def channel_role(channels):
    channel_set = set(channels)
    has_lower = bool(channel_set & LOWER_FUNNEL_CHANNELS)
    has_upper = bool(channel_set & UPPER_FUNNEL_CHANNELS)
    if has_lower and has_upper:
        return 'full_funnel'
    if has_lower:
        return 'lower_funnel'
    if has_upper:
        return 'upper_funnel'
    return 'unknown'


def parse_budget(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, '', 'missing'
    text = str(value).strip()
    if not text:
        return None, '', 'missing'
    currency = ''
    if re.search(r'\bchf\b', text, flags=re.I):
        currency = 'CHF'
    elif re.search(r'\beur\b|€', text, flags=re.I):
        currency = 'EUR'
    number_match = re.search(r"(\d+(?:[.,']\d+)*)\s*([kKmM])?", text)
    if not number_match:
        return None, currency, 'unparsed'
    number_text = number_match.group(1).replace("'", '').replace(',', '')
    try:
        amount = float(number_text)
    except ValueError:
        return None, currency, 'unparsed'
    suffix = (number_match.group(2) or '').lower()
    if suffix == 'k':
        amount *= 1_000
    elif suffix == 'm':
        amount *= 1_000_000
    return amount, currency, 'parsed'


def quality_score(data_quality, budget_quality):
    score = 100
    if data_quality != 'complete':
        score -= 45
    if budget_quality in {'missing', 'unparsed'}:
        score -= 15
    return max(score, 0)


def confidence_label(score):
    if score >= 80:
        return 'High'
    if score >= 55:
        return 'Medium'
    return 'Low'


def budget_direction(action, confidence, duplication_score):
    if confidence == 'Low':
        return 'Validate before moving budget'
    if action == 'Harvest local upper-funnel demand':
        return 'Upweight CTG lower-funnel'
    if action == 'CTG filling local gap':
        return 'Protect or upweight CTG'
    if action == 'Check duplication before upweighting':
        return 'Hold or reduce duplicated CTG'
    if action == 'Potential CTG whitespace':
        return 'Review for CTG test'
    if duplication_score > 0:
        return 'Coordinate with market'
    return 'Maintain'


def planning_bucket(direction):
    if direction in {'Upweight CTG lower-funnel', 'Protect or upweight CTG'}:
        return 'Scale'
    if direction == 'Hold or reduce duplicated CTG':
        return 'Reduce'
    if direction == 'Validate before moving budget':
        return 'Watch'
    if direction == 'Review for CTG test':
        return 'Fix'
    return 'Protect'


def recommendation_reason(row):
    market = row['market']
    model = row['model']
    week = row['week_start']
    ctg = row['ctg_channels'] or 'no CTG activity'
    local = row['local_channels'] or 'no confirmed local activity'
    action = row['recommended_action']
    confidence = row['confidence_level']
    if confidence == 'Low':
        return f"{market} / {model} in week {week}: validate local plan data before reallocating; current local activity fields are incomplete."
    if action == 'Harvest local upper-funnel demand':
        return f"{market} / {model} in week {week}: local activity is creating demand ({local}) and CTG has lower-funnel capture live ({ctg})."
    if action == 'CTG filling local gap':
        return f"{market} / {model} in week {week}: CTG is active ({ctg}) where no confirmed local activity is planned."
    if action == 'Check duplication before upweighting':
        return f"{market} / {model} in week {week}: CTG and local activity both cover lower-funnel channels, so check duplication before adding budget."
    if action == 'Potential CTG whitespace':
        return f"{market} / {model} in week {week}: local activity is live ({local}) but CTG has no matching support in the timeline."
    if action == 'Coordinated support':
        return f"{market} / {model} in week {week}: CTG and local activity are both live; keep coordination tight before changing budget."
    return f"{market} / {model} in week {week}: no active optimisation signal."


def read_local_activity():
    df = pd.read_excel(WORKBOOK_PATH, sheet_name='Survey_Rohdaten')
    records = []
    for _, row in df.iterrows():
        market = str(row.get('Market') or '').strip()
        model = normalize_model(row.get('Baureihe'))
        start = parse_date(row.get('Start'))
        end = parse_date(row.get('Ende'))
        channels = split_channels(row.get('Kanal'))
        if not market or model not in FOCUS_MODELS:
            continue
        data_quality = 'complete' if start and end and channels else 'needs_follow_up'
        budget_amount, budget_currency, budget_quality = parse_budget(row.get('Budget'))
        confidence_score = quality_score(data_quality, budget_quality)
        records.append({
            'market': market,
            'model': model,
            'activity_name': str(row.get('Maßnahme') or '').strip(),
            'start_date': start.date().isoformat() if start else '',
            'end_date': end.date().isoformat() if end else '',
            'channels': ', '.join(channels),
            'channel_role': channel_role(channels),
            'budget_raw': row.get('Budget'),
            'budget_amount': budget_amount,
            'budget_currency': budget_currency,
            'budget_quality': budget_quality,
            'kpi_raw': row.get('KPI'),
            'data_quality': data_quality,
            'confidence_score': confidence_score,
            'confidence_level': confidence_label(confidence_score),
        })
    return pd.DataFrame(records)


def read_ctg_timeline():
    ws = load_workbook(WORKBOOK_PATH, data_only=True)['Media Mix']
    records = []
    market_pairs = []
    for col in range(3, ws.max_column + 1, 2):
        market = ws.cell(1, col).value
        if market:
            market_pairs.append((str(market).strip(), col, col + 1))
    for row in range(3, ws.max_row + 1):
        model = normalize_model(ws.cell(row, 1).value)
        measure = str(ws.cell(row, 2).value or '').strip()
        if model not in FOCUS_MODELS or not measure:
            continue
        measure_channel = normalize_channel(measure)
        for market, start_col, end_col in market_pairs:
            start = parse_date(ws.cell(row, start_col).value)
            end = parse_date(ws.cell(row, end_col).value)
            if start and end:
                records.append({
                    'market': market,
                    'model': model,
                    'ctg_measure': measure,
                    'ctg_channel': measure_channel,
                    'start_date': start.date().isoformat(),
                    'end_date': end.date().isoformat(),
                    'channel_role': channel_role([measure_channel]),
                })
    return pd.DataFrame(records)


def build_weekly_signals(local_df, ctg_df):
    local_week_rows = []
    for _, row in local_df.iterrows():
        start = parse_date(row['start_date']) if row['start_date'] else None
        end = parse_date(row['end_date']) if row['end_date'] else None
        channels = split_channels(row['channels'])
        for wk in weeks_between(start, end):
            local_week_rows.append({
                'market': row['market'],
                'model': row['model'],
                'week_start': wk.date().isoformat(),
                'local_channels': channels,
                'local_activity_count': 1,
                'local_confidence_score': row.get('confidence_score', 0),
                'local_data_quality': row.get('data_quality', 'needs_follow_up'),
            })
    local_week_df = pd.DataFrame(local_week_rows)

    ctg_week_rows = []
    for _, row in ctg_df.iterrows():
        start = parse_date(row['start_date'])
        end = parse_date(row['end_date'])
        for wk in weeks_between(start, end):
            ctg_week_rows.append({
                'market': row['market'],
                'model': row['model'],
                'week_start': wk.date().isoformat(),
                'ctg_channels': [row['ctg_channel']],
                'ctg_activity_count': 1,
            })
    ctg_week_df = pd.DataFrame(ctg_week_rows)

    keys = ['market', 'model', 'week_start']
    base = pd.concat([
        local_week_df[keys] if not local_week_df.empty else pd.DataFrame(columns=keys),
        ctg_week_df[keys] if not ctg_week_df.empty else pd.DataFrame(columns=keys),
    ]).drop_duplicates()

    rows = []
    incomplete_pairs = {
        (row['market'], row['model'])
        for _, row in local_df.iterrows()
        if row.get('data_quality') != 'complete'
    }
    for _, key_row in base.iterrows():
        market, model, week = key_row['market'], key_row['model'], key_row['week_start']
        local_slice = local_week_df[
            (local_week_df['market'] == market)
            & (local_week_df['model'] == model)
            & (local_week_df['week_start'] == week)
        ] if not local_week_df.empty else pd.DataFrame()
        ctg_slice = ctg_week_df[
            (ctg_week_df['market'] == market)
            & (ctg_week_df['model'] == model)
            & (ctg_week_df['week_start'] == week)
        ] if not ctg_week_df.empty else pd.DataFrame()

        local_channels = sorted({ch for channels in local_slice.get('local_channels', []) for ch in channels})
        ctg_channels = sorted({ch for channels in ctg_slice.get('ctg_channels', []) for ch in channels})
        local_role = channel_role(local_channels)
        ctg_role = channel_role(ctg_channels)
        local_lower = bool(set(local_channels) & LOWER_FUNNEL_CHANNELS)
        local_upper = bool(set(local_channels) & UPPER_FUNNEL_CHANNELS)
        ctg_lower = bool(set(ctg_channels) & LOWER_FUNNEL_CHANNELS)
        ctg_upper = bool(set(ctg_channels) & UPPER_FUNNEL_CHANNELS)
        overlap = bool(set(local_channels) & set(ctg_channels))

        gap_score = int(bool(ctg_channels) and not local_channels) + int(ctg_lower and not local_lower)
        harvest_score = int(local_upper and ctg_lower) + int(bool(local_channels) and not local_lower and ctg_lower)
        duplication_score = int(overlap) + int(local_lower and ctg_lower)

        local_confidence = (
            float(local_slice['local_confidence_score'].max())
            if not local_slice.empty and 'local_confidence_score' in local_slice.columns
            else 100.0
        )
        data_quality_note = 'complete'
        if (market, model) in incomplete_pairs and not local_channels:
            local_confidence = 45.0
            data_quality_note = 'local_plan_incomplete'
        elif not local_slice.empty and 'local_data_quality' in local_slice.columns:
            qualities = set(local_slice['local_data_quality'].dropna().astype(str))
            if qualities - {'complete'}:
                data_quality_note = 'local_plan_incomplete'
        confidence = confidence_label(local_confidence)

        if confidence == 'Low':
            action = 'Validate data before optimisation'
        elif not local_channels and ctg_channels:
            action = 'CTG filling local gap'
        elif local_upper and ctg_lower and not overlap:
            action = 'Harvest local upper-funnel demand'
        elif duplication_score >= 2:
            action = 'Check duplication before upweighting'
        elif local_channels and not ctg_channels:
            action = 'Potential CTG whitespace'
        elif local_channels and ctg_channels:
            action = 'Coordinated support'
        else:
            action = 'No active signal'

        direction = budget_direction(action, confidence, duplication_score)

        signal_row = {
            'market': market,
            'model': model,
            'week_start': week,
            'ctg_channels': ', '.join(ctg_channels),
            'local_channels': ', '.join(local_channels),
            'ctg_role': ctg_role,
            'local_role': local_role,
            'local_activity_count': int(local_slice['local_activity_count'].sum()) if not local_slice.empty else 0,
            'ctg_activity_count': int(ctg_slice['ctg_activity_count'].sum()) if not ctg_slice.empty else 0,
            'gap_score': gap_score,
            'harvest_score': harvest_score,
            'duplication_score': duplication_score,
            'confidence_score': round(local_confidence, 0),
            'confidence_level': confidence,
            'data_quality_note': data_quality_note,
            'recommended_action': action,
            'budget_direction': direction,
            'planning_bucket': planning_bucket(direction),
        }
        signal_row['recommendation_reason'] = recommendation_reason(signal_row)
        rows.append(signal_row)
    return pd.DataFrame(rows).sort_values(['week_start', 'market', 'model'])


def build_channel_taxonomy():
    survey_df = pd.read_excel(WORKBOOK_PATH, sheet_name='Survey_Rohdaten')
    values = []
    for value in survey_df.get('Kanal', pd.Series(dtype=object)).dropna():
        for raw in re.split(r',|&', str(value)):
            raw = raw.strip()
            if raw:
                normalized = normalize_channel(raw)
                values.append({
                    'raw_channel': raw,
                    'normalized_channel': normalized,
                    'role': channel_role([normalized]),
                })
    media_ws = load_workbook(WORKBOOK_PATH, data_only=True)['Media Mix']
    for row in range(3, media_ws.max_row + 1):
        raw = str(media_ws.cell(row, 2).value or '').strip()
        if raw:
            normalized = normalize_channel(raw)
            values.append({
                'raw_channel': raw,
                'normalized_channel': normalized,
                'role': channel_role([normalized]),
            })
    taxonomy_df = pd.DataFrame(values).drop_duplicates().sort_values(['normalized_channel', 'raw_channel'])
    return taxonomy_df


def build_report(local_df, ctg_df, signals_df):
    market_count = local_df['market'].nunique()
    local_complete = int((local_df['data_quality'] == 'complete').sum())
    local_follow_up = int((local_df['data_quality'] != 'complete').sum())
    top_actions = signals_df['recommended_action'].value_counts().to_dict()
    gap_examples = signals_df[signals_df['recommended_action'] == 'CTG filling local gap'].head(8)
    harvest_examples = signals_df[signals_df['recommended_action'] == 'Harvest local upper-funnel demand'].head(8)
    duplicate_examples = signals_df[signals_df['recommended_action'] == 'Check duplication before upweighting'].head(8)

    def example_lines(df):
        if df.empty:
            return '- None identified in current cut.'
        return '\n'.join(
            f"- {r.market} / {r.model} / week {r.week_start}: CTG `{r.ctg_channels or 'none'}`, local `{r.local_channels or 'none'}`"
            for r in df.itertuples()
        )

    action_lines = ['| Signal | Rows |', '|---|---:|']
    for signal, count in top_actions.items():
        action_lines.append(f'| {signal} | {count} |')
    action_table = '\n'.join(action_lines)

    return f"""# CTG Market Activity Integration - Day 1 Output

## What Was Delivered Today

- Converted Nico's market activity workbook into three machine-readable interim files:
  - `market_activities/local_activity_normalized.csv`
  - `market_activities/ctg_timeline_normalized.csv`
  - `market_activities/market_activity_weekly_signals.csv`
- Added a first `CTG Market Activity` page to the Intelligence Console.
- Built the first weekly decision signals for gap-filling, harvesting, duplication checks, and potential CTG whitespace.

## Initial Findings

- Local survey coverage currently contains **{len(local_df)} activity rows** across **{market_count} markets**.
- **{local_complete} rows** have enough date/channel detail to use directly in weekly optimisation.
- **{local_follow_up} rows** need follow-up before they should influence decisions.
- CTG planning contributes **{len(ctg_df)} market/model/channel timeline rows** from the `Media Mix` / `Zeitstrahl` logic.
- The weekly signal layer currently contains **{len(signals_df)} market-model-week rows**.

## Decision Signal Counts

{action_table}

## Interesting Examples To Discuss

### CTG Filling Local Gaps

{example_lines(gap_examples)}

### Harvesting Local Upper-Funnel Demand

{example_lines(harvest_examples)}

### Duplication Checks

{example_lines(duplicate_examples)}

## Data Quality Notes

- `PCGB` is not yet usable as a local activity input: the survey row has no start date, end date, channel, budget, or KPI.
- Market channels need normalization before final scoring. Examples include `SEM`, `SEA`, `YouTube & CTV`, `Other Online Channel`, and marketplace-specific entries.
- Budgets are not yet comparable because they mix currencies and free text (`200K`, `CHF 380'000.-`, etc.). Day 1 uses presence/timing/channel signals, not budget weighting.
- `Zeitstrahl nach Markt` is the right planning view for stakeholders; the app should use a structured weekly version of it underneath.

## Draft Email To Debs

Subject: CTG market activity layer - Day 1 progress

Hi Debs,

I have started turning Nico's market activity timeline into a structured optimisation layer for the Intelligence Console.

The first version now converts the workbook into weekly market/model signals showing where CTG is filling local gaps, where we can harvest demand from local upper-funnel activity, and where we need to check for duplication before upweighting. I have also added a separate CTG Market Activity page in the app so this stays clean and does not interfere with the existing optimisation views.

The main Day 1 finding is that the data is already good enough to drive directional weekly recommendations, especially from the timeline view, but some fields still need clean-up before we should weight budget decisions from it. PCGB is the clearest missing market data point at the moment.

Tomorrow I will tighten the channel taxonomy and connect these market activity signals more directly to the weekly optimisation logic.

Thanks,  
Ali

## Draft Team Message

Quick update on the CTG market activity work: I have converted Nico's timeline into a first weekly signal layer for the Intelligence Console. It now flags where CTG fills a local market gap, where CTG search/inventory can harvest demand from local upper-funnel activity, and where there is a possible duplication risk. First app page is in place under `CTG Market Activity`; next step is channel taxonomy clean-up and tying the signals into optimisation priority.

## Suggested Nico Follow-Up

Hi Nico,

I have started using the `Zeitstrahl nach Markt` view as the basis for a weekly optimisation overlay in the Intelligence Console.

The structure is useful. The main thing I need next is completion/validation for missing or partial market rows, especially PCGB, plus confirmation that the channel labels in the survey can be grouped into search, social, programmatic, video/CTV, and other local channels for optimisation purposes.

Thanks,  
Ali
"""


def build_day2_report(local_df, ctg_df, signals_df, taxonomy_df):
    signal_counts = signals_df['recommended_action'].value_counts().to_dict()
    bucket_counts = signals_df['planning_bucket'].value_counts().to_dict()
    confidence_counts = signals_df['confidence_level'].value_counts().to_dict()
    directions = signals_df['budget_direction'].value_counts().to_dict()

    def table_from_dict(data, first_col, second_col='Rows'):
        lines = [f'| {first_col} | {second_col} |', '|---|---:|']
        for key, value in data.items():
            lines.append(f'| {key} | {value} |')
        return '\n'.join(lines)

    def top_reasons(action, limit=6):
        rows = signals_df[signals_df['recommended_action'] == action].head(limit)
        if rows.empty:
            return '- None identified in current cut.'
        return '\n'.join(f"- {reason}" for reason in rows['recommendation_reason'])

    def follow_up_table(df):
        if df.empty:
            return 'No incomplete rows found.'
        lines = ['| Market | Model | Activity | Data Quality | Confidence |', '|---|---|---|---|---|']
        for row in df[['market', 'model', 'activity_name', 'data_quality', 'confidence_level']].itertuples(index=False):
            activity = str(row.activity_name).replace('|', '/')
            lines.append(f'| {row.market} | {row.model} | {activity} | {row.data_quality} | {row.confidence_level} |')
        return '\n'.join(lines)

    follow_up = local_df[local_df['data_quality'] != 'complete']
    channel_summary = taxonomy_df.groupby(['normalized_channel', 'role']).size().reset_index(name='raw_labels')
    channel_lines = '\n'.join(
        f"- {r.normalized_channel}: {r.role}, {int(r.raw_labels)} raw label(s)"
        for r in channel_summary.itertuples()
    )

    return f"""# CTG Market Activity Integration - Day 2 Output

## What Changed Today

- Tightened the channel taxonomy so local market labels and CTG labels map into a common planning language.
- Added confidence handling so incomplete market rows are not treated as reliable optimisation evidence.
- Added `budget_direction`, `planning_bucket`, and `recommendation_reason` to the weekly signal table.
- Updated the app layer to support more explainable recommendations rather than only counts.

## New Files / Updated Outputs

- `market_activities/channel_taxonomy_day2.csv`
- `market_activities/local_activity_normalized.csv`
- `market_activities/ctg_timeline_normalized.csv`
- `market_activities/market_activity_weekly_signals.csv`
- `market_activities/market_activity_day2_report.md`

## Day 2 Data Summary

- Local activity rows: **{len(local_df)}**
- CTG timeline rows: **{len(ctg_df)}**
- Weekly signal rows: **{len(signals_df)}**
- Channel taxonomy mappings: **{len(taxonomy_df)}**
- Local rows needing follow-up: **{len(follow_up)}**

## Planning Buckets

{table_from_dict(bucket_counts, 'Planning Bucket')}

## Budget Directions

{table_from_dict(directions, 'Budget Direction')}

## Signal Counts

{table_from_dict(signal_counts, 'Signal')}

## Confidence Counts

{table_from_dict(confidence_counts, 'Confidence')}

## Channel Taxonomy Summary

{channel_lines}

## Strong Examples For Discussion

### Scale / Harvest

{top_reasons('Harvest local upper-funnel demand')}

### Protect Or Upweight CTG Gap-Fill

{top_reasons('CTG filling local gap')}

### Validate Before Optimisation

{top_reasons('Validate data before optimisation')}

## Follow-Up Needed

The main governance improvement today is that incomplete local market data now gets flagged rather than silently treated as absence of local activity.

Rows still needing validation:

{follow_up_table(follow_up)}

## Draft Email To Debs

Subject: CTG market activity layer - Day 2 progress

Hi Debs,

Today I tightened the market activity layer so the Intelligence Console is moving from a timeline view into a more explainable optimisation input.

The main improvement is that market and CTG activity are now being translated into common planning language: whether the activity is upper-funnel, lower-funnel, full-funnel, or incomplete. The weekly output now includes a recommended budget direction, a planning bucket, and a plain-English reason for each market/model/week. This means we can explain not just that a market is flagged, but why it should be protected, upweighted, reviewed, or validated before budget moves.

I have also added confidence handling. Incomplete market rows are now flagged as validation items rather than being treated as true gaps. This is important for PCGB-style cases where missing local activity data could otherwise create a misleading CTG opportunity.

Next, I would connect these planning buckets to recent performance so that the final recommendation combines both sides: where the market is active and where the media is actually responding.

Thanks,  
Ali

## Draft Team Message

Day 2 update on CTG market activity: the timeline is now mapped into common planning language and weekly rows have explainable budget directions. The app can now show whether a signal is a scale, protect, fix, reduce, or watch case, plus the plain reason behind it. I also added confidence handling so incomplete market data is not mistaken for a genuine gap.

## Suggested Nico Follow-Up

Hi Nico,

I have now mapped the survey and CTG timeline channels into a common taxonomy so we can use the timeline as a weekly optimisation input.

Could you help validate the remaining incomplete rows, especially where date/channel/budget is missing? The most important point is to avoid treating missing market data as a confirmed absence of market activity.

Thanks,  
Ali
"""


def main():
    local_df = read_local_activity()
    ctg_df = read_ctg_timeline()
    signals_df = build_weekly_signals(local_df, ctg_df)
    taxonomy_df = build_channel_taxonomy()

    local_df.to_csv(LOCAL_OUT, index=False)
    ctg_df.to_csv(CTG_OUT, index=False)
    signals_df.to_csv(SIGNALS_OUT, index=False)
    taxonomy_df.to_csv(CHANNEL_TAXONOMY_OUT, index=False)
    REPORT_OUT.write_text(build_report(local_df, ctg_df, signals_df), encoding='utf-8')
    DAY2_REPORT_OUT.write_text(build_day2_report(local_df, ctg_df, signals_df, taxonomy_df), encoding='utf-8')

    print(f'Wrote {LOCAL_OUT} ({len(local_df)} rows)')
    print(f'Wrote {CTG_OUT} ({len(ctg_df)} rows)')
    print(f'Wrote {SIGNALS_OUT} ({len(signals_df)} rows)')
    print(f'Wrote {CHANNEL_TAXONOMY_OUT} ({len(taxonomy_df)} rows)')
    print(f'Wrote {REPORT_OUT}')
    print(f'Wrote {DAY2_REPORT_OUT}')


if __name__ == '__main__':
    main()

from io import BytesIO
from pathlib import Path
import os
import json
import re
import shutil
import sqlite3
import subprocess
import sys
from datetime import datetime
from typing import Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import pythoncom
    import win32com.client
except Exception:
    pythoncom = None
    win32com = None

from opportunity import OPPORTUNITY_CONFIG, compute_headroom_scores
from taxonomy_analysis import build_taxonomy_analysis, load_taxonomy_data

try:
    import numpy as np
    from scipy.optimize import curve_fit
except Exception:
    np = None
    curve_fit = None

base_dir_env = os.getenv('APP_BASE_DIR')
BASE_DIR = Path(base_dir_env) if base_dir_env else Path(__file__).resolve().parent
CSV_PATH = BASE_DIR / 'pwc reports' / 'outputs' / 'python_output_all.csv'
LOGO_PATH = BASE_DIR / 'porsche_logo.png'
UTM_NOTES_PDF_PATH = BASE_DIR / 'UTM_data' / 'Porsche_UTM Adoption Notes_Feb2026.pdf'
UTM_NOTES_CSV_PATH = BASE_DIR / 'UTM_data' / 'Porsche_UTM Adoption Notes_Feb2026.csv'
ADDRESSBOOK_CSV_PATH = BASE_DIR / 'UTM_data' / 'PHD_Local_Market_Addressbook.csv'
UTM_OUTREACH_DB_PATH = BASE_DIR / 'UTM_data' / 'outreach_tracking.db'
TAXONOMY_HYGIENE_PATH = BASE_DIR / 'taxonomy_hygine' / 'Copy of VWG Taxonomy Output v1.3 8 Apr 2026 11 30 07.xlsx'
MANUAL_UTM_TEST_RECORDS = [
    {
        'market_code': 'test_market',
        'market_name': 'test_market',
        'status': 'Test market for Outlook workflow',
        'context': 'Manual test market added for local Outlook testing.',
        'next_steps': 'Send test outreach email to validate pywin32 integration.',
        'section': 'PlanIt Champion',
        'name': 'Ali Test',
        'email': 'ph.malek@gmail.com',
        'notes': 'Manual test contact added for Outlook workflow testing.',
    },
]
OPTIONAL_UTM_CC_CONTACTS = [
    {'label': 'Porsche Client', 'email': 'laura.esguerra1@porsche.de'},
    {'label': 'PHD internal: Adriana Conte', 'email': 'adriana.conte@omc.com'},
    {'label': 'PHD internal: Katharina Kleinke', 'email': 'katharina.kleinke@omc.com'},
    {'label': 'DTO: Chloe Wright', 'email': 'chloe.wright@omc.com'},
]
CONTACT_TAG_OPTIONS = ['account_lead', 'planit_champion']
UTM_ENGAGEMENT_STAGES = [
    'To Contact',
    'Sent',
    'Waiting for Reply',
    'In Progress',
    'Blocked',
    'Resolved',
    'Closed',
]
UTM_ENGAGEMENT_PRIORITIES = ['High', 'Medium', 'Low']
UTM_THREAD_SUBJECT_KEYWORD = 'UTM adoption follow-up'
# Emails to exclude when determining whether a market has replied (internal team / fixed CC contacts)
UTM_THREAD_EXCLUDED_REPLY_EMAILS = {c['email'].lower() for c in OPTIONAL_UTM_CC_CONTACTS}

st.set_page_config(page_title='Intelligence Console', layout='wide')

load_dotenv()


@st.cache_data(show_spinner=False)
def load_taxonomy_hygiene_workbook(path_str: str) -> pd.DataFrame:
    return load_taxonomy_data(Path(path_str))


@st.cache_data(show_spinner=False)
def run_taxonomy_hygiene_analysis(path_str: str) -> dict:
    df_tax = load_taxonomy_hygiene_workbook(path_str)
    return build_taxonomy_analysis(df_tax)


def format_meur(value: float) -> str:
    return f"EUR {value / 1_000_000:.1f}m"


def build_budget_sankey_figure(
    total_budget: float,
    base_highlight_share: float,
    offline_upweight: float,
    nonwebsite_upweight: float,
    click_credit_shift: float,
    highlight_upper_share: float,
    always_on_upper_share: float,
    upper_to_awareness_share: float,
    upper_to_consideration_share: float,
    lower_to_conversion_share: float,
):
    total_budget = float(total_budget)
    base_highlight_share = float(base_highlight_share)
    base_always_on_share = max(0.0, 1.0 - base_highlight_share)

    highlight_weight = base_highlight_share * (1.0 + offline_upweight + nonwebsite_upweight + click_credit_shift)
    always_on_weight = base_always_on_share
    recommended_highlight_share = highlight_weight / (highlight_weight + always_on_weight) if (highlight_weight + always_on_weight) > 0 else 0.0
    recommended_always_on_share = 1.0 - recommended_highlight_share

    highlight_budget = total_budget * recommended_highlight_share
    always_on_budget = total_budget * recommended_always_on_share

    upper_highlight_budget = highlight_budget * highlight_upper_share
    lower_highlight_budget = highlight_budget - upper_highlight_budget
    upper_always_on_budget = always_on_budget * always_on_upper_share
    lower_always_on_budget = always_on_budget - upper_always_on_budget

    upper_total_budget = upper_highlight_budget + upper_always_on_budget
    lower_total_budget = lower_highlight_budget + lower_always_on_budget

    awareness_budget = upper_total_budget * upper_to_awareness_share
    consideration_from_upper = upper_total_budget * upper_to_consideration_share
    consideration_from_lower = lower_total_budget * (1.0 - lower_to_conversion_share)
    conversion_budget = lower_total_budget * lower_to_conversion_share
    consideration_budget = consideration_from_upper + consideration_from_lower

    fig = go.Figure(
        go.Sankey(
            arrangement='fixed',
            node=dict(
                pad=28,
                thickness=22,
                line=dict(color='rgba(20,20,20,0.35)', width=1.0),
                label=[
                    f"Total Budget<br>{format_meur(total_budget)}",
                    f"Always On / Core<br>{format_meur(always_on_budget)}",
                    f"Highlight Activations<br>{format_meur(highlight_budget)}",
                    f"AO Upper / Demand Support<br>{format_meur(upper_always_on_budget)}",
                    f"AO Lower / Harvesting<br>{format_meur(lower_always_on_budget)}",
                    f"Highlight Upper / Launch Demand<br>{format_meur(upper_highlight_budget)}",
                    f"Highlight Lower / Launch Harvesting<br>{format_meur(lower_highlight_budget)}",
                    f"Awareness<br>{format_meur(awareness_budget)}",
                    f"Consideration<br>{format_meur(consideration_budget)}",
                    f"Conversion / Harvesting<br>{format_meur(conversion_budget)}",
                ],
                color=[
                    '#111111',
                    '#22313F',
                    '#F40437',
                    '#8E7C6E',
                    '#375D81',
                    '#D5A021',
                    '#5392C5',
                    '#D96C6C',
                    '#E2B84B',
                    '#5C8B5A',
                ],
                x=[0.02, 0.23, 0.23, 0.50, 0.50, 0.50, 0.50, 0.82, 0.82, 0.82],
                y=[0.46, 0.20, 0.70, 0.10, 0.34, 0.58, 0.82, 0.12, 0.45, 0.80],
                hovertemplate='%{label}<extra></extra>',
            ),
            link=dict(
                source=[
                    0, 0,      # total -> AO / Highlight
                    1, 1,      # AO -> upper/lower
                    2, 2,      # Highlight -> upper/lower
                    3, 3,      # AO upper -> awareness / consideration
                    4, 4,      # AO lower -> consideration / conversion
                    5, 5,      # Highlight upper -> awareness / consideration
                    6, 6,      # Highlight lower -> consideration / conversion
                ],
                target=[
                    1, 2,
                    3, 4,
                    5, 6,
                    7, 8,
                    8, 9,
                    7, 8,
                    8, 9,
                ],
                value=[
                    always_on_budget,
                    highlight_budget,
                    upper_always_on_budget,
                    lower_always_on_budget,
                    upper_highlight_budget,
                    lower_highlight_budget,
                    awareness_budget * (upper_always_on_budget / upper_total_budget) if upper_total_budget > 0 else 0.0,
                    consideration_from_upper * (upper_always_on_budget / upper_total_budget) if upper_total_budget > 0 else 0.0,
                    consideration_from_lower * (lower_always_on_budget / lower_total_budget) if lower_total_budget > 0 else 0.0,
                    conversion_budget * (lower_always_on_budget / lower_total_budget) if lower_total_budget > 0 else 0.0,
                    awareness_budget * (upper_highlight_budget / upper_total_budget) if upper_total_budget > 0 else 0.0,
                    consideration_from_upper * (upper_highlight_budget / upper_total_budget) if upper_total_budget > 0 else 0.0,
                    consideration_from_lower * (lower_highlight_budget / lower_total_budget) if lower_total_budget > 0 else 0.0,
                    conversion_budget * (lower_highlight_budget / lower_total_budget) if lower_total_budget > 0 else 0.0,
                ],
                color=[
                    'rgba(34,49,63,0.55)',
                    'rgba(244,4,55,0.46)',
                    'rgba(142,124,110,0.55)',
                    'rgba(55,93,129,0.52)',
                    'rgba(213,160,33,0.56)',
                    'rgba(83,146,197,0.54)',
                    'rgba(142,124,110,0.42)',
                    'rgba(142,124,110,0.30)',
                    'rgba(55,93,129,0.34)',
                    'rgba(55,93,129,0.26)',
                    'rgba(213,160,33,0.42)',
                    'rgba(213,160,33,0.32)',
                    'rgba(83,146,197,0.36)',
                    'rgba(83,146,197,0.28)',
                ],
                hovertemplate='Flow: %{value:$,.0f}<extra></extra>',
            ),
        )
    )
    fig.update_layout(
        font=dict(size=15, color='#111111', family='Arial Black'),
        paper_bgcolor='#F6F1EA',
        plot_bgcolor='#F6F1EA',
        margin=dict(l=20, r=20, t=20, b=20),
        height=680,
    )

    metrics = {
        'base_highlight_share': base_highlight_share,
        'recommended_highlight_share': recommended_highlight_share,
        'recommended_always_on_share': recommended_always_on_share,
        'highlight_budget': highlight_budget,
        'always_on_budget': always_on_budget,
        'upper_always_on_budget': upper_always_on_budget,
        'lower_always_on_budget': lower_always_on_budget,
        'upper_highlight_budget': upper_highlight_budget,
        'lower_highlight_budget': lower_highlight_budget,
        'awareness_budget': awareness_budget,
        'consideration_budget': consideration_budget,
        'conversion_budget': conversion_budget,
    }
    return fig, metrics

HEADROOM_SUMMARY_TEMPLATE = (
    "Headroom by [GROUP_BY] quantifies efficiency upside versus a benchmark CPL. "
    "Across all [GROUP_BY], headroom ranges from [BOTTOM_HEADROOM]% to [TOP_HEADROOM]%. "
    "[TOP_GROUP] leads with [TOP_HEADROOM]% headroom, indicating the strongest efficiency upside, "
    "while [BOTTOM_GROUP] is lowest at [BOTTOM_HEADROOM]%. "
    "Thresholds are set at [MED_THRESHOLD]% (MED) and [HIGH_THRESHOLD]% (HIGH). "
    "Summary: [TAKEAWAY]."
)

ALLOCATION_METHOD_TEMPLATE = (
    "Budget Allocation Methodology (Technical)\n"
    "Scope & Filters\n"
    "- Grouping level: [GROUP_BY]\n"
    "- Markets included: [MARKETS]\n"
    "- Channels included: [CHANNELS]\n"
    "- Models/Carline included: [MODELS]\n"
    "- Campaigns included: [CAMPAIGNS]\n"
    "- Curve time axis: [CURVE_GROUP]\n\n"
    "Definitions\n"
    "1) Spend response curve: we fit a saturation curve per group using DCFS vs Media Spend: "
    "f(x) = A*x/(B+x). This models diminishing returns and enables spend‑optimal allocation.\n"
    "2) Headroom (efficiency upside): for each group, headroom = (current CPL − benchmark CPL) / benchmark CPL. "
    "Positive headroom means CPL is worse than benchmark, indicating room to improve efficiency. "
    "Headroom is computed from recent periods and benchmarked against the 25th percentile CPL at the most granular "
    "available level (Market+Channel+Model, then Market+Channel, then Channel).\n"
    "3) Scale: we use recent DCFS volume as a scale proxy. Scale score is the percentile rank of recent DCFS "
    "within Channel (0–100), then normalized to 0–1 for weighting.\n"
    "4) Reverse‑funnel split: a user‑provided target % split by group (s_i). Inputs are normalized to sum to "
    "100% if needed; this produces reverse‑funnel shares s_i (Σ s_i = 1).\n\n"
    "Methodology\n"
    "Step A — Spend‑optimal allocation: compute x_i^spend that maximizes Σ f_i(x_i) subject to Σ x_i = Budget "
    "(closed‑form via equalized marginal returns).\n"
    "Step B — Driver allocation: compute weights per group using headroom/scale strengths:\n"
    "w_i = max(0, HeadroomStrength*headroom_i + ScaleStrength*scale_i). "
    "Normalize to shares p_i = w_i / Σ w.\n"
    "Step C — Blend spend vs drivers using ConstraintStrength:\n"
    "q_i = x_i^spend / Budget, r_i = (1−ConstraintStrength)*q_i + ConstraintStrength*p_i, "
    "x_i = Budget * r_i.\n"
    "Step D — Min/Max spend: if enabled, allocate minimums first, then distribute remaining budget by r_i. "
    "Finally, apply per‑group max caps and redistribute any overflow across uncapped groups.\n"
    "Step E — Reverse‑funnel blend: blend the risk‑aware allocation with the reverse‑funnel split using "
    "ReverseFunnelBlend (λ):\n"
    "x_i^blend = (1−λ) * x_i^risk + λ * (Budget * s_i).\n"
    "This blend is computed separately for unconstrained and constrained risk‑aware allocations. "
    "DCFS for blended scenarios is evaluated on the same response curve f_i(x).\n\n"
    "Parameters Used\n"
    "- HeadroomStrength = [HEADROOM_STRENGTH]\n"
    "- ScaleStrength = [SCALE_STRENGTH]\n"
    "- ConstraintStrength = [CONSTRAINT_STRENGTH] (0 = pure spend optimization, 1 = pure driver allocation)\n"
    "- ReverseFunnelBlend (λ) = [REVERSE_BLEND]\n"
    "- Minimum spend enabled = [MIN_CONSTRAINT_ENABLED]\n"
    "- Minimum spend (by group) = [MIN_BY_GROUP]\n"
    "- Maximum spend caps enabled = [MAX_CONSTRAINT_ENABLED]\n"
    "- Maximum spend (by group) = [MAX_BY_GROUP]\n"
    "- Reverse‑funnel split (by group) = [REVERSE_SPLIT_BY_GROUP]\n"
    "- Total budget = [BUDGET]\n"
    "- Number of fitted curves = [CURVE_COUNT]\n\n"
    "Results Summary\n"
    "- Total DCFS (unconstrained) = [TOTAL_DCFS_UNCONSTRAINED]\n"
    "- Total DCFS (constrained) = [TOTAL_DCFS_CONSTRAINED]\n\n"
    "- Total DCFS (blended, unconstrained) = [TOTAL_DCFS_BLEND_UNCONSTRAINED]\n"
    "- Total DCFS (blended, constrained) = [TOTAL_DCFS_BLEND_CONSTRAINED]\n\n"
    "Strategic Guidelines\n"
    "- ReverseFunnelBlend = 0 prioritizes risk‑aware optimization.\n"
    "- ReverseFunnelBlend = 1 enforces the reverse‑funnel target split.\n"
    "- Intermediate values linearly trade off optimization vs. strategic split at the group level.\n"
    "- Blending does not re‑apply min/max caps; if caps must be enforced post‑blend, apply an additional "
    "constraint pass.\n\n"
    "Group‑level Allocation Detail (per [GROUP_BY])\n"
    "[ALLOCATION_TABLE]\n"
)

INCENTIVE_METHOD_TEMPLATE = (
    "Performance‑Linked Remuneration Model (Technical)\n"
    "Scope\n"
    "- KPI: Dealer Contact Form Submissions (DCFS)\n"
    "- Measurement period: user‑defined (single target value)\n\n"
    "Definitions\n"
    "1) Target volume: T is an explicit expected DCFS total for the measurement period.\n"
    "2) Variable component: V = (alpha x C_BAH) + (beta x C_FTE).\n"
    "3) Deviation: delta = (A - T) / T.\n"
    "4) Adjustment rule:\n"
    "   - Penalty zone: delta < 0 => Adjustment = max(delta, -F_dir).\n"
    "   - No‑change zone: 0 <= delta <= H => Adjustment = 0.\n"
    "   - Reward zone: delta > H => Adjustment = min(delta - H, F_up).\n"
    "5) Adjusted variable fee: V x (1 + Adjustment).\n"
    "6) Total fee: Fixed Fee + Adjusted Variable Fee.\n\n"
    "Parameters Used\n"
    "- Target DCFS (T) = [T]\n"
    "- Actual DCFS (A) = [A]\n"
    "- alpha (BAH at‑risk share) = [ALPHA]\n"
    "- C_BAH = [C_BAH]\n"
    "- beta (FTE at‑risk share) = [BETA]\n"
    "- C_FTE = [C_FTE]\n"
    "- Variable component (V) = [V]\n"
    "- Hurdle (H) = [H]\n"
    "- Downside floor (F_dir) = [F_DIR]\n"
    "- Upside ceiling (F_up) = [F_UP]\n"
    "- Fixed fee = [FIXED]\n\n"
    "Results\n"
    "- Deviation (delta) = [DELTA]\n"
    "- Adjustment = [ADJUSTMENT] (zone: [ZONE])\n"
    "- Adjusted variable fee = [ADJ_VAR]\n"
    "- Total fee = [TOTAL_FEE]\n\n"
    "Guidance Notes\n"
    "- This model is deterministic and linear within zones; only the variable component is adjusted.\n"
    "- H sets the no‑change band on the upside; penalties apply immediately for any miss below target.\n"
    "- F_dir and F_up cap downside/upside adjustments on V; fixed fee remains intact.\n"
    "- If using DCFS totals from allocation outputs, ensure the period definition matches D and W.\n"
)

INCENTIVE_MARKET_NARRATIVE_TEMPLATE = (
    "Incentive Model Narrative (Technical, Reproducible)\n"
    "Scope\n"
    "- KPI: [KPI]\n"
    "- Time column: [TIME_COL]\n"
    "- Markets included for calculation: [CALC_MARKETS]\n"
    "- Markets shown in visualization: [VIZ_MARKETS]\n"
    "- KPI definition:\n"
    "  - CPL (Forms) = Media Spend / Forms Submission Started\n"
    "  - CPL (DCFS) = Media Spend / DCFS\n"
    "  - CPM = (Media Spend / Impressions) * 1000\n\n"
    "Time Series Construction\n"
    "For each market m and time period t:\n"
    "1) Aggregate spend and denominator:\n"
    "   spend_{m,t} = sum(Media Spend)\n"
    "   denom_{m,t} = sum(denominator)\n"
    "2) KPI time series point:\n"
    "   KPI_{m,t} = spend_{m,t} / denom_{m,t} (x1000 if CPM)\n"
    "3) Visualization uses KPI_{m,t} over time.\n"
    "Full‑period reference band per market:\n"
    "- Average over full period: avg_m = mean(KPI_{m,t}) across t\n"
    "- Volatility band: avg_m ± std(KPI_{m,t}) across t\n\n"
    "Volatility Scoring (Cross‑Market)\n"
    "1) Compute market CV from time series KPI:\n"
    "   CV_m = std(KPI_{m,t}) / mean(KPI_{m,t})\n"
    "2) Compute median CV across markets:\n"
    "   CV_med = median(CV_m)\n"
    "3) Relative volatility score:\n"
    "   vol_ratio_m = CV_m / CV_med\n"
    "4) Confidence mapping (clamped):\n"
    "   confidence_m = clamp(1 - k * (vol_ratio_m - 1), min_conf, max_conf)\n"
    "   k = [VOL_K]\n"
    "   min_conf = [MIN_CONF]\n"
    "   max_conf = [MAX_CONF]\n"
    "5) Effective at‑risk shares:\n"
    "   alpha_eff_m = alpha * confidence_m\n"
    "   beta_eff_m  = beta  * confidence_m\n"
    "   alpha = [ALPHA]\n"
    "   beta  = [BETA]\n\n"
    "Targets and Actuals\n"
    "Target per market is derived from the selected quantile of KPI_{m,t} across t:\n"
    "- Quantile: [TARGET_Q_LABEL] (q = [TARGET_Q])\n"
    "- Target adjustment multiplier: [TARGET_ADJ] (percent)\n"
    "- Target_T_m = quantile_q(KPI_{m,t}) * (TARGET_ADJ/100)\n"
    "Actual per market is user‑provided in the input table:\n"
    "- Actual_A_m = user input (defaults to avg KPI per market)\n\n"
    "Fee Construction (Per Market)\n"
    "Inputs:\n"
    "- C_BAH = [C_BAH]\n"
    "- C_FTE = [C_FTE]\n"
    "- H = [H]\n"
    "- F_dir = [F_DIR]\n"
    "- F_up = [F_UP]\n"
    "Variable and fixed fees per market:\n"
    "- Variable_V_m = alpha_eff_m * C_BAH + beta_eff_m * C_FTE\n"
    "- Fixed_Fee_m  = (1 - alpha_eff_m) * C_BAH + (1 - beta_eff_m) * C_FTE\n\n"
    "Performance Deviation (Cost KPI, lower is better)\n"
    "- delta_m = (Target_T_m - Actual_A_m) / Target_T_m\n\n"
    "Adjustment Rule\n"
    "- If delta_m < 0: Adjustment_m = max(delta_m, -F_dir)\n"
    "- If 0 <= delta_m <= H: Adjustment_m = 0\n"
    "- If delta_m > H: Adjustment_m = min(delta_m - H, F_up)\n\n"
    "Adjusted Fees\n"
    "- Adjusted_Variable_Fee_m = Variable_V_m * (1 + Adjustment_m)\n"
    "- Total_Fee_m = Fixed_Fee_m + Adjusted_Variable_Fee_m\n\n"
    "Market‑Level Outputs\n"
    "[MARKET_TABLE]\n"
)

CTG_PRE_POST_TEMPLATE = (
    "CTG Pre/Post Analysis Narrative (Technical, Reproducible)\n"
    "Scope\n"
    "- Data source: other_data/weekly_market_kpi_table.csv\n"
    "- KPI selection: [KPI]\n"
    "- Channel selection: [CHANNEL]\n"
    "- Pre‑CTG window: weeks [X_START]–23\n"
    "- Post‑CTG window: weeks 24–52\n"
    "- Aggregation method: [AGG_METHOD]\n\n"
    "Input Data Schema\n"
    "Columns used: market, channel, week, sessions, spend, icc_dcfs\n\n"
    "Filtering\n"
    "- If Channel = Paid Search/Paid Social, data is filtered to that channel.\n"
    "- If Channel = Both, all channels are included.\n\n"
    "Week Parsing\n"
    "- week label format: YYYY‑WW\n"
    "- week_num = integer WW extracted from week label\n"
    "- Pre‑CTG = week_num in [X_START, 23]\n"
    "- Post‑CTG = week_num in [24, 52]\n\n"
    "Metric Construction\n"
    "1) For each market m and window (pre/post), compute:\n"
    "   - spend_sum_m = sum(spend)\n"
    "   - sessions_sum_m = sum(sessions)\n"
    "   - leads_sum_m = sum(icc_dcfs)\n"
    "   - weeks_m = count(distinct week_num)\n"
    "2) Derived weekly metrics:\n"
    "   - average weekly spend = spend_sum_m / weeks_m\n"
    "   - average weekly leads = leads_sum_m / weeks_m\n"
    "   - average weekly sessions = sessions_sum_m / weeks_m\n"
    "   - average cpl in a week = spend_sum_m / leads_sum_m\n"
    "   - average cost per session in a week = spend_sum_m / sessions_sum_m\n"
    "   - leads per €1k = (leads_sum_m / spend_sum_m) * 1000\n"
    "   - sessions per €1k = (sessions_sum_m / spend_sum_m) * 1000\n\n"
    "Aggregation Method\n"
    "- Weighted (ratio of sums): compute totals across markets, then apply formulas once.\n"
    "- Unweighted (mean of markets): compute metrics per market, then average across markets.\n\n"
    "Pre/Post Summary Table\n"
    "[SUMMARY_TABLE]\n\n"
    "Market Breakdown (Sessions + CPL)\n"
    "[MARKET_BREAKDOWN]\n\n"
    "Delta Definition\n"
    "- delta = post - pre\n"
    "- delta % = (post - pre) / pre\n\n"
    "Visualization\n"
    "- Bar chart compares Pre vs Post for cost per session by market.\n"
    "- If KPI = sessions, cost per session = spend / sessions.\n"
    "- If KPI = icc_dcfs, cost per session = spend / icc_dcfs.\n\n"
    "Reproducibility Notes\n"
    "- Ensure identical filters and week windows.\n"
    "- Verify that week labels follow YYYY‑WW format.\n"
    "- For missing values, rows with null denominators are excluded from ratios.\n"
)

CTG_PRE_POST_TEMPLATE_COPY = (
    "CTG Pre/Post Analysis Narrative (ctg_pre + Overview Post)\n"
    "Scope\n"
    "- Pre data source: other_data/ctg_pre_02/matched_campaign_date_stats_pre_ctg.csv\n"
    "- Post data source: other_data/ctg_pre_02/matched_campaign_date_stats_post_ctg.csv\n"
    "- Market selection: [MARKETS]\n"
    "- Model selection: [MODEL]\n"
    "- Channel selection: [CHANNEL]\n"
    "- Pre‑CTG window: weeks [X_START]–23\n"
    "- Pre data cutoff: week start date before 2025-05-01\n"
    "- Post window: all weeks available in post CSV\n\n"
    "Input Data Schemas\n"
    "Pre columns used: market, model, channel, week, spend, conv\n"
    "Post columns used: market, model, channel, week, spend, conv\n"
    "- Conversion metric used in both pre/post = dcfs = conv\n\n"
    "Filtering\n"
    "- Pre: filter by selected markets, selected models, selected channel.\n"
    "- Post: filter by selected markets, selected models, selected channel.\n"
    "- Market mapping applied before filtering: CANADA→PCL, UK→PCGB, GERMANY→PD, FRANCE→POF, ITALY→PIT.\n\n"
    "Week Parsing (Pre only)\n"
    "- week label format: YYYY‑WW\n"
    "- week_num = integer WW extracted from week label\n"
    "- Pre‑CTG = week_num in [X_START, 23]\n\n"
    "- Pre data cutoff applied by converting ISO week to week start date\n"
    "  and keeping rows where week_start < 2025-05-01\n\n"
    "Metric Construction\n"
    "1) Pre (per market m):\n"
    "   - spend_sum_m = sum(spend) over pre weeks\n"
    "   - dcfs_sum_m = sum(dcfs) over pre weeks\n"
    "   - cost_per_dcfs_m = spend_sum_m / dcfs_sum_m\n"
    "2) Post (per market m, post CSV totals):\n"
    "   - spend_sum_m = sum(spend)\n"
    "   - dcfs_sum_m = sum(dcfs)\n"
    "   - cost_per_dcfs_m = spend_sum_m / dcfs_sum_m\n"
    "3) Global Post:\n"
    "   - total_spend = sum(spend) across selected markets\n"
    "   - total_dcfs = sum(dcfs) across selected markets\n"
    "   - global_cost_per_dcfs = total_spend / total_dcfs\n\n"
    "Pre/Post Summary Table (GLOBAL)\n"
    "[SUMMARY_TABLE]\n\n"
    "Market Breakdown (Pre = ctg_pre, Post = overview totals)\n"
    "[MARKET_BREAKDOWN]\n\n"
    "Model Breakdown (Pre = ctg_pre, Post = overview totals)\n"
    "[MODEL_BREAKDOWN]\n\n"
    "Delta Definition\n"
    "- Delta = Post - Pre\n"
    "- Delta % = (Post - Pre) / Pre\n"
)

INCENTIVE_METHOD_TEMPLATE = (
    "Incentive Methodology Report (Market CPL Banding)\n"
    "\n"
    "Scope & Purpose\n"
    "- Objective: quantify incentive impacts using historical KPI volatility and fixed incentive rules.\n"
    "- KPI used: [KPI]\n"
    "- Cadence (weeks per point): [CADENCE]\n"
    "- Benchmark (top 97th percentile across box plots): [BENCHMARK]\n"
    "\n"
    "Data Pipeline (Step‑by‑Step)\n"
    "1) Base dataset is the overview CSV used on the KPI pages.\n"
    "2) Filters applied:\n"
    "   - Markets: [MARKETS]\n"
    "   - Channels: [CHANNELS]\n"
    "   - Weeks: [WEEKS]\n"
    "3) Weekly aggregation:\n"
    "   - Per Market/Channel/Week: sum Media Spend, DCFS, Sessions, Forms.\n"
    "4) Cadence aggregation:\n"
    "   - cadence_bin = floor((week_index-1)/[CADENCE])\n"
    "   - Each bin aggregates consecutive weeks.\n"
    "   - Any bin containing a zero‑DCFS week is excluded to match cadence=1 behavior.\n"
    "5) KPI computation:\n"
    "   - CPL (DCFS) = Media Spend / DCFS\n"
    "\n"
    "Benchmarking\n"
    "- Selected percentile case: [BENCHMARK_CASE]\n"
    "- Method: for each group shown in the box plots (market, or market|channel when split),\n"
    "  compute the [BENCHMARK_CASE] percentile of KPI points, then take the maximum across\n"
    "  those groups to define the benchmark.\n"
    "- Rationale: use the top‑end percentile across groups to be conservative and ensure\n"
    "  hurdle/bonus/malus bands are anchored to the highest observed tail risk.\n"
    "- Benchmark value: [BENCHMARK]\n"
    "\n"
    "Incentive Rules (Fixed)\n"
    "- At‑risk share: 50% of BAH + 20% of FTE\n"
    "- Hurdle: +10% outperformance required for upside\n"
    "- Cap: ±50% on the variable component\n"
    "- For cost KPIs (lower is better), delta = (Benchmark - Actual) / Benchmark\n"
    "  - If delta < 0: malus = max(delta, -0.50)\n"
    "  - If 0 <= delta <= 0.10: no change\n"
    "  - If delta > 0.10: bonus = min(delta - 0.10, 0.50)\n"
    "\n"
    "Fees & Components\n"
    "- BAH fee input: [BAH_FEE]\n"
    "- FTE fee input: [FTE_FEE]\n"
    "- Variable BAH = 0.5 * BAH\n"
    "- Variable FTE = 0.2 * FTE\n"
    "- Variable total = Variable BAH + Variable FTE\n"
    "- Fixed fee = (BAH + FTE) - Variable total\n"
    "\n"
    "Band Population (Scorecards)\n"
    "[BAND_COUNTS]\n"
    "\n"
    "Band‑Level Adjustments (Scorecards)\n"
    "[BAND_FEES]\n"
    "\n"
    "Overall Expected Adjustment\n"
    "- Avg adjustment (all points): [AVG_ADJ]\n"
    "- Avg BAH adjustment: [AVG_BAH_ADJ]\n"
    "- Avg FTE adjustment: [AVG_FTE_ADJ]\n"
    "- Avg final fee: [AVG_FINAL_FEE]\n"
    "\n"
    "Notes\n"
    "- All numbers are derived from the currently filtered points.\n"
    "- Changing cadence, filters, or benchmark recomputes all metrics.\n"
)

@st.cache_data
def load_data(csv_path: Path, mtime: float):
    df = pd.read_csv(csv_path, low_memory=False)
    return normalize_data(df)


def normalize_data(df: pd.DataFrame) -> pd.DataFrame:
    df['report_date'] = pd.to_datetime(df.get('report_date'), errors='coerce')
    if 'Date' in df.columns:
        date_series = df['Date']
        if pd.api.types.is_datetime64_any_dtype(date_series):
            df['Date'] = date_series
        else:
            date_str = date_series.astype(str).str.strip()
            date_str = date_str.str.replace(r'\.0$', '', regex=True)
            date_ymd = date_str.where(date_str.str.match(r'^\d{8}$', na=False))
            parsed = pd.to_datetime(date_ymd, format='%Y%m%d', errors='coerce')
            parsed_fallback = pd.to_datetime(date_str, errors='coerce')
            df['Date'] = parsed.fillna(parsed_fallback)
    df['report_week'] = df.get('report_week', pd.Series(dtype=str)).astype(str).str.strip()
    report_week_clean = df['report_week'].replace({'nan': '', 'None': ''}).fillna('')
    report_week_clean = report_week_clean.str.upper().str.replace(' ', '', regex=False)

    report_date = df['report_date']
    iso = report_date.dt.isocalendar()
    iso_year = iso['year']
    iso_week = iso['week']

    week_match = report_week_clean.str.extract(r'CW(\d{1,2})', expand=False)
    week_num = pd.to_numeric(week_match, errors='coerce')

    df['report_cw'] = week_num
    df['report_cw_year'] = pd.Series(pd.NA, index=df.index, dtype='Int64')
    date_has_week = report_date.notna()
    df.loc[date_has_week, 'report_cw'] = df.loc[date_has_week, 'report_cw'].fillna(iso_week)
    df.loc[date_has_week, 'report_cw_year'] = iso_year.where(date_has_week)

    df['report_week_key'] = pd.Series(pd.NA, index=df.index, dtype='string')
    has_cw = df['report_cw'].notna() & df['report_cw_year'].notna()
    df.loc[has_cw, 'report_week_key'] = (
        df.loc[has_cw, 'report_cw_year'].astype(int).astype(str)
        + '-CW'
        + df.loc[has_cw, 'report_cw'].astype(int).astype(str).str.zfill(2)
    )
    df['report_week_sort'] = pd.Series(pd.NA, index=df.index, dtype='Int64')
    df.loc[has_cw, 'report_week_sort'] = (
        df.loc[has_cw, 'report_cw_year'].astype(int) * 100
        + df.loc[has_cw, 'report_cw'].astype(int)
    )

    if 'Date' in df.columns:
        date_base = df['Date']
        valid_date = date_base.notna()
        # Normalize existing calendar_week values and build sort keys from them.
        if 'calendar_week' not in df.columns:
            df['calendar_week'] = pd.Series(pd.NA, index=df.index, dtype=object)
        else:
            df['calendar_week'] = df['calendar_week'].astype(str).str.strip()
            df['calendar_week'] = df['calendar_week'].replace({'nan': pd.NA, 'None': pd.NA, '': pd.NA})
        if 'calendar_week_sort' not in df.columns:
            df['calendar_week_sort'] = pd.Series(pd.NA, index=df.index, dtype='Int64')
        week_match = df['calendar_week'].str.extract(r'(?:(\d{4})\s*-\s*)?CW\s*(\d{1,2})', expand=True)
        week_year = pd.to_numeric(week_match[0], errors='coerce')
        week_num = pd.to_numeric(week_match[1], errors='coerce')
        if 'report_cw_year' in df.columns:
            week_year = week_year.fillna(df['report_cw_year'])
        has_week = week_num.notna() & week_year.notna()
        df.loc[has_week, 'calendar_week_sort'] = (
            week_year[has_week].astype(int) * 100
            + week_num[has_week].astype(int)
        )

        # Forward-fill missing weeks with the last non-null week (per row order).
        df['calendar_week'] = df['calendar_week'].ffill()
        df['calendar_week_sort'] = df['calendar_week_sort'].ffill()

        week_start = date_base - pd.to_timedelta(date_base.dt.weekday, unit='D')
        week_end = week_start + pd.Timedelta(days=6)

        def _format_range(start, end):
            return f'{start.strftime("%B")} {start.day} - {end.strftime("%B")} {end.day}'

        df['week_text'] = df.get('week_text', pd.Series(dtype=object))
        df.loc[valid_date, 'week_text'] = [
            _format_range(start, end)
            for start, end in zip(week_start[valid_date], week_end[valid_date])
        ]

        unique_starts = sorted(week_start[valid_date].dropna().unique())
        week_index = {start: idx + 1 for idx, start in enumerate(unique_starts)}
        df['week_relative'] = df.get('week_relative', pd.Series(dtype=object))
        df.loc[valid_date, 'week_relative'] = (
            'BW '
            + week_start[valid_date].map(week_index).astype(int).astype(str)
        )
    for col in ['Media Spend', 'Number of Sessions', 'DCFS', 'Forms Submission Started', 'Impressions']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def get_calendar_week_options(df_in: pd.DataFrame) -> list:
    if 'calendar_week' not in df_in.columns:
        return []
    if 'calendar_week_sort' in df_in.columns:
        tmp = df_in[['calendar_week', 'calendar_week_sort']].dropna()
        if not tmp.empty:
            return (
                tmp.sort_values('calendar_week_sort')
                .drop_duplicates('calendar_week')['calendar_week']
                .tolist()
            )
    return sorted(df_in['calendar_week'].dropna().unique())


_DATE_RE = re.compile(r'^(\d{8})_')
_CW_RE = re.compile(r'\bCW\s*\d+\b', re.IGNORECASE)


def _parse_date_prefix(filename: str) -> Optional[str]:
    match = _DATE_RE.match(filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), '%d%m%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None


def _parse_report_week(text: str) -> str:
    match = _CW_RE.search(text or '')
    return match.group(0).replace(' ', '') if match else ''


def _find_python_output_sheet(wb):
    for name in wb.sheetnames:
        if re.search(r'python output', name, re.IGNORECASE):
            return name
    return None


@st.cache_data
def load_excel_python_output(file_bytes: bytes, filename: str) -> pd.DataFrame:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = _find_python_output_sheet(wb)
    if not sheet_name:
        wb.close()
        raise ValueError('Python Output sheet not found.')
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None and cell != '' for cell in row):
            rows.append(row)
    wb.close()

    if not rows:
        raise ValueError('Python Output sheet is empty.')

    header = [str(col).strip() if col is not None else '' for col in rows[0]]
    report_date = _parse_date_prefix(filename)
    report_week = _parse_report_week(filename)
    source_file = filename

    out_rows = []
    for row in rows[1:]:
        values = [row[idx] if idx < len(row) else None for idx in range(len(header))]
        out_rows.append([report_date, report_week, source_file] + values)

    df = pd.DataFrame(out_rows, columns=['report_date', 'report_week', 'source_file'] + header)
    return normalize_data(df)


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Python Output (cleaned)')
    return output.getvalue()


_UTM_MARKET_RE = re.compile(r'^([A-Z]{3,5})(?: \((.+)\))?$')
_UTM_SECTION_RE = re.compile(
    r'^(Status|Issues identified|Next steps|Context|Scope gaps|Observations):\s*(.*)$'
)
_UTM_SECTION_MAP = {
    'Status': 'Status',
    'Issues identified': 'Issues Identified',
    'Next steps': 'Next Steps',
    'Context': 'Context',
    'Scope gaps': 'Scope Gaps',
    'Observations': 'Observations',
}


def _utm_is_market_header(line: str) -> bool:
    match = _UTM_MARKET_RE.match(line)
    if not match:
        return False
    return match.group(1) not in {'Status', 'Context'}


def _utm_new_record(code: str, name: Optional[str] = None) -> dict:
    return {
        'Market Code': code,
        'Market Name': name or '',
        'Status': [],
        'Scope Gaps': [],
        'Observations': [],
        'Issues Identified': [],
        'Context': [],
        'Next Steps': [],
    }


def _utm_append_to_section(record: dict, section: str, text: str) -> None:
    cleaned = ' '.join(text.split()).strip()
    if not cleaned:
        return
    bucket = record[section]
    if bucket:
        bucket[-1] = f"{bucket[-1]} {cleaned}".strip()
    else:
        bucket.append(cleaned)


def _utm_parse_notes_text(raw_text: str) -> pd.DataFrame:
    lines = []
    for raw_line in raw_text.replace('\x0c', '\n').splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line == 'Confidential - Not for Public Consumption or Distribution':
            continue
        lines.append(line)

    records = []
    current = None
    current_section = None
    idx = 0

    while idx < len(lines):
        line = lines[idx]

        market_match = _UTM_MARKET_RE.match(line)
        if market_match and _utm_is_market_header(line):
            if current is not None:
                records.append(current)
            current = _utm_new_record(market_match.group(1), market_match.group(2))
            current_section = None
            idx += 1
            continue

        if current is None:
            idx += 1
            continue

        if line == '•':
            idx += 1
            continue

        section_match = _UTM_SECTION_RE.match(line)
        if section_match:
            current_section = _UTM_SECTION_MAP[section_match.group(1)]
            remainder = section_match.group(2).strip()
            if remainder:
                current[current_section].append(remainder)
            idx += 1
            continue

        if line == 'o' and current_section:
            bullet_parts = []
            idx += 1
            while idx < len(lines):
                candidate = lines[idx]
                if candidate in {'•', 'o'}:
                    break
                if _UTM_SECTION_RE.match(candidate):
                    break
                if _utm_is_market_header(candidate):
                    break
                bullet_parts.append(candidate)
                idx += 1
            if bullet_parts:
                current[current_section].append(' '.join(bullet_parts))
            continue

        if current_section:
            _utm_append_to_section(current, current_section, line)

        idx += 1

    if current is not None:
        records.append(current)

    formatted_rows = []
    for record in records:
        formatted_rows.append({
            'Market Code': record['Market Code'],
            'Market Name': record['Market Name'],
            'Status': ' '.join(record['Status']).strip(),
            'Scope Gaps': '\n'.join(f"- {item}" for item in record['Scope Gaps']),
            'Observations': '\n'.join(f"- {item}" for item in record['Observations']),
            'Issues Identified': '\n'.join(f"- {item}" for item in record['Issues Identified']),
            'Context': '\n'.join(f"- {item}" for item in record['Context']),
            'Next Steps': '\n'.join(f"- {item}" for item in record['Next Steps']),
        })

    return pd.DataFrame(formatted_rows)


@st.cache_data
def load_utm_adoption_notes(pdf_path: str, mtime: float, fallback_csv_path: Optional[str] = None) -> pd.DataFrame:
    mutool_path = shutil.which('mutool')
    if mutool_path is None:
        if fallback_csv_path and Path(fallback_csv_path).exists():
            return pd.read_csv(fallback_csv_path).fillna('')
        raise FileNotFoundError(
            "MuPDF 'mutool' is not installed and no fallback CSV was found for UTM adoption notes."
        )

    result = subprocess.run(
        [mutool_path, 'draw', '-F', 'txt', '-o', '-', pdf_path],
        check=True,
        capture_output=True,
        text=True,
    )
    return _utm_parse_notes_text(result.stdout)


def load_or_create_utm_adoption_notes_csv(pdf_path: Path, csv_path: Path) -> pd.DataFrame:
    if csv_path.exists():
        return pd.read_csv(csv_path).fillna('')

    base_df = load_utm_adoption_notes(
        str(pdf_path),
        pdf_path.stat().st_mtime,
        fallback_csv_path=str(csv_path),
    ).fillna('')
    base_df.to_csv(csv_path, index=False)
    return base_df


def save_utm_adoption_notes_csv(df: pd.DataFrame, csv_path: Path) -> None:
    df.fillna('').to_csv(csv_path, index=False)


@st.cache_data
def load_market_addressbook_csv(csv_path: str, mtime: float) -> pd.DataFrame:
    return pd.read_csv(csv_path).fillna('')


def _get_utm_outreach_db_connection(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def initialize_utm_outreach_db(db_path: Path) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS utm_adoption_notes (
                market_code TEXT PRIMARY KEY,
                market_name TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '',
                scope_gaps TEXT NOT NULL DEFAULT '',
                observations TEXT NOT NULL DEFAULT '',
                issues_identified TEXT NOT NULL DEFAULT '',
                context TEXT NOT NULL DEFAULT '',
                next_steps TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS utm_adoption_notes_base (
                market_code TEXT PRIMARY KEY,
                market_name TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '',
                scope_gaps TEXT NOT NULL DEFAULT '',
                observations TEXT NOT NULL DEFAULT '',
                issues_identified TEXT NOT NULL DEFAULT '',
                context TEXT NOT NULL DEFAULT '',
                next_steps TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                market TEXT NOT NULL,
                section TEXT NOT NULL,
                name TEXT NOT NULL DEFAULT '',
                title TEXT NOT NULL DEFAULT '',
                email TEXT NOT NULL,
                contact_tags TEXT NOT NULL DEFAULT '',
                notes TEXT NOT NULL DEFAULT '',
                source_sheet TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_contacts_market_section_email
            ON contacts (market, section, email);

            CREATE TABLE IF NOT EXISTS email_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                workflow TEXT NOT NULL,
                subject_template TEXT NOT NULL,
                body_template TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_email_templates_name_workflow
            ON email_templates (name, workflow);

            CREATE TABLE IF NOT EXISTS outreach_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                workflow TEXT NOT NULL,
                name TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT NOT NULL DEFAULT '',
                filters_json TEXT NOT NULL DEFAULT '',
                template_id INTEGER,
                FOREIGN KEY (template_id) REFERENCES email_templates(id)
            );

            CREATE TABLE IF NOT EXISTS outreach_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id INTEGER NOT NULL,
                market TEXT NOT NULL,
                status_snapshot TEXT NOT NULL DEFAULT '',
                to_emails TEXT NOT NULL DEFAULT '',
                cc_emails TEXT NOT NULL DEFAULT '',
                subject_rendered TEXT NOT NULL DEFAULT '',
                body_rendered TEXT NOT NULL DEFAULT '',
                sent_at TEXT,
                outlook_entry_id TEXT NOT NULL DEFAULT '',
                outlook_conversation_id TEXT NOT NULL DEFAULT '',
                sync_state TEXT NOT NULL DEFAULT 'pending',
                replied INTEGER NOT NULL DEFAULT 0,
                last_reply_at TEXT,
                last_reply_from TEXT NOT NULL DEFAULT '',
                last_reply_snippet TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (run_id) REFERENCES outreach_runs(id)
            );

            CREATE INDEX IF NOT EXISTS idx_outreach_messages_run_id
            ON outreach_messages (run_id);

            CREATE INDEX IF NOT EXISTS idx_outreach_messages_market
            ON outreach_messages (market);

            CREATE TABLE IF NOT EXISTS outreach_replies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                outreach_message_id INTEGER NOT NULL,
                received_at TEXT,
                sender_email TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                snippet TEXT NOT NULL DEFAULT '',
                outlook_entry_id TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (outreach_message_id) REFERENCES outreach_messages(id)
            );

            CREATE INDEX IF NOT EXISTS idx_outreach_replies_message_id
            ON outreach_replies (outreach_message_id);

            CREATE TABLE IF NOT EXISTS market_engagement (
                market_code TEXT PRIMARY KEY,
                market_name TEXT NOT NULL DEFAULT '',
                workflow_stage TEXT NOT NULL DEFAULT 'To Contact',
                owner TEXT NOT NULL DEFAULT '',
                priority TEXT NOT NULL DEFAULT 'Medium',
                issue_summary TEXT NOT NULL DEFAULT '',
                latest_note TEXT NOT NULL DEFAULT '',
                next_action TEXT NOT NULL DEFAULT '',
                next_action_due TEXT NOT NULL DEFAULT '',
                first_contact_at TEXT NOT NULL DEFAULT '',
                last_contact_at TEXT NOT NULL DEFAULT '',
                resolved_at TEXT NOT NULL DEFAULT '',
                closed_at TEXT NOT NULL DEFAULT '',
                touchpoints_count INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS market_engagement_updates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                market_code TEXT NOT NULL,
                event_type TEXT NOT NULL DEFAULT 'manual_update',
                workflow_stage TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT '',
                next_action TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (market_code) REFERENCES market_engagement(market_code)
            );

            CREATE INDEX IF NOT EXISTS idx_market_engagement_updates_market_code
            ON market_engagement_updates (market_code);

            CREATE TABLE IF NOT EXISTS utm_sent_threads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversation_id TEXT UNIQUE NOT NULL,
                subject TEXT NOT NULL DEFAULT '',
                first_sent_at TEXT NOT NULL DEFAULT '',
                last_activity_at TEXT NOT NULL DEFAULT '',
                message_count INTEGER NOT NULL DEFAULT 0,
                participant_emails TEXT NOT NULL DEFAULT '',
                fetched_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS utm_sent_thread_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                thread_id INTEGER NOT NULL,
                outlook_entry_id TEXT NOT NULL DEFAULT '',
                direction TEXT NOT NULL DEFAULT 'received',
                sender_email TEXT NOT NULL DEFAULT '',
                sender_name TEXT NOT NULL DEFAULT '',
                to_emails TEXT NOT NULL DEFAULT '',
                cc_emails TEXT NOT NULL DEFAULT '',
                timestamp TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                body_text TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (thread_id) REFERENCES utm_sent_threads(id)
            );

            CREATE INDEX IF NOT EXISTS idx_utm_sent_thread_messages_thread_id
            ON utm_sent_thread_messages (thread_id);
            """
        )
        contact_columns = {row[1] for row in conn.execute("PRAGMA table_info('contacts')").fetchall()}
        if 'contact_tags' not in contact_columns:
            conn.execute("ALTER TABLE contacts ADD COLUMN contact_tags TEXT NOT NULL DEFAULT ''")
        reply_columns = {row[1] for row in conn.execute("PRAGMA table_info('outreach_replies')").fetchall()}
        if 'sender_name' not in reply_columns:
            conn.execute("ALTER TABLE outreach_replies ADD COLUMN sender_name TEXT NOT NULL DEFAULT ''")
        if 'body_text' not in reply_columns:
            conn.execute("ALTER TABLE outreach_replies ADD COLUMN body_text TEXT NOT NULL DEFAULT ''")


def _normalize_utm_notes_df(notes_df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        'Market Code': 'market_code',
        'Market Name': 'market_name',
        'Status': 'status',
        'Scope Gaps': 'scope_gaps',
        'Observations': 'observations',
        'Issues Identified': 'issues_identified',
        'Context': 'context',
        'Next Steps': 'next_steps',
    }
    expected_columns = list(rename_map.values())
    normalized = notes_df.rename(columns=rename_map).copy()
    missing_columns = [column for column in expected_columns if column not in normalized.columns]
    if missing_columns:
        raise ValueError(f'UTM notes data missing required columns: {", ".join(missing_columns)}')
    normalized = normalized[expected_columns].fillna('')
    normalized = normalized.astype(str).apply(lambda column: column.str.strip())
    normalized = normalized[normalized['market_code'] != '']
    return normalized


def _normalize_contacts_df(contacts_df: pd.DataFrame) -> pd.DataFrame:
    expected_columns = ['market', 'section', 'name', 'title', 'email', 'notes', 'source_sheet']
    normalized = contacts_df.copy()
    missing_columns = [column for column in expected_columns if column not in normalized.columns]
    if missing_columns:
        raise ValueError(f'Addressbook CSV missing required columns: {", ".join(missing_columns)}')
    normalized = normalized[expected_columns].fillna('')
    normalized = normalized.astype(str).apply(lambda column: column.str.strip())
    normalized = normalized[normalized['email'] != '']
    normalized = normalized.drop_duplicates(subset=['market', 'section', 'email'], keep='first')
    return normalized


def bootstrap_utm_notes_db(db_path: Path, pdf_path: Path, csv_path: Path) -> dict:
    with _get_utm_outreach_db_connection(db_path) as conn:
        current_count = conn.execute('SELECT COUNT(*) FROM utm_adoption_notes').fetchone()[0]
        base_count = conn.execute('SELECT COUNT(*) FROM utm_adoption_notes_base').fetchone()[0]
        if current_count > 0 and base_count > 0:
            return {'bootstrapped': False, 'rows': int(current_count)}

    if csv_path.exists():
        source_df = pd.read_csv(csv_path).fillna('')
    else:
        source_df = load_utm_adoption_notes(
            str(pdf_path),
            pdf_path.stat().st_mtime,
            fallback_csv_path=None,
        ).fillna('')

    normalized = _normalize_utm_notes_df(source_df)
    note_records = [
        (
            row['market_code'],
            row['market_name'],
            row['status'],
            row['scope_gaps'],
            row['observations'],
            row['issues_identified'],
            row['context'],
            row['next_steps'],
        )
        for _, row in normalized.iterrows()
    ]

    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.execute('DELETE FROM utm_adoption_notes')
        conn.execute('DELETE FROM utm_adoption_notes_base')
        conn.executemany(
            '''
            INSERT INTO utm_adoption_notes (
                market_code,
                market_name,
                status,
                scope_gaps,
                observations,
                issues_identified,
                context,
                next_steps,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            note_records,
        )
        conn.executemany(
            '''
            INSERT INTO utm_adoption_notes_base (
                market_code,
                market_name,
                status,
                scope_gaps,
                observations,
                issues_identified,
                context,
                next_steps,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            note_records,
        )
    return {'bootstrapped': True, 'rows': len(note_records)}


def bootstrap_utm_contacts_db(db_path: Path, csv_path: Path) -> dict:
    expected_columns = ['market', 'section', 'name', 'title', 'email', 'notes', 'source_sheet']
    with _get_utm_outreach_db_connection(db_path) as conn:
        current_count = conn.execute('SELECT COUNT(*) FROM contacts').fetchone()[0]
        if current_count > 0:
            return {'bootstrapped': False, 'rows': int(current_count)}

    contacts_df = pd.read_csv(csv_path).fillna('')
    normalized = _normalize_contacts_df(contacts_df)

    contact_records = [
        tuple(row[column] for column in expected_columns)
        for _, row in normalized.iterrows()
    ]

    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.executemany(
            '''
            INSERT INTO contacts (
                market,
                section,
                name,
                title,
                email,
                notes,
                source_sheet,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            contact_records,
        )
    return {'bootstrapped': True, 'rows': len(contact_records)}


def load_utm_notes_from_db(db_path: Path) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        notes_df = pd.read_sql_query(
            '''
            SELECT
                market_code AS "Market Code",
                market_name AS "Market Name",
                status AS "Status",
                scope_gaps AS "Scope Gaps",
                observations AS "Observations",
                issues_identified AS "Issues Identified",
                context AS "Context",
                next_steps AS "Next Steps"
            FROM utm_adoption_notes
            ORDER BY market_code
            ''',
            conn,
        )
    return notes_df.fillna('')


def save_utm_notes_to_db(df: pd.DataFrame, db_path: Path) -> None:
    normalized = _normalize_utm_notes_df(df)
    note_records = [
        (
            row['market_code'],
            row['market_name'],
            row['status'],
            row['scope_gaps'],
            row['observations'],
            row['issues_identified'],
            row['context'],
            row['next_steps'],
        )
        for _, row in normalized.iterrows()
    ]

    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.execute('DELETE FROM utm_adoption_notes')
        conn.executemany(
            '''
            INSERT INTO utm_adoption_notes (
                market_code,
                market_name,
                status,
                scope_gaps,
                observations,
                issues_identified,
                context,
                next_steps,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            note_records,
        )


def reset_utm_notes_to_base(db_path: Path) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        base_rows = conn.execute(
            '''
            SELECT
                market_code,
                market_name,
                status,
                scope_gaps,
                observations,
                issues_identified,
                context,
                next_steps
            FROM utm_adoption_notes_base
            ORDER BY market_code
            '''
        ).fetchall()
        conn.execute('DELETE FROM utm_adoption_notes')
        conn.executemany(
            '''
            INSERT INTO utm_adoption_notes (
                market_code,
                market_name,
                status,
                scope_gaps,
                observations,
                issues_identified,
                context,
                next_steps,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            [tuple(row) for row in base_rows],
        )


def load_contacts_from_db(db_path: Path) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        contacts_df = pd.read_sql_query(
            '''
            SELECT market, section, name, title, email, contact_tags, notes, source_sheet
            FROM contacts
            ORDER BY market, section, name, email
            ''',
            conn,
        )
    return contacts_df.fillna('')


def annotate_addressbook_contacts(contacts_df: pd.DataFrame) -> pd.DataFrame:
    annotated = contacts_df.copy()
    title_series = annotated['title'].fillna('').astype(str)
    section_series = annotated['section'].fillna('').astype(str)
    email_series = annotated['email'].fillna('').astype(str)
    tag_series = annotated['contact_tags'].fillna('').astype(str)

    annotated['is_porsche_client'] = (
        section_series.str.contains('PAG Client Contact', case=False, na=False)
        | email_series.str.contains('@porsche\\.', case=False, na=False)
    )
    annotated['is_account_lead'] = tag_series.str.contains(r'(^|,)account_lead(,|$)', case=False, na=False)
    annotated['contact_type'] = 'Other'
    annotated.loc[annotated['is_porsche_client'], 'contact_type'] = 'Porsche Client'
    annotated.loc[annotated['is_account_lead'], 'contact_type'] = 'Account Lead'
    annotated.loc[
        annotated['is_porsche_client'] & annotated['is_account_lead'],
        'contact_type',
    ] = 'Porsche Client + Account Lead'
    return annotated


def load_utm_outreach_db_counts(db_path: Path) -> dict:
    with _get_utm_outreach_db_connection(db_path) as conn:
        counts = {
            'contacts': conn.execute('SELECT COUNT(*) FROM contacts').fetchone()[0],
            'email_templates': conn.execute('SELECT COUNT(*) FROM email_templates').fetchone()[0],
            'outreach_runs': conn.execute('SELECT COUNT(*) FROM outreach_runs').fetchone()[0],
            'outreach_messages': conn.execute('SELECT COUNT(*) FROM outreach_messages').fetchone()[0],
            'outreach_replies': conn.execute('SELECT COUNT(*) FROM outreach_replies').fetchone()[0],
        }
    return {key: int(value) for key, value in counts.items()}


def ensure_manual_utm_test_records(db_path: Path) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        for record in MANUAL_UTM_TEST_RECORDS:
            conn.execute(
                '''
                INSERT INTO contacts (
                    market,
                    section,
                    name,
                    title,
                    email,
                    notes,
                    source_sheet,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(market, section, email) DO UPDATE SET
                    name = excluded.name,
                    title = excluded.title,
                    notes = excluded.notes,
                    source_sheet = excluded.source_sheet,
                    updated_at = CURRENT_TIMESTAMP
                ''',
                (
                    record['market_code'],
                    record['section'],
                    record['name'],
                    '',
                    record['email'],
                    record['notes'],
                    'manual',
                ),
            )
            conn.execute(
                '''
                INSERT INTO utm_adoption_notes (
                    market_code,
                    market_name,
                    status,
                    scope_gaps,
                    observations,
                    issues_identified,
                    context,
                    next_steps,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(market_code) DO UPDATE SET
                    market_name = excluded.market_name,
                    status = excluded.status,
                    scope_gaps = excluded.scope_gaps,
                    observations = excluded.observations,
                    issues_identified = excluded.issues_identified,
                    context = excluded.context,
                    next_steps = excluded.next_steps,
                    updated_at = CURRENT_TIMESTAMP
                ''',
                (
                    record['market_code'],
                    record['market_name'],
                    record['status'],
                    '',
                    '',
                    '',
                    record['context'],
                    record['next_steps'],
                ),
            )


def bootstrap_contact_tags(db_path: Path) -> None:
    account_lead_pattern = r'account'
    with _get_utm_outreach_db_connection(db_path) as conn:
        contacts = conn.execute(
            '''
            SELECT id, title, section
            FROM contacts
            '''
        ).fetchall()
        for contact in contacts:
            title = str(contact['title'] or '')
            section = str(contact['section'] or '')
            updated_tags = ''
            if re.search(account_lead_pattern, title, flags=re.IGNORECASE):
                updated_tags = 'account_lead'
            elif section.strip().lower() == 'planit champion'.lower():
                updated_tags = 'planit_champion'
            conn.execute(
                '''
                UPDATE contacts
                SET contact_tags = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                ''',
                (updated_tags, int(contact['id'])),
            )


def ensure_default_utm_outreach_template(db_path: Path) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.execute(
            '''
            INSERT OR IGNORE INTO email_templates (
                name,
                workflow,
                subject_template,
                body_template
            ) VALUES (?, ?, ?, ?)
            ''',
            (
                'UTM Follow-up',
                'UTM Adoption',
                'UTM adoption follow-up for {market_name}',
                (
                    'Hello,\n\n'
                    'Following up on UTM adoption for {market_name} ({market_code}).\n\n'
                    'Current status:\n{status}\n\n'
                    'Next steps:\n{next_steps}\n\n'
                    'Please let us know if there are any blockers or updates we should capture.\n\n'
                    'Best,\nAli'
                ),
            ),
        )


def load_email_templates_df(db_path: Path, workflow: str) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        templates_df = pd.read_sql_query(
            '''
            SELECT id, name, workflow, subject_template, body_template, created_at, updated_at
            FROM email_templates
            WHERE workflow = ?
            ORDER BY name
            ''',
            conn,
            params=(workflow,),
        )
    return templates_df.fillna('')


def save_email_template(
    db_path: Path,
    workflow: str,
    name: str,
    subject_template: str,
    body_template: str,
) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.execute(
            '''
            INSERT INTO email_templates (
                name,
                workflow,
                subject_template,
                body_template,
                updated_at
            ) VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(name, workflow) DO UPDATE SET
                subject_template = excluded.subject_template,
                body_template = excluded.body_template,
                updated_at = CURRENT_TIMESTAMP
            ''',
            (name.strip(), workflow.strip(), subject_template, body_template),
        )


def render_email_template(template_text: str, context: dict) -> str:
    return template_text.format_map({key: context.get(key, '') for key in context})


def _market_context_from_row(row: pd.Series) -> dict:
    return {
        'market_code': str(row.get('Market Code', '')).strip(),
        'market_name': str(row.get('Market Name', '')).strip(),
        'status': str(row.get('Status', '')).strip(),
        'scope_gaps': str(row.get('Scope Gaps', '')).strip(),
        'observations': str(row.get('Observations', '')).strip(),
        'issues_identified': str(row.get('Issues Identified', '')).strip(),
        'context': str(row.get('Context', '')).strip(),
        'next_steps': str(row.get('Next Steps', '')).strip(),
    }


def _normalize_market_lookup_value(value: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', value.strip().lower())


def _build_market_lookup_keys(market_code: str, market_name: str) -> set[str]:
    keys = set()
    raw_values = {market_code, market_name}
    alias_map = {
        'great britain': ['uk', 'united kingdom', 'gb'],
        'uk': ['great britain', 'united kingdom', 'gb'],
        'united kingdom': ['uk', 'great britain', 'gb'],
        'north america': ['pcna'],
        'pcgb': ['uk', 'great britain', 'united kingdom', 'gb'],
        'pib': ['portugal'],
        'portugal': ['pib'],
    }

    for raw_value in list(raw_values):
        normalized = _normalize_market_lookup_value(raw_value)
        if normalized:
            keys.add(normalized)

    for raw_value in list(raw_values):
        normalized_raw = raw_value.strip().lower()
        if normalized_raw in alias_map:
            for alias in alias_map[normalized_raw]:
                normalized_alias = _normalize_market_lookup_value(alias)
                if normalized_alias:
                    keys.add(normalized_alias)
    return keys


def build_contact_rule_options(contacts_df: pd.DataFrame) -> list[str]:
    section_values = sorted(
        section for section in contacts_df['section'].dropna().astype(str).unique() if section.strip()
    )
    tag_values = sorted(
        tag
        for raw_tags in contacts_df['contact_tags'].dropna().astype(str)
        for tag in [tag.strip() for tag in raw_tags.split(',')]
        if tag
    )
    return [f"Section: {section}" for section in section_values] + [f"Tag: {tag}" for tag in sorted(set(tag_values))]


def merge_unique_emails(base_emails: list[str], extra_emails: list[str]) -> list[str]:
    merged = []
    seen = set()
    for email in base_emails + extra_emails:
        normalized = str(email).strip()
        if not normalized:
            continue
        email_key = normalized.lower()
        if email_key in seen:
            continue
        seen.add(email_key)
        merged.append(normalized)
    return merged


def resolve_market_recipients(
    contacts_df: pd.DataFrame,
    market_code: str,
    market_name: str,
    primary_rules: list[str],
    cc_rules: list[str],
) -> dict:
    lookup_keys = _build_market_lookup_keys(market_code, market_name)
    market_contacts = contacts_df.copy()
    market_contacts['market_lookup'] = market_contacts['market'].astype(str).map(_normalize_market_lookup_value)
    market_contacts = market_contacts[market_contacts['market_lookup'].isin(lookup_keys)].copy()
    market_contacts['email'] = market_contacts['email'].astype(str).str.strip()
    market_contacts['section'] = market_contacts['section'].astype(str).str.strip()
    market_contacts = market_contacts[market_contacts['email'] != '']
    market_contacts['contact_tags'] = market_contacts['contact_tags'].fillna('').astype(str)
    market_contacts = market_contacts.sort_values(['section', 'name', 'email'], na_position='last')

    def pick_emails_for_rules(rules: list[str], excluded_emails: set[str]) -> list[str]:
        picked = []
        excluded = set(excluded_emails)
        for rule in rules:
            selected = market_contacts.copy()
            if rule.startswith('Section: '):
                section_name = rule.replace('Section: ', '', 1).strip()
                selected = selected[selected['section'] == section_name]
            elif rule.startswith('Tag: '):
                tag_name = rule.replace('Tag: ', '', 1).strip()
                tag_pattern = rf'(^|,){re.escape(tag_name)}(,|$)'
                selected = selected[selected['contact_tags'].str.contains(tag_pattern, case=False, na=False)]
            else:
                continue

            for _, contact in selected.iterrows():
                email = str(contact['email']).strip()
                if not email or email.lower() in excluded:
                    continue
                picked.append(email)
                excluded.add(email.lower())
                break
        return picked

    to_emails = pick_emails_for_rules(primary_rules, set())
    excluded = {email.lower() for email in to_emails}
    cc_emails = pick_emails_for_rules(cc_rules, excluded)
    missing_primary = len(to_emails) == 0

    return {
        'to_emails': to_emails,
        'cc_emails': cc_emails,
        'missing_primary': missing_primary,
    }


def get_outlook_status() -> tuple[bool, str]:
    try:
        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'scripts' / 'outlook_worker.py'), 'check'],
            capture_output=True,
            text=True,
            timeout=10,
        )
    except subprocess.TimeoutExpired:
        return False, 'Outlook health check timed out after 10 seconds.'
    except Exception as exc:
        return False, f'Outlook health check failed: {exc}'

    try:
        payload = json.loads((result.stdout or '').strip() or '{}')
    except Exception:
        payload = {'ok': False, 'error': (result.stderr or result.stdout or 'Unknown Outlook worker error.').strip()}

    if result.returncode == 0 and payload.get('ok'):
        return True, payload.get('message', 'Outlook COM connection is available.')
    return False, payload.get('error', 'Unknown Outlook worker error.')


def create_outreach_run(
    db_path: Path,
    workflow: str,
    name: str,
    filters_payload: dict,
    template_id: Optional[int],
) -> int:
    with _get_utm_outreach_db_connection(db_path) as conn:
        cursor = conn.execute(
            '''
            INSERT INTO outreach_runs (
                workflow,
                name,
                filters_json,
                template_id
            ) VALUES (?, ?, ?, ?)
            ''',
            (workflow, name.strip(), json.dumps(filters_payload, ensure_ascii=True), template_id),
        )
        return int(cursor.lastrowid)


def create_outreach_message(
    db_path: Path,
    run_id: int,
    market: str,
    status_snapshot: str,
    to_emails: list[str],
    cc_emails: list[str],
    subject_rendered: str,
    body_rendered: str,
    sent_at: Optional[str],
    outlook_entry_id: str,
    outlook_conversation_id: str,
    sync_state: str,
) -> None:
    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.execute(
            '''
            INSERT INTO outreach_messages (
                run_id,
                market,
                status_snapshot,
                to_emails,
                cc_emails,
                subject_rendered,
                body_rendered,
                sent_at,
                outlook_entry_id,
                outlook_conversation_id,
                sync_state
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                run_id,
                market,
                status_snapshot,
                '; '.join(to_emails),
                '; '.join(cc_emails),
                subject_rendered,
                body_rendered,
                sent_at,
                outlook_entry_id,
                outlook_conversation_id,
                sync_state,
            ),
        )


def send_or_draft_outlook_email(
    to_emails: list[str],
    cc_emails: list[str],
    subject: str,
    body: str,
    send_mode: str,
) -> dict:
    try:
        payload = json.dumps(
            {
                'to_emails': to_emails,
                'cc_emails': cc_emails,
                'subject': subject,
                'body': body,
                'send_mode': send_mode,
            },
            ensure_ascii=True,
        )
        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'scripts' / 'outlook_worker.py'), 'send_or_draft', payload],
            capture_output=True,
            text=True,
            timeout=25,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f'Outlook {send_mode} timed out after 25 seconds.') from exc
    except Exception as exc:
        raise RuntimeError(f'Outlook {send_mode} failed to start: {exc}') from exc

    try:
        response = json.loads((result.stdout or '').strip() or '{}')
    except Exception as exc:
        raise RuntimeError((result.stderr or result.stdout or 'Unknown Outlook worker error.').strip()) from exc

    if result.returncode != 0 or not response.get('ok'):
        raise RuntimeError(response.get('error', f'Outlook {send_mode} failed.'))
    return {
        'entry_id': response.get('entry_id', ''),
        'conversation_id': response.get('conversation_id', ''),
        'sent_at': response.get('sent_at', datetime.now().isoformat(timespec='seconds')),
        'sync_state': response.get('sync_state', send_mode),
    }


def load_outreach_tracker_df(db_path: Path, workflow: str) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        tracker_df = pd.read_sql_query(
            '''
            SELECT
                om.id,
                om.market AS market,
                om.status_snapshot AS current_utm_status,
                om.sent_at AS sent_at,
                om.to_emails AS "to",
                om.cc_emails AS cc,
                om.subject_rendered AS subject,
                om.body_rendered AS body_rendered,
                om.sync_state AS sync_state,
                om.replied AS replied,
                om.last_reply_at AS last_reply_at,
                om.last_reply_from AS last_reply_from,
                om.last_reply_snippet AS last_reply_snippet,
                orun.name AS run_name
            FROM outreach_messages om
            INNER JOIN outreach_runs orun ON orun.id = om.run_id
            WHERE orun.workflow = ?
            ORDER BY om.id DESC
            ''',
            conn,
            params=(workflow,),
        )
    if not tracker_df.empty and 'replied' in tracker_df.columns:
        tracker_df['replied'] = tracker_df['replied'].astype(bool)
    return tracker_df.fillna('')


def load_outreach_thread_df(db_path: Path, outreach_message_id: int) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        thread_df = pd.read_sql_query(
            '''
            SELECT
                received_at,
                sender_name,
                sender_email,
                subject,
                snippet,
                body_text,
                outlook_entry_id
            FROM outreach_replies
            WHERE outreach_message_id = ?
            ORDER BY received_at ASC, id ASC
            ''',
            conn,
            params=(outreach_message_id,),
        )
    return thread_df.fillna('')


def ensure_market_engagement_records(db_path: Path, notes_df: pd.DataFrame) -> None:
    normalized_notes = _normalize_utm_notes_df(notes_df)
    records = [
        (row['market_code'], row['market_name'])
        for _, row in normalized_notes[['market_code', 'market_name']].drop_duplicates().iterrows()
    ]
    with _get_utm_outreach_db_connection(db_path) as conn:
        conn.executemany(
            '''
            INSERT INTO market_engagement (
                market_code,
                market_name
            ) VALUES (?, ?)
            ON CONFLICT(market_code) DO UPDATE SET
                market_name = excluded.market_name
            ''',
            records,
        )


def load_market_engagement_df(db_path: Path) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        engagement_df = pd.read_sql_query(
            '''
            SELECT
                market_code AS "Market Code",
                market_name AS "Market Name",
                workflow_stage AS "Stage",
                owner AS "Owner",
                priority AS "Priority",
                issue_summary AS "Issue Summary",
                latest_note AS "Latest Note",
                next_action AS "Next Action",
                next_action_due AS "Next Action Due",
                first_contact_at AS "First Contact At",
                last_contact_at AS "Last Contact At",
                resolved_at AS "Resolved At",
                closed_at AS "Closed At",
                touchpoints_count AS "Touchpoints",
                updated_at AS "Updated At"
            FROM market_engagement
            ORDER BY
                CASE workflow_stage
                    WHEN 'To Contact' THEN 1
                    WHEN 'Sent' THEN 2
                    WHEN 'Waiting for Reply' THEN 3
                    WHEN 'In Progress' THEN 4
                    WHEN 'Blocked' THEN 5
                    WHEN 'Resolved' THEN 6
                    WHEN 'Closed' THEN 7
                    ELSE 99
                END,
                market_code
            ''',
            conn,
        )
    return engagement_df.fillna('')


def load_market_engagement_history_df(db_path: Path, market_code: str) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        history_df = pd.read_sql_query(
            '''
            SELECT
                created_at AS "Created At",
                event_type AS "Event Type",
                workflow_stage AS "Stage",
                note AS "Note",
                next_action AS "Next Action"
            FROM market_engagement_updates
            WHERE market_code = ?
            ORDER BY id DESC
            ''',
            conn,
            params=(market_code,),
        )
    return history_df.fillna('')


def save_market_engagement_update(
    db_path: Path,
    market_code: str,
    market_name: str,
    workflow_stage: str,
    owner: str,
    priority: str,
    issue_summary: str,
    latest_note: str,
    next_action: str,
    next_action_due: str,
    event_type: str = 'manual_update',
) -> None:
    now_value = datetime.now().isoformat(timespec='seconds')
    with _get_utm_outreach_db_connection(db_path) as conn:
        existing = conn.execute(
            '''
            SELECT touchpoints_count, first_contact_at
            FROM market_engagement
            WHERE market_code = ?
            ''',
            (market_code,),
        ).fetchone()
        touchpoints_count = int(existing['touchpoints_count']) if existing else 0
        first_contact_at = existing['first_contact_at'] if existing else ''

        new_first_contact_at = first_contact_at
        new_last_contact_at = now_value
        new_resolved_at = now_value if workflow_stage == 'Resolved' else ''
        new_closed_at = now_value if workflow_stage == 'Closed' else ''

        if workflow_stage in {'Sent', 'Waiting for Reply', 'In Progress', 'Blocked', 'Resolved', 'Closed'} and not first_contact_at:
            new_first_contact_at = now_value

        conn.execute(
            '''
            INSERT INTO market_engagement (
                market_code,
                market_name,
                workflow_stage,
                owner,
                priority,
                issue_summary,
                latest_note,
                next_action,
                next_action_due,
                first_contact_at,
                last_contact_at,
                resolved_at,
                closed_at,
                touchpoints_count,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(market_code) DO UPDATE SET
                market_name = excluded.market_name,
                workflow_stage = excluded.workflow_stage,
                owner = excluded.owner,
                priority = excluded.priority,
                issue_summary = excluded.issue_summary,
                latest_note = excluded.latest_note,
                next_action = excluded.next_action,
                next_action_due = excluded.next_action_due,
                first_contact_at = excluded.first_contact_at,
                last_contact_at = excluded.last_contact_at,
                resolved_at = excluded.resolved_at,
                closed_at = excluded.closed_at,
                touchpoints_count = excluded.touchpoints_count,
                updated_at = excluded.updated_at
            ''',
            (
                market_code,
                market_name,
                workflow_stage,
                owner.strip(),
                priority,
                issue_summary.strip(),
                latest_note.strip(),
                next_action.strip(),
                next_action_due.strip(),
                new_first_contact_at,
                new_last_contact_at,
                new_resolved_at,
                new_closed_at,
                touchpoints_count + 1,
                now_value,
            ),
        )
        conn.execute(
            '''
            INSERT INTO market_engagement_updates (
                market_code,
                event_type,
                workflow_stage,
                note,
                next_action,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?)
            ''',
            (
                market_code,
                event_type,
                workflow_stage,
                latest_note.strip(),
                next_action.strip(),
                now_value,
            ),
        )


def load_market_engagement_kpis(db_path: Path) -> dict:
    with _get_utm_outreach_db_connection(db_path) as conn:
        summary = conn.execute(
            '''
            SELECT
                COUNT(*) AS total_markets,
                SUM(CASE WHEN workflow_stage NOT IN ('To Contact') THEN 1 ELSE 0 END) AS engaged_markets,
                SUM(CASE WHEN workflow_stage = 'Blocked' THEN 1 ELSE 0 END) AS blocked_markets,
                SUM(CASE WHEN workflow_stage = 'Resolved' THEN 1 ELSE 0 END) AS resolved_markets,
                SUM(CASE WHEN workflow_stage = 'Closed' THEN 1 ELSE 0 END) AS closed_markets,
                SUM(touchpoints_count) AS total_touchpoints
            FROM market_engagement
            '''
        ).fetchone()
    return {
        'total_markets': int(summary['total_markets'] or 0),
        'engaged_markets': int(summary['engaged_markets'] or 0),
        'blocked_markets': int(summary['blocked_markets'] or 0),
        'resolved_markets': int(summary['resolved_markets'] or 0),
        'closed_markets': int(summary['closed_markets'] or 0),
        'total_touchpoints': int(summary['total_touchpoints'] or 0),
    }


def sync_outlook_replies(db_path: Path, workflow: str) -> int:
    try:
        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'scripts' / 'outlook_worker.py'), 'sync_replies', str(db_path), workflow],
            capture_output=True,
            text=True,
            timeout=30,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError('Reply sync timed out after 30 seconds.') from exc
    except Exception as exc:
        raise RuntimeError(f'Reply sync failed to start: {exc}') from exc

    try:
        response = json.loads((result.stdout or '').strip() or '{}')
    except Exception as exc:
        raise RuntimeError((result.stderr or result.stdout or 'Unknown Outlook worker error.').strip()) from exc

    if result.returncode != 0 or not response.get('ok'):
        raise RuntimeError(response.get('error', 'Reply sync failed.'))
    return int(response.get('updated', 0))


def fetch_and_store_utm_threads(db_path: Path, subject_keyword: str) -> int:
    try:
        result = subprocess.run(
            [sys.executable, str(BASE_DIR / 'scripts' / 'outlook_worker.py'), 'fetch_utm_threads', subject_keyword],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError('Thread fetch timed out after 60 seconds.') from exc
    except Exception as exc:
        raise RuntimeError(f'Thread fetch failed to start: {exc}') from exc

    try:
        response = json.loads((result.stdout or '').strip() or '{}')
    except Exception as exc:
        raise RuntimeError((result.stderr or result.stdout or 'Unknown Outlook worker error.').strip()) from exc

    if result.returncode != 0 or not response.get('ok'):
        raise RuntimeError(response.get('error', 'Thread fetch failed.'))

    threads = response.get('threads', [])
    with _get_utm_outreach_db_connection(db_path) as conn:
        for thread in threads:
            conn.execute(
                '''
                INSERT INTO utm_sent_threads (conversation_id, subject, first_sent_at, last_activity_at, message_count, participant_emails, fetched_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(conversation_id) DO UPDATE SET
                    subject = excluded.subject,
                    first_sent_at = excluded.first_sent_at,
                    last_activity_at = excluded.last_activity_at,
                    message_count = excluded.message_count,
                    participant_emails = excluded.participant_emails,
                    fetched_at = excluded.fetched_at
                ''',
                (
                    thread['conversation_id'],
                    thread['subject'],
                    thread['first_sent_at'],
                    thread['last_activity_at'],
                    thread['message_count'],
                    thread['participant_emails'],
                    datetime.now().isoformat(timespec='seconds'),
                ),
            )
            thread_id = conn.execute(
                'SELECT id FROM utm_sent_threads WHERE conversation_id = ?',
                (thread['conversation_id'],),
            ).fetchone()[0]
            conn.execute('DELETE FROM utm_sent_thread_messages WHERE thread_id = ?', (thread_id,))
            for msg in thread.get('messages', []):
                conn.execute(
                    '''
                    INSERT INTO utm_sent_thread_messages
                        (thread_id, outlook_entry_id, direction, sender_email, sender_name, to_emails, cc_emails, timestamp, subject, body_text)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''',
                    (
                        thread_id,
                        msg.get('outlook_entry_id', ''),
                        msg.get('direction', 'received'),
                        msg.get('sender_email', ''),
                        msg.get('sender_name', ''),
                        msg.get('to_emails', ''),
                        msg.get('cc_emails', ''),
                        msg.get('timestamp', ''),
                        msg.get('subject', ''),
                        msg.get('body_text', ''),
                    ),
                )
    return len(threads)


def load_utm_threads_df(db_path: Path) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        rows = conn.execute(
            'SELECT id, subject, first_sent_at, last_activity_at, message_count, participant_emails, fetched_at FROM utm_sent_threads ORDER BY first_sent_at DESC'
        ).fetchall()
    if not rows:
        return pd.DataFrame(columns=['id', 'Subject', 'First Sent', 'Last Activity', 'Messages', 'Participants', 'Fetched At'])
    return pd.DataFrame(
        [dict(r) for r in rows],
        columns=['id', 'subject', 'first_sent_at', 'last_activity_at', 'message_count', 'participant_emails', 'fetched_at'],
    ).rename(columns={
        'subject': 'Subject',
        'first_sent_at': 'First Sent',
        'last_activity_at': 'Last Activity',
        'message_count': 'Messages',
        'participant_emails': 'Participants',
        'fetched_at': 'Fetched At',
    })


def load_utm_thread_contact_stats_df(db_path: Path, excluded_emails: Optional[set] = None) -> pd.DataFrame:
    """One row per thread: subject, sent-at, To, CC, all external repliers, reply count."""
    excluded = {e.lower() for e in (excluded_emails or set())}
    with _get_utm_outreach_db_connection(db_path) as conn:
        thread_rows = conn.execute(
            '''
            SELECT
                t.id,
                t.subject,
                t.first_sent_at,
                (
                    SELECT m.to_emails
                    FROM utm_sent_thread_messages m
                    WHERE m.thread_id = t.id AND m.direction = 'sent'
                    ORDER BY m.timestamp ASC
                    LIMIT 1
                ) AS to_emails,
                (
                    SELECT m.cc_emails
                    FROM utm_sent_thread_messages m
                    WHERE m.thread_id = t.id AND m.direction = 'sent'
                    ORDER BY m.timestamp ASC
                    LIMIT 1
                ) AS cc_emails
            FROM utm_sent_threads t
            ORDER BY t.first_sent_at DESC
            '''
        ).fetchall()
        received_rows = conn.execute(
            '''
            SELECT thread_id, sender_name, sender_email
            FROM utm_sent_thread_messages
            WHERE direction = 'received'
            ORDER BY timestamp ASC
            '''
        ).fetchall()

    if not thread_rows:
        return pd.DataFrame()

    # Group received messages by thread, filtering out excluded/internal senders
    thread_repliers: dict[int, list[str]] = {}
    for r in received_rows:
        tid = r['thread_id']
        email = (r['sender_email'] or '').strip()
        name = (r['sender_name'] or '').strip()
        if email.lower() in excluded:
            continue
        label = f'{name} <{email}>' if name and email else (name or email)
        if not label:
            continue
        if tid not in thread_repliers:
            thread_repliers[tid] = []
        # keep unique labels preserving order of first appearance
        if label not in thread_repliers[tid]:
            thread_repliers[tid].append(label)

    records = []
    for row in thread_rows:
        tid = row['id']
        repliers = thread_repliers.get(tid, [])
        records.append({
            'id': tid,
            'Subject': row['subject'] or '',
            'Sent At': row['first_sent_at'] or '',
            'To': row['to_emails'] or '',
            'CC': row['cc_emails'] or '',
            'Responded': len(repliers) > 0,
            'Replies': len(repliers),
            'Replied By': ', '.join(repliers),
        })

    return pd.DataFrame(records, columns=['id', 'Subject', 'Sent At', 'To', 'CC', 'Responded', 'Replies', 'Replied By'])


def load_utm_thread_messages_df(db_path: Path, thread_id: int) -> pd.DataFrame:
    with _get_utm_outreach_db_connection(db_path) as conn:
        rows = conn.execute(
            '''
            SELECT id, direction, sender_name, sender_email, to_emails, cc_emails, timestamp, subject, body_text
            FROM utm_sent_thread_messages
            WHERE thread_id = ?
            ORDER BY timestamp ASC
            ''',
            (thread_id,),
        ).fetchall()
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame([dict(r) for r in rows])


def _pdf_escape(text: str) -> str:
    sanitized = text.encode('latin-1', errors='replace').decode('latin-1')
    return sanitized.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _text_to_pdf_bytes(text: str) -> bytes:
    # Minimal PDF writer for plain text output.
    lines = text.splitlines() or ['']
    lines_per_page = 50
    pages = [lines[i:i + lines_per_page] for i in range(0, len(lines), lines_per_page)]

    objects = []
    xref_positions = []
    buffer = []

    def add_obj(obj_str: str):
        xref_positions.append(sum(len(s.encode('latin-1')) for s in buffer))
        buffer.append(obj_str)

    buffer.append('%PDF-1.4\n')
    add_obj('1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n')
    kids = ' '.join([f'{3 + i * 2} 0 R' for i in range(len(pages))])
    add_obj(f'2 0 obj\n<< /Type /Pages /Kids [{kids}] /Count {len(pages)} >>\nendobj\n')

    for idx, page_lines in enumerate(pages):
        content_stream = 'BT\n/F1 11 Tf\n72 740 Td\n'
        for line in page_lines:
            content_stream += f'({_pdf_escape(line)}) Tj\n0 -14 Td\n'
        content_stream += 'ET\n'
        content_bytes = content_stream.encode('latin-1')
        obj_num = 3 + idx * 2
        add_obj(
            f'{obj_num} 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] '
            f'/Contents {obj_num + 1} 0 R /Resources << /Font << /F1 5 0 R >> >> >>\nendobj\n'
        )
        add_obj(
            f'{obj_num + 1} 0 obj\n<< /Length {len(content_bytes)} >>\nstream\n'
            f'{content_stream}endstream\nendobj\n'
        )

    add_obj('5 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n')

    xref_start = sum(len(s.encode('latin-1')) for s in buffer)
    xref = ['xref\n0 {}\n'.format(len(xref_positions) + 1), '0000000000 65535 f \n']
    for pos in xref_positions:
        xref.append(f'{pos:010d} 00000 n \n')
    buffer.append(''.join(xref))
    buffer.append(
        f'trailer\n<< /Size {len(xref_positions) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n'
    )
    return ''.join(buffer).encode('latin-1')


def _build_step_payload(results: pd.DataFrame) -> dict:
    payload = {}
    payload['rows'] = len(results)
    payload['gate_passed_rate'] = float(results['gate_passed'].mean()) if 'gate_passed' in results else None
    payload['headroom'] = results['headroom_score'].describe().to_dict() if 'headroom_score' in results else {}
    payload['headroom_tier_counts'] = (
        results['headroom_tier'].value_counts().to_dict() if 'headroom_tier' in results else {}
    )
    payload['scale_score'] = results['scale_score'].describe().to_dict() if 'scale_score' in results else {}
    payload['curve_zone_counts'] = (
        results['curve_zone'].value_counts().to_dict() if 'curve_zone' in results else {}
    )
    payload['volatility'] = results['volatility'].describe().to_dict() if 'volatility' in results else {}
    payload['vol_tier_counts'] = (
        results['vol_tier'].value_counts().to_dict() if 'vol_tier' in results else {}
    )
    payload['opportunity_score'] = (
        results['opportunity_score'].describe().to_dict() if 'opportunity_score' in results else {}
    )
    payload['opportunity_tier_counts'] = (
        results['opportunity_tier'].value_counts().to_dict() if 'opportunity_tier' in results else {}
    )
    return payload


def _run_llm_report(step_payload: dict, step_text: str, final_text: str, progress=None) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    outputs = []
    total_steps = len(step_payload) + 1
    completed = 0
    for step_name, content in step_payload.items():
        prompt = (
            f'You are the analyst for {step_name}. Explain this step in simple language for marketers, '
            f'use the provided stats and the step definition.\n\n'
            f'Step definition:\n{step_text}\n\n'
            f'Stats:\n{content}\n'
        )
        response = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
        )
        outputs.append(f'## {step_name}\n{response.choices[0].message.content}')
        completed += 1
        if progress:
            progress.progress(completed / total_steps, text=f'Completed {step_name}')

    synthesis_prompt = (
        'You are the final reporting agent. Use the step summaries below to write a clear, '
        'non-technical report with a short executive summary, key findings per step, '
        'and a concise conclusion.\n\n'
        f'Step summaries:\n{chr(10).join(outputs)}\n\n'
        f'Final reporting guidance:\n{final_text}\n'
    )
    completed += 1
    if progress:
        progress.progress(completed / total_steps, text='Final report synthesis')
    final_response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': synthesis_prompt}],
    )
    report = '\n'.join(outputs) + '\n\n# Final Report\n' + final_response.choices[0].message.content
    return report


def _run_headroom_report(summary: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Write a short, tight headroom report (max 8 sentences). '
        'DCFS means Dealer Contact Form Submissions. '
        'Include: 1) the idea of headroom, 2) how the numbers are derived, '
        '3) how to use it strategically. Use plain language for marketers. '
        'Do not include any extra sections or headings.\n\n'
        f'Headroom summary:\n{summary}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content


def _run_scale_report(summary: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Write a short, tight scale report (max 8 sentences). '
        'DCFS means Dealer Contact Form Submissions. '
        'Include: 1) the idea of scale, 2) how the numbers are derived, '
        '3) how to use it strategically. Use plain language for marketers. '
        'Do not include any extra sections or headings.\n\n'
        f'Scale summary:\n{summary}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content


def _run_spend_distribution_report(summary: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Write a short, tight spend distribution report (max 8 sentences). '
        's50 is the saturation point captured from the media response curve. '
        'Explain exactly what is shown on the plot: the recent spend bars, the spend '
        'distribution boxplots, and the color-based zone classification. '
        'Use plain language for marketers. Do not add extra sections or headings.\n\n'
        f'Spend distribution summary:\n{summary}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content


def _run_predictability_report(summary: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Write a short, tight predictability report (max 8 sentences). '
        'CPL is cost per lead. Explain exactly what is shown on the plot: '
        'volatility bars by group and the LOW/MED/HIGH/VERY_HIGH thresholds. '
        'Use plain language for marketers. Do not add extra sections or headings.\n\n'
        f'Predictability summary:\n{summary}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content


def _run_opportunity_report(summary: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Write a short, tight opportunity score report (max 8 sentences). '
        'Explain exactly what is shown on the plot: average opportunity score by group '
        'and the 0–100 scale. Use plain language for marketers. '
        'Do not add extra sections or headings.\n\n'
        f'Opportunity summary:\n{summary}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content


def _run_final_conclusion(reports: dict) -> str:
    if OpenAI is None:
        return 'LLM client not available. Install openai package.'
    client = OpenAI()
    prompt = (
        'Using the reports below, write a concise conclusion and budget allocation strategy '
        'to win the incentive deal. Be specific about prioritization and guardrails. '
        'Do not invent metrics or data. Use plain language for marketers. '
        'Output two short paragraphs labeled "Conclusion" and "Strategy".\n\n'
        f'Reports:\n{reports}\n'
    )
    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
    )
    return response.choices[0].message.content

with st.sidebar:
    st.header('Data Ingestion')
    uploaded_excel = st.file_uploader(
        'Upload weekly Excel (Python Output sheet)',
        type=['xlsx'],
    )

data_source_label = f'CSV: {CSV_PATH.name}'
if uploaded_excel is not None:
    try:
        df = load_excel_python_output(uploaded_excel.getvalue(), uploaded_excel.name)
        data_source_label = f'Uploaded Excel: {uploaded_excel.name}'
    except Exception as exc:
        st.error(f'Unable to read the uploaded Excel file: {exc}')
        df = pd.DataFrame()
else:
    if CSV_PATH.exists():
        try:
            df = load_data(CSV_PATH, CSV_PATH.stat().st_mtime)
            data_source_label = f'CSV: {CSV_PATH.name}'
        except Exception as exc:
            st.warning(f'Unable to read default CSV: {exc}')
            df = pd.DataFrame()
    else:
        df = pd.DataFrame()
    if df.empty:
        st.info('No PWC data loaded. Some pages will be unavailable until a weekly Excel is uploaded.')

with st.sidebar.expander('Data diagnostics'):
    if 'Date' in df.columns:
        st.write('Date min:', df['Date'].min())
        st.write('Date max:', df['Date'].max())
    if 'calendar_week' in df.columns:
        week_list = get_calendar_week_options(df)
        st.write('Calendar weeks:', week_list[:5], '...', week_list[-5:])
        st.write('Total weeks:', len(week_list))

st.title('Intelligence Console')
st.caption(f'Data source: {data_source_label}')

numeric_cols = df.select_dtypes(include='number').columns.tolist()
numeric_cols = [col for col in numeric_cols if col not in {'Date'}]

categorical_cols = [
    col for col in [
        'Market', 'Model', 'Ad Type', 'Channel', 'Platform', 'Activation Group',
        'Campaign', 'calendar_week', 'week_relative', 'week_text', 'report_week'
    ]
    if col in df.columns
]

dual_selections = {}
dual_breakdown_dim = None
dual_aggregate = False
dual_aggregate_dims = {}
dual_left_kpi = None
dual_right_kpi = None

def _label_value(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 'not specified'
    if isinstance(value, str):
        text = value.strip()
        return text if text else 'not specified'
    return str(value)


def _safe_ratio(numerator, denominator):
    return numerator / denominator if denominator else None


def _aggregate_metrics(df_in):
    media = df_in['Media Spend'].sum()
    visits = df_in['Number of Sessions'].sum()
    forms = df_in['Forms Submission Started'].sum()
    dcfs = df_in['DCFS'].sum()
    v2l = _safe_ratio(dcfs, visits)
    return media, visits, forms, dcfs, v2l


def _add_section_header(rows, market, model, week_label):
    rows.append(['Markets', market, None, None, None, None])
    if model is not None:
        rows.append(['Models', model, None, None, None, None])
    rows.append(['calendar_week', week_label, None, None, None, None])
    rows.append([None, None, None, None, None, None])
    rows.append(['Row Labels', 'Media Spend', 'Vists (Sessions)', 'Forms Started', 'DCFS', 'Visits to Lead CR'])


def _add_channel_section(rows, df_in):
    base = df_in.copy()
    for col in ['Channel', 'Platform', 'Activation Group']:
        base[col] = base[col].apply(_label_value)

    grouped = (
        base.groupby(['Channel', 'Platform', 'Activation Group'], dropna=False)
        .agg({
            'Media Spend': 'sum',
            'Number of Sessions': 'sum',
            'Forms Submission Started': 'sum',
            'DCFS': 'sum',
        })
        .reset_index()
    )
    grouped['Visits to Lead CR'] = grouped.apply(
        lambda r: _safe_ratio(r['DCFS'], r['Number of Sessions']), axis=1
    )

    def sort_key(value):
        text = str(value).strip()
        return (text.lower() == 'not specified', text)

    channels = sorted(grouped['Channel'].unique(), key=sort_key)
    for channel in channels:
        rows.append([channel, None, None, None, None, None])
        channel_df = grouped[grouped['Channel'] == channel]
        platforms = sorted(channel_df['Platform'].unique(), key=sort_key)
        for platform in platforms:
            rows.append([platform, None, None, None, None, None])
            platform_df = channel_df[channel_df['Platform'] == platform]
            activations = sorted(platform_df['Activation Group'].unique(), key=sort_key)
            for activation in activations:
                row = platform_df[platform_df['Activation Group'] == activation].iloc[0]
                rows.append([
                    activation,
                    row['Media Spend'],
                    row['Number of Sessions'],
                    row['Forms Submission Started'],
                    row['DCFS'],
                    row['Visits to Lead CR'],
                ])

    total = _aggregate_metrics(base)
    rows.append(['Grand Total', *total])


def _add_model_summary(rows, df_in):
    base = df_in.copy()
    base['Model'] = base['Model'].apply(_label_value)
    grouped = (
        base.groupby('Model', dropna=False)
        .agg({
            'Media Spend': 'sum',
            'Number of Sessions': 'sum',
            'Forms Submission Started': 'sum',
            'DCFS': 'sum',
        })
        .reset_index()
    )
    grouped['Visits to Lead CR'] = grouped.apply(
        lambda r: _safe_ratio(r['DCFS'], r['Number of Sessions']), axis=1
    )

    for _, row in grouped.sort_values('Model').iterrows():
        rows.append([
            row['Model'],
            row['Media Spend'],
            row['Number of Sessions'],
            row['Forms Submission Started'],
            row['DCFS'],
            row['Visits to Lead CR'],
        ])

    total = _aggregate_metrics(base)
    rows.append(['Grand Total', *total])


def build_close_gap_workbook(df_in, market, week_label):
    rows = []
    _add_section_header(rows, market, 'All', week_label)
    _add_channel_section(rows, df_in)
    rows.append([None, None, None, None, None, None])
    rows.append([None, None, None, None, None, None])

    _add_section_header(rows, market, None, week_label)
    _add_model_summary(rows, df_in)
    rows.append([None, None, None, None, None, None])
    rows.append([None, None, None, None, None, None])

    for model in sorted(df_in['Model'].dropna().unique()):
        model_df = df_in[df_in['Model'] == model]
        _add_section_header(rows, market, model, week_label)
        _add_channel_section(rows, model_df)
        rows.append([None, None, None, None, None, None])
        rows.append([None, None, None, None, None, None])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    for row in rows:
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _saturation_curve(x, a, b):
    return a * x / (b + x)


def fit_saturation(x, y):
    if np is None or curve_fit is None:
        return None, None
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    mask = np.isfinite(x) & np.isfinite(y) & (x >= 0)
    x = x[mask]
    y = y[mask]
    if len(x) < 3:
        return None, None
    a0 = max(y.max(), 1.0)
    b0 = max(np.median(x), 1.0)
    try:
        params, _ = curve_fit(
            _saturation_curve,
            x,
            y,
            p0=[a0, b0],
            bounds=([0.0, 0.0], [np.inf, np.inf]),
            maxfev=20000,
        )
        return params[0], params[1]
    except Exception:
        return None, None


def compute_dynamic_s50(df_in):
    if np is None or curve_fit is None:
        return {}
    group_cols = [col for col in ['Market', 'Channel', 'Model'] if col in df_in.columns]
    if not group_cols:
        return {}
    curve_df = df_in.copy()
    curve_df['Media Spend'] = pd.to_numeric(curve_df['Media Spend'], errors='coerce')
    curve_df['DCFS'] = pd.to_numeric(curve_df['DCFS'], errors='coerce')
    curve_df = curve_df[(curve_df['Media Spend'] > 0) & (curve_df['DCFS'] >= 0)]
    if curve_df.empty:
        return {}
    s50_map = {}
    for key, group in curve_df.groupby(group_cols, dropna=False):
        a, b = fit_saturation(group['Media Spend'], group['DCFS'])
        if b is None or pd.isna(b) or b <= 0:
            continue
        s50_map[key] = float(b)
    return s50_map


with st.sidebar:
    page = st.radio(
        'Page',
        [
            'Overview',
            'Risk Analysis',
            'Market CPL',
            'Market Report - Excel Export',
            'KPI vs Investment',
            'Market Alignments',
            'UTM Adoption',
            'Taxonomy Hygiene',
            'Budget Setting Sankey',
            'Incentive Model',
            'Weekly Market KPIs',
            'CTG Pre/Post KPI per Session',
            'CTG Pre/Post KPI per Session (Copy)',
        ],
        horizontal=True,
    )
    if page == 'Overview':
        st.header('Plot Filters')
        st.caption('All includes every value. Use Aggregate to combine into one series.')
        dimension_candidates = [
            ('Market', 'Market'),
            ('Model', 'Model'),
            ('Campaign', 'Campaign'),
            ('Channel', 'Channel'),
            ('Platform', 'Platform'),
            ('Activation Group', 'Activation Group'),
        ]
        for label, col in dimension_candidates:
            if col in df.columns:
                options = ['All'] + sorted(df[col].dropna().unique())
                select_col, agg_col = st.columns([3, 1], vertical_alignment='center')
                with select_col:
                    dual_selections[col] = st.multiselect(f'{label}', options, default=['All'])
                with agg_col:
                    dual_aggregate_dims[col] = st.checkbox('Aggregate', value=True, key=f'agg_{col}')

        breakdown_dims = [
            col
            for col, selections in dual_selections.items()
            if selections and not dual_aggregate_dims.get(col, False)
        ]
        dual_breakdown_dim = breakdown_dims

        base_kpis = numeric_cols.copy()
        extra_kpis = [
            'Cost per Lead (Forms Submission Started)',
            'Cost per Lead (DCFS)',
            'CPM',
        ]
        kpi_choices = [k for k in base_kpis if k not in extra_kpis] + extra_kpis
        dual_left_kpi = st.selectbox('Left axis KPI', kpi_choices, index=0)
        compare_kpis = st.checkbox('Compare (add right axis)', value=False)
        if compare_kpis:
            dual_right_kpi = st.selectbox('Right axis KPI', kpi_choices, index=min(7, len(kpi_choices) - 1))
        else:
            dual_right_kpi = None

        metric = None
        agg_func = None
        filtered = df
        model_df = None
        market = None
        campaign = None
        top_n = None
        export_market = None
        export_weeks = None
    elif page == 'Market CPL':
        st.header('Filters')
        if 'Model' not in df.columns:
            st.warning('Model column not found in the dataset.')
            st.stop()

        market = None
        base_df = df

        model_options = ['All'] + sorted(base_df['Model'].dropna().unique())
        if not model_options:
            st.warning('No models available for the selected filters.')
            st.stop()
        model = st.selectbox('Model', model_options)
        model_df = base_df if model == 'All' else base_df[base_df['Model'] == model]

        campaign_options = []
        if 'Campaign' in model_df.columns:
            campaign_options = ['All'] + sorted(model_df['Campaign'].dropna().unique())
        campaign = st.selectbox('Campaign', campaign_options) if campaign_options else None

        filtered = model_df
        if campaign and campaign != 'All':
            filtered = filtered[filtered['Campaign'] == campaign]

        if filtered.empty:
            st.warning('No data available for the current filters.')
            st.stop()

        metric = None
        agg_func = None
        top_n = None
        export_market = None
        export_weeks = None
    elif page == 'Market Report - Excel Export':
        st.header('Filters')
        if 'Market' not in df.columns:
            st.warning('Market column not found in the dataset.')
            st.stop()
        export_market = st.selectbox('Market', sorted(df['Market'].dropna().unique()))
        campaign_options = ['All']
        if 'Campaign' in df.columns:
            campaign_options += sorted(df['Campaign'].dropna().unique())
        export_campaign = st.selectbox('Campaign', campaign_options)
        date_mode = st.radio('Filter by', ['Weeks', 'Date range'], horizontal=True)
        week_options = get_calendar_week_options(df)
        week_choices = ['All'] + week_options
        export_weeks = st.multiselect('Weeks', week_choices, default=['All'], disabled=date_mode == 'Date range')
        export_dates = None
        date_col = None
        if date_mode == 'Date range':
            date_col = 'Date' if 'Date' in df.columns else 'report_date'
            if date_col not in df.columns:
                st.warning('No date column found for date range filtering.')
                st.stop()
            date_series = pd.to_datetime(df[date_col], errors='coerce')
            min_date = date_series.min()
            max_date = date_series.max()
            if pd.isna(min_date) or pd.isna(max_date):
                st.warning('Date column has no valid values.')
                st.stop()
            export_dates = st.date_input(
                'Date range',
                value=(min_date.date(), max_date.date()),
                min_value=min_date.date(),
                max_value=max_date.date(),
            )
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'Risk Analysis':
        market_options = ['All'] + sorted(df['Market'].dropna().unique()) if 'Market' in df.columns else ['All']
        channel_options = ['All'] + sorted(df['Channel'].dropna().unique()) if 'Channel' in df.columns else ['All']
        model_options = ['All'] + sorted(df['Model'].dropna().unique()) if 'Model' in df.columns else ['All']
        campaign_options = ['All'] + sorted(df['Campaign'].dropna().unique()) if 'Campaign' in df.columns else ['All']

        headroom_high_input = st.number_input(
            'Headroom high threshold',
            min_value=0.01,
            max_value=5.0,
            value=float(OPPORTUNITY_CONFIG['headroom_high']),
            step=0.01,
            format='%.2f',
        )
        recent_periods_input = st.number_input(
            'Recent periods',
            min_value=1,
            max_value=52,
            value=int(OPPORTUNITY_CONFIG['recent_cpl_periods']),
            step=1,
        )
        growth_ratio_max_input = st.number_input(
            'Growth ratio max',
            min_value=0.01,
            max_value=5.0,
            value=float(OPPORTUNITY_CONFIG['growth_ratio_max']),
            step=0.01,
            format='%.2f',
        )
        mid_ratio_max_input = st.number_input(
            'Mid ratio max',
            min_value=0.01,
            max_value=5.0,
            value=float(OPPORTUNITY_CONFIG['mid_ratio_max']),
            step=0.01,
            format='%.2f',
        )
        curve_group_candidates = [col for col in ['calendar_week', 'Date', 'report_date'] if col in df.columns]
        curve_group_by = (
            st.selectbox('Curve group by', curve_group_candidates) if curve_group_candidates else None
        )
        group_by_candidates = [col for col in ['Market', 'Channel', 'Model'] if col in df.columns]
        group_by = st.selectbox('Group plots by', group_by_candidates, index=0) if group_by_candidates else None
        def _expand_all_markets():
            selected = st.session_state.get('risk_markets', [])
            if 'All' in selected:
                all_markets = [m for m in st.session_state.get('risk_market_options', []) if m != 'All']
                st.session_state['risk_markets'] = all_markets

        st.session_state['risk_market_options'] = market_options
        opp_markets = st.multiselect(
            'Markets',
            market_options,
            default=['All'],
            key='risk_markets',
            on_change=_expand_all_markets,
        )
        opp_channel = st.selectbox('Channel', channel_options)
        opp_model = st.selectbox('Model', model_options)
        opp_campaign = st.selectbox('Campaign', campaign_options)
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'KPI vs Investment':
        if 'Market' not in df.columns:
            st.warning('Market column not found in the dataset.')
            st.stop()
        kpi_market = st.selectbox('Market', ['All'] + sorted(df['Market'].dropna().unique()))
        kpi_channels = []
        if 'Channel' in df.columns:
            channel_choices = ['All'] + sorted(df['Channel'].dropna().unique())
            kpi_channels = st.multiselect('Channels', channel_choices, default=['All'])
        kpi_campaigns = []
        if 'Campaign' in df.columns:
            campaign_choices = ['All'] + sorted(df['Campaign'].dropna().unique())
            kpi_campaigns = st.multiselect('Campaigns', campaign_choices, default=['All'])
        kpi_platforms = []
        if 'Platform' in df.columns:
            platform_choices = ['All'] + sorted(df['Platform'].dropna().unique())
            kpi_platforms = st.multiselect('Platforms', platform_choices, default=['All'])
        kpi_models = []
        if 'Model' in df.columns:
            model_choices = ['All'] + sorted(df['Model'].dropna().unique())
            kpi_models = st.multiselect('Models', model_choices, default=['All'])
        kpi_activations = []
        if 'Activation Group' in df.columns:
            activation_choices = ['All'] + sorted(df['Activation Group'].dropna().unique())
            kpi_activations = st.multiselect('Activation Groups', activation_choices, default=['All'])

        kpi_options = [
            'Visits (Sessions)',
            'Dealer Contract Form Submissions',
            'DCFS',
            'Sessions to DCFS Conversion Rate',
            'Cost per Lead (Forms Submission Started)',
            'Cost per Lead (DCFS)',
        ]
        kpi_choice = st.selectbox('KPI', kpi_options)

        color_candidates = [None]
        for col in ['Channel', 'Campaign', 'Platform', 'Model', 'Activation Group']:
            if col in df.columns:
                color_candidates.append(col)
        color_by = st.selectbox('Color by', color_candidates, format_func=lambda x: x or 'None')

        group_candidates = []
        for col in ['calendar_week', 'Date', 'report_date']:
            if col in df.columns:
                group_candidates.append(col)
        group_by = st.selectbox('Group by', group_candidates) if group_candidates else None

        kpi_filters = {
            'market': kpi_market,
            'channels': kpi_channels,
            'campaigns': kpi_campaigns,
            'platforms': kpi_platforms,
            'models': kpi_models,
            'activations': kpi_activations,
            'kpi': kpi_choice,
            'color_by': color_by,
            'group_by': group_by,
        }
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
    elif page == 'Market Alignments':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'UTM Adoption':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'Incentive Model':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'Weekly Market KPIs':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'CTG Pre/Post KPI per Session':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None
    elif page == 'CTG Pre/Post KPI per Session (Copy)':
        metric = None
        agg_func = None
        filtered = None
        model_df = None
        market = None
        campaign = None
        top_n = None
        kpi_filters = None

if page == 'Risk Analysis':
    st.subheader('Risk Analysis')
    with st.popover('What is this?'):
        st.write(
            'This page scores markets/channels/models for opportunity and risk using recent '
            'performance and spend efficiency. It combines headroom, scale, curve position, '
            'and predictability into one opportunity score.'
        )
        st.write(
            'How to use it:\n'
            '1. Set thresholds in the left panel (or keep defaults).\n'
            '2. Filter by Market/Channel/Model if you want a narrower view.\n'
            '3. Review each step section to see how the score is built.\n'
            '4. Use the LLM report for a plain‑language summary.'
        )

    st.subheader('Efficiency headroom')
    config_override = dict(OPPORTUNITY_CONFIG)
    config_override['headroom_high'] = float(headroom_high_input)
    config_override['recent_cpl_periods'] = int(recent_periods_input)
    config_override['recent_scale_periods'] = int(recent_periods_input)
    config_override['recent_curve_periods'] = int(recent_periods_input)
    config_override['growth_ratio_max'] = float(growth_ratio_max_input)
    config_override['mid_ratio_max'] = float(mid_ratio_max_input)
    df_input = df.copy()
    if opp_markets and 'All' not in opp_markets:
        df_input = df_input[df_input['Market'].isin(opp_markets)]
    if opp_channel != 'All':
        df_input = df_input[df_input['Channel'] == opp_channel]
    if opp_model != 'All':
        df_input = df_input[df_input['Model'] == opp_model]
    if opp_campaign != 'All':
        df_input = df_input[df_input['Campaign'] == opp_campaign]

    if np is None or curve_fit is None:
        st.info('Install scipy to enable dynamic s50 curve fitting.')
    else:
        s50_map = compute_dynamic_s50(df_input)
        if s50_map:
            group_cols = [col for col in ['Market', 'Channel', 'Model'] if col in df_input.columns]
            if group_cols:
                df_input = df_input.copy()
                df_input['s50_spend'] = (
                    df_input[group_cols]
                    .apply(lambda r: s50_map.get(tuple(r.tolist())), axis=1)
                )

    results, missing = compute_headroom_scores(df_input, config_override)
    if missing:
        st.warning(f'Missing required columns: {", ".join(missing)}')
        st.stop()
    if results.empty:
        st.warning('No data available to compute headroom.')
        st.stop()

    st.subheader('LLM report')
    st.caption('LLM insights (wireframe)')
    if st.button('Generate LLM Report (Markdown)'):
        st.info('Coming soon...')

    st.subheader('Headroom process (selected group)')
    with st.popover('What is this?'):
        st.write(
            'Shows how headroom is derived for one Market/Channel/Model: '
            'current CPL, benchmark CPL (25th percentile), headroom %, and headroom score.'
        )
    pipeline_df = results[results['gate_passed']].copy()
    pipeline_df = pipeline_df.dropna(subset=['current_cpl', 'benchmark_cpl_p25'])
    if pipeline_df.empty:
        st.info('No headroom process data available for the current filters.')
    else:
        pipeline_df['group_label'] = (
            pipeline_df[['Market', 'Channel', 'Model']]
            .astype(str)
            .agg(' | '.join, axis=1)
        )
        selected_label = st.selectbox(
            'Select group',
            pipeline_df['group_label'].tolist(),
        )
        selected_row = pipeline_df[pipeline_df['group_label'] == selected_label].iloc[0]
        current_cpl = float(selected_row['current_cpl'])
        benchmark_cpl = float(selected_row['benchmark_cpl_p25'])
        headroom_pct = float(selected_row['headroom']) * 100 if pd.notna(selected_row['headroom']) else None
        headroom_score = float(selected_row['headroom_score']) if pd.notna(selected_row['headroom_score']) else None

        c1, c2, c3 = st.columns(3)
        c1.metric('Current CPL', f'{current_cpl:,.2f}')
        c2.metric('Benchmark CPL (P25)', f'{benchmark_cpl:,.2f}')
        c3.metric('Headroom %', f'{headroom_pct:.1f}%' if headroom_pct is not None else 'n/a')

        fig = make_subplots(
            rows=2,
            cols=1,
            vertical_spacing=0.18,
            specs=[[{}], [{'secondary_y': True}]],
            subplot_titles=('CPL vs Benchmark', 'Headroom % and Score'),
        )
        fig.add_trace(
            go.Bar(
                x=['Benchmark CPL (P25)', 'Current CPL'],
                y=[benchmark_cpl, current_cpl],
                marker_color=['#9DB2BF', '#1F77B4'],
            ),
            row=1,
            col=1,
        )
        fig.update_yaxes(title_text='CPL', row=1, col=1)

        if headroom_pct is not None:
            fig.add_trace(
                go.Bar(
                    x=['Headroom %'],
                    y=[headroom_pct],
                    marker_color='#2CA02C',
                    name='Headroom %',
                ),
                row=2,
                col=1,
                secondary_y=False,
            )
        if headroom_score is not None:
            fig.add_trace(
                go.Scatter(
                    x=['Headroom %'],
                    y=[headroom_score],
                    mode='markers+text',
                    text=[f'{headroom_score:.0f}'],
                    textposition='top center',
                    marker=dict(size=12, color='#FF7F0E'),
                    name='Headroom score',
                ),
                row=2,
                col=1,
                secondary_y=True,
            )
        fig.update_yaxes(title_text='Headroom %', row=2, col=1, secondary_y=False)
        fig.update_yaxes(title_text='Score (0–100)', range=[0, 100], row=2, col=1, secondary_y=True)
        fig.update_layout(showlegend=False, height=520)
        st.plotly_chart(fig, use_container_width=True)

    st.subheader('Headroom by group')
    with st.popover('What is this?'):
        st.write(
            'Compares current CPL vs. a benchmark to show efficiency headroom by group. '
            'Higher headroom % means more room to improve efficiency.'
        )
    base_df = results.copy()
    if base_df.empty:
        st.info('No headroom data for the current filters.')
    else:
        if group_by:
            agg_df = (
                base_df.groupby(group_by, dropna=False)
                .agg(
                    headroom_pct=('headroom', lambda s: float((s * 100).mean())),
                    scale_score=('scale_score', 'mean'),
                    curve_score=('curve_score', 'mean'),
                    spend_recent=('spend_recent', 'mean'),
                )
                .reset_index()
                .rename(columns={group_by: 'group'})
            )
        else:
            agg_df = pd.DataFrame({
                'group': ['All'],
                'headroom_pct': [(base_df['headroom'] * 100).mean()],
                'scale_score': [base_df['scale_score'].mean()],
                'curve_score': [base_df['curve_score'].mean()],
                'spend_recent': [base_df['spend_recent'].mean()],
            })
        group_order = (
            agg_df.sort_values('headroom_pct', ascending=False)['group'].tolist()
            if 'headroom_pct' in agg_df.columns
            else agg_df['group'].tolist()
        )
        palette = px.colors.qualitative.Safe
        group_color_map = {
            group: palette[idx % len(palette)] for idx, group in enumerate(group_order)
        }
        plot_df = agg_df.dropna(subset=['headroom_pct'])
        if plot_df.empty:
            st.info('No headroom data for the current filters.')
        else:
            high_pct = float(headroom_high_input) * 100
            med_pct = float(OPPORTUNITY_CONFIG['headroom_med']) * 100
            plot_df = plot_df.copy()
            plot_df['tier'] = 'LOW'
            plot_df.loc[plot_df['headroom_pct'] >= med_pct, 'tier'] = 'MED'
            plot_df.loc[plot_df['headroom_pct'] >= high_pct, 'tier'] = 'HIGH'
            fig = px.bar(
                plot_df.sort_values('headroom_pct', ascending=False),
                x='group',
                y='headroom_pct',
                text='tier',
                color='group',
                labels={'headroom_pct': 'Headroom %', 'group': group_by or 'Group'},
                color_discrete_map=group_color_map,
            )
            fig.update_xaxes(categoryorder='array', categoryarray=group_order)
            fig.update_traces(
                texttemplate='%{text}',
                textposition='outside',
            )
            threshold_pct = float(headroom_high_input) * 100
            fig.add_hline(
                y=threshold_pct,
                line_dash='dash',
                line_color='orange',
                annotation_text=f'High threshold ({threshold_pct:.0f}%)',
                annotation_position='top left',
            )
            fig.update_yaxes(title_text='Headroom %')
            st.plotly_chart(fig, use_container_width=True)
            headroom_table_df = plot_df.sort_values('headroom_pct', ascending=False).copy()
            show_headroom_table = st.checkbox(
                'Show headroom table',
                value=False,
                key='show_headroom_table',
                help='Display and export the aggregated table behind the headroom plot.',
            )
            if show_headroom_table:
                st.dataframe(headroom_table_df, use_container_width=True)
                st.download_button(
                    'Download headroom table (CSV)',
                    data=headroom_table_df.to_csv(index=False).encode('utf-8'),
                    file_name='headroom_table.csv',
                    mime='text/csv',
                    key='download_headroom_table',
                )
            if st.button('Generate headroom summary', key='headroom_summary'):
                ordered = plot_df.sort_values('headroom_pct', ascending=False)
                top_row = ordered.iloc[0]
                bottom_row = ordered.iloc[-1]
                summary = HEADROOM_SUMMARY_TEMPLATE
                summary = summary.replace('[GROUP_BY]', group_by or 'Group')
                summary = summary.replace('[RECENT_PERIODS]', str(int(recent_periods_input)))
                summary = summary.replace('[HIGH_THRESHOLD]', f'{high_pct:.0f}')
                summary = summary.replace('[MED_THRESHOLD]', f'{med_pct:.0f}')
                summary = summary.replace('[TOP_GROUP]', str(top_row['group']))
                summary = summary.replace('[TOP_HEADROOM]', f"{float(top_row['headroom_pct']):.1f}")
                summary = summary.replace('[BOTTOM_GROUP]', str(bottom_row['group']))
                summary = summary.replace('[BOTTOM_HEADROOM]', f"{float(bottom_row['headroom_pct']):.1f}")
                summary = summary.replace('[TAKEAWAY]', 'headroom is concentrated in the leading groups')
                st.text_area('Headroom summary (copy for report)', summary, height=140)
            if st.button('Generate headroom report', key='headroom_report'):
                st.info('Coming soon...')

        st.subheader('Scale')
        with st.popover('What is this?'):
            st.write(
                'Measures scalable volume using average DCFS over the most recent periods, '
                'then ranks it within the channel. Higher scores mean this group is already '
                'performing strongly on volume and has more headroom to scale investment.'
            )
        scale_df = plot_df.dropna(subset=['scale_score'])
        if scale_df.empty:
            st.info('No scale score data for the current filters.')
        else:
            scale_df = scale_df.copy()
            scale_df['scale_score'] = pd.to_numeric(scale_df['scale_score'], errors='coerce')
            scale_df = scale_df.dropna(subset=['scale_score'])
            fig = px.bar(
                scale_df.sort_values('scale_score', ascending=False),
                x='group',
                y='scale_score',
                text='scale_score',
                color='group',
                labels={'scale_score': 'Scale score (0–100)', 'group': group_by or 'Group'},
                color_discrete_map=group_color_map,
            )
            fig.update_xaxes(categoryorder='array', categoryarray=group_order)
            fig.update_traces(
                texttemplate='%{text:.0f}',
                textposition='outside',
            )
            fig.add_hline(
                y=75,
                line_dash='dash',
                line_color='orange',
                annotation_text='Top quartile (75th pct)',
                annotation_position='top left',
            )
            fig.update_yaxes(title_text='Scale score (0–100)', range=[0, 110])
            st.plotly_chart(fig, use_container_width=True)
            if st.button('Generate scale report', key='scale_report'):
                st.info('Coming soon...')

        st.subheader('Media response curve')
        with st.popover('What is this?'):
            st.write(
                'Plots Media Spend vs. DCFS to visualize response curves and fitted saturation trends. '
                'Used to infer growth vs. saturation.'
            )
        curve_data = df.copy()
        if opp_markets and 'All' not in opp_markets:
            curve_data = curve_data[curve_data['Market'].isin(opp_markets)]
        if opp_channel != 'All':
            curve_data = curve_data[curve_data['Channel'] == opp_channel]
        if opp_model != 'All':
            curve_data = curve_data[curve_data['Model'] == opp_model]
        curve_data['Media Spend'] = pd.to_numeric(curve_data['Media Spend'], errors='coerce')
        curve_data['DCFS'] = pd.to_numeric(curve_data['DCFS'], errors='coerce')
        curve_data = curve_data[(curve_data['Media Spend'] > 0) & (curve_data['DCFS'] >= 0)]
        if curve_data.empty:
            st.info('No spend/DCFS data available for the current filters.')
        else:
            time_col = curve_group_by
            if not time_col:
                st.info('No time column available for curve aggregation.')
            else:
                plot_df = (
                    curve_data.groupby([time_col, group_by], dropna=False)
                    .agg({'Media Spend': 'sum', 'DCFS': 'sum'})
                    .reset_index()
                )
                curve_fig = go.Figure()
                left_col, right_col = st.columns([4, 1])
                with right_col:
                    available_groups = sorted(plot_df[group_by].dropna().astype(str).unique().tolist())
                    selected_groups = st.multiselect(
                        f'Show {group_by} curves',
                        options=available_groups,
                        default=available_groups,
                        key='curve_groups_filter',
                    )
                    show_fit_points = st.checkbox(
                        'Show fit points',
                        value=True,
                        key='curve_show_points',
                    )
                if selected_groups:
                    plot_df = plot_df[plot_df[group_by].astype(str).isin(selected_groups)]
                color_map = {}
                if group_order:
                    palette = px.colors.qualitative.Safe
                    for idx, group in enumerate(group_order):
                        color_map[str(group)] = palette[idx % len(palette)]
                fit_rows = []
                fit_params = {}
                alloc_rows = []
                max_alloc = None
                if np is None or curve_fit is None:
                    st.info('Install scipy to enable curve fitting for Ax/(b+x).')
                else:
                    for group_key, group in plot_df.groupby(group_by, dropna=False):
                        group_label = str(group_key)
                        a, b = fit_saturation(group['Media Spend'], group['DCFS'])
                        if a is None or b is None:
                            continue
                        if a <= 0 or b <= 0:
                            continue
                        fit_params[group_label] = (float(a), float(b))
                        fit_rows.append({
                            'group': group_label,
                            'A': a,
                            'B': b,
                            'points': len(group),
                        })
                        if show_fit_points:
                            curve_fig.add_trace(
                                go.Scatter(
                                    x=group['Media Spend'],
                                    y=group['DCFS'],
                                    mode='markers',
                                    name=f'{group_label} points',
                                    marker=dict(
                                        size=7,
                                        color=color_map.get(group_label),
                                        opacity=0.6,
                                    ),
                                    showlegend=False,
                                )
                            )
                if fit_params:
                    use_min_constraints = st.checkbox(
                        'Use minimum spend constraints',
                        value=True,
                        key='use_min_constraints',
                    )
                    use_headroom_weighting = st.checkbox(
                        'Weight by headroom',
                        value=False,
                        key='use_headroom_weighting',
                    )
                    headroom_lambda = st.slider(
                        'Headroom strength',
                        min_value=0.0,
                        max_value=1.0,
                        value=1.0,
                        step=0.05,
                        disabled=not use_headroom_weighting,
                    )
                    if not use_headroom_weighting:
                        headroom_lambda = 0.0
                    use_scale_weighting = st.checkbox(
                        'Weight by scale',
                        value=False,
                        key='use_scale_weighting',
                    )
                    scale_lambda = st.slider(
                        'Scale strength',
                        min_value=0.0,
                        max_value=1.0,
                        value=1.0,
                        step=0.05,
                        disabled=not use_scale_weighting,
                    )
                    if not use_scale_weighting:
                        scale_lambda = 0.0
                    use_spend_weighting = st.checkbox(
                        'Use spend curve',
                        value=True,
                        key='use_spend_weighting',
                    )
                    constraint_strength = st.slider(
                        'Constraint strength',
                        min_value=0.0,
                        max_value=1.0,
                        value=0.0,
                        step=0.05,
                        disabled=not use_spend_weighting,
                    )
                    if not use_spend_weighting:
                        constraint_strength = 0.0
                    reverse_funnel_blend = st.slider(
                        'Reverse funnel blend',
                        min_value=0.0,
                        max_value=1.0,
                        value=0.0,
                        step=0.05,
                        help='0 = use risk-aware allocation only, 1 = use reverse-funnel allocation only.',
                    )
                    reverse_funnel_paste = st.text_area(
                        'Paste reverse funnel % split (Market <tab> %)',
                        value='',
                        placeholder='Market\t% Split\nPCGB\t22.5%\nPD\t18.8%\n...',
                        help='Paste a two-column list. Example: "PCGB<TAB>22.5%".',
                    )
                    hierarchical_market_channel = st.checkbox(
                        'Allocate market first, then split within each market by channel',
                        value=False,
                        key='hierarchical_market_channel',
                        help='Runs the market allocation first, then allocates each market budget across its channels.',
                    )

                    def _norm_label(label: str) -> str:
                        return ''.join(ch for ch in str(label).lower() if ch.isalnum())

                    def _parse_reverse_pct(text: str) -> dict:
                        out = {}
                        if not text:
                            return out
                        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                        if not lines:
                            return out
                        # Drop header line if it looks like it has words and a % label.
                        if 'split' in lines[0].lower() or 'market' in lines[0].lower():
                            lines = lines[1:]
                        for line in lines:
                            parts = [p for p in line.replace(',', ' ').split() if p]
                            if len(parts) < 2:
                                continue
                            pct_raw = parts[-1].strip()
                            name = ' '.join(parts[:-1]).strip()
                            pct_raw = pct_raw.replace('%', '')
                            try:
                                pct = float(pct_raw)
                            except ValueError:
                                continue
                            out[_norm_label(name)] = pct
                        return out

                    parsed_reverse = _parse_reverse_pct(reverse_funnel_paste)
                    reverse_overrides = {}
                    if parsed_reverse:
                        for g in fit_params.keys():
                            key = _norm_label(g)
                            if key in parsed_reverse:
                                reverse_overrides[g] = parsed_reverse[key]

                    st.subheader('Target DCFS to required spend (linear)')
                    if not time_col:
                        st.info('No time column available for weekly CPL selection.')
                    else:
                        week_values = (
                            curve_data[time_col]
                            .dropna()
                            .astype(str)
                            .unique()
                            .tolist()
                        )
                        week_values = sorted(week_values)
                        week_options = ['All'] + week_values
                        selected_weeks = st.multiselect(
                            'Weeks to average CPL',
                            week_options,
                            default=['All'],
                            key='required_spend_weeks',
                        )
                        if 'All' in selected_weeks or not selected_weeks:
                            selected_weeks = week_values
                        if not selected_weeks:
                            st.info('Select at least one week to compute CPL averages.')
                        else:
                            weekly = (
                                curve_data.groupby([time_col, group_by], dropna=False)
                                .agg({'Media Spend': 'sum', 'DCFS': 'sum'})
                                .reset_index()
                            )
                            weekly = weekly[weekly[time_col].astype(str).isin(selected_weeks)]
                            weekly['cpl'] = weekly.apply(
                                lambda r: (r['Media Spend'] / r['DCFS']) if r['DCFS'] and r['DCFS'] > 0 else None,
                                axis=1,
                            )
                            avg_cpl = (
                                weekly.groupby(group_by, dropna=False)['cpl']
                                .mean()
                                .to_dict()
                            )
                            global_target_dcfs = st.number_input(
                                'Global target DCFS (apply to all markets)',
                                min_value=0.0,
                                value=0.0,
                                step=1.0,
                                format='%.2f',
                                key='global_target_dcfs',
                            )
                            targets_df = pd.DataFrame({
                                'group': list(fit_params.keys()),
                                'avg_cpl_selected_weeks': [avg_cpl.get(g, None) for g in fit_params.keys()],
                                'target_dcfs': [global_target_dcfs] * len(fit_params),
                            })
                            targets_df = st.data_editor(
                                targets_df,
                                use_container_width=True,
                                num_rows='fixed',
                                key='target_dcfs_by_group',
                            )
                            targets_df['required_spend'] = targets_df.apply(
                                lambda r: (r['avg_cpl_selected_weeks'] * r['target_dcfs'])
                                if r['avg_cpl_selected_weeks'] is not None
                                and pd.notna(r['avg_cpl_selected_weeks'])
                                else None,
                                axis=1,
                            )
                            def _invert_curve(target_y: float, params):
                                if target_y is None or pd.isna(target_y):
                                    return None
                                a, b = params
                                if a <= 0 or b <= 0:
                                    return None
                                if target_y <= 0:
                                    return 0.0
                                if target_y >= a:
                                    return None
                                return (b * target_y) / (a - target_y)

                            targets_df['required_spend_curve'] = targets_df.apply(
                                lambda r: _invert_curve(
                                    r['target_dcfs'],
                                    fit_params.get(r['group'], (None, None)),
                                ),
                                axis=1,
                            )
                            st.dataframe(
                                targets_df.sort_values('required_spend', ascending=False, na_position='last'),
                                use_container_width=True,
                            )
                            plot_df_req = targets_df.dropna(subset=['required_spend'])
                            if not plot_df_req.empty:
                                plot_melt = plot_df_req[['group', 'required_spend', 'required_spend_curve']].copy()
                                plot_melt = plot_melt.melt(
                                    id_vars=['group'],
                                    value_vars=['required_spend', 'required_spend_curve'],
                                    var_name='method',
                                    value_name='required_spend_value',
                                )
                                plot_melt['method'] = plot_melt['method'].map({
                                    'required_spend': 'Linear (avg CPL)',
                                    'required_spend_curve': 'Curve-based',
                                })
                                req_fig = px.bar(
                                    plot_melt,
                                    x='group',
                                    y='required_spend_value',
                                    color='method',
                                    barmode='group',
                                    title='Required spend for target DCFS (linear vs curve-based)',
                                )
                                req_fig.update_layout(
                                    xaxis_title=group_by or 'Group',
                                    yaxis_title='Required spend',
                                    legend_title_text='Method',
                                )
                                st.plotly_chart(req_fig, use_container_width=True)
                            required_spend_map = {
                                str(row['group']): float(row['required_spend'])
                                for _, row in targets_df.dropna(subset=['required_spend']).iterrows()
                            }
                            st.session_state['required_spend_map'] = required_spend_map

                    current_groups = list(fit_params.keys())
                    min_df = pd.DataFrame({
                        'group': current_groups,
                        'min_spend': [500.0] * len(fit_params),
                        'max_spend': [0.0] * len(fit_params),
                        'reverse_funnel_pct': [100.0 / len(fit_params)] * len(fit_params),
                    })
                    if 'min_spend_per_curve_data' in st.session_state:
                        prior_df = st.session_state['min_spend_per_curve_data']
                        prior_groups = prior_df['group'].tolist() if 'group' in prior_df.columns else []
                        if set(prior_groups) == set(current_groups):
                            min_df = prior_df
                    if st.button('Use required spend as min budget', key='apply_required_min'):
                        required_spend_map = st.session_state.get('required_spend_map', {})
                        if required_spend_map:
                            min_df['min_spend'] = min_df['group'].map(
                                lambda g: required_spend_map.get(str(g), min_df.loc[min_df['group'] == g, 'min_spend'].iloc[0])
                            )
                            st.session_state['min_spend_per_curve_data'] = min_df
                        else:
                            st.warning('No required spend values available. Set target DCFS first.')
                    if reverse_overrides:
                        min_df['reverse_funnel_pct'] = min_df['group'].map(
                            lambda g: reverse_overrides.get(g, min_df.loc[min_df['group'] == g, 'reverse_funnel_pct'].iloc[0])
                        )
                    col_min, col_max = st.columns(2)
                    with col_min:
                        apply_min_all = st.number_input(
                            'Apply min to all',
                            min_value=0.0,
                            value=0.0,
                            step=100.0,
                            format='%.2f',
                            key='apply_min_all',
                        )
                    with col_max:
                        apply_max_all = st.number_input(
                            'Apply max to all',
                            min_value=0.0,
                            value=0.0,
                            step=100.0,
                            format='%.2f',
                            key='apply_max_all',
                        )
                    if apply_min_all > 0:
                        min_df['min_spend'] = apply_min_all
                    if apply_max_all > 0:
                        min_df['max_spend'] = apply_max_all
                    min_df = st.data_editor(
                        min_df,
                        use_container_width=True,
                        num_rows='fixed',
                        key='min_spend_per_curve_editor',
                        disabled=False,
                    )
                    st.session_state['min_spend_per_curve_data'] = min_df
                    min_total = float(min_df['min_spend'].fillna(0.0).sum())
                    if 'max_budget' not in st.session_state:
                        st.session_state['max_budget'] = min_total
                    budget = st.number_input(
                        'Max budget (total Media Spend)',
                        min_value=0.0,
                        value=float(st.session_state.get('max_budget', min_total)),
                        step=1000.0,
                        format='%.2f',
                        key='max_budget',
                    )
                    if budget < min_total:
                        st.error(
                            f"Max budget is below total minimum spend "
                            f"({min_total:,.2f}). Increase budget or reduce minimums."
                        )

                    def _allocate_from_fit_params(
                        local_fit_params,
                        local_budget,
                        local_weights=None,
                        local_min_map=None,
                        local_max_map=None,
                        use_constraints=False,
                    ):
                        if not local_fit_params or local_budget <= 0:
                            return pd.DataFrame(), pd.DataFrame()

                        local_weights = local_weights or {}
                        local_min_map = local_min_map or {label: 0.0 for label in local_fit_params.keys()}
                        local_max_map = local_max_map or {label: 0.0 for label in local_fit_params.keys()}

                        max_d0 = max((a / b) for a, b in local_fit_params.values())

                        def _total_spend_for_lambda(lam):
                            total = 0.0
                            for a, b in local_fit_params.values():
                                x = (a * b / lam) ** 0.5 - b
                                if x < 0:
                                    x = 0.0
                                total += x
                            return total

                        unconstrained_rows = []
                        if max_d0 > 0:
                            low_lam = 1e-12
                            high_lam = max_d0
                            for _ in range(80):
                                mid = (low_lam + high_lam) / 2
                                total = _total_spend_for_lambda(mid)
                                if total > local_budget:
                                    low_lam = mid
                                else:
                                    high_lam = mid
                            lam = high_lam
                            for group_label, (a, b) in local_fit_params.items():
                                x = (a * b / lam) ** 0.5 - b
                                if x < 0:
                                    x = 0.0
                                unconstrained_rows.append({
                                    'group': group_label,
                                    'min_spend': 0.0,
                                    'allocated_spend': x,
                                    'allocated_pct': (x / local_budget * 100.0) if local_budget > 0 else 0.0,
                                    'expected_dcfs': _saturation_curve(x, a, b) if x > 0 else 0.0,
                                })
                        spend_df_local = pd.DataFrame(unconstrained_rows)
                        if not spend_df_local.empty:
                            total_alloc = spend_df_local['allocated_spend'].sum()
                            if total_alloc > 0 and local_budget > 0:
                                scale_factor = local_budget / total_alloc
                                spend_df_local['allocated_spend'] = spend_df_local['allocated_spend'] * scale_factor
                                spend_df_local['allocated_pct'] = spend_df_local['allocated_spend'] / local_budget * 100.0
                                spend_df_local['expected_dcfs'] = spend_df_local.apply(
                                    lambda r: _saturation_curve(r['allocated_spend'], *local_fit_params[r['group']])
                                    if r['allocated_spend'] > 0 else 0.0,
                                    axis=1,
                                )

                        spend_shares = {
                            row['group']: (row['allocated_spend'] / local_budget) if local_budget > 0 else 0.0
                            for _, row in spend_df_local.iterrows()
                        }
                        total_weight = sum(max(0.0, float(local_weights.get(label, 0.0))) for label in local_fit_params.keys())
                        if total_weight <= 0:
                            total_weight = float(len(local_fit_params))
                            local_weights = {label: 1.0 for label in local_fit_params.keys()}
                        driver_shares = {
                            label: max(0.0, float(local_weights.get(label, 0.0))) / total_weight
                            for label in local_fit_params.keys()
                        }
                        spend_lambda = 1.0 - float(constraint_strength)
                        spend_lambda = min(1.0, max(0.0, spend_lambda))
                        blended_shares = {
                            label: spend_lambda * spend_shares.get(label, 0.0) + (1.0 - spend_lambda) * driver_shares.get(label, 0.0)
                            for label in local_fit_params.keys()
                        }

                        alloc_rows_local = []
                        if use_constraints:
                            local_min_total = sum(local_min_map.get(k, 0.0) for k in local_fit_params.keys())
                            remaining = local_budget - local_min_total
                            if remaining < 0:
                                return spend_df_local, pd.DataFrame()
                            adjustable = [k for k in local_fit_params.keys() if blended_shares.get(k, 0.0) > 0]
                            share_sum = sum(blended_shares.get(k, 0.0) for k in adjustable)
                            for label, (a, b) in local_fit_params.items():
                                base = remaining * (blended_shares[label] / share_sum) if share_sum > 0 and label in adjustable else 0.0
                                x = local_min_map.get(label, 0.0) + base
                                alloc_rows_local.append({
                                    'group': label,
                                    'min_spend': local_min_map.get(label, 0.0),
                                    'allocated_spend': x,
                                    'allocated_pct': (x / local_budget * 100.0) if local_budget > 0 else 0.0,
                                    'expected_dcfs': _saturation_curve(x, a, b) if x > 0 else 0.0,
                                })
                            alloc = {row['group']: row['allocated_spend'] for row in alloc_rows_local}
                            caps = {
                                k: (local_max_map.get(k, 0.0) if local_max_map.get(k, 0.0) > 0 else float('inf'))
                                for k in alloc.keys()
                            }
                            shares = {k: blended_shares.get(k, 0.0) for k in alloc.keys()}
                            while True:
                                overflow = 0.0
                                capped = set()
                                for k, v in alloc.items():
                                    cap = caps.get(k, float('inf'))
                                    if v > cap:
                                        overflow += v - cap
                                        alloc[k] = cap
                                        capped.add(k)
                                if overflow <= 1e-6:
                                    break
                                candidates = [k for k in alloc.keys() if k not in capped and caps.get(k, float('inf')) > alloc[k]]
                                if not candidates:
                                    break
                                share_sum = sum(shares.get(k, 0.0) for k in candidates)
                                if share_sum <= 0:
                                    break
                                for k in candidates:
                                    alloc[k] += overflow * (shares.get(k, 0.0) / share_sum)
                            alloc_rows_local = []
                            for label, (a, b) in local_fit_params.items():
                                x = float(alloc.get(label, 0.0))
                                alloc_rows_local.append({
                                    'group': label,
                                    'min_spend': local_min_map.get(label, 0.0),
                                    'allocated_spend': x,
                                    'allocated_pct': (x / local_budget * 100.0) if local_budget > 0 else 0.0,
                                    'expected_dcfs': _saturation_curve(x, a, b) if x > 0 else 0.0,
                                })
                        else:
                            for label, (a, b) in local_fit_params.items():
                                x = local_budget * blended_shares.get(label, 0.0)
                                alloc_rows_local.append({
                                    'group': label,
                                    'min_spend': 0.0,
                                    'allocated_spend': x,
                                    'allocated_pct': (x / local_budget * 100.0) if local_budget > 0 else 0.0,
                                    'expected_dcfs': _saturation_curve(x, a, b) if x > 0 else 0.0,
                                })
                        return spend_df_local, pd.DataFrame(alloc_rows_local)

                    def _channel_sort_order(channel_label):
                        channel_text = str(channel_label).strip().lower()
                        if 'search' in channel_text:
                            return 0
                        if 'social' in channel_text:
                            return 1
                        return 2

                    def _build_hierarchical_detail_table(source_df, spend_col, dcfs_col):
                        if (
                            not hierarchical_market_channel
                            or source_df is None
                            or source_df.empty
                            or 'Market' not in curve_data.columns
                            or 'Channel' not in curve_data.columns
                        ):
                            return pd.DataFrame()
                        detail_rows = []
                        total_budget_value = float(budget) if budget > 0 else 0.0
                        for _, market_row in source_df.iterrows():
                            market_label = str(market_row['group'])
                            market_budget = float(market_row.get(spend_col, 0.0) or 0.0)
                            if market_budget <= 0:
                                continue
                            market_curve = curve_data[curve_data['Market'].astype(str) == market_label].copy()
                            if market_curve.empty:
                                continue
                            channel_summary = (
                                market_curve.groupby('Channel', dropna=False)
                                .agg({'Media Spend': 'sum', 'DCFS': 'sum'})
                                .reset_index()
                            )
                            if channel_summary.empty:
                                continue
                            total_channel_dcfs = float(channel_summary['DCFS'].sum())
                            total_channel_spend = float(channel_summary['Media Spend'].sum())
                            if total_channel_dcfs > 0:
                                channel_summary['channel_share'] = channel_summary['DCFS'] / total_channel_dcfs
                            elif total_channel_spend > 0:
                                channel_summary['channel_share'] = channel_summary['Media Spend'] / total_channel_spend
                            else:
                                channel_summary['channel_share'] = 1.0 / len(channel_summary)
                            channel_summary['estimated_efficiency'] = channel_summary.apply(
                                lambda r: (float(r['DCFS']) / float(r['Media Spend']))
                                if pd.notna(r['Media Spend']) and float(r['Media Spend']) > 0
                                else 0.0,
                                axis=1,
                            )
                            for _, channel_row in channel_summary.iterrows():
                                allocated_spend = market_budget * float(channel_row['channel_share'])
                                detail_rows.append({
                                    'market': market_label,
                                    'channel': str(channel_row['Channel']),
                                    spend_col: allocated_spend,
                                    f'{spend_col}_pct_total': (allocated_spend / total_budget_value * 100.0) if total_budget_value > 0 else 0.0,
                                    f'{spend_col}_pct_within_market': (allocated_spend / market_budget * 100.0) if market_budget > 0 else 0.0,
                                    dcfs_col: allocated_spend * float(channel_row['estimated_efficiency']),
                                })
                        detail_df = pd.DataFrame(detail_rows)
                        if not detail_df.empty and 'channel' in detail_df.columns:
                            detail_df['_channel_sort'] = detail_df['channel'].map(_channel_sort_order)
                            detail_df = detail_df.sort_values(['market', '_channel_sort', 'channel'], na_position='last')
                            detail_df = detail_df.drop(columns=['_channel_sort'])
                        return detail_df
                    run_allocation = st.button('Run allocation', key='run_allocation')
                    if run_allocation:
                        if hierarchical_market_channel and ('Market' not in curve_data.columns or 'Channel' not in curve_data.columns):
                            st.warning('Hierarchical market-to-channel allocation requires Market and Channel columns.')
                            st.stop()
                        if hierarchical_market_channel and group_by != 'Market':
                            st.warning('Hierarchical allocation requires the page grouping to be set to Market.')
                            st.stop()
                        if use_min_constraints:
                            min_map = {
                                str(row['group']): float(row['min_spend'] or 0.0)
                                for _, row in min_df.iterrows()
                            }
                            max_map = {
                                str(row['group']): float(row.get('max_spend') or 0.0)
                                for _, row in min_df.iterrows()
                            }
                            reverse_pct_map = {
                                str(row['group']): float(row.get('reverse_funnel_pct') or 0.0)
                                for _, row in min_df.iterrows()
                            }
                        else:
                            min_map = {label: 0.0 for label in fit_params.keys()}
                            max_map = {label: 0.0 for label in fit_params.keys()}
                            reverse_pct_map = {
                                str(row['group']): float(row.get('reverse_funnel_pct') or 0.0)
                                for _, row in min_df.iterrows()
                            }
                        min_map_unconstrained = {label: 0.0 for label in fit_params.keys()}
                        if budget <= 0:
                            st.warning('Enter a max budget greater than 0.')
                            st.stop()
                        min_total = sum(min_map.get(k, 0.0) for k in fit_params.keys())
                        if min_total > budget:
                            st.warning('Total minimum spend exceeds the max budget.')
                            st.stop()

                        headroom_map = {}
                        if group_by and 'headroom' in results.columns:
                            headroom_map = (
                                results.groupby(group_by, dropna=False)['headroom']
                                .mean()
                                .to_dict()
                            )

                        scale_map = {}
                        if group_by and 'scale_score' in results.columns:
                            scale_map = (
                                results.groupby(group_by, dropna=False)['scale_score']
                                .mean()
                                .to_dict()
                            )
                        weights = {}
                        for label in fit_params.keys():
                            h = headroom_map.get(label, 0.0)
                            if h is None or pd.isna(h):
                                h = 0.0
                            s = scale_map.get(label, 0.0)
                            if s is None or pd.isna(s):
                                s = 0.0
                            s = float(s) / 100.0
                            weights[label] = max(
                                0.0,
                                (headroom_lambda if use_headroom_weighting else 0.0) * float(h)
                                + (scale_lambda if use_scale_weighting else 0.0) * s,
                            )

                        alloc_df_unconstrained, alloc_df_constrained = _allocate_from_fit_params(
                            fit_params,
                            float(budget),
                            local_weights=weights,
                            local_min_map=min_map,
                            local_max_map=max_map,
                            use_constraints=use_min_constraints,
                        )

                        reverse_total = sum(max(0.0, v) for v in reverse_pct_map.values())
                        if reverse_total <= 0:
                            reverse_shares = {k: 1.0 / len(fit_params) for k in fit_params.keys()}
                        else:
                            reverse_shares = {
                                k: max(0.0, reverse_pct_map.get(k, 0.0)) / reverse_total
                                for k in fit_params.keys()
                            }

                        if alloc_df_unconstrained is not None and not alloc_df_unconstrained.empty:
                            reverse_alloc_unconstrained = alloc_df_unconstrained[['group']].copy()
                            reverse_alloc_unconstrained['reverse_alloc'] = reverse_alloc_unconstrained['group'].map(
                                lambda g: budget * reverse_shares.get(g, 0.0)
                            )
                            reverse_alloc_unconstrained = reverse_alloc_unconstrained.set_index('group')
                        else:
                            reverse_alloc_unconstrained = None

                        if alloc_df_constrained is not None and not alloc_df_constrained.empty:
                            reverse_alloc_constrained = alloc_df_constrained[['group']].copy()
                            reverse_alloc_constrained['reverse_alloc'] = reverse_alloc_constrained['group'].map(
                                lambda g: budget * reverse_shares.get(g, 0.0)
                            )
                            reverse_alloc_constrained = reverse_alloc_constrained.set_index('group')
                        else:
                            reverse_alloc_constrained = None

                        if alloc_df_unconstrained is not None and not alloc_df_unconstrained.empty:
                            alloc_df_unconstrained = alloc_df_unconstrained.copy()
                            alloc_df_unconstrained['allocated_with_reverse'] = alloc_df_unconstrained['group'].apply(
                                lambda g: (
                                    (1.0 - reverse_funnel_blend) * float(
                                        alloc_df_unconstrained.loc[alloc_df_unconstrained['group'] == g, 'allocated_spend'].iloc[0]
                                    )
                                    + reverse_funnel_blend * float(reverse_alloc_unconstrained.loc[g, 'reverse_alloc'])
                                )
                            )
                            alloc_df_unconstrained['pct_with_reverse'] = (
                                alloc_df_unconstrained['allocated_with_reverse'] / budget * 100.0
                                if budget > 0 else 0.0
                            )
                            alloc_df_unconstrained['dcfs_with_reverse'] = alloc_df_unconstrained.apply(
                                lambda r: _saturation_curve(
                                    r['allocated_with_reverse'],
                                    *fit_params[r['group']],
                                ) if r['allocated_with_reverse'] > 0 else 0.0,
                                axis=1,
                            )

                        if alloc_df_constrained is not None and not alloc_df_constrained.empty:
                            alloc_df_constrained = alloc_df_constrained.copy()
                            alloc_df_constrained['allocated_with_reverse'] = alloc_df_constrained['group'].apply(
                                lambda g: (
                                    (1.0 - reverse_funnel_blend) * float(
                                        alloc_df_constrained.loc[alloc_df_constrained['group'] == g, 'allocated_spend'].iloc[0]
                                    )
                                    + reverse_funnel_blend * float(reverse_alloc_constrained.loc[g, 'reverse_alloc'])
                                )
                            )
                            alloc_df_constrained['pct_with_reverse'] = (
                                alloc_df_constrained['allocated_with_reverse'] / budget * 100.0
                                if budget > 0 else 0.0
                            )
                            alloc_df_constrained['dcfs_with_reverse'] = alloc_df_constrained.apply(
                                lambda r: _saturation_curve(
                                    r['allocated_with_reverse'],
                                    *fit_params[r['group']],
                                ) if r['allocated_with_reverse'] > 0 else 0.0,
                                axis=1,
                            )
                        detailed_alloc_df_unconstrained = _build_hierarchical_detail_table(
                            alloc_df_unconstrained,
                            'allocated_spend',
                            'expected_dcfs',
                        ).rename(columns={
                            'allocated_spend': 'allocated_without_constraint',
                            'allocated_spend_pct_total': 'allocated_without_constraint_pct_total',
                            'allocated_spend_pct_within_market': 'allocated_without_constraint_pct_within_market',
                            'expected_dcfs': 'dcfs_without_constraint',
                        })
                        detailed_alloc_df_constrained = _build_hierarchical_detail_table(
                            alloc_df_constrained,
                            'allocated_spend',
                            'expected_dcfs',
                        ).rename(columns={
                            'allocated_spend': 'allocated_with_constraint',
                            'allocated_spend_pct_total': 'allocated_with_constraint_pct_total',
                            'allocated_spend_pct_within_market': 'allocated_with_constraint_pct_within_market',
                            'expected_dcfs': 'dcfs_with_constraint',
                        })
                        detailed_alloc_df_blended_unconstrained = _build_hierarchical_detail_table(
                            alloc_df_unconstrained,
                            'allocated_with_reverse',
                            'dcfs_with_reverse',
                        ).rename(columns={
                            'allocated_with_reverse': 'blended_alloc_without_constraint',
                            'allocated_with_reverse_pct_total': 'blended_alloc_without_constraint_pct_total',
                            'allocated_with_reverse_pct_within_market': 'blended_alloc_without_constraint_pct_within_market',
                            'dcfs_with_reverse': 'blended_dcfs_without_constraint',
                        })
                        detailed_alloc_df_blended_constrained = _build_hierarchical_detail_table(
                            alloc_df_constrained,
                            'allocated_with_reverse',
                            'dcfs_with_reverse',
                        ).rename(columns={
                            'allocated_with_reverse': 'blended_alloc_with_constraint',
                            'allocated_with_reverse_pct_total': 'blended_alloc_with_constraint_pct_total',
                            'allocated_with_reverse_pct_within_market': 'blended_alloc_with_constraint_pct_within_market',
                            'dcfs_with_reverse': 'blended_dcfs_with_constraint',
                        })
                        st.session_state['alloc_state'] = {
                            'alloc_df_unconstrained': alloc_df_unconstrained,
                            'alloc_df_constrained': alloc_df_constrained,
                            'detailed_alloc_df_unconstrained': detailed_alloc_df_unconstrained,
                            'detailed_alloc_df_constrained': detailed_alloc_df_constrained,
                            'detailed_alloc_df_blended_unconstrained': detailed_alloc_df_blended_unconstrained,
                            'detailed_alloc_df_blended_constrained': detailed_alloc_df_blended_constrained,
                            'hierarchical_market_channel': hierarchical_market_channel,
                            'use_min_constraints': use_min_constraints,
                            'use_headroom_weighting': use_headroom_weighting,
                            'use_scale_weighting': use_scale_weighting,
                            'use_spend_weighting': use_spend_weighting,
                            'headroom_strength': headroom_lambda,
                            'scale_strength': scale_lambda,
                            'constraint_strength': constraint_strength,
                            'reverse_funnel_blend': reverse_funnel_blend,
                            'budget': budget,
                            'min_map': min_map,
                            'max_map': max_map,
                            'reverse_shares': reverse_shares,
                            'max_enabled': any(v > 0 for v in max_map.values()),
                            'group_by': group_by or 'Group',
                            'curve_group_by': curve_group_by or 'N/A',
                            'filters': {
                                'markets': opp_markets if isinstance(opp_markets, list) else [opp_markets],
                                'channel': opp_channel,
                                'model': opp_model,
                                'campaign': opp_campaign,
                            },
                        }

                    alloc_state = st.session_state.get('alloc_state')
                    if alloc_state:
                        alloc_df_unconstrained = alloc_state.get('alloc_df_unconstrained')
                        alloc_df_constrained = alloc_state.get('alloc_df_constrained')
                        detailed_alloc_df_unconstrained = alloc_state.get('detailed_alloc_df_unconstrained')
                        detailed_alloc_df_constrained = alloc_state.get('detailed_alloc_df_constrained')
                        detailed_alloc_df_blended_unconstrained = alloc_state.get('detailed_alloc_df_blended_unconstrained')
                        detailed_alloc_df_blended_constrained = alloc_state.get('detailed_alloc_df_blended_constrained')
                        hierarchical_market_channel = bool(alloc_state.get('hierarchical_market_channel'))
                        st.subheader('Optimal budget split')
                        if alloc_df_unconstrained is None or alloc_df_unconstrained.empty:
                            st.info('No allocation available.')
                        else:
                            display_compare_df = None
                            total_dcfs_unconstrained = float(alloc_df_unconstrained['expected_dcfs'].sum())
                            total_dcfs_constrained = (
                                float(alloc_df_constrained['expected_dcfs'].sum())
                                if alloc_df_constrained is not None and not alloc_df_constrained.empty
                                else None
                            )
                            total_dcfs_blended_unconstrained = float(
                                alloc_df_unconstrained['dcfs_with_reverse'].sum()
                            ) if 'dcfs_with_reverse' in alloc_df_unconstrained.columns else None
                            total_dcfs_blended_constrained = (
                                float(alloc_df_constrained['dcfs_with_reverse'].sum())
                                if alloc_df_constrained is not None
                                and not alloc_df_constrained.empty
                                and 'dcfs_with_reverse' in alloc_df_constrained.columns
                                else None
                            )
                            if hierarchical_market_channel and detailed_alloc_df_unconstrained is not None and not detailed_alloc_df_unconstrained.empty:
                                total_dcfs_unconstrained = float(detailed_alloc_df_unconstrained['dcfs_without_constraint'].sum()) if 'dcfs_without_constraint' in detailed_alloc_df_unconstrained.columns else total_dcfs_unconstrained
                                if detailed_alloc_df_constrained is not None and not detailed_alloc_df_constrained.empty and 'dcfs_with_constraint' in detailed_alloc_df_constrained.columns:
                                    total_dcfs_constrained = float(detailed_alloc_df_constrained['dcfs_with_constraint'].sum())
                                if detailed_alloc_df_blended_unconstrained is not None and not detailed_alloc_df_blended_unconstrained.empty and 'blended_dcfs_without_constraint' in detailed_alloc_df_blended_unconstrained.columns:
                                    total_dcfs_blended_unconstrained = float(detailed_alloc_df_blended_unconstrained['blended_dcfs_without_constraint'].sum())
                                if detailed_alloc_df_blended_constrained is not None and not detailed_alloc_df_blended_constrained.empty and 'blended_dcfs_with_constraint' in detailed_alloc_df_blended_constrained.columns:
                                    total_dcfs_blended_constrained = float(detailed_alloc_df_blended_constrained['blended_dcfs_with_constraint'].sum())

                            c1, c2, c3, c4 = st.columns(4)
                            c1.metric('Total DCFS (without constraints)', f'{total_dcfs_unconstrained:,.2f}')
                            if total_dcfs_constrained is not None:
                                c2.metric('Total DCFS (with constraints)', f'{total_dcfs_constrained:,.2f}')
                            else:
                                c2.metric('Total DCFS (with constraints)', 'n/a')
                            if total_dcfs_blended_unconstrained is not None:
                                c3.metric('Total DCFS (blended, no constraints)', f'{total_dcfs_blended_unconstrained:,.2f}')
                            else:
                                c3.metric('Total DCFS (blended, no constraints)', 'n/a')
                            if total_dcfs_blended_constrained is not None:
                                c4.metric('Total DCFS (blended, with constraints)', f'{total_dcfs_blended_constrained:,.2f}')
                            else:
                                c4.metric('Total DCFS (blended, with constraints)', 'n/a')
                            if hierarchical_market_channel or alloc_state.get('use_min_constraints') or alloc_state.get('use_headroom_weighting') or alloc_state.get('use_scale_weighting') or alloc_state.get('use_spend_weighting'):
                                min_map = alloc_state.get('min_map', {})
                                max_map = alloc_state.get('max_map', {})
                                if hierarchical_market_channel and detailed_alloc_df_unconstrained is not None and not detailed_alloc_df_unconstrained.empty:
                                    left = detailed_alloc_df_unconstrained.copy()
                                    if detailed_alloc_df_constrained is not None and not detailed_alloc_df_constrained.empty:
                                        right = detailed_alloc_df_constrained.copy()
                                    else:
                                        right = pd.DataFrame({
                                            'market': left['market'],
                                            'channel': left['channel'],
                                            'allocated_with_constraint': pd.NA,
                                            'allocated_with_constraint_pct_total': pd.NA,
                                            'allocated_with_constraint_pct_within_market': pd.NA,
                                            'dcfs_with_constraint': pd.NA,
                                        })
                                    if detailed_alloc_df_blended_unconstrained is not None and not detailed_alloc_df_blended_unconstrained.empty:
                                        blend_left = detailed_alloc_df_blended_unconstrained.copy()
                                    else:
                                        blend_left = pd.DataFrame({
                                            'market': left['market'],
                                            'channel': left['channel'],
                                            'blended_alloc_without_constraint': pd.NA,
                                            'blended_alloc_without_constraint_pct_total': pd.NA,
                                            'blended_alloc_without_constraint_pct_within_market': pd.NA,
                                            'blended_dcfs_without_constraint': pd.NA,
                                        })
                                    if detailed_alloc_df_blended_constrained is not None and not detailed_alloc_df_blended_constrained.empty:
                                        blend_right = detailed_alloc_df_blended_constrained.copy()
                                    else:
                                        blend_right = pd.DataFrame({
                                            'market': left['market'],
                                            'channel': left['channel'],
                                            'blended_alloc_with_constraint': pd.NA,
                                            'blended_alloc_with_constraint_pct_total': pd.NA,
                                            'blended_alloc_with_constraint_pct_within_market': pd.NA,
                                            'blended_dcfs_with_constraint': pd.NA,
                                        })
                                    compare_df = left.merge(right, on=['market', 'channel'], how='outer')
                                    compare_df = compare_df.merge(blend_left, on=['market', 'channel'], how='outer')
                                    compare_df = compare_df.merge(blend_right, on=['market', 'channel'], how='outer')
                                    compare_df['_channel_sort'] = compare_df['channel'].map(_channel_sort_order)
                                    display_compare_df = compare_df.sort_values(
                                        ['market', '_channel_sort', 'allocated_with_constraint', 'channel'],
                                        ascending=[True, True, False, True],
                                        na_position='last',
                                    )
                                    display_compare_df = display_compare_df.drop(columns=['_channel_sort'])
                                else:
                                    left = alloc_df_unconstrained.rename(columns={
                                        'allocated_spend': 'allocated_without_constraint',
                                        'allocated_pct': 'pct_without_constraint',
                                        'expected_dcfs': 'dcfs_without_constraint',
                                        'allocated_with_reverse': 'blended_alloc_without_constraint',
                                        'pct_with_reverse': 'blended_pct_without_constraint',
                                        'dcfs_with_reverse': 'blended_dcfs_without_constraint',
                                    })[['group', 'allocated_without_constraint', 'pct_without_constraint', 'dcfs_without_constraint', 'blended_alloc_without_constraint', 'blended_pct_without_constraint', 'blended_dcfs_without_constraint']]
                                    left['min_spend'] = left['group'].map(lambda g: min_map.get(g, 0.0))
                                    left['max_spend'] = left['group'].map(lambda g: max_map.get(g, 0.0))
                                    left = left[['group', 'min_spend', 'max_spend', 'allocated_without_constraint', 'pct_without_constraint', 'dcfs_without_constraint', 'blended_alloc_without_constraint', 'blended_pct_without_constraint', 'blended_dcfs_without_constraint']]
                                    if alloc_df_constrained is not None and not alloc_df_constrained.empty:
                                        right = alloc_df_constrained.rename(columns={
                                            'allocated_spend': 'allocated_with_constraint',
                                            'allocated_pct': 'pct_with_constraint',
                                            'expected_dcfs': 'dcfs_with_constraint',
                                            'allocated_with_reverse': 'blended_alloc_with_constraint',
                                            'pct_with_reverse': 'blended_pct_with_constraint',
                                            'dcfs_with_reverse': 'blended_dcfs_with_constraint',
                                        })[['group', 'allocated_with_constraint', 'pct_with_constraint', 'dcfs_with_constraint', 'blended_alloc_with_constraint', 'blended_pct_with_constraint', 'blended_dcfs_with_constraint']]
                                    else:
                                        right = pd.DataFrame({
                                            'group': left['group'],
                                            'allocated_with_constraint': pd.NA,
                                            'pct_with_constraint': pd.NA,
                                            'dcfs_with_constraint': pd.NA,
                                            'blended_alloc_with_constraint': pd.NA,
                                            'blended_pct_with_constraint': pd.NA,
                                            'blended_dcfs_with_constraint': pd.NA,
                                        })
                                    compare_df = left.merge(right, on='group', how='outer')
                                    display_compare_df = compare_df.sort_values('allocated_with_constraint', ascending=False, na_position='last')
                                st.dataframe(
                                    display_compare_df,
                                    use_container_width=True,
                                )
                            else:
                                st.dataframe(
                                    alloc_df_unconstrained.sort_values('allocated_spend', ascending=False),
                                    use_container_width=True,
                                )

                            if st.button('Generate allocation narrative', key='alloc_narrative'):
                                filters = alloc_state.get('filters', {})
                                markets = filters.get('markets') or []
                                channels = []
                                models = []
                                campaigns = []
                                if 'Market' in df.columns:
                                    channels = (
                                        df['Channel'].dropna().unique().tolist() if 'Channel' in df.columns else []
                                    )
                                    models = df['Model'].dropna().unique().tolist() if 'Model' in df.columns else []
                                    campaigns = (
                                        df['Campaign'].dropna().unique().tolist() if 'Campaign' in df.columns else []
                                    )
                                markets_text = ', '.join(map(str, markets)) if markets else ''
                                channels_text = ', '.join(sorted(map(str, channels))) if channels else ''
                                models_text = ', '.join(sorted(map(str, models))) if models else ''
                                campaigns_text = ', '.join(sorted(map(str, campaigns))) if campaigns else ''

                                table_df = display_compare_df
                                if table_df is None and alloc_df_unconstrained is not None and not alloc_df_unconstrained.empty:
                                    left = alloc_df_unconstrained.rename(columns={
                                        'allocated_spend': 'allocated_without_constraint',
                                        'allocated_pct': 'pct_without_constraint',
                                        'expected_dcfs': 'dcfs_without_constraint',
                                        'allocated_with_reverse': 'blended_alloc_without_constraint',
                                        'pct_with_reverse': 'blended_pct_without_constraint',
                                        'dcfs_with_reverse': 'blended_dcfs_without_constraint',
                                    })[['group', 'allocated_without_constraint', 'pct_without_constraint', 'dcfs_without_constraint', 'blended_alloc_without_constraint', 'blended_pct_without_constraint', 'blended_dcfs_without_constraint']]
                                    if alloc_df_constrained is not None and not alloc_df_constrained.empty:
                                        right = alloc_df_constrained.rename(columns={
                                            'allocated_spend': 'allocated_with_constraint',
                                            'allocated_pct': 'pct_with_constraint',
                                            'expected_dcfs': 'dcfs_with_constraint',
                                            'allocated_with_reverse': 'blended_alloc_with_constraint',
                                            'pct_with_reverse': 'blended_pct_with_constraint',
                                            'dcfs_with_reverse': 'blended_dcfs_with_constraint',
                                        })[['group', 'allocated_with_constraint', 'pct_with_constraint', 'dcfs_with_constraint', 'blended_alloc_with_constraint', 'blended_pct_with_constraint', 'blended_dcfs_with_constraint']]
                                        table_df = left.merge(right, on='group', how='outer')
                                    else:
                                        table_df = left
                                min_map = alloc_state.get('min_map', {})
                                max_map = alloc_state.get('max_map', {})
                                allocation_lines = []
                                if table_df is not None:
                                    if hierarchical_market_channel and 'market' in table_df.columns:
                                        table_df = table_df.copy()
                                        table_df['_channel_sort'] = table_df['channel'].map(_channel_sort_order)
                                        iter_df = table_df.sort_values(['market', '_channel_sort', 'channel'], na_position='last')
                                    else:
                                        iter_df = table_df.sort_values(['group'])
                                    for _, row in iter_df.iterrows():
                                        row_label = (
                                            f"{row['market']} | {row['channel']}"
                                            if hierarchical_market_channel and 'market' in row and 'channel' in row
                                            else row['group']
                                        )
                                        min_val = min_map.get(row['group'], 0.0) if 'group' in row else 0.0
                                        max_val = max_map.get(row['group'], 0.0) if 'group' in row else 0.0
                                        def _fmt(val):
                                            if val is None or (isinstance(val, float) and pd.isna(val)):
                                                return 'n/a'
                                            return f"{float(val):.2f}"
                                        allocation_lines.append(
                                            f"{row_label}: "
                                            f"min={min_val:.2f}, max={max_val:.2f}; "
                                            f"unconstrained={_fmt(row.get('allocated_without_constraint'))} "
                                            f"({_fmt(row.get('pct_without_constraint', row.get('allocated_without_constraint_pct_total')))}%), "
                                            f"dcfs_unconstrained={_fmt(row.get('dcfs_without_constraint'))}; "
                                            f"constrained={_fmt(row.get('allocated_with_constraint'))} "
                                            f"({_fmt(row.get('pct_with_constraint', row.get('allocated_with_constraint_pct_total')))}%), "
                                            f"dcfs_constrained={_fmt(row.get('dcfs_with_constraint'))}; "
                                            f"blend_unconstrained={_fmt(row.get('blended_alloc_without_constraint'))} "
                                            f"({_fmt(row.get('blended_pct_without_constraint', row.get('blended_alloc_without_constraint_pct_total')))}%), "
                                            f"dcfs_blend_unconstrained={_fmt(row.get('blended_dcfs_without_constraint'))}; "
                                            f"blend_constrained={_fmt(row.get('blended_alloc_with_constraint'))} "
                                            f"({_fmt(row.get('blended_pct_with_constraint', row.get('blended_alloc_with_constraint_pct_total')))}%), "
                                            f"dcfs_blend_constrained={_fmt(row.get('blended_dcfs_with_constraint'))}"
                                        )
                                allocation_table_text = '\n'.join(allocation_lines) if allocation_lines else 'n/a'

                                template = ALLOCATION_METHOD_TEMPLATE
                                template = template.replace('[GROUP_BY]', str(alloc_state.get('group_by', 'Group')))
                                template = template.replace('[MARKETS]', markets_text)
                                template = template.replace('[CHANNELS]', channels_text)
                                template = template.replace('[MODELS]', models_text)
                                template = template.replace('[CAMPAIGNS]', campaigns_text)
                                template = template.replace('[CURVE_GROUP]', str(alloc_state.get('curve_group_by', 'N/A')))
                                template = template.replace('[HEADROOM_STRENGTH]', f"{float(alloc_state.get('headroom_strength', 0.0)):.2f}")
                                template = template.replace('[SCALE_STRENGTH]', f"{float(alloc_state.get('scale_strength', 0.0)):.2f}")
                                template = template.replace('[CONSTRAINT_STRENGTH]', f"{float(alloc_state.get('constraint_strength', 0.0)):.2f}")
                                template = template.replace('[REVERSE_BLEND]', f"{float(alloc_state.get('reverse_funnel_blend', 0.0)):.2f}")
                                template = template.replace('[MIN_CONSTRAINT_ENABLED]', 'Yes' if alloc_state.get('use_min_constraints') else 'No')
                                min_lines = []
                                for k in sorted(min_map.keys()):
                                    min_lines.append(f"{k}: {min_map.get(k, 0.0):.2f}")
                                max_lines = []
                                for k in sorted(max_map.keys()):
                                    max_lines.append(f"{k}: {max_map.get(k, 0.0):.2f}")
                                template = template.replace('[MIN_BY_GROUP]', ', '.join(min_lines) if min_lines else 'n/a')
                                template = template.replace('[MAX_CONSTRAINT_ENABLED]', 'Yes' if alloc_state.get('max_enabled') else 'No')
                                template = template.replace('[MAX_BY_GROUP]', ', '.join(max_lines) if max_lines else 'n/a')
                                reverse_shares = alloc_state.get('reverse_shares', {}) or {}
                                reverse_lines = []
                                for k in sorted(reverse_shares.keys()):
                                    reverse_lines.append(f"{k}: {reverse_shares.get(k, 0.0) * 100.0:.2f}%")
                                template = template.replace('[REVERSE_SPLIT_BY_GROUP]', ', '.join(reverse_lines) if reverse_lines else 'n/a')
                                template = template.replace('[BUDGET]', f"{float(alloc_state.get('budget', 0.0)):.2f}")
                                template = template.replace('[TOTAL_DCFS_UNCONSTRAINED]', f"{total_dcfs_unconstrained:,.2f}")
                                if total_dcfs_constrained is not None:
                                    template = template.replace('[TOTAL_DCFS_CONSTRAINED]', f"{total_dcfs_constrained:,.2f}")
                                else:
                                    template = template.replace('[TOTAL_DCFS_CONSTRAINED]', 'n/a')
                                total_dcfs_blend_unconstrained = (
                                    float(alloc_df_unconstrained['dcfs_with_reverse'].sum())
                                    if alloc_df_unconstrained is not None
                                    and not alloc_df_unconstrained.empty
                                    and 'dcfs_with_reverse' in alloc_df_unconstrained.columns
                                    else None
                                )
                                total_dcfs_blend_constrained = (
                                    float(alloc_df_constrained['dcfs_with_reverse'].sum())
                                    if alloc_df_constrained is not None
                                    and not alloc_df_constrained.empty
                                    and 'dcfs_with_reverse' in alloc_df_constrained.columns
                                    else None
                                )
                                template = template.replace(
                                    '[TOTAL_DCFS_BLEND_UNCONSTRAINED]',
                                    f"{total_dcfs_blend_unconstrained:,.2f}" if total_dcfs_blend_unconstrained is not None else 'n/a'
                                )
                                template = template.replace(
                                    '[TOTAL_DCFS_BLEND_CONSTRAINED]',
                                    f"{total_dcfs_blend_constrained:,.2f}" if total_dcfs_blend_constrained is not None else 'n/a'
                                )
                                template = template.replace('[CURVE_COUNT]', str(len(fit_params)))
                                template = template.replace('[ALLOCATION_TABLE]', allocation_table_text)
                                st.text_area('Allocation narrative (copy)', template, height=420)

                        max_alloc = None
                        curve_choice = st.selectbox(
                            'Allocation overlay',
                            [
                                'Unconstrained (risk-aware)',
                                'Constrained (risk-aware)',
                                'Unconstrained (blended)',
                                'Constrained (blended)',
                            ],
                        )

                        def _alloc_rows_from(df_in: pd.DataFrame, spend_col: str, dcfs_col: str):
                            if df_in is None or df_in.empty:
                                return []
                            rows = []
                            for _, row in df_in.iterrows():
                                rows.append({
                                    'group': row['group'],
                                    'allocated_spend': float(row[spend_col]) if pd.notna(row[spend_col]) else 0.0,
                                    'expected_dcfs': float(row[dcfs_col]) if pd.notna(row[dcfs_col]) else 0.0,
                                })
                            return rows

                        if curve_choice == 'Unconstrained (risk-aware)':
                            alloc_rows = _alloc_rows_from(alloc_df_unconstrained, 'allocated_spend', 'expected_dcfs')
                        elif curve_choice == 'Constrained (risk-aware)':
                            alloc_rows = _alloc_rows_from(alloc_df_constrained, 'allocated_spend', 'expected_dcfs')
                        elif curve_choice == 'Unconstrained (blended)':
                            alloc_rows = _alloc_rows_from(alloc_df_unconstrained, 'allocated_with_reverse', 'dcfs_with_reverse')
                        else:
                            alloc_rows = _alloc_rows_from(alloc_df_constrained, 'allocated_with_reverse', 'dcfs_with_reverse')

                        for df_alloc in [alloc_df_unconstrained, alloc_df_constrained]:
                            if df_alloc is not None and not df_alloc.empty:
                                max_alloc = max(
                                    max_alloc or 0.0,
                                    float(df_alloc['allocated_spend'].max()),
                                )
                        if alloc_rows:
                            max_alloc = max(
                                max_alloc or 0.0,
                                max(float(r['allocated_spend']) for r in alloc_rows),
                            )
                max_x = float(plot_df['Media Spend'].max()) if not plot_df.empty else 0.0
                if max_alloc is not None:
                    max_x = max_alloc
                if fit_params:
                    for group_label, (a, b) in fit_params.items():
                        x_fit = np.linspace(0, max_x, 150)
                        y_fit = _saturation_curve(x_fit, a, b)
                        curve_fig.add_trace(
                            go.Scatter(
                                x=x_fit,
                                y=y_fit,
                                mode='lines',
                                name=f'{group_label}',
                                line=dict(width=3, color=color_map.get(group_label)),
                                showlegend=True,
                            )
                        )
                if alloc_rows:
                    for row in alloc_rows:
                        curve_fig.add_trace(
                            go.Scatter(
                                x=[row['allocated_spend']],
                                y=[row['expected_dcfs']],
                                mode='markers',
                                name=f"{row['group']} allocation",
                                marker=dict(
                                    size=10,
                                    symbol='x',
                                    color=color_map.get(row['group']),
                                ),
                                showlegend=False,
                            )
                        )
                curve_fig.update_layout(
                    xaxis_title='Media Spend',
                    yaxis_title='DCFS',
                    legend_title_text=group_by or 'Group',
                )
                curve_fig.update_xaxes(range=[0, max_x])
                curve_fig.update_yaxes(range=[0, None])
                with left_col:
                    st.plotly_chart(curve_fig, use_container_width=True)
                if fit_rows:
                    st.subheader('Media response fit parameters (A, B)')
                    st.dataframe(pd.DataFrame(fit_rows).sort_values('A', ascending=False))

        st.subheader('Spend distribution')
        with st.popover('What is this?'):
            st.write(
                'Compares recent spend to s50 benchmarks and shows spend distributions by group '
                'to classify growth/mid/saturated zones.'
            )
        curve_plot = curve_data.copy()
        if curve_plot.empty or results.empty:
            st.info('No spend data available for the current filters.')
        else:
            if group_by:
                recent_map = (
                    results.groupby(group_by, dropna=False)['spend_recent']
                    .mean()
                    .rename_axis('group')
                    .to_dict()
                )
                s50_map = (
                    results.groupby(group_by, dropna=False)['s50_spend']
                    .mean()
                    .rename_axis('group')
                    .to_dict()
                )
            else:
                recent_map = {'All': results['spend_recent'].mean()}
                s50_map = {'All': results['s50_spend'].mean()}
            groups = [g for g in group_order if g in recent_map] if group_order else list(recent_map.keys())

            zone_colors = {
                'GROWTH': '#2ca02c',
                'MID': '#f2c744',
                'SATURATED': '#d62728',
                'UNKNOWN': '#9e9e9e',
            }
            zones_by_group = {}
            for group in groups:
                spend_recent = recent_map.get(group)
                s50_spend = s50_map.get(group)
                if spend_recent is None or s50_spend is None or s50_spend <= 0:
                    zones_by_group[group] = 'UNKNOWN'
                    continue
                ratio = spend_recent / s50_spend
                growth_ratio_max = float(growth_ratio_max_input)
                mid_ratio_max = float(mid_ratio_max_input)
                if ratio <= growth_ratio_max:
                    zones_by_group[group] = 'GROWTH'
                elif ratio <= mid_ratio_max:
                    zones_by_group[group] = 'MID'
                else:
                    zones_by_group[group] = 'SATURATED'

            fig = go.Figure()
            for zone, color in zone_colors.items():
                zone_groups = [g for g in groups if zones_by_group.get(g) == zone]
                if not zone_groups:
                    continue
                zone_custom = [
                    [recent_map.get(g), s50_map.get(g)]
                    for g in zone_groups
                ]
                fig.add_trace(
                    go.Bar(
                        x=zone_groups,
                        y=[recent_map.get(g) for g in zone_groups],
                        marker=dict(color=color, opacity=0.5),
                        name=f'Recent spend ({zone})',
                        customdata=zone_custom,
                        hovertemplate=(
                            'Group: %{x}<br>'
                            'Current spend: %{customdata[0]:,.2f}<br>'
                            'Saturation point: %{customdata[1]:,.2f}<extra></extra>'
                        ),
                    )
                )
            for group in groups:
                group_mask = curve_plot[group_by] == group if group_by else pd.Series([True] * len(curve_plot))
                y_vals = curve_plot.loc[group_mask, 'Media Spend']
                if y_vals.empty:
                    continue
                zone = zones_by_group.get(group, 'UNKNOWN')
                fig.add_trace(
                    go.Box(
                        x=[group] * len(y_vals),
                        y=y_vals,
                        boxpoints=False,
                        marker=dict(color='rgba(0,0,0,0)'),
                        line=dict(color='#444444'),
                        name=f'{group} ({zone})',
                        showlegend=False,
                    )
                )
            if group_order:
                fig.update_xaxes(categoryorder='array', categoryarray=group_order)
            fig.update_layout(yaxis_title='Media Spend')
            st.plotly_chart(fig, use_container_width=True)
            if st.button('Generate spend distribution report', key='spend_distribution_report'):
                st.info('Coming soon...')

        st.subheader('Predictability')
        with st.popover('What is this?'):
            st.write(
                'Shows CPL volatility (IQR/median) and the resulting predictability tier. '
                'Higher volatility means less predictable performance.'
            )
        vol_df = base_df.dropna(subset=['volatility'])
        if vol_df.empty:
            st.info('No volatility data for the current filters.')
        else:
            if group_by:
                vol_agg = (
                    vol_df.groupby(group_by, dropna=False)
                    .agg(
                        volatility=('volatility', 'median'),
                        predictability_penalty=('predictability_penalty', 'mean'),
                    )
                    .reset_index()
                    .rename(columns={group_by: 'group'})
                )
            else:
                vol_agg = vol_df.assign(group='All').agg(
                    volatility=('volatility', 'median'),
                    predictability_penalty=('predictability_penalty', 'mean'),
                ).to_frame().T
            vol_agg = vol_agg.dropna(subset=['volatility'])
            if vol_agg.empty:
                st.info('No volatility data for the current filters.')
            else:
                vol_low = float(OPPORTUNITY_CONFIG['vol_low'])
                vol_med = float(OPPORTUNITY_CONFIG['vol_med'])
                vol_high = float(OPPORTUNITY_CONFIG['vol_high'])
                vol_agg['vol_tier'] = 'VERY_HIGH'
                vol_agg.loc[vol_agg['volatility'] <= vol_high, 'vol_tier'] = 'HIGH'
                vol_agg.loc[vol_agg['volatility'] <= vol_med, 'vol_tier'] = 'MED'
                vol_agg.loc[vol_agg['volatility'] <= vol_low, 'vol_tier'] = 'LOW'
                fig = px.bar(
                    vol_agg.sort_values('volatility', ascending=False),
                    x='group',
                    y='volatility',
                    color='vol_tier',
                    labels={'volatility': 'Volatility (IQR / median)', 'group': group_by or 'Group'},
                    color_discrete_map={
                        'LOW': '#2ca02c',
                        'MED': '#f2c744',
                        'HIGH': '#ff7f0e',
                        'VERY_HIGH': '#d62728',
                    },
                )
                fig.update_xaxes(categoryorder='array', categoryarray=group_order)
                fig.add_hline(y=vol_low, line_dash='dash', line_color='#2ca02c', annotation_text='LOW')
                fig.add_hline(y=vol_med, line_dash='dash', line_color='#f2c744', annotation_text='MED')
                fig.add_hline(y=vol_high, line_dash='dash', line_color='#ff7f0e', annotation_text='HIGH')
                fig.update_yaxes(title_text='Volatility (IQR / median)')
                st.plotly_chart(fig, use_container_width=True)
                if st.button('Generate predictability report', key='predictability_report'):
                    st.info('Coming soon...')

        st.subheader('Opportunity score')
        with st.popover('What is this?'):
            st.write(
                'Combines headroom, scale, and curve scores minus volatility penalties into a '
                '0–100 opportunity score.'
            )
        opp_df = base_df.dropna(subset=['opportunity_score'])
        if opp_df.empty:
            st.info('No opportunity score data for the current filters.')
        else:
            if group_by:
                opp_agg = (
                    opp_df.groupby(group_by, dropna=False)['opportunity_score']
                    .mean()
                    .reset_index()
                    .rename(columns={group_by: 'group'})
                )
            else:
                opp_agg = pd.DataFrame({'group': ['All'], 'opportunity_score': [opp_df['opportunity_score'].mean()]})
            fig = px.bar(
                opp_agg.sort_values('opportunity_score', ascending=False),
                x='group',
                y='opportunity_score',
                text='opportunity_score',
                labels={'opportunity_score': 'Opportunity score (0–100)', 'group': group_by or 'Group'},
            )
            fig.update_xaxes(categoryorder='array', categoryarray=group_order)
            fig.update_traces(texttemplate='%{text:.0f}', textposition='outside')
            fig.update_yaxes(title_text='Opportunity score (0–100)', range=[0, 110])
            st.plotly_chart(fig, use_container_width=True)
            if st.button('Generate opportunity score report', key='opportunity_report'):
                st.info('Coming soon...')

        st.subheader('Conclusion and budget strategy')
        st.caption('LLM conclusion (wireframe)')
        if st.button('Generate final conclusion', key='final_conclusion'):
            st.info('Coming soon...')

    st.caption('Headroom is based on median CPL of the most recent 3 valid periods vs. historical P25 CPL.')

    display_cols = [
        'Market',
        'Channel',
        'Model',
        'current_cpl',
        'benchmark_cpl_p25',
        'benchmark_source',
        'headroom',
        'headroom_score',
        'headroom_tier',
        'avg_dcfs_recent',
        'avg_spend_recent',
        'scale_score',
        'scale_dist_n',
        'scale_dist_p25',
        'scale_dist_p50',
        'scale_dist_p75',
        'volatility',
        'vol_tier',
        'predictability_penalty',
        'raw_opportunity_score',
        'opportunity_score',
        'opportunity_tier',
        'tier_override_notes',
        'spend_recent',
        'k_used',
        's50_spend',
        'curve_ratio',
        'curve_zone_raw',
        'curve_score_raw',
        'curve_zone',
        'curve_score',
        'curve_worthy',
        'curve_worthiness_notes',
        'gate_passed',
        'gate_reasons',
        'audit',
    ]
    display_cols = [col for col in display_cols if col in results.columns]
    st.dataframe(results[display_cols].sort_values(['headroom_score'], ascending=False), use_container_width=True)
    st.download_button(
        'Download headroom results (CSV)',
        data=results.to_csv(index=False),
        file_name='opportunity_headroom_step1.csv',
        mime='text/csv',
    )
    st.stop()

if page == 'Market CPL':
    st.subheader('Average KPI by market')
    if 'Market' not in df.columns:
        st.warning('Market column not found in the dataset.')
        st.stop()
    required = {
        'media': 'Media Spend',
        'sessions': 'Number of Sessions',
        'dcfs': 'DCFS',
        'forms': 'Forms Submission Started',
    }
    for key, col in required.items():
        if col not in df.columns:
            required[key] = None

    week_options = get_calendar_week_options(df)
    week_choices = ['All'] + week_options
    selected_weeks = st.multiselect('Weeks', week_choices, default=['All'])
    cadence = st.slider('Cadence (weeks per point)', min_value=1, max_value=8, value=1, step=1)

    market_options = sorted(df['Market'].dropna().unique())
    market_choices = ['All'] + market_options
    m_col, agg_col = st.columns([4, 1], vertical_alignment='center')
    def _expand_market_cpl_markets():
        selected = st.session_state.get('market_cpl_markets', [])
        if 'All' in selected:
            st.session_state['market_cpl_markets'] = [m for m in market_choices if m != 'All']
    with m_col:
        selected_markets = st.multiselect(
            'Markets',
            market_choices,
            default=['All'],
            key='market_cpl_markets',
            on_change=_expand_market_cpl_markets,
        )
    with agg_col:
        include_all_markets = st.checkbox('All markets', value=False, key='market_cpl_all_markets')

    channel_options = sorted(df['Channel'].dropna().unique()) if 'Channel' in df.columns else []
    channel_choices = ['All'] + channel_options
    c_col, c_agg_col = st.columns([4, 1], vertical_alignment='center')
    with c_col:
        selected_channels = st.multiselect('Channels', channel_choices, default=['All'])
    with c_agg_col:
        include_all_channels = st.checkbox('All channels', value=False, key='market_cpl_all_channels')

    kpi_options = [
        'Media Invest',
        'Visits (Sessions)',
        'Dealer Contract Form Submissions',
        'DCFS',
        'Sessions to DCFS Conversion Rate',
        'Cost per Lead (Forms Submission Started)',
        'Cost per Lead (DCFS)',
        'Cost per Lead (both)',
    ]
    kpi_choice = st.selectbox('KPI', kpi_options)

    kpi_df = model_df.copy()
    if campaign and campaign != 'All':
        kpi_df = kpi_df[kpi_df['Campaign'] == campaign]
    if selected_weeks and 'All' not in selected_weeks:
        kpi_df = kpi_df[kpi_df['calendar_week'].isin(selected_weeks)]
    if selected_markets and 'All' not in selected_markets:
        kpi_df = kpi_df[kpi_df['Market'].isin(selected_markets)]
    if selected_channels and 'All' not in selected_channels and 'Channel' in kpi_df.columns:
        kpi_df = kpi_df[kpi_df['Channel'].isin(selected_channels)]

    if kpi_df.empty:
        st.warning('No data available for the selected weeks/markets.')
        st.stop()

    if cadence == 1:
        weekly_base = kpi_df.copy()
    else:
        weekly_base = (
            kpi_df.groupby(['Market', 'Channel', 'calendar_week'], dropna=False)
            .agg({
                'Media Spend': 'sum',
                'Number of Sessions': 'sum',
                'Forms Submission Started': 'sum',
                'DCFS': 'sum',
            })
            .reset_index()
        )
        week_order = get_calendar_week_options(kpi_df)
        week_idx_map = {w: i for i, w in enumerate(week_order, start=1)}
        weekly_base['week_idx'] = weekly_base['calendar_week'].map(week_idx_map)
        weekly_base = weekly_base.dropna(subset=['week_idx'])
        weekly_base['cadence_bin'] = ((weekly_base['week_idx'] - 1) // cadence).astype(int)
        if st.checkbox('Show cadence bins', value=False, key='market_cpl_show_cadence_bins'):
            bin_debug = (
                weekly_base.groupby(['Market', 'Channel', 'cadence_bin'], dropna=False)
                .agg(
                    weeks_in_bin=('week_idx', 'nunique'),
                    min_week=('week_idx', 'min'),
                    max_week=('week_idx', 'max'),
                )
                .reset_index()
                .sort_values(['Market', 'Channel', 'cadence_bin'])
            )
            st.dataframe(bin_debug, use_container_width=True)
        if st.checkbox('Show cadence calc debug', value=False, key='market_cpl_show_cadence_calc'):
            calc_debug = (
                weekly_base.groupby(['Market', 'Channel', 'cadence_bin'], dropna=False)
                .agg(
                    weeks=('calendar_week', lambda s: ','.join(sorted(set(s.astype(str))))),
                    spend_sum=('Media Spend', 'sum'),
                    dcfs_sum=('DCFS', 'sum'),
                )
                .reset_index()
            )
            calc_debug['cpl_bin'] = calc_debug.apply(
                lambda r: (r['spend_sum'] / r['dcfs_sum']) if r['dcfs_sum'] else None,
                axis=1,
            )
            st.dataframe(calc_debug, use_container_width=True)
        # drop any bin that included a zero-DCFS week (to mirror cadence=1 behavior)
        zero_dcfs = (
            weekly_base.groupby(['Market', 'Channel', 'cadence_bin'], dropna=False)['DCFS']
            .apply(lambda s: (s == 0).any())
            .reset_index()
            .rename(columns={'DCFS': 'has_zero_dcfs'})
        )
        weekly_base = weekly_base.merge(zero_dcfs, on=['Market', 'Channel', 'cadence_bin'], how='left')
        weekly_base = weekly_base[weekly_base['has_zero_dcfs'] != True].drop(columns=['has_zero_dcfs'])
        weekly_base = (
            weekly_base.groupby(['Market', 'Channel', 'cadence_bin'], dropna=False)
            .agg({
                'Media Spend': 'sum',
                'Number of Sessions': 'sum',
                'Forms Submission Started': 'sum',
                'DCFS': 'sum',
            })
            .reset_index()
        )
        weekly_base['calendar_week'] = weekly_base['cadence_bin'].apply(lambda b: f'bin_{b}')
    points_base = weekly_base.copy()
    if include_all_channels and 'Channel' in points_base.columns:
        points_base = points_base.copy()
        points_base['Channel'] = 'All Selected Channels'
    if include_all_markets:
        points_base = points_base.copy()
        points_base['Market'] = 'All Selected Markets'

    def safe_ratio(num, denom):
        return num / denom if denom else None

    if kpi_choice == 'Media Invest':
        points_base['kpi_value'] = points_base['Media Spend']
    elif kpi_choice == 'Visits (Sessions)':
        points_base['kpi_value'] = points_base['Number of Sessions']
    elif kpi_choice == 'Dealer Contract Form Submissions':
        points_base['kpi_value'] = points_base['Forms Submission Started']
    elif kpi_choice == 'DCFS':
        points_base['kpi_value'] = points_base['DCFS']
    elif kpi_choice == 'Sessions to DCFS Conversion Rate':
        points_base['kpi_value'] = points_base.apply(
            lambda r: safe_ratio(r['DCFS'], r['Number of Sessions']), axis=1
        )
    elif kpi_choice == 'Cost per Lead (Forms Submission Started)':
        points_base['kpi_value'] = points_base.apply(
            lambda r: safe_ratio(r['Media Spend'], r['Forms Submission Started']), axis=1
        )
    elif kpi_choice == 'Cost per Lead (DCFS)':
        points_base['kpi_value'] = points_base.apply(
            lambda r: safe_ratio(r['Media Spend'], r['DCFS']), axis=1
        )
    else:
        weekly_cpl_forms = points_base.copy()
        weekly_cpl_forms['kpi'] = 'CPL (Forms Submission Started)'
        weekly_cpl_forms['kpi_value'] = weekly_cpl_forms.apply(
            lambda r: safe_ratio(r['Media Spend'], r['Forms Submission Started']), axis=1
        )
        weekly_cpl_dcfs = points_base.copy()
        weekly_cpl_dcfs['kpi'] = 'CPL (DCFS)'
        weekly_cpl_dcfs['kpi_value'] = weekly_cpl_dcfs.apply(
            lambda r: safe_ratio(r['Media Spend'], r['DCFS']), axis=1
        )
        points_base = pd.concat([weekly_cpl_forms, weekly_cpl_dcfs], ignore_index=True)

    x_dim = 'Market'
    if 'Channel' in points_base.columns and not include_all_channels and not include_all_markets:
        points_base = points_base.copy()
        points_base['Market_Channel'] = points_base['Market'].astype(str) + ' | ' + points_base['Channel'].astype(str)
        x_dim = 'Market_Channel'
    if kpi_choice == 'Cost per Lead (both)':
        avg_kpi = (
            points_base.groupby([x_dim, 'kpi'], dropna=False)['kpi_value']
            .mean()
            .reset_index()
            .sort_values('kpi_value', ascending=False)
        )
    else:
        avg_kpi = (
            points_base.groupby(x_dim, dropna=False)['kpi_value']
            .mean()
            .reset_index()
            .sort_values('kpi_value', ascending=False)
        )
    points_base['week'] = points_base['calendar_week']
    points_base['week'] = points_base['week'].fillna('Unknown').astype(str)

    pct97 = (
        points_base.groupby([x_dim], dropna=False)['kpi_value']
        .quantile(0.97)
        .max()
    )
    benchmark_default = float(pct97) if pct97 is not None and not pd.isna(pct97) else 0.0
    filter_signature = {
        'campaign': campaign,
        'weeks': tuple(sorted(selected_weeks)) if selected_weeks else (),
        'markets': tuple(sorted(selected_markets)) if selected_markets else (),
        'channels': tuple(sorted(selected_channels)) if selected_channels else (),
        'all_markets': include_all_markets,
        'all_channels': include_all_channels,
        'cadence': cadence,
        'kpi_choice': kpi_choice,
    }
    if st.session_state.get('market_cpl_filter_signature') != filter_signature:
        st.session_state['market_cpl_benchmark'] = benchmark_default
        st.session_state['market_cpl_filter_signature'] = filter_signature
    elif 'market_cpl_benchmark' not in st.session_state:
        st.session_state['market_cpl_benchmark'] = benchmark_default
    benchmark_value = st.number_input(
        'Benchmark (KPI value)',
        key='market_cpl_benchmark',
        min_value=0.0,
        step=1.0,
    )

    st.subheader('Average + volatility (box plot)')
    show_points = st.checkbox('Show individual points', value=True, key='market_cpl_show_points')
    if kpi_choice == 'Cost per Lead (both)':
        box_fig = px.box(
            points_base,
            x=x_dim,
            y='kpi_value',
            facet_col='kpi',
            points=False,
            labels={'kpi_value': kpi_choice, 'Market': 'Market', 'Channel': 'Channel', 'Market_Channel': 'Market | Channel'},
        )
        pct_df = (
            points_base.groupby([x_dim, 'kpi'], dropna=False)['kpi_value']
            .quantile([0.9, 0.95, 0.97])
            .reset_index()
            .rename(columns={'level_2': 'percentile', 'kpi_value': 'value'})
        )
        pct_df['percentile'] = pct_df['percentile'].map({0.9: 'P90', 0.95: 'P95', 0.97: 'P97'})
        pct_fig = px.scatter(
            pct_df,
            x=x_dim,
            y='value',
            color='percentile',
            symbol='percentile',
            facet_col='kpi',
        )
        scatter_fig = px.strip(
            points_base,
            x=x_dim,
            y='kpi_value',
            color='week',
            facet_col='kpi',
        )
    else:
        box_fig = px.box(
            points_base,
            x=x_dim,
            y='kpi_value',
            points=False,
            labels={'kpi_value': kpi_choice, 'Market': 'Market', 'Channel': 'Channel', 'Market_Channel': 'Market | Channel'},
        )
        pct_df = (
            points_base.groupby([x_dim], dropna=False)['kpi_value']
            .quantile([0.9, 0.95, 0.97])
            .reset_index()
            .rename(columns={'level_1': 'percentile', 'kpi_value': 'value'})
        )
        pct_df['percentile'] = pct_df['percentile'].map({0.9: 'P90', 0.95: 'P95', 0.97: 'P97'})
        pct_fig = px.scatter(
            pct_df,
            x=x_dim,
            y='value',
            color='percentile',
            symbol='percentile',
        )
        scatter_fig = px.strip(
            points_base,
            x=x_dim,
            y='kpi_value',
            color='week',
        )

    if show_points:
        for trace in scatter_fig.data:
            trace.marker.size = 6
            trace.marker.opacity = 0.6
            box_fig.add_trace(trace)

    for trace in pct_fig.data:
        trace.marker.size = 9
        box_fig.add_trace(trace)

    box_fig.update_layout(height=520, boxmode='overlay')
    if benchmark_value and benchmark_value > 0:
        box_fig.add_hline(
            y=benchmark_value,
            line_dash='dash',
            line_color='orange',
            annotation_text='Benchmark (Malus starts)',
            annotation_position='top left',
        )
        bonus_start = benchmark_value * 0.9
        bonus_cap = benchmark_value * 0.5
        malus_cap = benchmark_value * 1.5
        box_fig.add_hline(
            y=bonus_start,
            line_dash='dash',
            line_color='green',
            annotation_text='Hurdle / Bonus starts (+10%)',
            annotation_position='top left',
        )
        box_fig.add_hline(
            y=bonus_cap,
            line_dash='dash',
            line_color='green',
            annotation_text='Bonus cap (+50%)',
            annotation_position='top left',
        )
        box_fig.add_hline(
            y=malus_cap,
            line_dash='dash',
            line_color='red',
            annotation_text='Malus cap (-50%)',
            annotation_position='top left',
        )
    st.plotly_chart(box_fig, use_container_width=True)

    if benchmark_value and benchmark_value > 0:
        f1, f2 = st.columns(2)
        with f1:
            bah_fee = st.number_input('BAH fee (€)', min_value=0.0, value=0.0, step=1000.0)
        with f2:
            fte_fee = st.number_input('FTE fee (€)', min_value=0.0, value=0.0, step=1000.0)

        bonus_start = benchmark_value * 0.9
        bonus_cap = benchmark_value * 0.5
        malus_cap = benchmark_value * 1.5
        values = points_base['kpi_value'].dropna()
        total_points = len(values)
        if total_points:
            bonus_count = (values <= bonus_start).sum()
            neutral_count = ((values > bonus_start) & (values < benchmark_value)).sum()
            malus_count = (values >= benchmark_value).sum()
            bonus_cap_count = (values <= bonus_cap).sum()
            malus_cap_count = (values >= malus_cap).sum()

            st.subheader('Incentive band population')
            def _card(label, count, pct, bah_adj=None, fte_adj=None, final_fee=None):
                return f"""
                <div style="padding:10px 14px; border:1px solid #E6E6E6; border-radius:12px; background:#FAFAFA;">
                  <div style="font-size:11px; color:#666; text-transform:uppercase; letter-spacing:0.06em;">{label}</div>
                  <div style="font-size:26px; font-weight:700; margin-top:4px;">{count:,}</div>
                  <div style="font-size:12px; color:#888; margin-top:2px;">{pct:.1f}% of points</div>
                  <div style="font-size:12px; color:#444; margin-top:6px;">BAH adj: {bah_adj if bah_adj is not None else 'n/a'}</div>
                  <div style="font-size:12px; color:#444;">FTE adj: {fte_adj if fte_adj is not None else 'n/a'}</div>
                  <div style="font-size:12px; color:#444;">Final fee: {final_fee if final_fee is not None else 'n/a'}</div>
                </div>
                """
            def _adjustment(val):
                if val is None or benchmark_value <= 0:
                    return None
                delta = (benchmark_value - val) / benchmark_value
                if delta < 0:
                    return max(delta, -0.5)
                if delta <= 0.10:
                    return 0.0
                return min(delta - 0.10, 0.5)

            adjustments = values.apply(_adjustment)
            variable_bah = 0.5 * bah_fee
            variable_fte = 0.2 * fte_fee
            variable_fee = variable_bah + variable_fte
            fixed_fee = (bah_fee + fte_fee) - variable_fee

            def _band_stats(mask):
                subset = adjustments[mask]
                if subset.empty:
                    return None, None, None
                adj = subset.mean()
                bah_adj = variable_bah * adj
                fte_adj = variable_fte * adj
                total_fee = fixed_fee + variable_fee * (1 + adj)
                return f"{bah_adj:,.2f}", f"{fte_adj:,.2f}", f"{total_fee:,.2f}"

            bonus_bah, bonus_fte, bonus_total = _band_stats(values <= bonus_start)
            neutral_bah, neutral_fte, neutral_total = _band_stats((values > bonus_start) & (values < benchmark_value))
            malus_bah, malus_fte, malus_total = _band_stats(values >= benchmark_value)
            bonus_cap_bah, bonus_cap_fte, bonus_cap_total = _band_stats(values <= bonus_cap)
            malus_cap_bah, malus_cap_fte, malus_cap_total = _band_stats(values >= malus_cap)
            cards = [
                ('Bonus band', bonus_count, bonus_count / total_points * 100.0, bonus_bah, bonus_fte, bonus_total),
                ('Neutral band', neutral_count, neutral_count / total_points * 100.0, neutral_bah, neutral_fte, neutral_total),
                ('Malus band', malus_count, malus_count / total_points * 100.0, malus_bah, malus_fte, malus_total),
                ('Bonus cap hit', bonus_cap_count, bonus_cap_count / total_points * 100.0, bonus_cap_bah, bonus_cap_fte, bonus_cap_total),
                ('Malus cap hit', malus_cap_count, malus_cap_count / total_points * 100.0, malus_cap_bah, malus_cap_fte, malus_cap_total),
            ]
            c1, c2, c3, c4, c5 = st.columns(5)
            for col, (label, count, pct, bah_adj, fte_adj, total_fee) in zip([c1, c2, c3, c4, c5], cards):
                with col:
                    st.markdown(_card(label, count, pct, bah_adj, fte_adj, total_fee), unsafe_allow_html=True)

            overall_adj = adjustments.mean() if not adjustments.empty else None
            if overall_adj is not None:
                overall_bah = variable_bah * overall_adj
                overall_fte = variable_fte * overall_adj
                overall_total = fixed_fee + variable_fee * (1 + overall_adj)
                st.markdown(
                    f"""
                    <div style="margin-top:12px; padding:12px 16px; border:1px solid #E6E6E6; border-radius:14px; background:#F7F7F7;">
                      <div style="font-size:12px; color:#666; text-transform:uppercase; letter-spacing:0.06em;">Average adjustment (all points)</div>
                      <div style="display:flex; gap:20px; margin-top:6px; font-size:14px;">
                        <div><strong>BAH adj:</strong> {overall_bah:,.2f}</div>
                        <div><strong>FTE adj:</strong> {overall_fte:,.2f}</div>
                        <div><strong>Final fee:</strong> {overall_total:,.2f}</div>
                      </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            st.caption('Narrative benchmark cases')
            p1, p2, p3, p4, p5 = st.columns(5)
            with p1:
                use_p99 = st.checkbox('P99', value=False, key='market_cpl_narr_p99')
            with p2:
                use_p97 = st.checkbox('P97', value=False, key='market_cpl_narr_p97')
            with p3:
                use_p90 = st.checkbox('P90', value=False, key='market_cpl_narr_p90')
            with p4:
                use_p67 = st.checkbox('P67', value=False, key='market_cpl_narr_p67')
            with p5:
                use_p50 = st.checkbox('P50', value=False, key='market_cpl_narr_p50')

            def _benchmark_for_percentile(p):
                value = (
                    points_base.groupby([x_dim], dropna=False)['kpi_value']
                    .quantile(p)
                    .max()
                )
                return float(value) if value is not None and not pd.isna(value) else 0.0

            if st.button('Generate Incentive Narrative', key='market_cpl_incentive_narrative'):
                percentile_cases = []
                if use_p99:
                    percentile_cases.append(('P99', _benchmark_for_percentile(0.99)))
                if use_p97:
                    percentile_cases.append(('P97', _benchmark_for_percentile(0.97)))
                if use_p90:
                    percentile_cases.append(('P90', _benchmark_for_percentile(0.90)))
                if use_p67:
                    percentile_cases.append(('P67', _benchmark_for_percentile(0.67)))
                if use_p50:
                    percentile_cases.append(('P50', _benchmark_for_percentile(0.50)))
                if not percentile_cases:
                    st.warning('Select at least one percentile to generate the narrative.')
                else:
                    reports = []
                    for label, bench in percentile_cases:
                        if bench <= 0:
                            continue
                        bonus_start_local = bench * 0.9
                        bonus_cap_local = bench * 0.5
                        malus_cap_local = bench * 1.5
                        vals = values.copy()
                        total = len(vals)
                        bonus_count_local = (vals <= bonus_start_local).sum()
                        neutral_count_local = ((vals > bonus_start_local) & (vals < bench)).sum()
                        malus_count_local = (vals >= bench).sum()
                        bonus_cap_count_local = (vals <= bonus_cap_local).sum()
                        malus_cap_count_local = (vals >= malus_cap_local).sum()

                        def _adjustment_local(val):
                            if val is None or bench <= 0:
                                return None
                            delta = (bench - val) / bench
                            if delta < 0:
                                return max(delta, -0.5)
                            if delta <= 0.10:
                                return 0.0
                            return min(delta - 0.10, 0.5)

                        adjustments_local = vals.apply(_adjustment_local)

                        def _band_stats_local(mask):
                            subset = adjustments_local[mask]
                            if subset.empty:
                                return None, None, None
                            adj = subset.mean()
                            bah_adj = variable_bah * adj
                            fte_adj = variable_fte * adj
                            total_fee = fixed_fee + variable_fee * (1 + adj)
                            return f"{bah_adj:,.2f}", f"{fte_adj:,.2f}", f"{total_fee:,.2f}"

                        bonus_bah_l, bonus_fte_l, bonus_total_l = _band_stats_local(vals <= bonus_start_local)
                        neutral_bah_l, neutral_fte_l, neutral_total_l = _band_stats_local((vals > bonus_start_local) & (vals < bench))
                        malus_bah_l, malus_fte_l, malus_total_l = _band_stats_local(vals >= bench)
                        bonus_cap_bah_l, bonus_cap_fte_l, bonus_cap_total_l = _band_stats_local(vals <= bonus_cap_local)
                        malus_cap_bah_l, malus_cap_fte_l, malus_cap_total_l = _band_stats_local(vals >= malus_cap_local)

                        band_counts = [
                            f"Bonus band: {bonus_count_local} ({bonus_count_local/total*100:.1f}%)",
                            f"Neutral band: {neutral_count_local} ({neutral_count_local/total*100:.1f}%)",
                            f"Malus band: {malus_count_local} ({malus_count_local/total*100:.1f}%)",
                            f"Bonus cap hit: {bonus_cap_count_local} ({bonus_cap_count_local/total*100:.1f}%)",
                            f"Malus cap hit: {malus_cap_count_local} ({malus_cap_count_local/total*100:.1f}%)",
                        ]
                        band_fees = [
                            f"Bonus band: BAH adj={bonus_bah_l}, FTE adj={bonus_fte_l}, Final fee={bonus_total_l}",
                            f"Neutral band: BAH adj={neutral_bah_l}, FTE adj={neutral_fte_l}, Final fee={neutral_total_l}",
                            f"Malus band: BAH adj={malus_bah_l}, FTE adj={malus_fte_l}, Final fee={malus_total_l}",
                            f"Bonus cap hit: BAH adj={bonus_cap_bah_l}, FTE adj={bonus_cap_fte_l}, Final fee={bonus_cap_total_l}",
                            f"Malus cap hit: BAH adj={malus_cap_bah_l}, FTE adj={malus_cap_fte_l}, Final fee={malus_cap_total_l}",
                        ]

                        overall_adj_l = adjustments_local.mean() if not adjustments_local.empty else None
                        overall_bah_l = variable_bah * overall_adj_l if overall_adj_l is not None else None
                        overall_fte_l = variable_fte * overall_adj_l if overall_adj_l is not None else None
                        overall_total_l = fixed_fee + variable_fee * (1 + overall_adj_l) if overall_adj_l is not None else None

                        report = INCENTIVE_METHOD_TEMPLATE
                        report = report.replace('[KPI]', kpi_choice)
                        report = report.replace('[CADENCE]', str(cadence))
                        report = report.replace('[BENCHMARK_CASE]', label)
                        report = report.replace('[BENCHMARK]', f"{bench:,.2f}")
                        report = report.replace('[MARKETS]', ', '.join(selected_markets) if selected_markets else 'All')
                        report = report.replace('[CHANNELS]', ', '.join(selected_channels) if selected_channels else 'All')
                        report = report.replace('[WEEKS]', ', '.join(selected_weeks) if selected_weeks else 'All')
                        report = report.replace('[BAH_FEE]', f"{bah_fee:,.2f}")
                        report = report.replace('[FTE_FEE]', f"{fte_fee:,.2f}")
                        report = report.replace('[BAND_COUNTS]', '\n'.join(band_counts))
                        report = report.replace('[BAND_FEES]', '\n'.join(band_fees))
                        report = report.replace('[AVG_ADJ]', f"{overall_adj_l:,.4f}" if overall_adj_l is not None else 'n/a')
                        report = report.replace('[AVG_BAH_ADJ]', f"{overall_bah_l:,.2f}" if overall_adj_l is not None else 'n/a')
                        report = report.replace('[AVG_FTE_ADJ]', f"{overall_fte_l:,.2f}" if overall_adj_l is not None else 'n/a')
                        report = report.replace('[AVG_FINAL_FEE]', f"{overall_total_l:,.2f}" if overall_adj_l is not None else 'n/a')
                        reports.append(f"## Benchmark case: {label}\n\n{report}")
                    st.text_area('Incentive Methodology Narrative', '\n\n---\n\n'.join(reports), height=520)

    st.stop()

if page == 'Market Report - Excel Export':
    required_cols = [
        'Market',
        'Model',
        'Channel',
        'Platform',
        'Activation Group',
        'Media Spend',
        'Number of Sessions',
        'Forms Submission Started',
        'DCFS',
        'calendar_week',
    ]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        st.warning(f'Missing required columns: {", ".join(missing)}')
        st.stop()

    export_df = df[df['Market'] == export_market].copy()
    if export_campaign != 'All' and 'Campaign' in export_df.columns:
        export_df = export_df[export_df['Campaign'] == export_campaign]
    if date_mode == 'Date range':
        date_col = 'Date' if 'Date' in export_df.columns else 'report_date'
        export_df[date_col] = pd.to_datetime(export_df[date_col], errors='coerce')
        if not export_dates or len(export_dates) != 2:
            st.warning('Select a start and end date.')
            st.stop()
        start_date, end_date = export_dates
        export_df = export_df[
            (export_df[date_col] >= pd.Timestamp(start_date))
            & (export_df[date_col] <= pd.Timestamp(end_date))
        ]
    elif export_weeks and 'All' not in export_weeks:
        export_df = export_df[export_df['calendar_week'].isin(export_weeks)]

    if export_df.empty:
        st.warning('No data available for the selected market/weeks.')
        st.stop()

    if date_mode == 'Date range':
        week_label = f'{export_dates[0]} to {export_dates[1]}'
    else:
        week_label = 'All' if not export_weeks or 'All' in export_weeks else ', '.join(export_weeks)
    st.subheader('Market Report - Excel Export')
    st.caption('Exports the same stacked tables as the shared PCL Excel file.')

    workbook = build_close_gap_workbook(export_df, export_market, week_label)
    st.download_button(
        'Download Excel',
        data=workbook,
        file_name=f'Close_the_Gap_{export_market}_2025.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    st.stop()

if page == 'KPI vs Investment':
    required_cols = [
        'Media Spend',
        'Number of Sessions',
        'Forms Submission Started',
        'DCFS',
    ]
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        st.warning(f'Missing required columns: {", ".join(missing)}')
        st.stop()

    data = df.copy()
    if kpi_filters['market'] != 'All':
        data = data[data['Market'] == kpi_filters['market']]
    if kpi_filters['channels'] and 'All' not in kpi_filters['channels']:
        data = data[data['Channel'].isin(kpi_filters['channels'])]
    if kpi_filters['campaigns'] and 'All' not in kpi_filters['campaigns']:
        data = data[data['Campaign'].isin(kpi_filters['campaigns'])]
    if kpi_filters['platforms'] and 'All' not in kpi_filters['platforms']:
        data = data[data['Platform'].isin(kpi_filters['platforms'])]
    if kpi_filters['models'] and 'All' not in kpi_filters['models']:
        data = data[data['Model'].isin(kpi_filters['models'])]
    if kpi_filters['activations'] and 'All' not in kpi_filters['activations']:
        data = data[data['Activation Group'].isin(kpi_filters['activations'])]

    if data.empty:
        st.warning('No data available for the selected filters.')
        st.stop()

    group_by = kpi_filters['group_by']
    if not group_by:
        st.warning('No grouping column available for the selected dataset.')
        st.stop()

    agg = (
        data.groupby(group_by, dropna=False)
        .agg({
            'Media Spend': 'sum',
            'Number of Sessions': 'sum',
            'Forms Submission Started': 'sum',
            'DCFS': 'sum',
        })
        .reset_index()
    )
    if kpi_filters['kpi'] == 'Visits (Sessions)':
        agg['kpi_value'] = agg['Number of Sessions']
    elif kpi_filters['kpi'] == 'Dealer Contract Form Submissions':
        agg['kpi_value'] = agg['Forms Submission Started']
    elif kpi_filters['kpi'] == 'DCFS':
        agg['kpi_value'] = agg['DCFS']
    elif kpi_filters['kpi'] == 'Sessions to DCFS Conversion Rate':
        agg['kpi_value'] = agg.apply(lambda r: _safe_ratio(r['DCFS'], r['Number of Sessions']), axis=1)
    elif kpi_filters['kpi'] == 'Cost per Lead (Forms Submission Started)':
        agg['kpi_value'] = agg.apply(
            lambda r: _safe_ratio(r['Media Spend'], r['Forms Submission Started']), axis=1
        )
    elif kpi_filters['kpi'] == 'Cost per Lead (DCFS)':
        agg['kpi_value'] = agg.apply(lambda r: _safe_ratio(r['Media Spend'], r['DCFS']), axis=1)
    else:
        agg['kpi_value'] = None

    color_by = kpi_filters['color_by']
    if color_by:
        color_map = (
            data.groupby([group_by, color_by], dropna=False)
            .agg({
                'Media Spend': 'sum',
                'Number of Sessions': 'sum',
                'Forms Submission Started': 'sum',
                'DCFS': 'sum',
            })
            .reset_index()
        )
        if kpi_filters['kpi'] == 'Visits (Sessions)':
            color_map['kpi_value'] = color_map['Number of Sessions']
        elif kpi_filters['kpi'] == 'Dealer Contract Form Submissions':
            color_map['kpi_value'] = color_map['Forms Submission Started']
        elif kpi_filters['kpi'] == 'DCFS':
            color_map['kpi_value'] = color_map['DCFS']
        elif kpi_filters['kpi'] == 'Sessions to DCFS Conversion Rate':
            color_map['kpi_value'] = color_map.apply(
                lambda r: _safe_ratio(r['DCFS'], r['Number of Sessions']), axis=1
            )
        elif kpi_filters['kpi'] == 'Cost per Lead (Forms Submission Started)':
            color_map['kpi_value'] = color_map.apply(
                lambda r: _safe_ratio(r['Media Spend'], r['Forms Submission Started']), axis=1
            )
        elif kpi_filters['kpi'] == 'Cost per Lead (DCFS)':
            color_map['kpi_value'] = color_map.apply(
                lambda r: _safe_ratio(r['Media Spend'], r['DCFS']), axis=1
            )
        plot_df = color_map
    else:
        plot_df = agg

    st.subheader('KPI vs investment')
    fig = px.scatter(
        plot_df,
        x='Media Spend',
        y='kpi_value',
        color=color_by,
        hover_name=group_by,
        labels={'Media Spend': 'Media Spend', 'kpi_value': kpi_filters['kpi'], group_by: group_by},
    )
    fig.update_layout(height=520)

    fit_rows = []
    if np is None or curve_fit is None:
        st.info('Install scipy to enable curve fitting for Ax/(b+x).')
    else:
        if color_by:
            for key, group in plot_df.groupby(color_by, dropna=False):
                a, b = fit_saturation(group['Media Spend'], group['kpi_value'])
                if a is None or b is None:
                    continue
                fit_rows.append({color_by: key, 'A': a, 'B': b, 'points': len(group)})
                x_fit = np.linspace(group['Media Spend'].min(), group['Media Spend'].max(), 100)
                y_fit = _saturation_curve(x_fit, a, b)
                fig.add_trace(
                    go.Scatter(
                        x=x_fit,
                        y=y_fit,
                        mode='lines',
                        name=f'{key} fit',
                        line=dict(dash='solid'),
                        showlegend=True,
                    )
                )
        else:
            a, b = fit_saturation(plot_df['Media Spend'], plot_df['kpi_value'])
            if a is not None and b is not None:
                fit_rows.append({'A': a, 'B': b, 'points': len(plot_df)})
                x_fit = np.linspace(plot_df['Media Spend'].min(), plot_df['Media Spend'].max(), 100)
                y_fit = _saturation_curve(x_fit, a, b)
                fig.add_trace(
                    go.Scatter(
                        x=x_fit,
                        y=y_fit,
                        mode='lines',
                        name='Fit',
                        line=dict(dash='solid'),
                        showlegend=True,
                    )
                )
    st.plotly_chart(fig, use_container_width=True)

    if fit_rows:
        st.subheader('Saturation fit parameters')
        st.dataframe(pd.DataFrame(fit_rows))

    st.subheader('Aggregated data')
    st.dataframe(plot_df)
    st.stop()

if page == 'Market Alignments':
    st.subheader('Market Alignments')
    st.caption('Workspace for market coordination and contacts.')
    st.write('Addressbook (placeholder)')
    if 'market_addressbook' not in st.session_state:
        st.session_state['market_addressbook'] = pd.DataFrame({
            'Market': [],
            'Contact Name': [],
            'Role': [],
            'Email': [],
            'Notes': [],
        })
    addressbook_df = st.data_editor(
        st.session_state['market_addressbook'],
        use_container_width=True,
        num_rows='dynamic',
        key='market_addressbook_editor',
    )
    st.session_state['market_addressbook'] = addressbook_df
    st.stop()

if page == 'UTM Adoption':
    st.subheader('UTM Adoption')
    st.caption('Workspace for tracking market-level UTM rollout and compliance.')
    initialize_utm_outreach_db(UTM_OUTREACH_DB_PATH)
    bootstrap_utm_notes_db(UTM_OUTREACH_DB_PATH, UTM_NOTES_PDF_PATH, UTM_NOTES_CSV_PATH)
    bootstrap_utm_contacts_db(UTM_OUTREACH_DB_PATH, ADDRESSBOOK_CSV_PATH)
    ensure_manual_utm_test_records(UTM_OUTREACH_DB_PATH)
    bootstrap_contact_tags(UTM_OUTREACH_DB_PATH)

    utm_notes_df = load_utm_notes_from_db(UTM_OUTREACH_DB_PATH)
    ensure_market_engagement_records(UTM_OUTREACH_DB_PATH, utm_notes_df)
    if 'utm_adoption_notes_working' not in st.session_state:
        st.session_state['utm_adoption_notes_working'] = utm_notes_df.copy()
    else:
        current_working = st.session_state['utm_adoption_notes_working']
        if len(current_working) != len(utm_notes_df) or list(current_working.columns) != list(utm_notes_df.columns):
            st.session_state['utm_adoption_notes_working'] = utm_notes_df.copy()
    utm_notes_df = st.session_state['utm_adoption_notes_working']

    top_left, top_mid, top_right, top_reset = st.columns([1, 1, 3, 1])
    with top_left:
        st.metric('Markets', len(utm_notes_df))
    with top_mid:
        populated_status = utm_notes_df['Status'].fillna('').astype(str).str.strip().ne('').sum()
        st.metric('Statuses Captured', int(populated_status))
    with top_right:
        st.caption(f'Source: `{UTM_OUTREACH_DB_PATH}`')
        st.caption('Edits are written to SQLite when you press `Save Changes`.')
    with top_reset:
        if st.button('Reset Table', key='utm_reset_table'):
            reset_utm_notes_to_base(UTM_OUTREACH_DB_PATH)
            st.session_state['utm_adoption_notes_working'] = load_utm_notes_from_db(UTM_OUTREACH_DB_PATH)
            st.rerun()

    market_options = ['All'] + utm_notes_df['Market Code'].dropna().astype(str).tolist()
    selected_market = st.selectbox('Market', market_options, index=0, key='utm_adoption_market')
    search_term = st.text_input('Search notes', value='', key='utm_adoption_search')

    filtered_notes = utm_notes_df.copy()
    if selected_market != 'All':
        filtered_notes = filtered_notes[filtered_notes['Market Code'] == selected_market]

    search_term = search_term.strip()
    if search_term:
        search_mask = filtered_notes.fillna('').apply(
            lambda column: column.astype(str).str.contains(search_term, case=False, na=False)
        ).any(axis=1)
        filtered_notes = filtered_notes[search_mask]

    filtered_notes = filtered_notes.copy()
    editable_cols = ['Market Name', 'Status', 'Scope Gaps', 'Observations', 'Issues Identified', 'Context', 'Next Steps']
    with st.form('utm_adoption_form'):
        edited_notes = st.data_editor(
            filtered_notes,
            use_container_width=True,
            hide_index=True,
            height=700,
            disabled=['Market Code'],
            column_config={
                'Market Code': st.column_config.TextColumn('Market', width='small'),
                'Market Name': st.column_config.TextColumn('Market Name', width='medium'),
                'Status': st.column_config.TextColumn('Status', width='large'),
                'Scope Gaps': st.column_config.TextColumn('Scope Gaps', width='large'),
                'Observations': st.column_config.TextColumn('Observations', width='large'),
                'Issues Identified': st.column_config.TextColumn('Issues Identified', width='large'),
                'Context': st.column_config.TextColumn('Context', width='large'),
                'Next Steps': st.column_config.TextColumn('Next Steps', width='large'),
            },
            key='utm_adoption_editor',
        )
        save_submitted = st.form_submit_button('Save Changes')

    if save_submitted:
        utm_notes_df.loc[edited_notes.index, editable_cols] = edited_notes[editable_cols]
        st.session_state['utm_adoption_notes_working'] = utm_notes_df
        save_utm_notes_to_db(utm_notes_df, UTM_OUTREACH_DB_PATH)
        st.success('Saved to SQLite.')
        st.rerun()

    st.download_button(
        'Download table as CSV',
        utm_notes_df.to_csv(index=False).encode('utf-8'),
        file_name='utm_adoption_notes_export.csv',
        mime='text/csv',
        key='utm_adoption_download_csv',
    )

    st.divider()
    st.subheader('Addressbook')
    st.caption('Market contacts consolidated from the PHD local market workbook.')

    addressbook_df = annotate_addressbook_contacts(load_contacts_from_db(UTM_OUTREACH_DB_PATH))
    outreach_db_counts = load_utm_outreach_db_counts(UTM_OUTREACH_DB_PATH)

    addr_left, addr_mid, addr_right = st.columns([1, 1, 3])
    with addr_left:
        st.metric('Contacts', len(addressbook_df))
    with addr_mid:
        st.metric('Markets Covered', int(addressbook_df['market'].nunique()))
    with addr_right:
        st.caption(f'Source: `{UTM_OUTREACH_DB_PATH}`')

    addressbook_market_options = ['All'] + sorted(
        market for market in addressbook_df['market'].dropna().astype(str).unique() if market.strip()
    )
    addressbook_section_options = ['All'] + sorted(
        section for section in addressbook_df['section'].dropna().astype(str).unique() if section.strip()
    )
    addressbook_contact_type_options = ['All', 'Porsche Client', 'Account Lead', 'Porsche Client + Account Lead']

    addr_filter_left, addr_filter_mid, addr_filter_right, addr_filter_far_right = st.columns([1, 1, 1, 2])
    with addr_filter_left:
        selected_addressbook_market = st.selectbox(
            'Addressbook Market',
            addressbook_market_options,
            index=0,
            key='utm_addressbook_market',
        )
    with addr_filter_mid:
        selected_addressbook_section = st.selectbox(
            'Section',
            addressbook_section_options,
            index=0,
            key='utm_addressbook_section',
        )
    with addr_filter_right:
        selected_contact_type = st.selectbox(
            'Contact Type',
            addressbook_contact_type_options,
            index=0,
            key='utm_addressbook_contact_type',
        )
    with addr_filter_far_right:
        addressbook_search_term = st.text_input(
            'Search contacts',
            value='',
            key='utm_addressbook_search',
        ).strip()

    filtered_addressbook = addressbook_df.copy()
    if selected_addressbook_market != 'All':
        filtered_addressbook = filtered_addressbook[filtered_addressbook['market'] == selected_addressbook_market]
    if selected_addressbook_section != 'All':
        filtered_addressbook = filtered_addressbook[filtered_addressbook['section'] == selected_addressbook_section]
    if selected_contact_type != 'All':
        filtered_addressbook = filtered_addressbook[filtered_addressbook['contact_type'] == selected_contact_type]
    if addressbook_search_term:
        addressbook_search_mask = filtered_addressbook.fillna('').apply(
            lambda column: column.astype(str).str.contains(addressbook_search_term, case=False, na=False)
        ).any(axis=1)
        filtered_addressbook = filtered_addressbook[addressbook_search_mask]

    if 'utm_manual_to_emails' not in st.session_state:
        st.session_state['utm_manual_to_emails'] = []
    if 'utm_manual_cc_emails' not in st.session_state:
        st.session_state['utm_manual_cc_emails'] = []

    contact_picker_options = {}
    for _, row in filtered_addressbook.iterrows():
        label = f"{row['market']} | {row['section']} | {row['name'] or row['email']} | {row['email']}"
        contact_picker_options[label] = str(row['email']).strip()

    picked_contact_labels = st.multiselect(
        'Pick contacts for the draft',
        list(contact_picker_options.keys()),
        default=[],
        key='utm_addressbook_contact_multiselect',
        help='Filter the addressbook first, then pick the exact contacts you want.',
    )
    picked_emails = [contact_picker_options[label] for label in picked_contact_labels]

    st.dataframe(
        filtered_addressbook,
        use_container_width=True,
        hide_index=True,
        column_config={
            'market': st.column_config.TextColumn('Market', width='small'),
            'section': st.column_config.TextColumn('Section', width='small'),
            'name': st.column_config.TextColumn('Name', width='medium'),
            'title': st.column_config.TextColumn('Title', width='medium'),
            'email': st.column_config.TextColumn('Email', width='medium'),
            'contact_tags': st.column_config.TextColumn('Tags', width='small'),
            'notes': st.column_config.TextColumn('Notes', width='medium'),
            'source_sheet': st.column_config.TextColumn('Source Sheet', width='small'),
            'contact_type': st.column_config.TextColumn('Contact Type', width='small'),
        },
    )

    add_to_left, add_to_mid, add_to_right = st.columns([1, 1, 2])
    with add_to_left:
        if st.button('Add Selected to Draft To', key='utm_add_selected_to_to'):
            st.session_state['utm_manual_to_emails'] = merge_unique_emails(
                st.session_state['utm_manual_to_emails'],
                picked_emails,
            )
            st.success('Selected contacts added to draft To.')
            st.rerun()
    with add_to_mid:
        if st.button('Add Selected to Draft CC', key='utm_add_selected_to_cc'):
            st.session_state['utm_manual_cc_emails'] = merge_unique_emails(
                st.session_state['utm_manual_cc_emails'],
                picked_emails,
            )
            st.success('Selected contacts added to draft CC.')
            st.rerun()
    with add_to_right:
        if st.button('Clear Manual Draft Recipients', key='utm_clear_manual_recipients'):
            st.session_state['utm_manual_to_emails'] = []
            st.session_state['utm_manual_cc_emails'] = []
            st.session_state['utm_addressbook_contact_multiselect'] = []
            st.success('Manual draft recipients cleared.')
            st.rerun()

    st.caption(
        f"Manual draft To: {', '.join(st.session_state['utm_manual_to_emails']) or 'None'}\n\n"
        f"Manual draft CC: {', '.join(st.session_state['utm_manual_cc_emails']) or 'None'}"
    )

    st.download_button(
        'Download addressbook as CSV',
        filtered_addressbook.to_csv(index=False).encode('utf-8'),
        file_name='utm_addressbook_export.csv',
        mime='text/csv',
        key='utm_addressbook_download_csv',
    )

    st.divider()
    st.subheader('Outreach')
    st.caption('Create market-specific Outlook drafts or sends and track them in SQLite.')

    ensure_default_utm_outreach_template(UTM_OUTREACH_DB_PATH)
    templates_df = load_email_templates_df(UTM_OUTREACH_DB_PATH, 'UTM Adoption')
    outlook_status_left, outlook_status_right = st.columns([1, 4])
    with outlook_status_left:
        check_outlook_clicked = st.button('Check Outlook Connection', key='utm_outlook_check')
    with outlook_status_right:
        st.caption('Outlook checks now run on demand in a worker process with a timeout.')

    if check_outlook_clicked:
        outlook_ready, outlook_message = get_outlook_status()
        if outlook_ready:
            st.success(outlook_message)
        else:
            st.warning(outlook_message)

    outreach_status_options = sorted(
        status for status in utm_notes_df['Status'].dropna().astype(str).unique() if status.strip()
    )
    outreach_filter_left, outreach_filter_mid = st.columns([2, 3])
    with outreach_filter_left:
        selected_outreach_statuses = st.multiselect(
            'Statuses',
            outreach_status_options,
            default=[],
            key='utm_outreach_statuses',
        )
    with outreach_filter_mid:
        candidate_markets_df = utm_notes_df.copy()
        if selected_outreach_statuses:
            candidate_markets_df = candidate_markets_df[
                candidate_markets_df['Status'].astype(str).isin(selected_outreach_statuses)
            ]
        candidate_markets = candidate_markets_df['Market Code'].dropna().astype(str).tolist()
        selected_outreach_markets = st.multiselect(
            'Markets to contact',
            candidate_markets,
            default=candidate_markets[: min(5, len(candidate_markets))],
            key='utm_outreach_markets',
        )

    recipient_rule_options = build_contact_rule_options(addressbook_df)
    recipient_left, recipient_mid, recipient_right = st.columns([2, 2, 2])
    with recipient_left:
        primary_rules = st.multiselect(
            'Primary recipient rules',
            recipient_rule_options,
            default=[
                rule for rule in ['Section: PlanIt Champion', 'Section: Market Key Contact']
                if rule in recipient_rule_options
            ],
            key='utm_outreach_primary_rules',
            help='Each selected rule contributes one recipient.',
        )
    with recipient_mid:
        cc_rules = st.multiselect(
            'CC recipient rules',
            recipient_rule_options,
            default=[
                rule for rule in ['Section: Digital Contact', 'Section: Search Contact']
                if rule in recipient_rule_options
            ],
            key='utm_outreach_cc_rules',
            help='Each selected rule contributes one CC recipient.',
        )
    with recipient_right:
        run_name = st.text_input(
            'Run name',
            value=f'UTM Adoption {datetime.now().strftime("%Y-%m-%d %H:%M")}',
            key='utm_outreach_run_name',
        )

    st.caption('Optional fixed CC contacts')
    optional_cc_cols = st.columns(len(OPTIONAL_UTM_CC_CONTACTS))
    selected_optional_cc = []
    for idx, contact in enumerate(OPTIONAL_UTM_CC_CONTACTS):
        checkbox_key = f"utm_optional_cc_{re.sub(r'[^a-z0-9]+', '_', contact['email'].lower())}"
        with optional_cc_cols[idx]:
            checked = st.checkbox(
                contact['label'],
                value=False,
                key=checkbox_key,
                help=contact['email'],
            )
        if checked:
            selected_optional_cc.append(contact['email'])

    template_options = {
        f"{row['name']} (#{int(row['id'])})": row
        for _, row in templates_df.iterrows()
    }
    selected_template_label = st.selectbox(
        'Email template',
        list(template_options.keys()),
        index=0 if template_options else None,
        key='utm_outreach_template_select',
    )
    selected_template_row = template_options[selected_template_label] if selected_template_label else None
    selected_template_id = int(selected_template_row['id']) if selected_template_row is not None else None

    if selected_template_row is not None:
        loaded_template_id = st.session_state.get('utm_outreach_loaded_template_id')
        if loaded_template_id != selected_template_id:
            st.session_state['utm_outreach_template_name'] = str(selected_template_row['name'])
            st.session_state['utm_outreach_subject_template'] = str(selected_template_row['subject_template'])
            st.session_state['utm_outreach_body_template'] = str(selected_template_row['body_template'])
            st.session_state['utm_outreach_loaded_template_id'] = selected_template_id

    template_name = st.text_input('Template name', key='utm_outreach_template_name')
    subject_template = st.text_input('Subject template', key='utm_outreach_subject_template')
    body_template = st.text_area('Body template', height=220, key='utm_outreach_body_template')

    template_action_left, template_action_right = st.columns([1, 4])
    with template_action_left:
        if st.button('Save Template', key='utm_outreach_save_template'):
            if not template_name.strip():
                st.error('Template name is required.')
            else:
                save_email_template(
                    UTM_OUTREACH_DB_PATH,
                    'UTM Adoption',
                    template_name,
                    subject_template,
                    body_template,
                )
                st.success('Template saved to SQLite.')
                st.rerun()
    with template_action_right:
        st.caption(
            'Available placeholders: '
            '`{market_code}`, `{market_name}`, `{status}`, `{scope_gaps}`, '
            '`{observations}`, `{issues_identified}`, `{context}`, `{next_steps}`.'
        )

    preview_rows = []
    manual_to_emails = st.session_state.get('utm_manual_to_emails', [])
    manual_cc_emails = st.session_state.get('utm_manual_cc_emails', [])
    for market_code in selected_outreach_markets:
        matching_rows = utm_notes_df[utm_notes_df['Market Code'].astype(str) == market_code]
        if matching_rows.empty:
            continue
        row = matching_rows.iloc[0]
        market_context = _market_context_from_row(row)
        recipient_info = resolve_market_recipients(
            addressbook_df,
            market_context['market_code'],
            market_context['market_name'],
            primary_rules,
            cc_rules,
        )
        combined_to_emails = merge_unique_emails(recipient_info['to_emails'], manual_to_emails)
        combined_cc_emails = merge_unique_emails(recipient_info['cc_emails'], manual_cc_emails)
        existing_cc = {email.lower() for email in combined_cc_emails}
        existing_to = {email.lower() for email in combined_to_emails}
        for extra_email in selected_optional_cc:
            if extra_email.lower() in existing_cc or extra_email.lower() in existing_to:
                continue
            combined_cc_emails.append(extra_email)
            existing_cc.add(extra_email.lower())
        rendered_subject = render_email_template(subject_template, market_context)
        rendered_body = render_email_template(body_template, market_context)
        preview_rows.append(
            {
                'market': market_code,
                'market_name': market_context['market_name'],
                'status': market_context['status'],
                'to': '; '.join(combined_to_emails),
                'cc': '; '.join(combined_cc_emails),
                'subject': rendered_subject,
                'body': rendered_body,
                'primary_count': len(combined_to_emails),
                'invalid_primary_count': len(combined_to_emails) != 2,
                'to_emails': combined_to_emails,
                'cc_emails': combined_cc_emails,
            }
        )

    st.caption('Preview')
    preview_df = pd.DataFrame(
        [
            {
                'Market': row['market'],
                'Market Name': row['market_name'],
                'Status': row['status'],
                'To': row['to'],
                'CC': row['cc'],
                'Subject': row['subject'],
                'Primary Count': row['primary_count'],
                'Ready': not row['invalid_primary_count'],
            }
            for row in preview_rows
        ]
    )
    if preview_df.empty:
        st.info('Select at least one market to preview outreach emails.')
    else:
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
        st.caption('Use `Create Outlook Drafts` first when testing. Direct send depends on desktop Outlook being open and responsive.')

    action_left, action_mid, action_right = st.columns([1, 1, 1])
    with action_left:
        create_drafts_clicked = st.button(
            'Create Outlook Drafts',
            key='utm_outreach_create_drafts',
        )
    with action_mid:
        send_emails_clicked = st.button(
            'Send Emails',
            key='utm_outreach_send_emails',
        )
    with action_right:
        sync_replies_clicked = st.button(
            'Sync Replies',
            key='utm_outreach_sync_replies',
        )

    if create_drafts_clicked or send_emails_clicked:
        ready_rows = preview_rows[:]
        warning_rows = [
            f"{row['market']} ({row['primary_count']} primary contacts)"
            for row in preview_rows
            if row['invalid_primary_count']
        ]
        if not ready_rows:
            st.error('No markets selected for draft creation.')
        else:
            send_mode = 'send' if send_emails_clicked else 'draft'
            filters_payload = {
                'statuses': selected_outreach_statuses,
                'markets': selected_outreach_markets,
                'primary_rules': primary_rules,
                'cc_rules': cc_rules,
                'optional_cc_emails': selected_optional_cc,
                'send_mode': send_mode,
            }
            run_id = create_outreach_run(
                UTM_OUTREACH_DB_PATH,
                'UTM Adoption',
                run_name,
                filters_payload,
                selected_template_id,
            )
            completed_count = 0
            failed_markets = []
            for row in ready_rows:
                try:
                    outlook_result = send_or_draft_outlook_email(
                        row['to_emails'],
                        row['cc_emails'],
                        row['subject'],
                        row['body'],
                        send_mode=send_mode,
                    )
                    create_outreach_message(
                        UTM_OUTREACH_DB_PATH,
                        run_id,
                        row['market'],
                        row['status'],
                        row['to_emails'],
                        row['cc_emails'],
                        row['subject'],
                        row['body'],
                        outlook_result['sent_at'],
                        outlook_result['entry_id'],
                        outlook_result['conversation_id'],
                        outlook_result['sync_state'],
                    )
                    completed_count += 1
                except Exception as exc:
                    failed_markets.append(f"{row['market']}: {exc}")

            if completed_count:
                outcome = 'sent' if send_mode == 'send' else 'drafted'
                st.success(f'{completed_count} market emails {outcome} and tracked in SQLite.')
            if warning_rows:
                st.warning('These markets did not resolve to exactly 2 primary contacts and may need manual fixes in Outlook: ' + ', '.join(warning_rows))
            if failed_markets:
                st.error('Outlook failures: ' + ' | '.join(failed_markets))
            st.rerun()

    if sync_replies_clicked:
        try:
            synced_count = sync_outlook_replies(UTM_OUTREACH_DB_PATH, 'UTM Adoption')
            st.success(f'Synced reply status for {synced_count} tracked messages.')
            st.rerun()
        except Exception as exc:
            st.error(f'Reply sync failed: {exc}')

    tracker_df = load_outreach_tracker_df(UTM_OUTREACH_DB_PATH, 'UTM Adoption')
    st.caption('Tracker')
    if tracker_df.empty:
        st.info('No outreach messages have been tracked yet.')
    else:
        st.dataframe(
            tracker_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                'market': st.column_config.TextColumn('Market', width='small'),
                'current_utm_status': st.column_config.TextColumn('UTM Status', width='medium'),
                'sent_at': st.column_config.TextColumn('Sent At', width='medium'),
                'to': st.column_config.TextColumn('To', width='medium'),
                'cc': st.column_config.TextColumn('CC', width='medium'),
                'subject': st.column_config.TextColumn('Subject', width='large'),
                'sync_state': st.column_config.TextColumn('Sync State', width='small'),
                'replied': st.column_config.CheckboxColumn('Replied'),
                'last_reply_at': st.column_config.TextColumn('Last Reply At', width='medium'),
                'last_reply_from': st.column_config.TextColumn('Last Reply From', width='medium'),
                'last_reply_snippet': st.column_config.TextColumn('Last Reply Snippet', width='large'),
                'run_name': st.column_config.TextColumn('Run Name', width='medium'),
            },
        )
        thread_options = {
            f"{row['market']} | {row['subject']} | {row['sent_at']}": int(row['id'])
            for _, row in tracker_df.iterrows()
        }
        selected_thread_label = st.selectbox(
            'Thread viewer',
            list(thread_options.keys()),
            index=0,
            key='utm_outreach_thread_select',
        )
        selected_thread_id = thread_options[selected_thread_label]
        selected_thread_row = tracker_df[tracker_df['id'] == selected_thread_id].iloc[0]
        thread_df = load_outreach_thread_df(UTM_OUTREACH_DB_PATH, selected_thread_id)

        st.caption('Original tracked message')
        st.markdown(
            f"**Subject:** {selected_thread_row['subject']}\n\n"
            f"**To:** {selected_thread_row['to']}\n\n"
            f"**CC:** {selected_thread_row['cc']}\n\n"
            f"**Sent At:** {selected_thread_row['sent_at']}\n\n"
            f"**Body:**\n\n{selected_thread_row['body_rendered']}"
        )

        st.caption('Conversation items from Outlook')
        if thread_df.empty:
            st.info('No synced thread items yet for this tracked message.')
        else:
            for _, thread_row in thread_df.iterrows():
                sender_label = thread_row['sender_name'] or thread_row['sender_email'] or 'Unknown sender'
                with st.expander(f"{thread_row['received_at']} | {sender_label} | {thread_row['subject']}", expanded=False):
                    st.write(f"From: {sender_label}")
                    if thread_row['sender_email']:
                        st.write(f"Email: {thread_row['sender_email']}")
                    st.write(f"Received: {thread_row['received_at']}")
                    st.write(thread_row['body_text'] or thread_row['snippet'])

    st.divider()
    st.subheader('Market Engagement')
    st.caption('Manual workflow board for market communications. All inputs are stored in SQLite.')

    engagement_df = load_market_engagement_df(UTM_OUTREACH_DB_PATH)
    engagement_kpis = load_market_engagement_kpis(UTM_OUTREACH_DB_PATH)

    kpi_col_1, kpi_col_2, kpi_col_3, kpi_col_4, kpi_col_5, kpi_col_6 = st.columns(6)
    with kpi_col_1:
        st.metric('Markets', engagement_kpis['total_markets'])
    with kpi_col_2:
        st.metric('Engaged', engagement_kpis['engaged_markets'])
    with kpi_col_3:
        st.metric('Touchpoints', engagement_kpis['total_touchpoints'])
    with kpi_col_4:
        st.metric('Blocked', engagement_kpis['blocked_markets'])
    with kpi_col_5:
        st.metric('Resolved', engagement_kpis['resolved_markets'])
    with kpi_col_6:
        st.metric('Closed', engagement_kpis['closed_markets'])

    st.caption('Kanban board')
    board_columns = st.columns(len(UTM_ENGAGEMENT_STAGES))
    for idx, stage in enumerate(UTM_ENGAGEMENT_STAGES):
        stage_df = engagement_df[engagement_df['Stage'] == stage]
        with board_columns[idx]:
            st.markdown(f"**{stage}**")
            st.caption(f"{len(stage_df)} markets")
            if stage_df.empty:
                st.write('No markets')
            else:
                for _, stage_row in stage_df.iterrows():
                    due_label = stage_row['Next Action Due'] or 'No due date'
                    st.markdown(
                        f"**{stage_row['Market Code']}**  \n"
                        f"{stage_row['Issue Summary'] or stage_row['Market Name']}  \n"
                        f"`{stage_row['Priority']}` | {due_label}"
                    )

    engagement_market_options = [
        f"{row['Market Code']} | {row['Market Name']}"
        for _, row in engagement_df.iterrows()
    ]
    if engagement_market_options:
        selected_engagement_label = st.selectbox(
            'Engagement detail',
            engagement_market_options,
            index=0,
            key='utm_engagement_market_select',
        )
        selected_market_code = selected_engagement_label.split(' | ', 1)[0]
        selected_engagement_row = engagement_df[engagement_df['Market Code'] == selected_market_code].iloc[0]

        detail_left, detail_right = st.columns([2, 1])
        with detail_left:
            engagement_stage = st.selectbox(
                'Stage',
                UTM_ENGAGEMENT_STAGES,
                index=UTM_ENGAGEMENT_STAGES.index(selected_engagement_row['Stage']) if selected_engagement_row['Stage'] in UTM_ENGAGEMENT_STAGES else 0,
                key=f"utm_engagement_stage_{selected_market_code}",
            )
            engagement_owner = st.text_input(
                'Owner',
                value=selected_engagement_row['Owner'],
                key=f"utm_engagement_owner_{selected_market_code}",
            )
            engagement_issue_summary = st.text_input(
                'Issue summary',
                value=selected_engagement_row['Issue Summary'],
                key=f"utm_engagement_issue_summary_{selected_market_code}",
            )
            engagement_latest_note = st.text_area(
                'Latest note',
                value=selected_engagement_row['Latest Note'],
                height=180,
                key=f"utm_engagement_latest_note_{selected_market_code}",
            )
        with detail_right:
            engagement_priority = st.selectbox(
                'Priority',
                UTM_ENGAGEMENT_PRIORITIES,
                index=UTM_ENGAGEMENT_PRIORITIES.index(selected_engagement_row['Priority']) if selected_engagement_row['Priority'] in UTM_ENGAGEMENT_PRIORITIES else 1,
                key=f"utm_engagement_priority_{selected_market_code}",
            )
            engagement_next_action = st.text_area(
                'Next action',
                value=selected_engagement_row['Next Action'],
                height=120,
                key=f"utm_engagement_next_action_{selected_market_code}",
            )
            engagement_next_action_due = st.text_input(
                'Next action due',
                value=selected_engagement_row['Next Action Due'],
                key=f"utm_engagement_next_action_due_{selected_market_code}",
                help='Free-text date or deadline, for example 2026-03-24 or Friday EOD.',
            )
            st.caption(f"First contact: {selected_engagement_row['First Contact At'] or 'n/a'}")
            st.caption(f"Last contact: {selected_engagement_row['Last Contact At'] or 'n/a'}")
            st.caption(f"Touchpoints: {selected_engagement_row['Touchpoints']}")

        save_engagement_clicked = st.button('Save Engagement Update', key='utm_engagement_save')
        if save_engagement_clicked:
            save_market_engagement_update(
                UTM_OUTREACH_DB_PATH,
                selected_engagement_row['Market Code'],
                selected_engagement_row['Market Name'],
                engagement_stage,
                engagement_owner,
                engagement_priority,
                engagement_issue_summary,
                engagement_latest_note,
                engagement_next_action,
                engagement_next_action_due,
            )
            st.success('Engagement update saved to SQLite.')
            st.rerun()

        history_df = load_market_engagement_history_df(UTM_OUTREACH_DB_PATH, selected_engagement_row['Market Code'])
        st.caption('Update history')
        if history_df.empty:
            st.info('No manual updates logged yet for this market.')
        else:
            st.dataframe(history_df, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader('Outreach Database')
    st.caption('Local SQLite store for Outlook outreach state on the UTM Adoption workflow.')

    db_left, db_mid, db_right = st.columns([1, 1, 3])
    with db_left:
        st.metric('DB Contacts', outreach_db_counts['contacts'])
    with db_mid:
        st.metric('Templates', outreach_db_counts['email_templates'])
    with db_right:
        st.caption(f'Database: `{UTM_OUTREACH_DB_PATH}`')
        st.caption('UTM notes and address book are now served directly from SQLite on this page.')

    db_stat_left, db_stat_mid, db_stat_right = st.columns([1, 1, 1])
    with db_stat_left:
        st.metric('Runs', outreach_db_counts['outreach_runs'])
    with db_stat_mid:
        st.metric('Messages', outreach_db_counts['outreach_messages'])
    with db_stat_right:
        st.metric('Replies', outreach_db_counts['outreach_replies'])

    # -------------------------------------------------------------------------
    # Sent Thread Viewer
    # -------------------------------------------------------------------------
    st.divider()
    st.subheader('Sent Thread Viewer')
    st.caption(
        f'Fetches all Outlook emails whose subject contains "{UTM_THREAD_SUBJECT_KEYWORD}", '
        'retrieves every message in those conversation threads, and stores them locally for review.'
    )

    fetch_col, _ = st.columns([1, 3])
    with fetch_col:
        if st.button('Fetch / Refresh Threads from Outlook', key='utm_fetch_threads_btn'):
            with st.spinner('Connecting to Outlook and fetching threads…'):
                try:
                    fetched_count = fetch_and_store_utm_threads(UTM_OUTREACH_DB_PATH, UTM_THREAD_SUBJECT_KEYWORD)
                    st.success(f'Done — {fetched_count} thread(s) stored.')
                    st.rerun()
                except Exception as fetch_exc:
                    st.error(str(fetch_exc))

    utm_threads_df = load_utm_threads_df(UTM_OUTREACH_DB_PATH)

    if utm_threads_df.empty:
        st.info('No threads stored yet. Click "Fetch / Refresh Threads from Outlook" above.')
    else:
        thread_kpi_a, thread_kpi_b, thread_kpi_c = st.columns(3)
        with thread_kpi_a:
            st.metric('Threads', len(utm_threads_df))
        with thread_kpi_b:
            st.metric('Total Messages', int(utm_threads_df['Messages'].sum()))
        with thread_kpi_c:
            last_fetch = utm_threads_df['Fetched At'].max()
            st.metric('Last Fetched', last_fetch or '—')

        # --- Per-market contact stats ---
        st.markdown('**Markets contacted**')
        contact_stats_df = load_utm_thread_contact_stats_df(UTM_OUTREACH_DB_PATH, UTM_THREAD_EXCLUDED_REPLY_EMAILS)
        if not contact_stats_df.empty:
            responded_count = int(contact_stats_df['Responded'].sum())
            no_response_count = len(contact_stats_df) - responded_count
            stat_c1, stat_c2, stat_c3 = st.columns(3)
            with stat_c1:
                st.metric('Contacted', len(contact_stats_df))
            with stat_c2:
                st.metric('Responded', responded_count)
            with stat_c3:
                st.metric('No Response', no_response_count)

            display_df = contact_stats_df.drop(columns=['id']).copy()
            display_df['Responded'] = display_df['Responded'].map({True: 'Yes', False: 'No'})
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Subject': st.column_config.TextColumn('Subject', width='medium'),
                    'Sent At': st.column_config.TextColumn('Sent At', width='small'),
                    'To': st.column_config.TextColumn('To', width='large'),
                    'CC': st.column_config.TextColumn('CC', width='large'),
                    'Responded': st.column_config.TextColumn('Responded', width='small'),
                    'Replies': st.column_config.NumberColumn('Replies', width='small'),
                    'Replied By': st.column_config.TextColumn('Replied By', width='large'),
                },
            )
        st.divider()

        st.markdown('**View thread**')
        thread_labels = [
            f"{row['Subject']}  |  {row['First Sent']}  |  {row['Messages']} msg(s)"
            for _, row in utm_threads_df.iterrows()
        ]
        selected_thread_label = st.selectbox(
            'Select a thread to view',
            options=thread_labels,
            key='utm_thread_viewer_select',
            label_visibility='collapsed',
        )

        if selected_thread_label is not None:
            selected_thread_idx = thread_labels.index(selected_thread_label)
            selected_thread_id = int(utm_threads_df.iloc[selected_thread_idx]['id'])
            selected_thread_subject = utm_threads_df.iloc[selected_thread_idx]['Subject']
            selected_thread_participants = utm_threads_df.iloc[selected_thread_idx]['Participants']

            st.markdown(f'**Subject:** {selected_thread_subject}')
            st.caption(f'Participants: {selected_thread_participants}')

            thread_messages_df = load_utm_thread_messages_df(UTM_OUTREACH_DB_PATH, selected_thread_id)

            if thread_messages_df.empty:
                st.info('No messages found for this thread.')
            else:
                for _, msg_row in thread_messages_df.iterrows():
                    is_sent = msg_row['direction'] == 'sent'
                    avatar_name = 'You' if is_sent else (msg_row['sender_name'] or msg_row['sender_email'] or 'Them')
                    with st.chat_message(name=avatar_name):
                        header_parts = []
                        if msg_row['sender_name']:
                            header_parts.append(f"**{msg_row['sender_name']}**")
                        if msg_row['sender_email']:
                            header_parts.append(f"`{msg_row['sender_email']}`")
                        direction_tag = 'Sent' if is_sent else 'Reply'
                        st.markdown(' '.join(header_parts) + f'  —  _{direction_tag}_')
                        meta_parts = []
                        if msg_row['timestamp']:
                            meta_parts.append(msg_row['timestamp'])
                        if msg_row['to_emails']:
                            meta_parts.append(f"To: {msg_row['to_emails']}")
                        if msg_row['cc_emails']:
                            meta_parts.append(f"CC: {msg_row['cc_emails']}")
                        if meta_parts:
                            st.caption('  |  '.join(meta_parts))
                        body = (msg_row['body_text'] or '').strip()
                        if body:
                            with st.expander('Message body', expanded=is_sent):
                                st.text(body)
                        else:
                            st.caption('_(no body)_')

    st.stop()

if page == 'Taxonomy Hygiene':
    st.subheader('Taxonomy Hygiene')
    st.caption('Governance-focused review of PlanIT taxonomy structure, values, and conditional logic.')

    if not TAXONOMY_HYGIENE_PATH.exists():
        st.error(f'Taxonomy workbook not found: {TAXONOMY_HYGIENE_PATH}')
        st.stop()

    analysis = run_taxonomy_hygiene_analysis(str(TAXONOMY_HYGIENE_PATH))
    taxonomy_df = load_taxonomy_hygiene_workbook(str(TAXONOMY_HYGIENE_PATH))
    summary = analysis['executive_summary']

    top1, top2, top3, top4 = st.columns(4)
    top1.metric('Rows reviewed', f"{len(taxonomy_df):,}")
    top2.metric('Critical dimensions', summary['critical_dimensions'])
    top3.metric('High severity dimensions', summary['high_dimensions'])
    top4.metric('Design-smell values', summary['design_smell_values'])

    st.markdown('**Executive Summary**')
    st.write(summary['headline'])
    st.write(summary['practical_priority'])
    st.info(
        'This is a provisional governance assessment based on the workbook `Taxonomy Outputs` sheet. '
        'It treats placeholders as invalid input by default and flags workaround-driven values as design smells.'
    )

    st.markdown('**Main Findings**')
    findings = analysis['grouped_findings']
    for label, frame in [
        ('Excessive dropdown complexity', findings['excessive_dropdown_complexity']),
        ('Irrelevant cross-channel values', findings['irrelevant_cross_channel_values']),
        ('Bad fallback values', findings['bad_fallback_values']),
        ('Weak validation logic', findings['weak_validation_logic']),
        ('Future-state structural issues', findings['future_state_structural_issues']),
    ]:
        with st.expander(label, expanded=False):
            if frame.empty:
                st.write('No rows flagged in this category from the first-pass rules.')
            else:
                st.dataframe(frame, use_container_width=True, hide_index=True)

    with st.expander('Missing dimensions', expanded=True):
        st.dataframe(analysis['missing_dimensions'], use_container_width=True, hide_index=True)

    st.markdown('**Dimension Review Table**')
    severity_filter = st.multiselect(
        'Severity filter',
        options=['Critical', 'High', 'Medium', 'Low'],
        default=['Critical', 'High', 'Medium', 'Low'],
        key='taxonomy_severity_filter',
    )
    dim_df = analysis['dimension_review']
    if severity_filter:
        dim_df = dim_df[dim_df['Severity'].isin(severity_filter)]
    st.dataframe(dim_df, use_container_width=True, hide_index=True)

    st.markdown('**Value Review Table**')
    dimension_options = sorted(analysis['value_review']['Dimension'].dropna().unique().tolist())
    selected_dimension = st.selectbox('Value review dimension', dimension_options, index=0, key='taxonomy_value_dimension')
    value_df = analysis['value_review']
    value_df = value_df[value_df['Dimension'] == selected_dimension]
    st.dataframe(value_df, use_container_width=True, hide_index=True)

    st.markdown('**Proposed Validation Rules**')
    st.dataframe(analysis['validation_rules'], use_container_width=True, hide_index=True)

    st.markdown('**Current-State Recommendations**')
    st.dataframe(analysis['current_state'], use_container_width=True, hide_index=True)

    st.markdown('**Future-State Recommendations**')
    st.dataframe(analysis['future_state'], use_container_width=True, hide_index=True)

    st.markdown('**Stakeholder Questions / Sign-off Needs**')
    st.dataframe(analysis['stakeholder_questions'], use_container_width=True, hide_index=True)

    st.markdown('**Appendices**')
    st.dataframe(analysis['appendix'], use_container_width=True, hide_index=True)

    st.stop()

if page == 'Budget Setting Sankey':
    st.subheader('Budget Setting Sankey')
    st.caption('Interactive budget-flow view built from the budget-setting discussion: fixed budget, explicit highlight split, and adjustable uplift / reallocation assumptions.')

    default_total_budget = 123_343_635.95
    default_base_highlight_share = 19_149_513.34 / 123_343_635.95

    chart_col, control_col = st.columns([3.2, 1.2], vertical_alignment='top')

    with control_col:
        st.markdown('**Controls**')
        total_budget = st.number_input(
            'Total budget (EUR)',
            min_value=1_000_000.0,
            max_value=500_000_000.0,
            value=float(default_total_budget),
            step=1_000_000.0,
            format='%.2f',
            key='budget_sankey_total_budget',
        )
        base_highlight_share = st.slider(
            'Base highlight share',
            min_value=0.05,
            max_value=0.40,
            value=float(round(default_base_highlight_share, 4)),
            step=0.005,
            key='budget_sankey_base_highlight_share',
        )
        offline_upweight = st.slider(
            'Offline upweight',
            min_value=0.0,
            max_value=0.50,
            value=0.187,
            step=0.01,
            key='budget_sankey_offline_upweight',
            help='Reflects the point that non-digital investment also drives sessions but is not captured in digital impression counts.',
        )
        nonwebsite_upweight = st.slider(
            'Beyond-website upweight',
            min_value=0.0,
            max_value=0.60,
            value=0.20,
            step=0.01,
            key='budget_sankey_nonwebsite_upweight',
            help='Planning assumption to reflect sales influence beyond the OGS / website-touch path.',
        )
        click_credit_shift = st.slider(
            'Highlight credit shift',
            min_value=0.0,
            max_value=0.30,
            value=0.08,
            step=0.01,
            key='budget_sankey_click_credit_shift',
            help='Planning assumption to reflect click / session contribution from highlight activity that otherwise burdens Always On.',
        )
        highlight_upper_share = st.slider(
            'Highlight upper share',
            min_value=0.0,
            max_value=1.0,
            value=0.65,
            step=0.01,
            key='budget_sankey_highlight_upper_share',
            help='Explicit classification assumption. Highlight is not automatically all upper funnel.',
        )
        always_on_upper_share = st.slider(
            'Always On upper share',
            min_value=0.0,
            max_value=1.0,
            value=0.20,
            step=0.01,
            key='budget_sankey_always_on_upper_share',
            help='Allows Always On to contain both demand support and lower-funnel harvesting rather than behaving as a single block.',
        )
        upper_to_awareness_share = st.slider(
            'Upper to Awareness',
            min_value=0.0,
            max_value=1.0,
            value=0.60,
            step=0.01,
            key='budget_sankey_upper_to_awareness_share',
        )
        upper_to_consideration_share = 1.0 - upper_to_awareness_share
        lower_to_conversion_share = st.slider(
            'Lower to Conversion',
            min_value=0.0,
            max_value=1.0,
            value=0.70,
            step=0.01,
            key='budget_sankey_lower_to_conversion_share',
        )

    fig, sankey_metrics = build_budget_sankey_figure(
        total_budget=total_budget,
        base_highlight_share=base_highlight_share,
        offline_upweight=offline_upweight,
        nonwebsite_upweight=nonwebsite_upweight,
        click_credit_shift=click_credit_shift,
        highlight_upper_share=highlight_upper_share,
        always_on_upper_share=always_on_upper_share,
        upper_to_awareness_share=upper_to_awareness_share,
        upper_to_consideration_share=upper_to_consideration_share,
        lower_to_conversion_share=lower_to_conversion_share,
    )

    with chart_col:
        st.markdown('### Budget Decision Engine')
        st.caption('From budget envelope to full-funnel planning')
        top_a, top_b, top_c, top_d = st.columns(4)
        top_a.metric('Highlight share', f"{sankey_metrics['recommended_highlight_share'] * 100:.1f}%")
        top_b.metric('Awareness', format_meur(sankey_metrics['awareness_budget']))
        top_c.metric('Consideration', format_meur(sankey_metrics['consideration_budget']))
        top_d.metric('Conversion', format_meur(sankey_metrics['conversion_budget']))
        st.plotly_chart(fig, use_container_width=True)

    st.markdown('**Assumption logic used in the diagram**')
    st.write(
        'The current workbook implies a base highlight share. This page turns that into a fuller decision engine by: '
        '1) reweighting Highlight upward using the challenge levers, 2) explicitly classifying both Highlight and Always On into upper vs lower funnel, '
        'and 3) landing the full budget into Awareness, Consideration, and Conversion / Harvesting.'
    )

    assumption_df = pd.DataFrame([
        {
            'Assumption': 'Base highlight share from current model',
            'Value': f"{base_highlight_share * 100:.1f}%",
            'Interpretation': 'Current workbook-style starting point before challenge adjustments.',
        },
        {
            'Assumption': 'Offline / non-digital upweight',
            'Value': f"{offline_upweight * 100:.1f}%",
            'Interpretation': 'Reflects untracked non-digital contribution to session generation; 18.7% is the reference discussed.',
        },
        {
            'Assumption': 'Beyond-website sales influence upweight',
            'Value': f"{nonwebsite_upweight * 100:.1f}%",
            'Interpretation': 'Illustrative planning factor because the current session logic is anchored too narrowly on OGS / web-touch journeys.',
        },
        {
            'Assumption': 'Highlight-to-Always-On credit shift',
            'Value': f"{click_credit_shift * 100:.1f}%",
            'Interpretation': 'Reflects the point that highlight activity also contributes clicks / sessions that reduce lower-funnel burden.',
        },
        {
            'Assumption': 'Upper-funnel share within Highlight',
            'Value': f"{highlight_upper_share * 100:.1f}%",
            'Interpretation': 'Explicit classification assumption because Highlight is not automatically the same as Upper Funnel.',
        },
        {
            'Assumption': 'Upper-funnel share within Always On / Core',
            'Value': f"{always_on_upper_share * 100:.1f}%",
            'Interpretation': 'Prevents Always On from being treated as purely lower funnel and allows demand support within core activity.',
        },
        {
            'Assumption': 'Upper-funnel split to Awareness',
            'Value': f"{upper_to_awareness_share * 100:.1f}%",
            'Interpretation': 'Remaining upper-funnel budget lands in Consideration.',
        },
        {
            'Assumption': 'Lower-funnel split to Conversion / Harvesting',
            'Value': f"{lower_to_conversion_share * 100:.1f}%",
            'Interpretation': 'Remaining lower-funnel budget lands in Consideration.',
        },
    ])
    st.dataframe(assumption_df, use_container_width=True, hide_index=True)

    output_df = pd.DataFrame([
        {'Budget Bucket': 'Always On / Core', 'EUR': sankey_metrics['always_on_budget'], 'Share %': sankey_metrics['recommended_always_on_share'] * 100},
        {'Budget Bucket': 'Highlight Activations', 'EUR': sankey_metrics['highlight_budget'], 'Share %': sankey_metrics['recommended_highlight_share'] * 100},
        {'Budget Bucket': 'Always On Upper', 'EUR': sankey_metrics['upper_always_on_budget'], 'Share %': (sankey_metrics['upper_always_on_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Always On Lower', 'EUR': sankey_metrics['lower_always_on_budget'], 'Share %': (sankey_metrics['lower_always_on_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Upper Funnel Highlight', 'EUR': sankey_metrics['upper_highlight_budget'], 'Share %': (sankey_metrics['upper_highlight_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Lower Funnel Highlight', 'EUR': sankey_metrics['lower_highlight_budget'], 'Share %': (sankey_metrics['lower_highlight_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Awareness', 'EUR': sankey_metrics['awareness_budget'], 'Share %': (sankey_metrics['awareness_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Consideration', 'EUR': sankey_metrics['consideration_budget'], 'Share %': (sankey_metrics['consideration_budget'] / total_budget) * 100},
        {'Budget Bucket': 'Conversion / Harvesting', 'EUR': sankey_metrics['conversion_budget'], 'Share %': (sankey_metrics['conversion_budget'] / total_budget) * 100},
    ])
    st.markdown('**Numeric output**')
    st.dataframe(output_df.style.format({'EUR': '€{:,.0f}', 'Share %': '{:.1f}%'}), use_container_width=True)
    st.download_button(
        'Download Sankey output (CSV)',
        data=output_df.to_csv(index=False).encode('utf-8'),
        file_name='budget_setting_sankey_output.csv',
        mime='text/csv',
        key='download_budget_setting_sankey_output',
    )

    st.warning(
        'This view is an explicit planning model, not a claim that the workbook already calculates the split this way. '
        'It is intended to make the budget-setting assumptions visible and challengeable in one place.'
    )

    st.stop()

if page == 'Incentive Model':
    st.subheader('Incentive Model - Market Volatility & Remuneration')
    kpi_options = [
        'Cost per Lead (Forms Submission Started)',
        'Cost per Lead (DCFS)',
        'CPM',
    ]
    selected_kpi = st.selectbox('KPI', kpi_options, index=0, key='incentive_kpi')

    top_left, top_right = st.columns([3, 1])
    with top_left:
        with st.popover('How it works'):
            st.markdown(
                """
This flow uses cost KPIs only. Lower is better.

- Target: `T` is the expected cost KPI for a market.
- Actual: `A` is the realized cost KPI for a market.
- Deviation (cost): `delta = (T - A) / T`
- Adjustment:
    - Penalty: `max(delta, -F_dir)` when `delta < 0`
    - No change: `0` when `0 <= delta <= H`
    - Reward: `min(delta - H, F_up)` when `delta > H`
- Variable component: `V = (alpha_eff x C_BAH) + (beta_eff x C_FTE)`
- Adjusted variable fee: `V x (1 + Adjustment)`
- Total fee: `Fixed Fee + Adjusted Variable Fee`
"""
            )
    with top_right:
        st.caption('Lower cost = better performance.')

    if 'Market' not in df.columns:
        st.warning('Market column not found in the dataset.')
        st.stop()

    time_candidates = [c for c in ['calendar_week', 'Date', 'report_date'] if c in df.columns]
    if not time_candidates:
        st.warning('No time column found for time series.')
        st.stop()

    control_col, _ = st.columns([2, 1])
    with control_col:
        time_col = st.selectbox('Time column', time_candidates, key='incentive_time_col')
        market_options = ['All'] + sorted(df['Market'].dropna().unique())
        def _expand_calc_markets():
            selected = st.session_state.get('incentive_calc_markets', [])
            if 'All' in selected:
                st.session_state['incentive_calc_markets'] = [m for m in market_options if m != 'All']

        def _expand_viz_markets():
            selected = st.session_state.get('incentive_viz_markets', [])
            if 'All' in selected:
                current_calc = st.session_state.get('incentive_calc_markets', [])
                calc_only = [m for m in current_calc if m != 'All']
                st.session_state['incentive_viz_markets'] = calc_only

        calc_markets = st.multiselect(
            'Markets for calculation',
            market_options,
            default=['All'],
            key='incentive_calc_markets',
            on_change=_expand_calc_markets,
        )
        calc_selected = st.session_state.get('incentive_calc_markets', calc_markets)
        calc_list = [m for m in calc_selected if m != 'All'] if calc_selected else []
        viz_options = ['All'] + calc_list
        if calc_list:
            default_viz = ['All']
        else:
            default_viz = []
        viz_markets = st.multiselect(
            'Markets for visualisation (subset)',
            viz_options,
            default=default_viz,
            key='incentive_viz_markets',
            on_change=_expand_viz_markets,
        )
        aggregate_markets = st.checkbox('Aggregate all markets into one series', value=False)
        vol_method = 'CV (std/mean)'

    base = df.copy()
    if selected_kpi == 'Cost per Lead (Forms Submission Started)':
        base['denom'] = base.get('Forms Submission Started')
    elif selected_kpi == 'Cost per Lead (DCFS)':
        base['denom'] = base.get('DCFS')
    elif selected_kpi == 'CPM':
        base['denom'] = base.get('Impressions')

    if 'denom' not in base.columns:
        st.info('Selected KPI is not available for the dataset.')
        st.stop()
    base['row_kpi'] = base.apply(
        lambda r: _safe_ratio(r.get('Media Spend'), r.get('denom')) * (1000.0 if selected_kpi == 'CPM' else 1.0)
        if r.get('denom') else None,
        axis=1,
    )

    if calc_markets:
        if 'All' not in calc_markets:
            base = base[base['Market'].isin(calc_markets)]
    else:
        st.info('Select at least one market for calculation.')
        st.stop()

    if viz_markets:
        if 'All' not in viz_markets:
            viz_set = set(viz_markets)
            calc_set = set(base['Market'].dropna().astype(str).unique())
            if not viz_set.issubset(calc_set):
                st.warning('Visualisation markets must be a subset of calculation markets.')
            viz_base = base[base['Market'].isin(viz_markets)]
        else:
            viz_base = base
    else:
        viz_base = base

    if base.empty:
        st.info('No data available for the selected filters.')
        st.stop()

    def _order_time(frame, col):
        if col == 'calendar_week':
            ordered = get_calendar_week_options(frame)
            return ordered
        if col in ['Date', 'report_date']:
            return sorted(pd.to_datetime(frame[col], errors='coerce').dropna().unique())
        return sorted(frame[col].dropna().unique())

    def _compute_series(df_in, group_cols, multiplier=1.0):
        agg = df_in.groupby(group_cols, dropna=False).agg(
            spend_sum=('Media Spend', 'sum'),
            denom_sum=('denom', 'sum'),
            std_kpi=('row_kpi', 'std'),
            n=('row_kpi', 'count'),
        ).reset_index()
        agg['kpi_value'] = agg.apply(
            lambda r: ((r['spend_sum'] / r['denom_sum']) * multiplier) if r['denom_sum'] and r['denom_sum'] > 0 else None,
            axis=1,
        )
        agg['ci'] = 1.96 * agg['std_kpi'] / agg['n'].clip(lower=1) ** 0.5
        return agg

    multiplier = 1000.0 if selected_kpi == 'CPM' else 1.0
    if aggregate_markets:
        series = _compute_series(viz_base, [time_col], multiplier=multiplier)
        overall_mean = viz_base['row_kpi'].mean()
        overall_std = viz_base['row_kpi'].std()
        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=series[time_col],
                y=series['kpi_value'],
                mode='lines+markers',
                name='Time series',
                line=dict(width=2),
            )
        )
        if overall_mean is not None and not pd.isna(overall_mean):
            fig.add_trace(
                go.Scatter(
                    x=series[time_col],
                    y=[overall_mean] * len(series),
                    mode='lines',
                    name='Average (full period)',
                    line=dict(width=2, dash='dash'),
                )
            )
        if overall_mean is not None and overall_std is not None and not pd.isna(overall_std):
            upper = [overall_mean + overall_std] * len(series)
            lower = [overall_mean - overall_std] * len(series)
            fig.add_trace(
                go.Scatter(
                    x=series[time_col],
                    y=upper,
                    mode='lines',
                    line=dict(width=0),
                    showlegend=False,
                )
            )
            fig.add_trace(
                go.Scatter(
                    x=series[time_col],
                    y=lower,
                    mode='lines',
                    line=dict(width=0),
                    fill='tonexty',
                    fillcolor='rgba(0, 100, 200, 0.15)',
                    name='Volatility band (±1σ)',
                )
            )
        fig.update_layout(
            title=f'{selected_kpi} time series (aggregate) with full-period band',
            xaxis_title=time_col,
            yaxis_title=selected_kpi,
        )
        order = _order_time(series, time_col)
        fig.update_xaxes(categoryorder='array', categoryarray=order)
        st.plotly_chart(fig, use_container_width=True)
    else:
        series = _compute_series(viz_base, [time_col, 'Market'], multiplier=multiplier)
        fig = go.Figure()
        palette = px.colors.qualitative.Safe
        for idx, (market, group) in enumerate(series.groupby('Market', dropna=False)):
            color = palette[idx % len(palette)]
            fig.add_trace(
                go.Scatter(
                    x=group[time_col],
                    y=group['kpi_value'],
                    mode='lines+markers',
                    name=str(market),
                    line=dict(width=2, color=color),
                )
            )
            market_rows = viz_base[viz_base['Market'] == market]['row_kpi']
            m_mean = market_rows.mean()
            m_std = market_rows.std()
            if m_mean is not None and not pd.isna(m_mean):
                fig.add_trace(
                    go.Scatter(
                        x=group[time_col],
                        y=[m_mean] * len(group),
                        mode='lines',
                        line=dict(width=1, dash='dash', color=color),
                        showlegend=False,
                    )
                )
            if m_mean is not None and m_std is not None and not pd.isna(m_std):
                upper = [m_mean + m_std] * len(group)
                lower = [m_mean - m_std] * len(group)
                fig.add_trace(
                    go.Scatter(
                        x=group[time_col],
                        y=upper,
                        mode='lines',
                        line=dict(width=0),
                        showlegend=False,
                    )
                )
                fig.add_trace(
                    go.Scatter(
                        x=group[time_col],
                        y=lower,
                        mode='lines',
                        line=dict(width=0),
                        fill='tonexty',
                        fillcolor='rgba(0, 100, 200, 0.10)',
                        showlegend=False,
                    )
                )
        fig.update_layout(
            title=f'{selected_kpi} time series with full-period band',
            xaxis_title=time_col,
            yaxis_title=selected_kpi,
        )
        order = _order_time(series, time_col)
        fig.update_xaxes(categoryorder='array', categoryarray=order)
        st.plotly_chart(fig, use_container_width=True)

    st.subheader('Volatility to alpha/beta (by market)')
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        alpha = st.number_input('Base alpha', min_value=0.0, max_value=1.0, value=0.50, step=0.01, format='%.2f')
    with c2:
        beta = st.number_input('Base beta', min_value=0.0, max_value=1.0, value=0.20, step=0.01, format='%.2f')
    with c3:
        vol_scale = st.number_input('Volatility sensitivity (k)', min_value=0.0, value=1.0, step=0.1, format='%.2f')
    with c4:
        min_conf = st.number_input('Min confidence', min_value=0.0, max_value=1.0, value=0.60, step=0.05, format='%.2f')
    max_conf = st.number_input('Max confidence', min_value=0.0, max_value=1.0, value=1.00, step=0.05, format='%.2f')

    market_series = (
        base.groupby([time_col, 'Market'], dropna=False)
        .agg(mean_kpi=('row_kpi', 'mean'))
        .reset_index()
    )
    vol_series = market_series.groupby('Market')['mean_kpi']
    vol_df = vol_series.agg(mean='mean', std='std').reset_index()
    vol_df['volatility'] = vol_df.apply(
        lambda r: (r['std'] / r['mean']) if r['mean'] and r['mean'] != 0 else None,
        axis=1,
    )
    median_cv = vol_df['volatility'].median(skipna=True)
    vol_df['vol_ratio'] = vol_df['volatility'].apply(
        lambda v: None if v is None or pd.isna(v) or not median_cv else (v / median_cv)
    )
    vol_df['confidence'] = vol_df['vol_ratio'].apply(
        lambda r: None if r is None or pd.isna(r) else max(min_conf, min(max_conf, 1.0 - vol_scale * (r - 1.0)))
    )
    vol_df['alpha_eff'] = vol_df['confidence'].apply(lambda c: None if c is None else alpha * c)
    vol_df['beta_eff'] = vol_df['confidence'].apply(lambda c: None if c is None else beta * c)
    avg_series = market_series.groupby('Market')['mean_kpi'].mean()
    vol_df['avg_kpi'] = vol_df['Market'].map(avg_series)
    st.dataframe(
        vol_df[['Market', 'avg_kpi', 'volatility', 'vol_ratio', 'confidence', 'alpha_eff', 'beta_eff']]
        .sort_values('volatility', ascending=False, na_position='last'),
        use_container_width=True,
    )
    chart_df = vol_df.dropna(subset=['vol_ratio'])
    if not chart_df.empty:
        vol_fig = px.bar(
            chart_df.sort_values('vol_ratio', ascending=False),
            x='Market',
            y='vol_ratio',
            title='Volatility score by market (CV / median CV)',
            labels={'vol_ratio': 'Volatility score', 'Market': 'Market'},
        )
        vol_fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(vol_fig, use_container_width=True)

    st.subheader('Incentive model by market')
    f1, f2 = st.columns(2)
    with f1:
        C_BAH = st.number_input('BAH commission (C_BAH)', min_value=0.0, value=100000.0, step=1000.0, format='%.2f')
    with f2:
        C_FTE = st.number_input('FTE cost (C_FTE)', min_value=0.0, value=50000.0, step=1000.0, format='%.2f')
    g1, g2, g3 = st.columns(3)
    with g1:
        H = st.number_input('Hurdle H', min_value=0.0, value=0.10, step=0.01, format='%.2f')
    with g2:
        F_dir = st.number_input('Downside floor F_dir', min_value=0.0, value=0.50, step=0.01, format='%.2f')
    with g3:
        F_up = st.number_input('Upside ceiling F_up', min_value=0.0, value=0.50, step=0.01, format='%.2f')

    markets_list = vol_df['Market'].dropna().astype(str).tolist()
    target_quantile = st.selectbox(
        'Target quantile',
        ['25th percentile', 'Median', '75th percentile'],
        index=1,
        key='incentive_target_quantile',
    )
    target_adjust = st.number_input(
        'Target adjustment (%)',
        min_value=0.0,
        value=100.0,
        step=1.0,
        format='%.1f',
        help='100 = baseline target, 110 = +10%, 90 = -10%',
    )
    quantile_map = {
        '25th percentile': 0.25,
        'Median': 0.50,
        '75th percentile': 0.75,
    }
    q = quantile_map[target_quantile]
    target_series = market_series.groupby('Market')['mean_kpi'].quantile(q)
    target_series = target_series * (target_adjust / 100.0)
    input_df = pd.DataFrame({
        'Market': markets_list,
        'Target_T': vol_df['Market'].map(target_series).values,
        'Actual_A': vol_df['avg_kpi'].values,
    })
    input_df = st.data_editor(
        input_df,
        use_container_width=True,
        num_rows='fixed',
        key='incentive_market_inputs',
    )

    results = []
    alpha_map = dict(zip(vol_df['Market'].astype(str), vol_df['alpha_eff']))
    beta_map = dict(zip(vol_df['Market'].astype(str), vol_df['beta_eff']))
    for _, row in input_df.iterrows():
        market = str(row['Market'])
        T = float(row['Target_T']) if pd.notna(row['Target_T']) else None
        A = float(row['Actual_A']) if pd.notna(row['Actual_A']) else None
        if T is None or T <= 0 or A is None:
            continue
        alpha_eff = alpha_map.get(market, alpha)
        beta_eff = beta_map.get(market, beta)
        V = (alpha_eff * C_BAH) + (beta_eff * C_FTE)
        fixed_fee = ((1.0 - alpha_eff) * C_BAH) + ((1.0 - beta_eff) * C_FTE)
        delta = (T - A) / T
        if delta < 0:
            adjustment = max(delta, -F_dir)
            zone = 'Penalty'
        elif delta <= H:
            adjustment = 0.0
            zone = 'No change'
        else:
            adjustment = min(delta - H, F_up)
            zone = 'Reward'
        adjusted_variable_fee = V * (1.0 + adjustment)
        total_fee = fixed_fee + adjusted_variable_fee
        results.append({
            'Market': market,
            'Target_T': T,
            'Actual_A': A,
            'delta': delta,
            'Adjustment': adjustment,
            'alpha_eff': alpha_eff,
            'beta_eff': beta_eff,
            'Fixed_Fee': fixed_fee,
            'Variable_V': V,
            'Adjusted_Variable_Fee': adjusted_variable_fee,
            'Total_Fee': total_fee,
            'Zone': zone,
        })
    if results:
        results_df = pd.DataFrame(results)
        results_df['delta'] = results_df['delta'].map(lambda x: f'{x*100.0:.2f}%')
        results_df['Adjustment'] = results_df['Adjustment'].map(lambda x: f'{x*100.0:.2f}%')
        st.dataframe(results_df, use_container_width=True)

    if st.button('Generate incentive narrative (technical)', key='incentive_narrative'):
        calc_markets_list = [m for m in (calc_markets or []) if m != 'All']
        viz_markets_list = [m for m in (viz_markets or []) if m != 'All']
        report = INCENTIVE_MARKET_NARRATIVE_TEMPLATE
        report = report.replace('[KPI]', selected_kpi)
        report = report.replace('[TIME_COL]', str(time_col))
        report = report.replace('[CALC_MARKETS]', ', '.join(map(str, calc_markets_list)) if calc_markets_list else 'All')
        report = report.replace('[VIZ_MARKETS]', ', '.join(map(str, viz_markets_list)) if viz_markets_list else 'All')
        report = report.replace('[VOL_K]', f"{vol_scale:.2f}")
        report = report.replace('[MIN_CONF]', f"{min_conf:.2f}")
        report = report.replace('[MAX_CONF]', f"{max_conf:.2f}")
        report = report.replace('[ALPHA]', f"{alpha:.2f}")
        report = report.replace('[BETA]', f"{beta:.2f}")
        report = report.replace('[TARGET_Q_LABEL]', target_quantile)
        report = report.replace('[TARGET_Q]', f"{q:.2f}")
        report = report.replace('[TARGET_ADJ]', f"{target_adjust:.1f}%")
        report = report.replace('[C_BAH]', f"{C_BAH:,.2f}")
        report = report.replace('[C_FTE]', f"{C_FTE:,.2f}")
        report = report.replace('[H]', f"{H:.2f}")
        report = report.replace('[F_DIR]', f"{F_dir:.2f}")
        report = report.replace('[F_UP]', f"{F_up:.2f}")
        if results:
            lines = []
            for _, row in results_df.iterrows():
                lines.append(
                    f"{row['Market']}: "
                    f"Target={row['Target_T']:.2f}, Actual={row['Actual_A']:.2f}, "
                    f"delta={row['delta']}, adjustment={row['Adjustment']}, "
                    f"alpha_eff={row['alpha_eff']:.2f}, beta_eff={row['beta_eff']:.2f}, "
                    f"Variable_V={row['Variable_V']:.2f}, Fixed_Fee={row['Fixed_Fee']:.2f}, "
                    f"Adjusted_Variable_Fee={row['Adjusted_Variable_Fee']:.2f}, "
                    f"Total_Fee={row['Total_Fee']:.2f}, Zone={row['Zone']}"
                )
            report = report.replace('[MARKET_TABLE]', '\n'.join(lines))
        else:
            report = report.replace('[MARKET_TABLE]', 'n/a')
        st.text_area('Incentive narrative (copy)', report, height=520)
    st.stop()

if page == 'Weekly Market KPIs':
    st.subheader('Weekly Market KPIs')
    data_path = BASE_DIR / 'other_data' / 'Extracted_weekly_market_data__approx_.csv'
    new_data_path = BASE_DIR / 'other_data' / 'weekly_market_kpi_table.csv'
    if new_data_path.exists():
        weekly_df = pd.read_csv(new_data_path)
        required_cols = {'market', 'week', 'channel'}
        if not required_cols.issubset(set(weekly_df.columns)):
            st.warning('CSV missing required columns: market, week, channel')
            st.stop()
        weekly_df['channel'] = weekly_df['channel'].astype(str)
        weekly_df.loc[
            weekly_df['channel'].str.contains('cross channel', case=False, na=False)
            | weekly_df['channel'].str.contains('kpi selection', case=False, na=False),
            'channel',
        ] = 'Cross Channel'
        kpi_cols = [c for c in weekly_df.columns if c not in ['market', 'week', 'channel']]
    else:
        if not data_path.exists():
            st.warning(f'CSV not found: {data_path}')
            st.stop()
        weekly_df = pd.read_csv(data_path)
        required_cols = {'market', 'week'}
        if not required_cols.issubset(set(weekly_df.columns)):
            st.warning('CSV missing required columns: market, week')
            st.stop()
        kpi_cols = [c for c in weekly_df.columns if c not in ['market', 'week']]
    if not kpi_cols:
        st.warning('No KPI columns found in the CSV.')
        st.stop()

    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
    with c1:
        kpi_choice = st.selectbox('KPI', kpi_cols)
    with c2:
        market_options = ['All'] + sorted(weekly_df['market'].dropna().unique())
        selected_markets = st.multiselect('Markets', market_options, default=['All'])
    with c3:
        if 'channel' in weekly_df.columns:
            channel_options = ['All'] + sorted(weekly_df['channel'].dropna().unique())
            selected_channels = st.multiselect('Channels', channel_options, default=['All'])
        else:
            selected_channels = ['All']
    with c4:
        x_start = st.slider('Pre CTG start week (x)', min_value=1, max_value=22, value=1, step=1)

    plot_df = weekly_df.copy()
    if selected_markets and 'All' not in selected_markets:
        plot_df = plot_df[plot_df['market'].isin(selected_markets)]
    if 'channel' in plot_df.columns and selected_channels and 'All' not in selected_channels:
        normalized_channels = list(selected_channels)
        if 'Paid Search' in selected_channels and 'Paid Social' in selected_channels and 'Cross Channel' not in selected_channels:
            normalized_channels.append('Cross Channel')
        plot_df = plot_df[plot_df['channel'].isin(normalized_channels)]
        if selected_markets and 'All' not in selected_markets:
            desired_markets = set(selected_markets)
        else:
            desired_markets = set(weekly_df['market'].dropna().unique())
        present_markets = set(plot_df['market'].dropna().unique())
        missing_markets = desired_markets - present_markets
        if missing_markets:
            fallback = weekly_df[
                (weekly_df['market'].isin(missing_markets))
                & (weekly_df['channel'] == 'Cross Channel')
            ]
            if not fallback.empty:
                plot_df = pd.concat([plot_df, fallback], ignore_index=True)
        if (
            'Paid Search' in selected_channels
            and 'Paid Social' in selected_channels
            and 'GLOBAL' in set(weekly_df['market'].dropna().unique())
            and 'GLOBAL' not in set(plot_df['market'].dropna().unique())
        ):
            global_fallback = weekly_df[
                (weekly_df['market'] == 'GLOBAL') & (weekly_df['channel'] == 'Cross Channel')
            ]
            if not global_fallback.empty:
                plot_df = pd.concat([plot_df, global_fallback], ignore_index=True)
    if plot_df.empty:
        st.info('No data for the selected filters.')
        st.stop()

    group_cols = ['week', 'market'] + (['channel'] if 'channel' in plot_df.columns else [])
    agg_df = plot_df.groupby(group_cols, dropna=False)[kpi_choice].sum().reset_index()
    agg_df['week_str'] = agg_df['week'].astype(str)
    def _week_key(val: str):
        try:
            year, wk = str(val).split('-')
            return int(year) * 100 + int(wk)
        except Exception:
            return None
    week_order = (
        agg_df['week_str']
        .dropna()
        .unique()
        .tolist()
    )
    week_order = sorted(week_order, key=_week_key)
    color_col = 'market'
    if 'channel' in agg_df.columns and selected_channels and 'All' not in selected_channels:
        color_col = 'channel'
    fig = px.bar(
        agg_df,
        x='week_str',
        y=kpi_choice,
        color=color_col,
        barmode='group',
        category_orders={'week_str': week_order},
        title=f'{kpi_choice} by week',
    )
    fig.update_layout(xaxis_title='Week', yaxis_title=kpi_choice)
    fig.update_xaxes(type='category')
    st.plotly_chart(fig, use_container_width=True)

    for col in ['spend', 'icc_dcfs', 'sessions']:
        if col in plot_df.columns:
            plot_df[col] = pd.to_numeric(plot_df[col], errors='coerce')

    def _week_num(w):
        try:
            parts = str(w).split('-')
            return int(parts[1])
        except Exception:
            return None

    if 'spend' in plot_df.columns and 'icc_dcfs' in plot_df.columns:
        dist_df = plot_df.copy()
        dist_df['week_num'] = dist_df['week'].apply(_week_num)
        dist_df = dist_df.dropna(subset=['week_num'])
        dist_df['cost_per_dcfs'] = dist_df.apply(
            lambda r: (r['spend'] / r['icc_dcfs']) if r.get('icc_dcfs') and r['icc_dcfs'] > 0 else None,
            axis=1,
        )
        dist_df = dist_df.dropna(subset=['cost_per_dcfs'])
        dist_df['period'] = dist_df['week_num'].apply(
            lambda w: 'Pre CTG' if w is not None and w >= x_start and w <= 23 else 'Post CTG'
        )
        if not dist_df.empty:
            period_colors = {'Pre CTG': '#1f77b4', 'Post CTG': '#7fc7ff'}
            def _box_overlay(frame, title):
                avg_df = (
                    frame.groupby(['market', 'period'], dropna=False)['cost_per_dcfs']
                    .median()
                    .reset_index()
                )
                fig = go.Figure()
                for period in ['Pre CTG', 'Post CTG']:
                    sub = avg_df[avg_df['period'] == period]
                    fig.add_trace(
                        go.Bar(
                            x=sub['market'],
                            y=sub['cost_per_dcfs'],
                            name=period,
                            marker_color=period_colors.get(period),
                            opacity=0.25,
                            offsetgroup=period,
                        )
                    )
                for period in ['Pre CTG', 'Post CTG']:
                    sub = frame[frame['period'] == period]
                    fig.add_trace(
                        go.Box(
                            x=sub['market'],
                            y=sub['cost_per_dcfs'],
                            name=period,
                            marker_color=period_colors.get(period),
                            boxpoints='all',
                            jitter=0.3,
                            pointpos=0.0,
                            opacity=0.8,
                            offsetgroup=period,
                        )
                    )
                fig.update_layout(
                    barmode='group',
                    boxmode='group',
                    xaxis_title='Market',
                    yaxis_title='Cost per DCFS',
                    legend_title_text='Period',
                )
                st.plotly_chart(fig, use_container_width=True)

            non_global = dist_df[dist_df['market'] != 'GLOBAL']
            box_market_options = ['All'] + sorted(non_global['market'].dropna().unique().tolist())
            selected_box_markets = st.multiselect('Boxplot markets', box_market_options, default=['All'])
            if selected_box_markets and 'All' not in selected_box_markets:
                non_global = non_global[non_global['market'].isin(selected_box_markets)]
            if not non_global.empty:
                st.subheader('Cost per DCFS distributions (markets only)')
                _box_overlay(non_global, 'Cost per DCFS distributions (markets only)')
            global_only = dist_df[dist_df['market'] == 'GLOBAL']
            if not global_only.empty:
                st.subheader('Cost per DCFS distributions (GLOBAL only)')
                _box_overlay(global_only, 'Cost per DCFS distributions (GLOBAL only)')
    st.stop()

if page == 'CTG Pre/Post KPI per Session':
    st.subheader('CTG Pre vs Post: Cost per Session')
    data_path = BASE_DIR / 'other_data' / 'weekly_market_kpi_table.csv'
    if not data_path.exists():
        st.warning('Missing weekly_market_kpi_table.csv. Run the build script first.')
        st.stop()

    ctg_df = pd.read_csv(data_path)
    required_cols = {'market', 'channel', 'week', 'sessions', 'spend', 'icc_dcfs'}
    if not required_cols.issubset(set(ctg_df.columns)):
        st.warning('CSV missing required columns: market, channel, week, sessions, spend, icc_dcfs')
        st.stop()

    def _week_num(w):
        try:
            parts = str(w).split('-')
            return int(parts[1])
        except Exception:
            return None

    ctg_df['week_num'] = ctg_df['week'].apply(_week_num)
    ctg_df = ctg_df.dropna(subset=['week_num'])

    c1, c2, c3 = st.columns(3)
    with c1:
        kpi_choice = st.selectbox('KPI', ['sessions', 'icc_dcfs'])
    with c2:
        channel_choice = st.selectbox('Channel', ['Paid Search', 'Paid Social', 'Both'])
    with c3:
        x_start = st.slider('Pre CTG start week (x)', min_value=1, max_value=22, value=1, step=1)

    channel_map = {
        'Paid Search': 'Paid Search',
        'Paid Social': 'Paid Social',
    }
    ctg_df['channel'] = ctg_df['channel'].astype(str)
    ctg_df.loc[ctg_df['channel'].str.contains('cross channel', case=False, na=False), 'channel'] = 'Cross Channel'
    ctg_df.loc[ctg_df['channel'].str.strip().str.lower() == 'weeks', 'channel'] = 'Cross Channel'
    for col in ['sessions', 'spend', 'icc_dcfs']:
        ctg_df[col] = pd.to_numeric(ctg_df[col], errors='coerce')

    ctg_all = ctg_df.copy()
    cross_sessions = ctg_all[ctg_all['channel'] == 'Cross Channel']

    if channel_choice != 'Both':
        spend_base = ctg_df[ctg_df['channel'] == channel_map[channel_choice]]
    else:
        spend_base = ctg_df[ctg_df['channel'].isin(['Paid Search', 'Paid Social'])]

    if spend_base.empty:
        st.info('No spend data for the selected filters.')
        st.stop()

    def _period(df_in, start_week, end_week):
        return df_in[(df_in['week_num'] >= start_week) & (df_in['week_num'] <= end_week)]

    pre_spend = _period(spend_base, x_start, 23)
    post_spend = _period(spend_base, 24, 52)
    pre_sessions = _period(cross_sessions, x_start, 23)
    post_sessions = _period(cross_sessions, 24, 52)

    def _ratio(spend_df, sessions_df):
        if kpi_choice == 'sessions':
            spend_grouped = (
                spend_df.groupby('market', dropna=False)
                .agg(spend_sum=('spend', 'sum'))
                .reset_index()
            )
            sessions_grouped = (
                sessions_df.groupby('market', dropna=False)
                .agg(ses_sum=('sessions', 'sum'))
                .reset_index()
            )
            grouped = spend_grouped.merge(sessions_grouped, on='market', how='left')
            grouped['cost_per_session'] = grouped.apply(
                lambda r: (r['spend_sum'] / r['ses_sum']) if r['ses_sum'] and r['ses_sum'] > 0 else None,
                axis=1,
            )
        else:
            grouped = (
                spend_df.groupby('market', dropna=False)
                .agg(kpi_sum=(kpi_choice, 'sum'), spend_sum=('spend', 'sum'))
                .reset_index()
            )
            grouped['cost_per_session'] = grouped.apply(
                lambda r: (r['spend_sum'] / r['kpi_sum']) if r['kpi_sum'] and r['kpi_sum'] > 0 else None,
                axis=1,
            )
        return grouped[['market', 'cost_per_session']]

    agg_method = st.selectbox(
        'Aggregation method',
        ['Weighted (ratio of sums)', 'Unweighted (mean of markets)'],
        index=0,
    )

    st.subheader('Weekly cost distributions (raw points)')
    if kpi_choice == 'sessions':
        spend_week = (
            spend_base.groupby(['market', 'week', 'week_num'], dropna=False)
            .agg(spend_sum=('spend', 'sum'))
            .reset_index()
        )
        sessions_week = (
            cross_sessions.groupby(['market', 'week', 'week_num'], dropna=False)
            .agg(sessions_sum=('sessions', 'sum'))
            .reset_index()
        )
        dist_df = spend_week.merge(sessions_week, on=['market', 'week', 'week_num'], how='inner')
        dist_df['cost_per_session'] = dist_df.apply(
            lambda r: (r['spend_sum'] / r['sessions_sum']) if r['sessions_sum'] and r['sessions_sum'] > 0 else None,
            axis=1,
        )
        dist_df['cost_per_dcfs'] = None
    else:
        dist_df = spend_base.copy()
        dist_df['cost_per_session'] = dist_df.apply(
            lambda r: (r['spend'] / r['sessions']) if r.get('sessions') and r['sessions'] > 0 else None,
            axis=1,
        )
        dist_df['cost_per_dcfs'] = dist_df.apply(
            lambda r: (r['spend'] / r['icc_dcfs']) if r.get('icc_dcfs') and r['icc_dcfs'] > 0 else None,
            axis=1,
        )
    dist_df = dist_df.dropna(subset=['cost_per_session', 'cost_per_dcfs'], how='all')
    dist_df['period'] = dist_df['week_num'].apply(
        lambda w: 'Pre CTG' if w is not None and w >= x_start and w <= 23 else 'Post CTG'
    )
    cpl_pre = None
    cpl_post = None
    if dist_df.empty:
        st.info('No weekly points available for cost distributions.')
    else:
        period_colors = {'Pre CTG': '#1f77b4', 'Post CTG': '#7fc7ff'}
        if kpi_choice != 'sessions':
            cpl_pre = (
                dist_df[dist_df['period'] == 'Pre CTG']
                .groupby('market', dropna=False)['cost_per_dcfs']
                .median()
                .reset_index()
                .rename(columns={'cost_per_dcfs': 'avg_cpl_box'})
            )
            cpl_post = (
                dist_df[dist_df['period'] == 'Post CTG']
                .groupby('market', dropna=False)['cost_per_dcfs']
                .median()
                .reset_index()
                .rename(columns={'cost_per_dcfs': 'avg_cpl_box'})
            )
        def _overlay_box(metric_col, title, y_label):
            avg_df = (
                dist_df.groupby(['market', 'period'], dropna=False)[metric_col]
                .median()
                .reset_index()
            )
            fig = go.Figure()
            for period in ['Pre CTG', 'Post CTG']:
                sub = avg_df[avg_df['period'] == period]
                fig.add_trace(
                    go.Bar(
                        x=sub['market'],
                        y=sub[metric_col],
                        name=period,
                        marker_color=period_colors.get(period),
                        opacity=0.25,
                        offsetgroup=period,
                    )
                )
            for period in ['Pre CTG', 'Post CTG']:
                sub = dist_df[dist_df['period'] == period]
                fig.add_trace(
                    go.Box(
                        x=sub['market'],
                        y=sub[metric_col],
                        name=period,
                        marker_color=period_colors.get(period),
                        boxpoints='all',
                        jitter=0.3,
                        pointpos=0.0,
                        opacity=0.8,
                        offsetgroup=period,
                    )
                )
            fig.update_layout(
                barmode='group',
                boxmode='group',
                title=title,
                xaxis_title='Market',
                yaxis_title=y_label,
                legend_title_text='Period',
            )
            st.plotly_chart(fig, use_container_width=True)

        if kpi_choice == 'sessions':
            _overlay_box('cost_per_session', 'Cost per session (bars = average, boxes = distribution)', 'Cost per session')
        else:
            _overlay_box('cost_per_dcfs', 'Cost per DCFS (bars = average, boxes = distribution)', 'Cost per DCFS')

    if agg_method == 'Unweighted (mean of markets)':
        metric_col = 'cost_per_session' if kpi_choice == 'sessions' else 'cost_per_dcfs'
        avg_weekly = (
            dist_df.groupby(['market', 'period'], dropna=False)[metric_col]
            .mean()
            .reset_index()
            .rename(columns={metric_col: 'kpi_per_session'})
        )
        combined = avg_weekly
    else:
        pre_ratio = _ratio(pre_spend, pre_sessions).rename(columns={'cost_per_session': 'pre_ctg'})
        post_ratio = _ratio(post_spend, post_sessions).rename(columns={'cost_per_session': 'post_ctg'})
        combined = pre_ratio.merge(post_ratio, on='market', how='outer')
        combined = combined.melt(id_vars=['market'], value_vars=['pre_ctg', 'post_ctg'], var_name='period', value_name='kpi_per_session')
        combined['period'] = combined['period'].map({'pre_ctg': 'Pre CTG', 'post_ctg': 'Post CTG'})

    fig = px.bar(
        combined,
        x='market',
        y='kpi_per_session',
        color='period',
        barmode='group',
        title=f'Cost per session: Pre (x–23) vs Post (24–52)',
    )
    fig.update_layout(legend_title_text='Period')
    fig.update_layout(xaxis_title='Market', yaxis_title='Cost per session')
    st.plotly_chart(fig, use_container_width=True)

    st.subheader('Pre vs Post summary')
    st.caption('Delta is Post - Pre, with % change = (Post - Pre) / Pre.')

    def _period_metrics(spend_df, sessions_df):
        if agg_method == 'Unweighted (mean of markets)':
            spend_week = (
                spend_df.groupby(['market', 'week_num'], dropna=False)
                .agg(spend=('spend', 'sum'), leads=('icc_dcfs', 'sum'))
                .reset_index()
            )
            sessions_week = (
                sessions_df.groupby(['market', 'week_num'], dropna=False)
                .agg(sessions=('sessions', 'sum'))
                .reset_index()
            )
            weekly = spend_week.merge(sessions_week, on=['market', 'week_num'], how='left')
            weekly['cpl'] = weekly.apply(
                lambda r: (r['spend'] / r['leads']) if r['leads'] and r['leads'] > 0 else None, axis=1
            )
            weekly['cps'] = weekly.apply(
                lambda r: (r['spend'] / r['sessions']) if r['sessions'] and r['sessions'] > 0 else None, axis=1
            )
            weekly['leads_per_1k'] = weekly.apply(
                lambda r: (r['leads'] / r['spend'] * 1000.0) if r['spend'] else None, axis=1
            )
            weekly['sessions_per_1k'] = weekly.apply(
                lambda r: (r['sessions'] / r['spend'] * 1000.0) if r['spend'] else None, axis=1
            )
            grouped = (
                weekly.groupby('market', dropna=False)
                .agg(
                    avg_weekly_spend=('spend', 'mean'),
                    avg_weekly_leads=('leads', 'mean'),
                    avg_weekly_sessions=('sessions', 'mean'),
                    avg_cpl=('cpl', 'mean'),
                    avg_cps=('cps', 'mean'),
                    leads_per_1k=('leads_per_1k', 'mean'),
                    sessions_per_1k=('sessions_per_1k', 'mean'),
                )
                .reset_index()
            )
            return {
                'average weekly spend': grouped['avg_weekly_spend'].mean(),
                'average weekly leads': grouped['avg_weekly_leads'].mean(),
                'average weekly sessions': grouped['avg_weekly_sessions'].mean(),
                'average cpl in a week': grouped['avg_cpl'].mean(),
                'average cost per session in a week': grouped['avg_cps'].mean(),
                'leads per €1k': grouped['leads_per_1k'].mean(),
                'sessions per €1k': grouped['sessions_per_1k'].mean(),
            }

        sessions_by_market = (
            sessions_df.groupby('market', dropna=False)
            .agg(sessions_sum=('sessions', 'sum'))
            .reset_index()
        )
        weeks = spend_df['week_num'].nunique()
        spend_sum = spend_df['spend'].sum()
        leads_sum = spend_df['icc_dcfs'].sum()
        sessions_sum = sessions_by_market['sessions_sum'].sum()
        avg_weekly_spend = (spend_sum / weeks) if weeks else None
        avg_weekly_leads = (leads_sum / weeks) if weeks else None
        avg_weekly_sessions = (sessions_sum / weeks) if weeks else None
        avg_cpl = (spend_sum / leads_sum) if leads_sum else None
        avg_cps = (spend_sum / sessions_sum) if sessions_sum else None
        leads_per_1k = (leads_sum / spend_sum * 1000.0) if spend_sum else None
        sessions_per_1k = (sessions_sum / spend_sum * 1000.0) if spend_sum else None
        return {
            'average weekly spend': avg_weekly_spend,
            'average weekly leads': avg_weekly_leads,
            'average weekly sessions': avg_weekly_sessions,
            'average cpl in a week': avg_cpl,
            'average cost per session in a week': avg_cps,
            'leads per €1k': leads_per_1k,
            'sessions per €1k': sessions_per_1k,
        }

    # Overall summary uses GLOBAL only
    pre_metrics = _period_metrics(
        pre_spend[pre_spend['market'] == 'GLOBAL'],
        pre_sessions[pre_sessions['market'] == 'GLOBAL'],
    )
    post_metrics = _period_metrics(
        post_spend[post_spend['market'] == 'GLOBAL'],
        post_sessions[post_sessions['market'] == 'GLOBAL'],
    )
    if dist_df is not None and not dist_df.empty:
        global_dist = dist_df[dist_df['market'] == 'GLOBAL']
        if not global_dist.empty:
            pre_box = global_dist[global_dist['period'] == 'Pre CTG']
            post_box = global_dist[global_dist['period'] == 'Post CTG']
            if kpi_choice == 'sessions':
                if not pre_box.empty:
                    pre_metrics['average cost per session in a week'] = pre_box['cost_per_session'].median()
                if not post_box.empty:
                    post_metrics['average cost per session in a week'] = post_box['cost_per_session'].median()
            else:
                if not pre_box.empty:
                    pre_metrics['average cpl in a week'] = pre_box['cost_per_dcfs'].median()
                if not post_box.empty:
                    post_metrics['average cpl in a week'] = post_box['cost_per_dcfs'].median()
    # Average weekly spend/leads for GLOBAL
    def _avg_weekly(df_in, col):
        weekly = (
            df_in.groupby('week_num', dropna=False)
            .agg(val=(col, 'sum'))
            .reset_index()
        )
        return weekly['val'].mean() if not weekly.empty else None

    pre_metrics['average weekly spend'] = _avg_weekly(
        pre_spend[pre_spend['market'] == 'GLOBAL'], 'spend'
    )
    post_metrics['average weekly spend'] = _avg_weekly(
        post_spend[post_spend['market'] == 'GLOBAL'], 'spend'
    )
    # Derive average weekly leads from average spend / average CPL
    if pre_metrics.get('average weekly spend') and pre_metrics.get('average cpl in a week'):
        pre_metrics['average weekly leads'] = (
            pre_metrics['average weekly spend'] / pre_metrics['average cpl in a week']
        )
    if post_metrics.get('average weekly spend') and post_metrics.get('average cpl in a week'):
        post_metrics['average weekly leads'] = (
            post_metrics['average weekly spend'] / post_metrics['average cpl in a week']
        )
    if pre_metrics.get('average cpl in a week'):
        pre_metrics['leads per €1k'] = 1000.0 / pre_metrics['average cpl in a week']
    if post_metrics.get('average cpl in a week'):
        post_metrics['leads per €1k'] = 1000.0 / post_metrics['average cpl in a week']
    rows = []
    for metric, pre_val in pre_metrics.items():
        post_val = post_metrics.get(metric)
        if pre_val is None or post_val is None:
            delta = None
            pct = None
        else:
            delta = post_val - pre_val
            pct = (delta / pre_val) if pre_val else None
        rows.append({
            'metric': metric,
            'pre (week x-23)': pre_val,
            'post (week 24-52)': post_val,
            'delta (post-pre)': delta,
            'delta %': pct,
        })
    summary_df = pd.DataFrame(rows)
    summary_df['delta %'] = summary_df['delta %'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    st.dataframe(summary_df, use_container_width=True)

    st.subheader('Market breakdown (Sessions + CPL)')
    def _market_breakdown(spend_df, sessions_df, cpl_override=None):
        if agg_method == 'Unweighted (mean of markets)':
            spend_week = (
                spend_df.groupby(['market', 'week_num'], dropna=False)
                .agg(spend=('spend', 'sum'), leads=('icc_dcfs', 'sum'))
                .reset_index()
            )
            sessions_week = (
                sessions_df.groupby(['market', 'week_num'], dropna=False)
                .agg(sessions=('sessions', 'sum'))
                .reset_index()
            )
            weekly = spend_week.merge(sessions_week, on=['market', 'week_num'], how='left')
            weekly['avg_weekly_sessions'] = weekly['sessions']
            weekly['avg_cpl'] = weekly.apply(
                lambda r: (r['spend'] / r['leads']) if r['leads'] and r['leads'] > 0 else None, axis=1
            )
            grouped = (
                weekly.groupby('market', dropna=False)
                .agg(
                    avg_weekly_sessions=('avg_weekly_sessions', 'mean'),
                    avg_cpl=('avg_cpl', 'mean'),
                    avg_weekly_spend=('spend', 'mean'),
                )
                .reset_index()
            )
            if cpl_override is not None:
                grouped = grouped.drop(columns=['avg_cpl']).merge(cpl_override, on='market', how='left')
                grouped = grouped.rename(columns={'avg_cpl_box': 'avg_cpl'})
            grouped['avg_weekly_leads'] = grouped.apply(
                lambda r: (r['avg_weekly_spend'] / r['avg_cpl']) if r['avg_cpl'] else None, axis=1
            )
            return grouped[['market', 'avg_weekly_sessions', 'avg_cpl', 'avg_weekly_spend', 'avg_weekly_leads']]

        grouped = (
            spend_df.groupby('market', dropna=False)
            .agg(spend_sum=('spend', 'sum'), leads_sum=('icc_dcfs', 'sum'), weeks=('week_num', 'nunique'))
            .reset_index()
        )
        sessions_by_market = (
            sessions_df.groupby('market', dropna=False)
            .agg(sessions_sum=('sessions', 'sum'))
            .reset_index()
        )
        grouped = grouped.merge(sessions_by_market, on='market', how='left')
        grouped['avg_weekly_sessions'] = grouped.apply(
            lambda r: (r['sessions_sum'] / r['weeks']) if r['weeks'] else None, axis=1
        )
        grouped['avg_weekly_spend'] = grouped.apply(
            lambda r: (r['spend_sum'] / r['weeks']) if r['weeks'] else None, axis=1
        )
        grouped['avg_cpl'] = grouped.apply(
            lambda r: (r['spend_sum'] / r['leads_sum']) if r['leads_sum'] else None, axis=1
        )
        if cpl_override is not None:
            grouped = grouped.drop(columns=['avg_cpl']).merge(cpl_override, on='market', how='left')
            grouped = grouped.rename(columns={'avg_cpl_box': 'avg_cpl'})
        grouped['avg_weekly_leads'] = grouped.apply(
            lambda r: (r['avg_weekly_spend'] / r['avg_cpl']) if r['avg_cpl'] else None, axis=1
        )
        return grouped[['market', 'avg_weekly_sessions', 'avg_cpl', 'avg_weekly_spend', 'avg_weekly_leads']]

    pre_mkt = _market_breakdown(pre_spend, pre_sessions, cpl_override=cpl_pre).rename(columns={
        'avg_weekly_sessions': 'pre_avg_weekly_sessions',
        'avg_cpl': 'pre_avg_cpl',
        'avg_weekly_spend': 'pre_avg_weekly_spend',
        'avg_weekly_leads': 'pre_avg_weekly_leads',
    })
    post_mkt = _market_breakdown(post_spend, post_sessions, cpl_override=cpl_post).rename(columns={
        'avg_weekly_sessions': 'post_avg_weekly_sessions',
        'avg_cpl': 'post_avg_cpl',
        'avg_weekly_spend': 'post_avg_weekly_spend',
        'avg_weekly_leads': 'post_avg_weekly_leads',
    })
    mkt = pre_mkt.merge(post_mkt, on='market', how='outer')
    mkt['delta_sessions'] = mkt.apply(
        lambda r: (r['post_avg_weekly_sessions'] - r['pre_avg_weekly_sessions'])
        if pd.notna(r['post_avg_weekly_sessions']) and pd.notna(r['pre_avg_weekly_sessions']) else None,
        axis=1,
    )
    mkt['delta_sessions_%'] = mkt.apply(
        lambda r: ((r['post_avg_weekly_sessions'] - r['pre_avg_weekly_sessions']) / r['pre_avg_weekly_sessions'])
        if pd.notna(r['post_avg_weekly_sessions']) and pd.notna(r['pre_avg_weekly_sessions']) and r['pre_avg_weekly_sessions'] else None,
        axis=1,
    )
    mkt['delta_cpl'] = mkt.apply(
        lambda r: (r['post_avg_cpl'] - r['pre_avg_cpl'])
        if pd.notna(r['post_avg_cpl']) and pd.notna(r['pre_avg_cpl']) else None,
        axis=1,
    )
    mkt['delta_cpl_%'] = mkt.apply(
        lambda r: ((r['post_avg_cpl'] - r['pre_avg_cpl']) / r['pre_avg_cpl'])
        if pd.notna(r['post_avg_cpl']) and pd.notna(r['pre_avg_cpl']) and r['pre_avg_cpl'] else None,
        axis=1,
    )
    mkt['delta_sessions_%'] = mkt['delta_sessions_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    mkt['delta_cpl_%'] = mkt['delta_cpl_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    mkt['delta_spend'] = mkt.apply(
        lambda r: (r['post_avg_weekly_spend'] - r['pre_avg_weekly_spend'])
        if pd.notna(r['post_avg_weekly_spend']) and pd.notna(r['pre_avg_weekly_spend']) else None,
        axis=1,
    )
    mkt['delta_spend_%'] = mkt.apply(
        lambda r: ((r['post_avg_weekly_spend'] - r['pre_avg_weekly_spend']) / r['pre_avg_weekly_spend'])
        if pd.notna(r['post_avg_weekly_spend']) and pd.notna(r['pre_avg_weekly_spend']) and r['pre_avg_weekly_spend'] else None,
        axis=1,
    )
    mkt['delta_leads'] = mkt.apply(
        lambda r: (r['post_avg_weekly_leads'] - r['pre_avg_weekly_leads'])
        if pd.notna(r['post_avg_weekly_leads']) and pd.notna(r['pre_avg_weekly_leads']) else None,
        axis=1,
    )
    mkt['delta_leads_%'] = mkt.apply(
        lambda r: ((r['post_avg_weekly_leads'] - r['pre_avg_weekly_leads']) / r['pre_avg_weekly_leads'])
        if pd.notna(r['post_avg_weekly_leads']) and pd.notna(r['pre_avg_weekly_leads']) and r['pre_avg_weekly_leads'] else None,
        axis=1,
    )
    mkt['delta_spend_%'] = mkt['delta_spend_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    mkt['delta_leads_%'] = mkt['delta_leads_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    st.dataframe(mkt, use_container_width=True)

    if st.button('Generate CTG narrative (technical)', key='ctg_narrative'):
        report = CTG_PRE_POST_TEMPLATE
        report = report.replace('[KPI]', kpi_choice)
        report = report.replace('[CHANNEL]', channel_choice)
        report = report.replace('[X_START]', str(x_start))
        report = report.replace('[AGG_METHOD]', agg_method)
        table_lines = []
        for _, row in summary_df.iterrows():
            table_lines.append(
                f"{row['metric']}: pre={row['pre (week x-23)']}, "
                f"post={row['post (week 24-52)']}, "
                f"delta={row['delta (post-pre)']}, "
                f"delta%={row['delta %']}"
            )
        report = report.replace('[SUMMARY_TABLE]', '\n'.join(table_lines) if table_lines else 'n/a')
        mkt_lines = []
        for _, row in mkt.iterrows():
            mkt_lines.append(
                f"{row['market']}: "
                f"pre_spend={row['pre_avg_weekly_spend']}, post_spend={row['post_avg_weekly_spend']}, "
                f"delta_spend={row['delta_spend']}, delta_spend%={row['delta_spend_%']}; "
                f"pre_leads={row['pre_avg_weekly_leads']}, post_leads={row['post_avg_weekly_leads']}, "
                f"delta_leads={row['delta_leads']}, delta_leads%={row['delta_leads_%']}; "
                f"pre_sessions={row['pre_avg_weekly_sessions']}, post_sessions={row['post_avg_weekly_sessions']}, "
                f"delta_sessions={row['delta_sessions']}, delta_sessions%={row['delta_sessions_%']}; "
                f"pre_cpl={row['pre_avg_cpl']}, post_cpl={row['post_avg_cpl']}, "
                f"delta_cpl={row['delta_cpl']}, delta_cpl%={row['delta_cpl_%']}"
            )
        report = report.replace('[MARKET_BREAKDOWN]', '\n'.join(mkt_lines) if mkt_lines else 'n/a')
        st.text_area('CTG narrative (copy)', report, height=520)
    st.stop()

if page == 'CTG Pre/Post KPI per Session (Copy)':
    st.subheader('CTG Pre vs Post: Cost per DCFS (ctg_pre)')
    data_path = BASE_DIR / 'other_data' / 'ctg_pre_02' / 'matched_campaign_date_stats_pre_ctg.csv'
    post_data_path = BASE_DIR / 'other_data' / 'ctg_pre_02' / 'matched_campaign_date_stats_post_ctg.csv'
    if not data_path.exists():
        st.warning('Missing matched_campaign_date_stats_pre_ctg.csv.')
        st.stop()
    if not post_data_path.exists():
        st.warning('Missing matched_campaign_date_stats_post_ctg.csv.')
        st.stop()

    ctg_df = pd.read_csv(data_path)
    required_cols = {'market', 'model', 'week', 'spend', 'conv'}
    if not required_cols.issubset(set(ctg_df.columns)):
        st.warning('Pre CSV missing required columns: market, model, week, spend, conv')
        st.stop()
    if 'channel' not in ctg_df.columns:
        ctg_df['channel'] = 'Paid Search'
    ctg_df['dcfs'] = ctg_df['conv']

    def _week_num(w):
        try:
            parts = str(w).split('-')
            return int(parts[1])
        except Exception:
            return None

    ctg_df['week_num'] = ctg_df['week'].apply(_week_num)
    def _week_start_date(w):
        try:
            parts = str(w).split('-')
            year = int(parts[0])
            week = int(parts[1])
            return datetime.fromisocalendar(year, week, 1).date()
        except Exception:
            return None

    ctg_df['week_start'] = ctg_df['week'].apply(_week_start_date)
    cutoff_date = datetime(2025, 5, 1).date()
    ctg_df = ctg_df[ctg_df['week_start'].notna() & (ctg_df['week_start'] < cutoff_date)]
    ctg_df = ctg_df.dropna(subset=['week_num'])
    for col in ['spend', 'dcfs']:
        ctg_df[col] = pd.to_numeric(ctg_df[col], errors='coerce')

    market_map = {
        'CANADA': 'PCL',
        'UK': 'PCGB',
        'GERMANY': 'PD',
        'FRANCE': 'POF',
        'ITALY': 'PIT',
    }
    ctg_df['market'] = ctg_df['market'].astype(str)
    ctg_df['market'] = ctg_df['market'].apply(
        lambda m: market_map.get(m.strip().upper(), m)
    )
    allowed_markets = {'POF', 'PCL', 'PCGB', 'PD'}
    ctg_df = ctg_df[ctg_df['market'].isin(allowed_markets)]

    market_values = sorted(ctg_df['market'].dropna().unique().tolist())
    market_options = ['All'] + market_values
    if 'ctg_copy_markets' not in st.session_state:
        st.session_state['ctg_copy_markets'] = market_values

    def _expand_all_ctg_copy_markets():
        selected = st.session_state.get('ctg_copy_markets', [])
        if 'All' in selected:
            st.session_state['ctg_copy_markets'] = [m for m in market_options if m != 'All']

    model_values = sorted(ctg_df['model'].dropna().unique().tolist())
    model_options = ['All'] + model_values
    if 'ctg_copy_models' not in st.session_state:
        st.session_state['ctg_copy_models'] = model_values

    def _expand_all_ctg_copy_models():
        selected = st.session_state.get('ctg_copy_models', [])
        if 'All' in selected:
            st.session_state['ctg_copy_models'] = [m for m in model_options if m != 'All']

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        selected_models_ui = st.multiselect(
            'Model',
            model_options,
            key='ctg_copy_models',
            on_change=_expand_all_ctg_copy_models,
        )
    with c2:
        channel_choice = st.selectbox('Channel', ['Paid Search', 'Paid Social', 'Both'])
    with c3:
        x_start = st.slider('Pre CTG start week (x)', min_value=1, max_value=22, value=1, step=1)
    with c4:
        kpi_choice = st.selectbox('KPI', ['dcfs'])
    with c5:
        selected_markets_ui = st.multiselect(
            'Markets',
            market_options,
            key='ctg_copy_markets',
            on_change=_expand_all_ctg_copy_markets,
        )

    selected_markets = [m for m in selected_markets_ui if m != 'All']
    if not selected_markets:
        st.info('Choose at least one market.')
        st.stop()
    selected_models = [m for m in selected_models_ui if m != 'All']
    if not selected_models:
        st.info('Choose at least one model.')
        st.stop()

    plot_df = ctg_df.copy()
    selected_models_norm = {str(m).strip().lower() for m in selected_models}
    plot_df = plot_df[plot_df['model'].astype(str).str.strip().str.lower().isin(selected_models_norm)]

    channel_map = {
        'Paid Search': 'Paid Search',
        'Paid Social': 'Paid Social',
    }
    if channel_choice != 'Both':
        plot_df = plot_df[plot_df['channel'] == channel_map[channel_choice]]
    else:
        plot_df = plot_df[plot_df['channel'].isin(['Paid Search', 'Paid Social'])]

    if plot_df.empty:
        st.info('No data for the selected filters.')
        st.stop()

    plot_df = plot_df[plot_df['market'].isin(selected_markets)]

    def _period(df_in, start_week, end_week):
        return df_in[(df_in['week_num'] >= start_week) & (df_in['week_num'] <= end_week)]

    pre_spend = _period(plot_df, x_start, 23)
    # Use only valid rows where both spend and selected KPI are available.
    pre_spend = pre_spend[pre_spend['spend'].notna() & pre_spend[kpi_choice].notna()]
    post_spend = _period(plot_df, 24, 52)

    def _ratio(spend_df):
        grouped = (
            spend_df.groupby('market', dropna=False)
            .agg(kpi_sum=(kpi_choice, 'sum'), spend_sum=('spend', 'sum'))
            .reset_index()
        )
        grouped['cost_per_dcfs'] = grouped.apply(
            lambda r: (r['spend_sum'] / r['kpi_sum']) if r['kpi_sum'] and r['kpi_sum'] > 0 else None,
            axis=1,
        )
        return grouped[['market', 'cost_per_dcfs']]

    pre_ratio = _ratio(pre_spend).rename(columns={'cost_per_dcfs': 'pre_ctg'})
    pre_ratio = pre_ratio[pre_ratio['pre_ctg'].notna()]

    # Post source: matched campaign-date CSV (same schema treatment as pre)
    overview_df = pd.read_csv(post_data_path)
    post_required_cols = {'market', 'model', 'week', 'spend', 'conv'}
    if not post_required_cols.issubset(set(overview_df.columns)):
        st.warning('Post CSV missing required columns: market, model, week, spend, conv')
        st.stop()
    if 'channel' not in overview_df.columns:
        overview_df['channel'] = 'Paid Search'
    overview_df['dcfs'] = overview_df['conv']
    for col in ['spend', 'dcfs']:
        overview_df[col] = pd.to_numeric(overview_df[col], errors='coerce')
    overview_df['Model_norm'] = overview_df['model'].astype(str).str.strip().str.lower()
    overview_df['Market'] = overview_df['market'].astype(str).apply(
        lambda m: market_map.get(m.strip().upper(), m)
    )
    overview_df['Channel'] = overview_df['channel'].astype(str)
    overview_df = overview_df[overview_df['Model_norm'].isin(selected_models_norm)]
    if channel_choice != 'Both':
        overview_df = overview_df[overview_df['Channel'].astype(str).str.strip() == channel_choice]
    else:
        overview_df = overview_df[overview_df['Channel'].astype(str).str.strip().isin(['Paid Search', 'Paid Social'])]
    overview_df = overview_df[overview_df['Market'].isin(selected_markets)]

    spend_col = 'spend'
    kpi_col = 'dcfs'
    post_ratio = pd.DataFrame(columns=['market', 'post_ctg'])
    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        overview_df = overview_df[overview_df[spend_col].notna() & overview_df[kpi_col].notna()]
        by_market = (
            overview_df.groupby('Market', dropna=False)
            .agg(
                spend_sum=(spend_col, lambda s: s.sum(min_count=1)),
                dcfs_sum=(kpi_col, lambda s: s.sum(min_count=1)),
            )
            .reset_index()
        )
        by_market['post_ctg'] = by_market.apply(
            lambda r: (r['spend_sum'] / r['dcfs_sum']) if r['dcfs_sum'] and r['dcfs_sum'] > 0 else None,
            axis=1,
        )
        post_ratio = by_market.rename(columns={'Market': 'market'})[['market', 'post_ctg']]
        post_ratio = post_ratio[post_ratio['post_ctg'].notna()]
        total_spend = by_market['spend_sum'].sum(min_count=1)
        total_dcfs = by_market['dcfs_sum'].sum(min_count=1)
        global_post = (total_spend / total_dcfs) if total_dcfs else None
        post_ratio = pd.concat(
            [
                post_ratio[post_ratio['market'] != 'GLOBAL'],
                pd.DataFrame([{'market': 'GLOBAL', 'post_ctg': global_post}]),
            ],
            ignore_index=True,
        )

    pre_markets = set(pre_ratio['market'].dropna().unique())
    post_markets = set(post_ratio['market'].dropna().unique()) if not post_ratio.empty else set()
    common_markets = pre_markets & post_markets
    if common_markets:
        pre_ratio = pre_ratio[pre_ratio['market'].isin(common_markets)]
        post_ratio = post_ratio[post_ratio['market'].isin(common_markets)]
    combined = pre_ratio.merge(post_ratio, on='market', how='inner')
    combined = combined.melt(
        id_vars=['market'],
        value_vars=['pre_ctg', 'post_ctg'],
        var_name='period',
        value_name='kpi_per_session',
    )
    combined['period'] = combined['period'].map({'pre_ctg': 'Pre CTG', 'post_ctg': 'Post CTG'})

    show_log = st.checkbox('Log scale (y-axis)', value=False, key='ctg_pre_log_scale')
    fig = px.bar(
        combined,
        x='market',
        y='kpi_per_session',
        color='period',
        barmode='group',
        title='Pre CTG vs Post CTG',
        text='kpi_per_session',
    )
    fig.update_layout(legend_title_text='Period')
    fig.update_layout(xaxis_title='Market', yaxis_title='Cost per DCFS')
    fig.update_traces(texttemplate='%{text:.2f}', textposition='outside')
    if show_log:
        fig.update_yaxes(type='log')
    st.plotly_chart(fig, use_container_width=True)

    st.subheader('Pre vs Post summary (GLOBAL only)')
    def _avg_weekly(df_in, col):
        weekly = (
            df_in.groupby('week_num', dropna=False)
            .agg(val=(col, 'sum'))
            .reset_index()
        )
        return weekly['val'].mean() if not weekly.empty else None

    global_pre = pre_spend[pre_spend['market'] == 'GLOBAL']
    pre_avg_spend = _avg_weekly(global_pre, 'spend')
    pre_cpl = None
    if not pre_ratio.empty and 'GLOBAL' in pre_ratio['market'].values:
        pre_cpl = pre_ratio[pre_ratio['market'] == 'GLOBAL']['pre_ctg'].iloc[0]

    post_total_spend = None
    post_total_dcfs = None
    post_cpl = None
    post_avg_weekly_spend = None
    post_avg_weekly_dcfs = None
    post_weeks = None
    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        post_total_spend = overview_df[spend_col].sum(min_count=1)
        post_total_dcfs = overview_df[kpi_col].sum(min_count=1)
        post_cpl = (post_total_spend / post_total_dcfs) if post_total_dcfs else None
        week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        if week_col:
            post_weeks = overview_df[week_col].nunique()
            post_avg_weekly_spend = (post_total_spend / post_weeks) if post_weeks else None
            post_avg_weekly_dcfs = (post_total_dcfs / post_weeks) if post_weeks else None

    summary_rows = []
    def _add_row(metric, pre_val, post_val):
        if pre_val is None or post_val is None:
            delta = None
            pct = None
        else:
            delta = post_val - pre_val
            pct = (delta / pre_val) if pre_val else None
        summary_rows.append({
            'metric': metric,
            'pre (week x-23)': pre_val,
            'post (weekly avg over all weeks)': post_val,
            'delta (post-pre)': delta,
            'delta %': pct,
        })

    _add_row('average weekly spend', pre_avg_spend, post_avg_weekly_spend)
    if pre_avg_spend is not None and pre_cpl is not None:
        _add_row('average weekly leads', pre_avg_spend / pre_cpl, post_avg_weekly_spend / post_cpl if post_cpl else None)
    _add_row('average cpl in a week', pre_cpl, post_cpl)
    if pre_cpl is not None and post_cpl is not None:
        _add_row('leads per €1k', 1000.0 / pre_cpl, 1000.0 / post_cpl)

    summary_df = pd.DataFrame(summary_rows)
    summary_df['delta %'] = summary_df['delta %'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    st.dataframe(summary_df, use_container_width=True)

    st.subheader('Market breakdown (Pre = ctg_pre, Post = overview totals)')
    # Pre market metrics (ctg_pre): avg weekly spend + CPL from sums
    pre_weekly = (
        pre_spend.groupby(['market', 'week_num'], dropna=False)
        .agg(spend=('spend', 'sum'), dcfs=(kpi_choice, 'sum'))
        .reset_index()
    )
    pre_mkt = (
        pre_weekly.groupby('market', dropna=False)
        .agg(
            pre_spend_sum=('spend', 'sum'),
            pre_dcfs_sum=('dcfs', 'sum'),
            weeks_m=('week_num', 'nunique'),
        )
        .reset_index()
    )
    pre_mkt['pre_avg_weekly_spend'] = pre_mkt.apply(
        lambda r: (r['pre_spend_sum'] / r['weeks_m']) if r['weeks_m'] else None,
        axis=1,
    )
    pre_mkt['pre_avg_weekly_leads'] = pre_mkt.apply(
        lambda r: (r['pre_dcfs_sum'] / r['weeks_m']) if r['weeks_m'] else None,
        axis=1,
    )
    pre_mkt['pre_cpl'] = pre_mkt.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_dcfs_sum']) if r['pre_dcfs_sum'] else None,
        axis=1,
    )

    # Post market metrics (overview totals)
    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        agg_spec = {
            'post_spend_sum': (spend_col, lambda s: s.sum(min_count=1)),
            'post_dcfs_sum': (kpi_col, lambda s: s.sum(min_count=1)),
        }
        if week_col:
            agg_spec['post_weeks'] = (week_col, 'nunique')
        post_mkt = (
            overview_df.groupby('Market', dropna=False)
            .agg(**agg_spec)
            .reset_index()
            .rename(columns={'Market': 'market'})
        )
        post_mkt = post_mkt[post_mkt['market'].isin(pre_mkt['market'])]
        post_mkt['post_cpl'] = post_mkt.apply(
            lambda r: (r['post_spend_sum'] / r['post_dcfs_sum']) if r['post_dcfs_sum'] else None, axis=1
        )
        post_mkt['post_avg_weekly_spend'] = post_mkt.apply(
            lambda r: (r['post_spend_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
            axis=1,
        )
        post_mkt['post_avg_weekly_leads'] = post_mkt.apply(
            lambda r: (r['post_dcfs_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
            axis=1,
        )
    else:
        post_mkt = pd.DataFrame(columns=['market', 'post_spend_sum', 'post_dcfs_sum', 'post_cpl', 'post_avg_weekly_spend', 'post_avg_weekly_leads'])

    mkt = pre_mkt.merge(post_mkt, on='market', how='inner')
    mkt = mkt[mkt['pre_cpl'].notna() & mkt['post_cpl'].notna()]
    mkt['delta_cpl'] = mkt.apply(
        lambda r: (r['post_cpl'] - r['pre_cpl']) if pd.notna(r['post_cpl']) and pd.notna(r['pre_cpl']) else None,
        axis=1,
    )
    mkt['delta_cpl_%'] = mkt.apply(
        lambda r: ((r['post_cpl'] - r['pre_cpl']) / r['pre_cpl']) if pd.notna(r['post_cpl']) and pd.notna(r['pre_cpl']) and r['pre_cpl'] else None,
        axis=1,
    )
    mkt['delta_spend'] = mkt.apply(
        lambda r: (r['post_avg_weekly_spend'] - r['pre_avg_weekly_spend'])
        if pd.notna(r['post_avg_weekly_spend']) and pd.notna(r['pre_avg_weekly_spend']) else None,
        axis=1,
    )
    mkt['delta_spend_%'] = mkt.apply(
        lambda r: ((r['post_avg_weekly_spend'] - r['pre_avg_weekly_spend']) / r['pre_avg_weekly_spend'])
        if pd.notna(r['post_avg_weekly_spend']) and pd.notna(r['pre_avg_weekly_spend']) and r['pre_avg_weekly_spend'] else None,
        axis=1,
    )
    mkt['delta_leads'] = mkt.apply(
        lambda r: (r['post_avg_weekly_leads'] - r['pre_avg_weekly_leads'])
        if pd.notna(r['post_avg_weekly_leads']) and pd.notna(r['pre_avg_weekly_leads']) else None,
        axis=1,
    )
    mkt['delta_leads_%'] = mkt.apply(
        lambda r: ((r['post_avg_weekly_leads'] - r['pre_avg_weekly_leads']) / r['pre_avg_weekly_leads'])
        if pd.notna(r['post_avg_weekly_leads']) and pd.notna(r['pre_avg_weekly_leads']) and r['pre_avg_weekly_leads'] else None,
        axis=1,
    )
    mkt['delta_cpl_%'] = mkt['delta_cpl_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    mkt['delta_spend_%'] = mkt['delta_spend_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    mkt['delta_leads_%'] = mkt['delta_leads_%'].map(lambda v: f'{v*100.0:.2f}%' if v is not None else None)
    st.dataframe(mkt, use_container_width=True)

    st.subheader('Selected markets')
    # Pre totals by market (ctg_pre)
    pre_totals = (
        pre_spend.groupby('market', dropna=False)
        .agg(pre_spend_sum=('spend', 'sum'), pre_dcfs_sum=(kpi_choice, 'sum'), pre_weeks=('week_num', 'nunique'))
        .reset_index()
    )
    pre_totals['cpl_pre'] = pre_totals.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_dcfs_sum']) if r['pre_dcfs_sum'] and r['pre_dcfs_sum'] > 0 else None,
        axis=1,
    )
    pre_totals['pre_avg_weekly_spend'] = pre_totals.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_weeks']) if r['pre_weeks'] else None,
        axis=1,
    )
    pre_totals['pre_avg_weekly_dcfs'] = pre_totals.apply(
        lambda r: (r['pre_dcfs_sum'] / r['pre_weeks']) if r['pre_weeks'] else None,
        axis=1,
    )
    # Post totals by market (overview)
    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        agg_spec = {
            'post_spend_sum': (spend_col, lambda s: s.sum(min_count=1)),
            'post_dcfs_sum': (kpi_col, lambda s: s.sum(min_count=1)),
        }
        if week_col:
            agg_spec['post_weeks'] = (week_col, 'nunique')
        post_totals = (
            overview_df.groupby('Market', dropna=False)
            .agg(**agg_spec)
            .reset_index()
            .rename(columns={'Market': 'market'})
        )
    else:
        post_totals = pd.DataFrame(columns=['market', 'post_spend_sum', 'post_dcfs_sum'])
    post_totals['cpl_post'] = post_totals.apply(
        lambda r: (r['post_spend_sum'] / r['post_dcfs_sum']) if r['post_dcfs_sum'] and r['post_dcfs_sum'] > 0 else None,
        axis=1,
    )
    post_totals['post_avg_weekly_spend'] = post_totals.apply(
        lambda r: (r['post_spend_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
        axis=1,
    )
    post_totals['post_avg_weekly_dcfs'] = post_totals.apply(
        lambda r: (r['post_dcfs_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
        axis=1,
    )
    model_label = ', '.join(selected_models) if selected_models else 'All'
    market_table = (
        pre_totals.merge(post_totals, on='market', how='inner')
        .assign(model=model_label)
    )
    market_table = market_table.rename(columns={
        'market': 'market',
        'model': 'model',
        'pre_avg_weekly_dcfs': 'dcfs_pre',
        'post_avg_weekly_dcfs': 'dcfs_post',
        'pre_avg_weekly_spend': 'spend_pre',
        'post_avg_weekly_spend': 'spend_post',
    })[
        ['market', 'model', 'dcfs_pre', 'dcfs_post', 'spend_pre', 'spend_post', 'cpl_pre', 'cpl_post']
    ]
    market_table = market_table[market_table['cpl_pre'].notna() & market_table['cpl_post'].notna()]
    st.dataframe(market_table, use_container_width=True)

    st.subheader('Model breakdown (by market + model)')
    pre_totals_model = (
        pre_spend.groupby(['market', 'model'], dropna=False)
        .agg(pre_spend_sum=('spend', 'sum'), pre_dcfs_sum=(kpi_choice, 'sum'), pre_weeks=('week_num', 'nunique'))
        .reset_index()
    )
    pre_totals_model['cpl_pre'] = pre_totals_model.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_dcfs_sum']) if r['pre_dcfs_sum'] and r['pre_dcfs_sum'] > 0 else None,
        axis=1,
    )
    pre_totals_model['pre_avg_weekly_spend'] = pre_totals_model.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_weeks']) if r['pre_weeks'] else None,
        axis=1,
    )
    pre_totals_model['pre_avg_weekly_dcfs'] = pre_totals_model.apply(
        lambda r: (r['pre_dcfs_sum'] / r['pre_weeks']) if r['pre_weeks'] else None,
        axis=1,
    )
    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        agg_spec = {
            'post_spend_sum': (spend_col, lambda s: s.sum(min_count=1)),
            'post_dcfs_sum': (kpi_col, lambda s: s.sum(min_count=1)),
        }
        if week_col:
            agg_spec['post_weeks'] = (week_col, 'nunique')
        post_totals_model = (
            overview_df.groupby(['Market', 'Model_norm'], dropna=False)
            .agg(**agg_spec)
            .reset_index()
            .rename(columns={'Market': 'market', 'Model_norm': 'model'})
        )
    else:
        post_totals_model = pd.DataFrame(columns=['market', 'model', 'post_spend_sum', 'post_dcfs_sum'])
    post_totals_model['cpl_post'] = post_totals_model.apply(
        lambda r: (r['post_spend_sum'] / r['post_dcfs_sum']) if r['post_dcfs_sum'] and r['post_dcfs_sum'] > 0 else None,
        axis=1,
    )
    post_totals_model['post_avg_weekly_spend'] = post_totals_model.apply(
        lambda r: (r['post_spend_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
        axis=1,
    )
    post_totals_model['post_avg_weekly_dcfs'] = post_totals_model.apply(
        lambda r: (r['post_dcfs_sum'] / r['post_weeks']) if r.get('post_weeks') else None,
        axis=1,
    )
    model_table = (
        pre_totals_model.merge(post_totals_model, on=['market', 'model'], how='inner')
        .rename(columns={
            'pre_avg_weekly_dcfs': 'dcfs_pre',
            'post_avg_weekly_dcfs': 'dcfs_post',
            'pre_avg_weekly_spend': 'spend_pre',
            'post_avg_weekly_spend': 'spend_post',
        })[
            ['market', 'model', 'dcfs_pre', 'dcfs_post', 'spend_pre', 'spend_post', 'cpl_pre', 'cpl_post']
        ]
    )
    model_table = model_table[model_table['cpl_pre'].notna() & model_table['cpl_post'].notna()]
    st.dataframe(model_table, use_container_width=True)

    st.subheader('Carline breakdown (aggregated across all markets)')
    pre_totals_carline = (
        pre_spend.groupby('model', dropna=False)
        .agg(
            pre_spend_sum=('spend', 'sum'),
            pre_dcfs_sum=(kpi_choice, 'sum'),
            pre_weeks=('week_num', 'nunique'),
        )
        .reset_index()
    )
    pre_totals_carline['cpl_pre'] = pre_totals_carline.apply(
        lambda r: (r['pre_spend_sum'] / r['pre_dcfs_sum']) if r['pre_dcfs_sum'] and r['pre_dcfs_sum'] > 0 else None,
        axis=1,
    )
    pre_totals_carline['spend_pre'] = pre_totals_carline['pre_spend_sum']
    pre_totals_carline['dcfs_pre'] = pre_totals_carline['pre_dcfs_sum']

    if not overview_df.empty and spend_col in overview_df.columns and kpi_col in overview_df.columns:
        week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        agg_spec = {
            'post_spend_sum': (spend_col, lambda s: s.sum(min_count=1)),
            'post_dcfs_sum': (kpi_col, lambda s: s.sum(min_count=1)),
        }
        if week_col:
            agg_spec['post_weeks'] = (week_col, 'nunique')
        post_totals_carline = (
            overview_df.groupby('Model_norm', dropna=False)
            .agg(**agg_spec)
            .reset_index()
            .rename(columns={'Model_norm': 'model'})
        )
    else:
        post_totals_carline = pd.DataFrame(columns=['model', 'post_spend_sum', 'post_dcfs_sum', 'post_weeks'])

    post_totals_carline['cpl_post'] = post_totals_carline.apply(
        lambda r: (r['post_spend_sum'] / r['post_dcfs_sum']) if r['post_dcfs_sum'] and r['post_dcfs_sum'] > 0 else None,
        axis=1,
    )
    post_totals_carline['spend_post'] = post_totals_carline['post_spend_sum']
    post_totals_carline['dcfs_post'] = post_totals_carline['post_dcfs_sum']

    carline_table = (
        pre_totals_carline.merge(post_totals_carline, on='model', how='inner')
        [['model', 'dcfs_pre', 'dcfs_post', 'spend_pre', 'spend_post', 'cpl_pre', 'cpl_post']]
    )
    carline_table = carline_table[carline_table['cpl_pre'].notna() & carline_table['cpl_post'].notna()]
    if not carline_table.empty:
        total_dcfs_pre = carline_table['dcfs_pre'].sum(min_count=1)
        total_dcfs_post = carline_table['dcfs_post'].sum(min_count=1)
        total_spend_pre = carline_table['spend_pre'].sum(min_count=1)
        total_spend_post = carline_table['spend_post'].sum(min_count=1)
        total_row = {
            'model': 'TOTAL',
            'dcfs_pre': total_dcfs_pre,
            'dcfs_post': total_dcfs_post,
            'spend_pre': total_spend_pre,
            'spend_post': total_spend_post,
            'cpl_pre': (total_spend_pre / total_dcfs_pre) if total_dcfs_pre else None,
            'cpl_post': (total_spend_post / total_dcfs_post) if total_dcfs_post else None,
        }
        carline_table = pd.concat([carline_table, pd.DataFrame([total_row])], ignore_index=True)
    st.dataframe(carline_table, use_container_width=True)

    st.subheader('Model breakdown chart (all markets)')
    model_delta = model_table.copy()
    model_delta['delta_pct'] = model_delta.apply(
        lambda r: ((r['cpl_post'] - r['cpl_pre']) / r['cpl_pre'])
        if r['cpl_pre'] and r['cpl_post'] else None,
        axis=1,
    )
    heat_df = (
        model_delta.pivot(index='model', columns='market', values='delta_pct')
        .sort_index()
    )
    heat_text = heat_df.apply(lambda col: col.map(lambda v: f"{v*100:.1f}%" if pd.notna(v) else ""))
    fig_models = px.imshow(
        heat_df,
        color_continuous_scale=['#1a9850', '#f7f7f7', '#b2182b'],
        zmin=-0.5,
        zmax=0.5,
        text_auto=False,
        aspect='auto',
    )
    fig_models.update_traces(text=heat_text.values, texttemplate="%{text}", hovertemplate="Market=%{x}<br>Model=%{y}<br>Δ%=%{z:.2%}<extra></extra>")
    fig_models.update_layout(
        xaxis_title='Market',
        yaxis_title='Model',
        coloraxis_colorbar_title='Δ% (Post vs Pre)',
        height=420 + 40 * len(heat_df.index),
    )
    st.plotly_chart(fig_models, use_container_width=True)

    if st.button('Generate CTG narrative (copy)', key='ctg_pre_copy_narrative'):
        report = CTG_PRE_POST_TEMPLATE_COPY
        report = report.replace('[MARKETS]', ', '.join(selected_markets))
        report = report.replace('[MODEL]', model_label)
        report = report.replace('[CHANNEL]', channel_choice)
        report = report.replace('[X_START]', str(x_start))
        table_lines = []
        for _, row in summary_df.iterrows():
            table_lines.append(
                f"{row['metric']}: pre={row['pre (week x-23)']}, "
                f"post={row['post (weekly avg over all weeks)']}, "
                f"delta={row['delta (post-pre)']}, "
                f"delta%={row['delta %']}"
            )
        report = report.replace('[SUMMARY_TABLE]', '\n'.join(table_lines) if table_lines else 'n/a')
        mkt_lines = []
        for _, row in mkt.iterrows():
            mkt_lines.append(
                f"{row['market']}: "
                f"pre_spend={row.get('pre_avg_weekly_spend')}, post_spend={row.get('post_avg_weekly_spend')}, "
                f"delta_spend={row.get('delta_spend')}, delta_spend%={row.get('delta_spend_%')}; "
                f"pre_leads={row.get('pre_avg_weekly_leads')}, post_leads={row.get('post_avg_weekly_leads')}, "
                f"delta_leads={row.get('delta_leads')}, delta_leads%={row.get('delta_leads_%')}; "
                f"pre_cpl={row.get('pre_cpl')}, post_cpl={row.get('post_cpl')}, "
                f"delta_cpl={row.get('delta_cpl')}, delta_cpl%={row.get('delta_cpl_%')}"
            )
        report = report.replace('[MARKET_BREAKDOWN]', '\n'.join(mkt_lines) if mkt_lines else 'n/a')
        model_lines = []
        for _, row in model_table.iterrows():
            model_lines.append(
                f"{row['market']} / {row['model']}: "
                f"pre_spend={row['spend_pre']}, post_spend={row['spend_post']}, "
                f"pre_dcfs={row['dcfs_pre']}, post_dcfs={row['dcfs_post']}, "
                f"pre_cpl={row['cpl_pre']}, post_cpl={row['cpl_post']}"
            )
        report = report.replace('[MODEL_BREAKDOWN]', '\n'.join(model_lines) if model_lines else 'n/a')

        context_lines = []
        pre_weeks = pre_spend['week_num'].nunique() if 'week_num' in pre_spend.columns else None
        post_week_col = 'calendar_week' if 'calendar_week' in overview_df.columns else ('week' if 'week' in overview_df.columns else None)
        post_weeks = overview_df[post_week_col].nunique() if post_week_col else None
        context_lines.append(f"Pre weeks included: {pre_weeks if pre_weeks is not None else 'n/a'}")
        context_lines.append(f"Post weeks included: {post_weeks if post_weeks is not None else 'n/a'}")
        pre_total_spend = pre_spend['spend'].sum(min_count=1) if 'spend' in pre_spend.columns else None
        pre_total_dcfs = pre_spend[kpi_choice].sum(min_count=1) if kpi_choice in pre_spend.columns else None
        post_total_spend = overview_df[spend_col].sum(min_count=1) if spend_col in overview_df.columns else None
        post_total_dcfs = overview_df[kpi_col].sum(min_count=1) if kpi_col in overview_df.columns else None
        context_lines.append(f"Pre totals: spend={pre_total_spend}, conversions={pre_total_dcfs}")
        context_lines.append(f"Post totals: spend={post_total_spend}, conversions={post_total_dcfs}")
        if 'campaign_name' in pre_spend.columns:
            context_lines.append(f"Pre unique campaigns used (after filters): {pre_spend['campaign_name'].dropna().nunique()}")
        else:
            context_lines.append("Pre unique campaigns used (after filters): n/a")
        if 'campaign_name' in overview_df.columns:
            context_lines.append(f"Post unique campaigns used (after filters): {overview_df['campaign_name'].dropna().nunique()}")
        else:
            context_lines.append("Post unique campaigns used (after filters): n/a")
        pre_model_totals = (
            pre_spend.groupby('model', dropna=False)[kpi_choice]
            .sum(min_count=1)
            .reset_index()
        )
        post_model_totals = (
            overview_df.groupby('Model_norm', dropna=False)[kpi_col]
            .sum(min_count=1)
            .reset_index()
            .rename(columns={'Model_norm': 'model'})
        )
        context_lines.append("Pre conversions by model:")
        for _, row in pre_model_totals.iterrows():
            context_lines.append(f"  - {row['model']}: {row[kpi_choice]}")
        context_lines.append("Post conversions by model:")
        for _, row in post_model_totals.iterrows():
            context_lines.append(f"  - {row['model']}: {row[kpi_col]}")
        report = report + "\n\nData Context\n" + "\n".join(context_lines)
        st.text_area('CTG narrative (copy)', report, height=520)
    st.stop()

if page == 'Overview':
    st.subheader('KPI chart')
    with st.popover('What is this?'):
        st.write(
            'This page shows how your key performance indicators (KPIs) change by calendar week. '
            'Use the filters to pick which data you want to include, and the chart will update.'
        )
        st.write(
            'How to use it (beginner-friendly):\n'
            '1. Start with everything aggregated (default). This gives one clean line per KPI.\n'
            '2. Pick a KPI on the left axis (for example, Media Spend or DCFS).\n'
            '3. If you want to compare two KPIs, turn on Compare and pick a right-axis KPI.\n'
            '4. To focus on a subset, use the dropdowns. Select specific values instead of All.\n'
            '5. To split the line by a dimension, uncheck Aggregate next to that dropdown.'
        )
    dual_base = df.copy()
    breakdown_dims = dual_breakdown_dim or []

    for col, selections in dual_selections.items():
        if not selections:
            continue
        if 'All' in selections:
            continue
        value_selections = [s for s in selections if s != 'All']
        if value_selections:
            dual_base = dual_base[dual_base[col].isin(value_selections)]

    if dual_base.empty:
        st.warning('No data available for the selected filters.')
        st.stop()

    left_kpi = dual_left_kpi or (numeric_cols[0] if numeric_cols else None)
    right_kpi = dual_right_kpi

    group_cols = ['calendar_week'] + breakdown_dims

    agg_spec = {}
    for col in numeric_cols:
        if col in dual_base.columns:
            agg_spec[col] = 'sum'

    weekly = dual_base.groupby(group_cols, dropna=False).agg(agg_spec).reset_index()

    def _compute_kpi(frame, label):
        if label in frame.columns:
            return frame[label]
        if label == 'Cost per Lead (Forms Submission Started)':
            return frame.apply(lambda r: _safe_ratio(r['Media Spend'], r['Forms Submission Started']), axis=1)
        if label == 'Cost per Lead (DCFS)':
            return frame.apply(lambda r: _safe_ratio(r['Media Spend'], r['DCFS']), axis=1)
        if label == 'CPM':
            if 'Impressions' not in frame.columns:
                return pd.Series([None] * len(frame))
            return frame.apply(
                lambda r: (_safe_ratio(r['Media Spend'], r['Impressions']) or 0) * 1000
                if _safe_ratio(r['Media Spend'], r['Impressions']) is not None
                else None,
                axis=1,
            )
        return pd.Series([None] * len(frame))

    weekly['left_kpi'] = _compute_kpi(weekly, left_kpi)
    if right_kpi:
        weekly['right_kpi'] = _compute_kpi(weekly, right_kpi)

    week_order = get_calendar_week_options(dual_base)
    fig = make_subplots(specs=[[{"secondary_y": bool(right_kpi)}]])

    if breakdown_dims:
        palette = px.colors.qualitative.Plotly
        for idx, (key, group) in enumerate(weekly.groupby(breakdown_dims, dropna=False)):
            color = palette[idx % len(palette)]
            if isinstance(key, tuple):
                parts = [str(part) if part not in [None, ''] else 'Not specified' for part in key]
                name = ' | '.join(parts)
            else:
                name = str(key) if key not in [None, ''] else 'Not specified'
            fig.add_trace(
                go.Scatter(
                    x=group['calendar_week'],
                    y=group['left_kpi'],
                    mode='lines+markers',
                    name=f'{name} — {left_kpi}',
                    line=dict(color=color),
                ),
                secondary_y=False,
            )
            if right_kpi:
                fig.add_trace(
                    go.Scatter(
                        x=group['calendar_week'],
                        y=group['right_kpi'],
                        mode='lines+markers',
                        name=f'{name} — {right_kpi}',
                        line=dict(color=color, dash='dash'),
                    ),
                    secondary_y=True,
                )
    else:
        fig.add_trace(
            go.Scatter(
                x=weekly['calendar_week'],
                y=weekly['left_kpi'],
                mode='lines+markers',
                name=left_kpi,
            ),
            secondary_y=False,
        )
        if right_kpi:
            fig.add_trace(
                go.Scatter(
                    x=weekly['calendar_week'],
                    y=weekly['right_kpi'],
                    mode='lines+markers',
                    name=right_kpi,
                    line=dict(dash='dash'),
                ),
                secondary_y=True,
            )
    fig.update_layout(
        height=420,
        xaxis=dict(categoryorder='array', categoryarray=week_order),
        legend_title_text='KPI',
    )
    fig.update_yaxes(title_text=left_kpi, secondary_y=False)
    if right_kpi:
        fig.update_yaxes(title_text=right_kpi, secondary_y=True)
    st.plotly_chart(fig, use_container_width=True)

    def _total_kpi_value(label, series_name):
        # For cost-style KPIs, compute total as ratio of totals from filtered base data.
        if label == 'Cost per Lead (Forms Submission Started)':
            spend_total = pd.to_numeric(dual_base.get('Media Spend'), errors='coerce').sum(min_count=1)
            denom_total = pd.to_numeric(dual_base.get('Forms Submission Started'), errors='coerce').sum(min_count=1)
            return _safe_ratio(spend_total, denom_total)
        if label == 'Cost per Lead (DCFS)':
            spend_total = pd.to_numeric(dual_base.get('Media Spend'), errors='coerce').sum(min_count=1)
            denom_total = pd.to_numeric(dual_base.get('DCFS'), errors='coerce').sum(min_count=1)
            return _safe_ratio(spend_total, denom_total)
        if label == 'CPM':
            spend_total = pd.to_numeric(dual_base.get('Media Spend'), errors='coerce').sum(min_count=1)
            impressions_total = pd.to_numeric(dual_base.get('Impressions'), errors='coerce').sum(min_count=1)
            ratio = _safe_ratio(spend_total, impressions_total)
            return (ratio * 1000.0) if ratio is not None else None
        return pd.to_numeric(weekly[series_name], errors='coerce').sum(skipna=True)

    total_left = _total_kpi_value(left_kpi, 'left_kpi')
    if right_kpi and 'right_kpi' in weekly.columns:
        total_right = _total_kpi_value(right_kpi, 'right_kpi')
        c1, c2 = st.columns(2)
        with c1:
            st.metric(f'Total {left_kpi}', f'{total_left:,.2f}' if total_left is not None else 'n/a')
        with c2:
            st.metric(f'Total {right_kpi}', f'{total_right:,.2f}' if total_right is not None else 'n/a')
    else:
        st.metric(f'Total {left_kpi}', f'{total_left:,.2f}' if total_left is not None else 'n/a')

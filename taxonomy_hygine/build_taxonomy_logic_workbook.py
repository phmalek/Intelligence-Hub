from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "taxonomy_hygiene_logic_collaboration.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F1A24")
HEADER_FONT = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="EEE7F7")
SECTION_FONT = Font(name="Aptos", size=10, bold=True, color="4F2D7F")
BODY_FONT = Font(name="Aptos", size=10, color="1F1F1F")
MUTED_FONT = Font(name="Aptos", size=10, color="6B6B6B")
EXAMPLE_FILL = PatternFill("solid", fgColor="F8F4FC")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E8")
GUIDE_FILL = PatternFill("solid", fgColor="F4F6F8")
THIN_GREY = Side(style="thin", color="D6D6D6")
THIN_PURPLE = Side(style="thin", color="D9CCE8")


LOGIC_ROWS = [
    {
        "logic_area": "Placeholder values are bad inputs",
        "human_logic": "Values like Mixed, Other, Not Used, Unknown, N/A, or blank-like placeholders are weak input taxonomy choices by default. In most cases they should behave as reporting states or data-quality flags, not user-selectable inputs.",
        "why_it_matters": "These values weaken governance, reduce analytical value, and usually indicate a design gap or forced workaround.",
        "example_1": "Format contains Mixed (5,074) and Not Used (358).",
        "example_2": "Match Type contains Not Used (10,151), which behaves like a fallback rather than a meaningful input choice.",
        "typical_action": "Restrict or remove placeholder values; convert them to reporting-only or data-quality states.",
        "applies_to": "Format; Match Type; Dimensions Mix; Format Mix; other fields with fallback values",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "Critical",
    },
    {
        "logic_area": "Channel relevance matters",
        "human_logic": "Fields should only be used where they make business sense. Search-style fields should not behave like universal taxonomy dimensions across all channels.",
        "why_it_matters": "If a field is visible or mandatory where it does not belong, users will fill it with nonsense and reporting quality drops.",
        "example_1": "Match Type contains Broad, Exact, and Phrase, which are clearly search-style concepts.",
        "example_2": "Keyword Type / Messaging contains Brand, Generic, and Competitor, which again points to a search-specific field rather than a universal field.",
        "typical_action": "Make the field conditional by channel or sub-channel instead of globally available.",
        "applies_to": "Match Type; Keyword Type / Messaging",
        "channel_scope": "Search-led",
        "current_or_future": "Current-state feasible",
        "priority": "High",
    },
    {
        "logic_area": "Too-generic cross-channel design is a problem",
        "human_logic": "If a field is used globally but only has real meaning in a narrow set of cases, the field design is too generic and should be restricted or redesigned.",
        "why_it_matters": "Over-generic design creates long dropdowns, workaround values, and inconsistent user choices.",
        "example_1": "Buying Mode is almost entirely Not Used (11,216) with only Programmatic Open Inventory (43) showing meaningful use.",
        "example_2": "Dimensions Mix contains huge catch-all strings like 160x600 | 300x250 | 300x600 | 320x100 | 728x90 | 800x250 | 970x250, suggesting the field is trying to work too broadly across different execution realities.",
        "typical_action": "Restrict field use to relevant contexts or move redesign into future-state work.",
        "applies_to": "Buying Mode; Dimensions Mix",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "High",
    },
    {
        "logic_area": "Long dropdowns are suspicious unless clearly justified",
        "human_logic": "Very long option lists are a hygiene risk unless the values are truly distinct and operationally necessary.",
        "why_it_matters": "Long lists encourage user confusion, inconsistent naming, and weak downstream grouping.",
        "example_1": "Buying Platform includes a long list including Google - DV360, Facebook Business Manager, Google Ads, Direct Buy, LinkedIn Ads, Pinterest Ads, Reddit Ads, Amazon DSP, TikTok Ads, The Trade Desk, Microsoft Ads, Bing, Quantcast, YouTube Ads, SA360, GumGum, Snapchat Ads, and Vistar.",
        "example_2": "Supplier is similarly long and overlapping, including Google-DV360, Meta, Facebook & Instagram, Google Ads, Google, Instagram, LinkedIn, Pinterest, Facebook, Amazon, TikTok, Google-YouTube, and Bing.",
        "typical_action": "Review for consolidation, synonym cleanup, or clearer hierarchy.",
        "applies_to": "Buying Platform; Supplier",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "Medium",
    },
    {
        "logic_area": "Low-clarity values are a governance risk",
        "human_logic": "If users are unlikely to understand the difference between values, the field will not be populated consistently enough to support good reporting or governance.",
        "why_it_matters": "Ambiguous or overlapping values create unreliable segmentation and weak comparability.",
        "example_1": "Targeting mixes Keyword, Keyword Targeting, Audience, Custom Audience, Multiple, and Mixed, which are not cleanly distinct.",
        "example_2": "Audience Segment mixes precise values like Luxury Cars Interest and Website Engagement with vague values like Multiple, In-Market, Brands, and Social Media.",
        "typical_action": "Rename, collapse, or redefine overlapping values; add guidance or channel conditions if needed.",
        "applies_to": "Targeting; Audience Segment",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "High",
    },
    {
        "logic_area": "Workarounds are not valid structure",
        "human_logic": "If a value exists mainly because the taxonomy forced users into a workaround, it should be flagged as a design smell rather than accepted as legitimate structure.",
        "why_it_matters": "Workarounds hide design problems and make bad inputs look normal.",
        "example_1": "Format Mix contains - (4,783) and blank-like states (1,303), which look like workflow placeholders rather than real taxonomy choices.",
        "example_2": "Dimensions Mix contains - (4,601) and N/A (679), suggesting users are filling the field even when it is not meaningful.",
        "typical_action": "Flag as design smell; review whether the field should be optional, conditional, or redesigned.",
        "applies_to": "Format Mix; Dimensions Mix",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "High",
    },
    {
        "logic_area": "Missing dimensions matter as much as bad values",
        "human_logic": "Hygiene is not only about removal. If important business concepts exist operationally but are not captured cleanly, the taxonomy is still incomplete.",
        "why_it_matters": "Missing structure forces users into bad proxies and limits governance, reporting, and modelling value.",
        "example_1": "The data contains audience-source concepts across Targeting and Audience Segment such as 1st Party, 2nd Party, Look-A-Like, Retargeting, and Website Engagement, but there is no explicit Audience Source or Targeting Source field.",
        "example_2": "The data contains funnel-like intent signals across Planning Principle (Awareness, Consideration, Purchase / Leads) and Campaign Type & Phase (UPPER, MIDDLE, LOWER), but there is no single clean funnel-stage governance field.",
        "typical_action": "Add missing dimension or redesign related fields in future-state planning.",
        "applies_to": "Targeting; Audience Segment; Planning Principle; Campaign Type & Phase",
        "channel_scope": "Cross-channel",
        "current_or_future": "Future-state structural",
        "priority": "High",
    },
    {
        "logic_area": "Governance usefulness matters, not just frequency",
        "human_logic": "A value should not be removed only because it is low volume. Some niche values are operationally important even if they appear rarely.",
        "why_it_matters": "Frequency alone is a weak removal rule; business relevance still matters.",
        "example_1": "Buying Type includes low-volume but meaningful values like Cost Per Completed View (27) and Real Cost Per Click (12).",
        "example_2": "Buying Platform includes niche but plausible values like SA360 (27), GumGum (12), and Vistar (6), which may still matter operationally.",
        "typical_action": "Keep niche values if they are real and useful; require sign-off before removal.",
        "applies_to": "Buying Type; Buying Platform",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "Medium",
    },
    {
        "logic_area": "Current-state and future-state should be separated",
        "human_logic": "Quick wins should be separated from structural redesign so the team can move now without pretending all issues can be solved in the current system.",
        "why_it_matters": "Mixing easy fixes with large redesign makes the action plan unclear and slows down delivery.",
        "example_1": "Current-state cleanup is visible in fields like Format (Mixed, Not Used) and Dimensions Mix (-, N/A, Mixed).",
        "example_2": "Future-state redesign is more likely needed where the structure itself is weak, for example Audience Segment vs Targeting, or Planning Principle vs Campaign Type & Phase.",
        "typical_action": "Label recommendation as current-state feasible or future-state structural before implementation planning.",
        "applies_to": "Format; Dimensions Mix; Audience Segment; Targeting; Planning Principle; Campaign Type & Phase",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "Medium",
    },
    {
        "logic_area": "The output should lead to action",
        "human_logic": "The purpose of the dashboard is not to describe the taxonomy. It should point toward specific actions such as keep, restrict, remove, rename, make conditional, or add missing dimension.",
        "why_it_matters": "If the output is only descriptive, it will not help PlanIT owners or the wider team improve the taxonomy.",
        "example_1": "Match Type and Keyword Type / Messaging are strong candidates to become conditional fields rather than effectively universal fields.",
        "example_2": "Buying Mode, Format Mix, and Dimensions Mix are strong candidates for restriction, cleanup, or redesign because they carry many fallback or mixed-use values.",
        "typical_action": "Translate logic into clear implementation actions and assign owners / sign-off needs.",
        "applies_to": "Match Type; Keyword Type / Messaging; Buying Mode; Format Mix; Dimensions Mix",
        "channel_scope": "Cross-channel",
        "current_or_future": "Current-state feasible",
        "priority": "Critical",
    },
]


LIST_VALUES = {
    "channel_scope": [
        "Cross-channel",
        "Search-led",
        "Social-led",
        "Programmatic-led",
        "Display-led",
        "Inventory Ads-led",
        "Video-led",
        "Other",
    ],
    "current_or_future": [
        "Current-state feasible",
        "Future-state structural",
    ],
    "priority": [
        "Critical",
        "High",
        "Medium",
        "Low",
    ],
    "development_stage": [
        "Seeded from current dashboard logic",
        "Needs team review",
        "Agreed for implementation",
        "Needs stakeholder sign-off",
        "Future-state only",
        "Parked",
    ],
    "status": [
        "Open",
        "In review",
        "Refining",
        "Ready for implementation",
        "Waiting for sign-off",
        "Done",
        "Parked",
    ],
}


HEADERS = [
    "Logic Area",
    "Human Logic",
    "Why It Matters",
    "Concrete Example 1",
    "Concrete Example 2",
    "Typical Action",
    "Applies To Fields",
    "Channel Scope",
    "Current-State or Future-State",
    "Priority",
    "Development Stage",
    "Status",
    "Team Additions / New Thoughts",
    "Potential New Rule or Recommendation",
    "Owner / Reviewer",
    "Notes",
]


WIDTHS = {
    "A": 28,
    "B": 52,
    "C": 34,
    "D": 44,
    "E": 44,
    "F": 32,
    "G": 30,
    "H": 18,
    "I": 24,
    "J": 12,
    "K": 28,
    "L": 18,
    "M": 34,
    "N": 34,
    "O": 18,
    "P": 24,
}


def style_cell(cell, *, fill=None, font=None, border=None, alignment=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if alignment is not None:
        cell.alignment = alignment


def build_logic_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Logic Tracker"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws["A1"] = "Taxonomy Hygiene Logic Collaboration Tracker"
    ws["A2"] = "Broad human logic seeded from the current dashboard approach. Add new rules, concerns, and examples in the yellow collaboration columns."
    ws.merge_cells("A1:P1")
    ws.merge_cells("A2:P2")
    style_cell(
        ws["A1"],
        font=Font(name="Aptos", size=15, bold=True, color="2E1A47"),
        fill=PatternFill("solid", fgColor="F4EFFA"),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    style_cell(
        ws["A2"],
        font=Font(name="Aptos", size=10, italic=True, color="5C4C73"),
        fill=PatternFill("solid", fgColor="F9F6FC"),
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22

    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        style_cell(
            cell,
            fill=HEADER_FILL,
            font=HEADER_FONT,
            border=Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )
    ws.row_dimensions[3].height = 34

    start_row = 4
    for row_idx, item in enumerate(LOGIC_ROWS, start=start_row):
        values = [
            item["logic_area"],
            item["human_logic"],
            item["why_it_matters"],
            item["example_1"],
            item["example_2"],
            item["typical_action"],
            item["applies_to"],
            item["channel_scope"],
            item["current_or_future"],
            item["priority"],
            "Seeded from current dashboard logic",
            "Open",
            "",
            "",
            "",
            "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            fill = None
            font = BODY_FONT
            border = Border(left=THIN_PURPLE, right=THIN_PURPLE, top=THIN_PURPLE, bottom=THIN_PURPLE)
            if col_idx == 1:
                fill = SECTION_FILL
                font = SECTION_FONT
            elif col_idx in (4, 5):
                fill = EXAMPLE_FILL
            elif col_idx >= 13:
                fill = INPUT_FILL
            style_cell(
                cell,
                fill=fill,
                font=font,
                border=border,
                alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
            )
        ws.row_dimensions[row_idx].height = 72

    end_row = start_row + len(LOGIC_ROWS) - 1
    table = Table(displayName="TaxonomyLogicTracker", ref=f"A3:P{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)

    for col, width in WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = f"A3:P{end_row}"


def build_lists_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"
    columns = [
        ("A", "channel_scope"),
        ("B", "current_or_future"),
        ("C", "priority"),
        ("D", "development_stage"),
        ("E", "status"),
    ]
    for col_letter, key in columns:
        ws[f"{col_letter}1"] = key
        for idx, value in enumerate(LIST_VALUES[key], start=2):
            ws[f"{col_letter}{idx}"] = value


def add_validations(ws):
    end_row = 3 + len(LOGIC_ROWS)
    validations = {
        "H": "Lists!$A$2:$A$9",
        "I": "Lists!$B$2:$B$3",
        "J": "Lists!$C$2:$C$5",
        "K": "Lists!$D$2:$D$7",
        "L": "Lists!$E$2:$E$8",
    }
    for col, formula in validations.items():
        dv = DataValidation(type="list", formula1=f"={formula}", allow_blank=True)
        dv.prompt = "Choose from the dropdown."
        dv.error = "Select a value from the list."
        ws.add_data_validation(dv)
        dv.add(f"{col}4:{col}{end_row}")


def build_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "How to use this workbook"
    ws["A2"] = "Use each row as one broad logic area. Add your own rules, examples, and concerns in the collaboration columns."
    style_cell(ws["A1"], fill=GUIDE_FILL, font=Font(name="Aptos", size=14, bold=True, color="25324A"))
    style_cell(ws["A2"], fill=GUIDE_FILL, font=Font(name="Aptos", size=10, color="4A5568"))
    notes = [
        "Logic Area: short label for the governance idea.",
        "Human Logic: plain-English explanation of the logic, not code.",
        "Concrete Example 1 / 2: use real fields and values from the taxonomy input where possible.",
        "Team Additions / New Thoughts: use this to extend, challenge, or refine the existing logic.",
        "Potential New Rule or Recommendation: write the practical outcome you think this logic should lead to.",
        "Development Stage: use this like a ticketing-progress field.",
        "Status: overall state of the row right now.",
    ]
    for idx, text in enumerate(notes, start=4):
        ws[f"A{idx}"] = text
        style_cell(
            ws[f"A{idx}"],
            font=BODY_FONT if idx < 7 else MUTED_FONT,
            alignment=Alignment(wrap_text=True, vertical="top"),
        )
    ws.column_dimensions["A"].width = 120


def main():
    wb = Workbook()
    build_logic_sheet(wb)
    build_lists_sheet(wb)
    build_guide_sheet(wb)
    add_validations(wb["Logic Tracker"])
    wb.save(OUTPUT_PATH)
    print(f"Saved {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

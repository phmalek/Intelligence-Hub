from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


path = Path(r"C:\Users\ali\repos\porsche\taxonomy_hygine\_tmp_taxonomy_readcopy.xlsx")
wb = load_workbook(path, data_only=True, read_only=True)
print("worksheets", wb.sheetnames)

ws = wb["Taxonomy Outputs"]
header_row = 9
headers = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(header_row, c).value
    if val:
        headers[str(val)] = c
        print(c, val)

interesting = [
    "Planning Principle",
    "Campaign Type & Phase",
    "KPI Objective",
    "Buying Mode",
    "Format",
    "Format Mix",
    "Dimensions",
    "Dimensions Mix",
    "Buying Platform",
    "Supplier",
    "Buying Type",
    "Audience Segment",
    "Targeting",
    "Demographic",
    "Device",
    "Language",
    "Match Type",
    "Keyword Type / Messaging",
]

selected = {field: headers[field] for field in interesting if field in headers}
counters = {field: Counter() for field in selected}

for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    for field, col in selected.items():
        val = row[col - 1]
        if val is None:
            continue
        text = str(val).strip()
        if text:
            counters[field][text] += 1

for field in interesting:
    if field not in counters:
        continue
    print(f"--- {field} ---")
    for value, count in counters[field].most_common(20):
        print(count, value)

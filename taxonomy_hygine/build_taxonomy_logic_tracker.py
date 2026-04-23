from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = Path(__file__).resolve().parent
REPO_ROOT = BASE_DIR.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from taxonomy_analysis import (
    current_state_recommendations,
    future_state_recommendations,
    missing_dimensions,
    proposed_validation_rules,
)

OUTPUT_PATH = BASE_DIR / "taxonomy_hygiene_logic_tracker_v2.xlsx"


HEADERS = [
    "Ticket ID",
    "Dimension / Field",
    "Value / Logic Area",
    "Issue Type",
    "Current Problem",
    "Proposed Logic / Rule",
    "Channel Scope",
    "Current-State or Future-State",
    "Priority",
    "Owner",
    "Development Stage",
    "Status",
    "Dependencies / Risks",
    "Stakeholder Input Needed",
    "Dashboard Implementation Notes",
]


COLUMN_WIDTHS = {
    "A": 12,
    "B": 22,
    "C": 24,
    "D": 20,
    "E": 34,
    "F": 42,
    "G": 18,
    "H": 22,
    "I": 12,
    "J": 18,
    "K": 18,
    "L": 16,
    "M": 24,
    "N": 28,
    "O": 30,
}


ISSUE_TYPES = [
    "Too many values",
    "Irrelevant cross-channel usage",
    "Redundant with another field",
    "Ambiguous naming",
    "Free-text risk",
    "Low usage but important",
    "Placeholder / bad fallback value",
    "Missing controlled vocabulary",
    "Missing dimension",
    "Wrong granularity",
    "Reporting-only value used as input",
    "Design smell workaround",
]

CHANNEL_SCOPES = [
    "All",
    "Search",
    "Social",
    "Programmatic",
    "Display",
    "CTV",
    "Video",
    "Inventory Ads",
    "Direct Buy",
    "Multi-channel",
]

CURRENT_FUTURE = [
    "Current-State",
    "Future-State",
]

PRIORITIES = [
    "Critical",
    "High",
    "Medium",
    "Low",
]

DEV_STAGES = [
    "Backlog",
    "To Review",
    "Ready for Spec",
    "Ready for Build",
    "In Development",
    "In UAT",
    "Blocked",
    "Done",
]

STATUSES = [
    "Open",
    "In Progress",
    "Waiting for Input",
    "Blocked",
    "Closed",
]


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F1F1F")
    header_font = Font(color="FFFFFF", bold=True, name="Aptos")
    thin_grey = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)

    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 28

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, 202):
        ws.row_dimensions[row].height = 36
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(name="Aptos", size=11)


def build_seed_rows():
    rows = []
    ticket = 1

    for _, rule in proposed_validation_rules(None).iterrows():
        rows.append([
            f"TX-{ticket:03d}",
            str(rule["Controlled Field"]),
            f"{rule['Driver Field']} = {rule['Driver Value']}",
            "Irrelevant cross-channel usage" if "filter" in str(rule["Logic Type"]).lower() or "forbidden" in str(rule["Logic Type"]).lower() else "Missing controlled vocabulary",
            str(rule["Reason"]),
            f"{rule['Logic Type']}: allow {rule['Allowed Values']}",
            "Multi-channel" if "," in str(rule["Driver Value"]) or "!=" in str(rule["Driver Value"]) else str(rule["Driver Value"]),
            "Current-State",
            "High" if str(rule["Confidence"]) == "High" else "Medium",
            "",
            "Ready for Spec",
            "Open",
            "Needs confirmation that PlanIT can enforce this validation pattern.",
            "Channel lead / taxonomy owner",
            "Mirror the same rule in dashboard quality checks and filter logic.",
        ])
        ticket += 1

    for _, rec in current_state_recommendations().iterrows():
        rows.append([
            f"TX-{ticket:03d}",
            "Cross-field governance",
            str(rec["Recommendation"]),
            "Missing controlled vocabulary",
            str(rec["Why it matters"]),
            str(rec["Recommendation"]),
            "All",
            "Current-State",
            str(rec["Priority"]),
            "",
            "Backlog",
            "Open",
            "May affect templates, reporting and PlanIT dropdown configuration.",
            "Taxonomy owner / analytics owner",
            "Translate into specific dashboard rules or warnings once finalised.",
        ])
        ticket += 1

    for _, rec in future_state_recommendations().iterrows():
        rows.append([
            f"TX-{ticket:03d}",
            "Structural redesign",
            str(rec["Recommendation"]),
            "Design smell workaround",
            str(rec["Why it matters"]),
            str(rec["Recommendation"]),
            "All",
            "Future-State",
            str(rec["Priority"]),
            "",
            "Backlog",
            "Open",
            "Likely requires template and historical compatibility review.",
            "Chloe / Sam Tait / channel leads",
            "Not for immediate dashboard implementation; keep visible for roadmap planning.",
        ])
        ticket += 1

    for _, miss in missing_dimensions().iterrows():
        rows.append([
            f"TX-{ticket:03d}",
            str(miss["Proposed Dimension Name"]),
            "Missing dimension",
            "Missing dimension",
            str(miss["Why current taxonomy fails"]),
            f"Add dimension with example values: {miss['Example Values']}",
            str(miss["Applies To"]),
            str(miss["Whether current-state feasible or future-state only"]),
            "High",
            "",
            "Backlog",
            "Open",
            "Needs agreement on field ownership, level, and allowed values.",
            "Laura / Chloe / analytics / channel owners",
            "Add as new controlled field or derived dashboard dimension depending on implementation route.",
        ])
        ticket += 1

    return rows


def add_seed_rows(ws):
    rows = build_seed_rows()
    start_row = 2
    for row_idx, row_values in enumerate(rows, start=start_row):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def add_table(ws):
    table = Table(displayName="TaxonomyLogicTracker", ref="A1:O200")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def add_validations(ws, list_ws):
    validations = {
        "D": ISSUE_TYPES,
        "G": CHANNEL_SCOPES,
        "H": CURRENT_FUTURE,
        "I": PRIORITIES,
        "K": DEV_STAGES,
        "L": STATUSES,
    }

    start_col = 1
    for target_col, values in validations.items():
        col_letter = chr(64 + start_col)
        for idx, value in enumerate(values, start=2):
            list_ws.cell(row=idx, column=start_col, value=value)
        formula = f"=Lists!${col_letter}$2:${col_letter}${len(values) + 1}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.prompt = "Select from dropdown"
        dv.error = "Choose a value from the approved dropdown list."
        ws.add_data_validation(dv)
        dv.add(f"{target_col}2:{target_col}200")
        start_col += 1


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Logic Tracker"
    style_sheet(ws)
    add_seed_rows(ws)
    add_table(ws)

    guide = wb.create_sheet("Guide")
    guide["A1"] = "How to use"
    guide["A1"].font = Font(bold=True, size=14, name="Aptos")
    guide["A3"] = "1. Add one row per logic issue, rule, missing field, or governance ticket."
    guide["A4"] = "2. Use dropdowns for Issue Type, Channel Scope, Current-State or Future-State, Priority, Development Stage, and Status."
    guide["A5"] = "3. Keep Proposed Logic / Rule specific enough to implement in the dashboard or PlanIT validation layer."
    guide["A6"] = "4. Use Dashboard Implementation Notes to describe how the logic should surface in the app."
    guide["A8"] = "Suggested workflow"
    guide["A8"].font = Font(bold=True, size=12, name="Aptos")
    guide["A9"] = "Backlog -> To Review -> Ready for Spec -> Ready for Build -> In Development -> In UAT -> Done"
    guide.column_dimensions["A"].width = 110
    for row in range(1, 12):
        guide.row_dimensions[row].height = 22
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Aptos", size=11)

    list_ws = wb.create_sheet("Lists")
    add_validations(ws, list_ws)
    list_ws.sheet_state = "hidden"

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(f"Saved {OUTPUT_PATH}")

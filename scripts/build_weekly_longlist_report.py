#!/usr/bin/env python3
import csv
import re
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
WEEKLY_DIR = BASE_DIR / 'pwc reports' / 'Weekly'
OUTPUT_DIR = BASE_DIR / 'pwc reports' / 'outputs'
OUTPUT_PATH = OUTPUT_DIR / 'weekly_longlist_report.xlsx'
CSV_DIR = OUTPUT_DIR / 'weekly_longlist_csv'
CSV_ALL_PATH = OUTPUT_DIR / 'weekly_longlist_all.csv'

WEEK_RE = re.compile(r'\bCW\s*\d+\b', re.IGNORECASE)
WEEK_DIR_RE = re.compile(r'^CW(\d+)$', re.IGNORECASE)
GRANULAR_HEADERS = ['markets', 'platform', 'ad type', 'models']
HEADER_ALIASES = {
    'ad type': ['activation group', 'activation', 'activation type'],
    'models': ['model'],
}


def normalize_text(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def find_longlist_sheet(wb):
    for name in wb.sheetnames:
        if re.search(r'longlist per week', name, re.IGNORECASE):
            return name
    return None


def row_contains(row, token):
    token_lower = token.lower()
    for cell in row:
        if isinstance(cell, str) and token_lower in cell.lower():
            return True
    return False


def row_is_blank(row):
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def normalize_header(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip().lower()
    return None


def find_week_labels(rows: List[List], header_idx: int) -> List[Tuple[int, str]]:
    for offset in range(1, 4):
        idx = header_idx - offset
        if idx < 0:
            break
        labels = parse_week_labels(rows[idx])
        if labels:
            labels.sort(key=lambda x: x[0])
            return labels
    return []


def find_week_row(rows: List[List], start_idx: int, lookback: int = 3) -> Optional[List]:
    for i in range(start_idx - 1, max(-1, start_idx - lookback - 1), -1):
        if i < 0:
            break
        if any(isinstance(cell, str) and WEEK_RE.search(cell) for cell in rows[i]):
            return rows[i]
    return None


def parse_week_labels(week_row: List) -> List[Tuple[int, str]]:
    labels = []
    for idx, cell in enumerate(week_row):
        if isinstance(cell, str):
            match = WEEK_RE.search(cell)
            if match:
                labels.append((idx, match.group(0).replace(' ', '')))
    return labels


def find_report_week_from_text(text: str) -> Optional[str]:
    match = WEEK_RE.search(text or '')
    if not match:
        return None
    return match.group(0).replace(' ', '')


def find_report_week(filename: str, source_path: Path) -> Optional[str]:
    for candidate in (filename, str(source_path)):
        week = find_report_week_from_text(candidate)
        if week:
            return week
    return None


def parse_sheet(sheet, source_file: Path):
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([normalize_text(cell) for cell in row])

    report_week = find_report_week(source_file.name, source_file)
    if report_week is None:
        report_week = 'Unknown'

    data_rows = []
    idx = 0
    while idx < len(rows):
        row = rows[idx]

        headers = [normalize_header(cell) for cell in row]
        header_positions = {}
        for header in GRANULAR_HEADERS:
            candidates = [header] + HEADER_ALIASES.get(header, [])
            for candidate in candidates:
                if candidate in headers:
                    header_positions[header] = headers.index(candidate)
                    break
        if len(header_positions) == len(GRANULAR_HEADERS):
            header_idx = idx
            week_labels = find_week_labels(rows, header_idx)
            if not week_labels:
                idx += 1
                continue
            report_week_value = report_week if report_week != 'Unknown' else week_labels[0][1]

            dims = header_positions
            model_col = dims['models']

            def week_for_col(col):
                applicable = None
                for start_col, label in week_labels:
                    if col >= start_col:
                        applicable = label
                return applicable

            metric_columns = []
            for col_idx, cell in enumerate(row):
                if col_idx <= model_col:
                    continue
                if isinstance(cell, str) and cell.strip():
                    metric_columns.append((col_idx, cell.strip()))

            idx = header_idx + 1
            while idx < len(rows):
                data_row = rows[idx]
                if row_is_blank(data_row):
                    break
                if row_contains(data_row, 'Column Labels') or row_contains(data_row, 'Row Labels'):
                    break
                if row_contains(data_row, 'Markets Comparison'):
                    break

                market = data_row[dims['markets']] if dims['markets'] < len(data_row) else None
                platform = data_row[dims['platform']] if dims['platform'] < len(data_row) else None
                ad_type = data_row[dims['ad type']] if dims['ad type'] < len(data_row) else None
                model = data_row[dims['models']] if dims['models'] < len(data_row) else None

                if not any([market, platform, ad_type, model]):
                    idx += 1
                    continue

                for col_idx, metric_name in metric_columns:
                    if col_idx >= len(data_row):
                        continue
                    week_label = week_for_col(col_idx)
                    if not week_label:
                        continue
                    if week_label != report_week_value:
                        continue
                    value = data_row[col_idx]
                    if isinstance(value, str) and value.strip().lower() in {'na', 'not live'}:
                        value = None
                    if value is None:
                        continue
                    data_rows.append({
                        'report_week': report_week if report_week != 'Unknown' else week_label,
                        'week': week_label,
                        'market': market,
                        'platform': platform,
                        'ad_type': ad_type,
                        'model': model,
                        'metric': metric_name,
                        'value': value,
                        'source_file': str(source_file.relative_to(BASE_DIR)),
                    })
                idx += 1
            continue

        idx += 1

    return data_rows


def main():
    raise RuntimeError(
        'Deprecated pipeline. Use scripts/extract_python_output_all.py to build '
        '`pwc reports/outputs/python_output_all.csv` from the "Python Output" sheet.'
    )

    weekly_dirs = [
        d for d in WEEKLY_DIR.iterdir()
        if d.is_dir() and WEEK_DIR_RE.match(d.name)
    ]
    if not weekly_dirs:
        raise FileNotFoundError(f'No CW folders found in {WEEKLY_DIR}')

    latest_dir = max(weekly_dirs, key=lambda d: int(WEEK_DIR_RE.match(d.name).group(1)))
    files = sorted(
        p
        for p in latest_dir.rglob('*.xlsx')
        if ':Zone.Identifier' not in p.name and not p.name.startswith('~$')
    )
    all_rows = []

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    CSV_DIR.mkdir(parents=True, exist_ok=True)

    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_name = find_longlist_sheet(wb)
        if not sheet_name:
            wb.close()
            continue
        sheet = wb[sheet_name]
        file_rows = parse_sheet(sheet, path)
        all_rows.extend(file_rows)
        wb.close()

        csv_path = CSV_DIR / f'{path.stem}.csv'
        with csv_path.open('w', newline='', encoding='utf-8') as handle:
            writer = csv.DictWriter(handle, fieldnames=[
                'report_week',
                'week',
                'market',
                'platform',
                'ad_type',
                'model',
                'metric',
                'value',
                'source_file',
            ])
            writer.writeheader()
            writer.writerows(file_rows)

    with CSV_ALL_PATH.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.DictWriter(handle, fieldnames=[
            'report_week',
            'week',
            'market',
            'platform',
            'ad_type',
            'model',
            'metric',
            'value',
            'source_file',
        ])
        writer.writeheader()
        writer.writerows(all_rows)

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = 'Raw_Longlist'

    headers = ['report_week', 'week', 'market', 'platform', 'ad_type', 'model', 'metric', 'value', 'source_file']
    ws.append(headers)
    for row in all_rows:
        ws.append([row[h] for h in headers])

    wb_out.save(OUTPUT_PATH)
    print(f'Wrote {len(all_rows)} rows to {OUTPUT_PATH}')
    print(f'Wrote per-file CSVs to {CSV_DIR}')
    print(f'Wrote consolidated CSV to {CSV_ALL_PATH}')


if __name__ == '__main__':
    main()

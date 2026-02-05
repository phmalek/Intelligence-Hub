#!/usr/bin/env python3
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
WEEKLY_DIR = BASE_DIR / 'pwc reports' / 'Weekly'
OUTPUT_PATH = BASE_DIR / 'pwc reports' / 'outputs' / 'python_output_all.csv'

DATE_RE = re.compile(r'^(\d{8})_')
CW_RE = re.compile(r'\bCW\s*\d+\b', re.IGNORECASE)
WEEK_DIR_RE = re.compile(r'^CW(\d+)$', re.IGNORECASE)


def parse_date_prefix(filename: str) -> Optional[str]:
    match = DATE_RE.match(filename)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), '%d%m%Y').strftime('%Y-%m-%d')
    except ValueError:
        return None


def parse_report_week(path: Path) -> Optional[str]:
    for text in (path.name, str(path)):
        match = CW_RE.search(text)
        if match:
            return match.group(0).replace(' ', '')
    return None


def find_python_output_sheet(wb):
    for name in wb.sheetnames:
        if re.search(r'python output', name, re.IGNORECASE):
            return name
    return None


def main():
    weekly_dirs = [
        d for d in WEEKLY_DIR.iterdir()
        if d.is_dir() and WEEK_DIR_RE.match(d.name)
    ]
    if not weekly_dirs:
        raise FileNotFoundError(f'No CW folders found in {WEEKLY_DIR}')

    latest_dir = max(weekly_dirs, key=lambda d: int(WEEK_DIR_RE.match(d.name).group(1)))
    files = sorted(
        p
        for p in latest_dir.rglob('*.xlsx')
        if ':Zone.Identifier' not in p.name and not p.name.startswith('~$')
    )
    all_rows = []
    header = None

    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_name = find_python_output_sheet(wb)
        if not sheet_name:
            wb.close()
            continue
        ws = wb[sheet_name]

        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and cell != '' for cell in row):
                rows.append(row)

        wb.close()
        if not rows:
            continue

        file_header = [str(col).strip() if col is not None else '' for col in rows[0]]
        if header is None:
            header = file_header
        elif file_header != header:
            # Build a superset header to handle schema drift across weeks.
            merged = list(header)
            for col in file_header:
                if col not in merged:
                    merged.append(col)
            header = merged

        report_date = parse_date_prefix(path.name)
        report_week = parse_report_week(path) or ''

        file_index = {col: idx for idx, col in enumerate(file_header)}
        for row in rows[1:]:
            out_row = []
            for col in header:
                idx = file_index.get(col)
                out_row.append(row[idx] if idx is not None and idx < len(row) else None)
            all_rows.append([report_date, report_week, str(path.relative_to(BASE_DIR))] + out_row)

    if header is None:
        raise RuntimeError('No Python Output sheets found.')

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.writer(handle)
        writer.writerow(['report_date', 'report_week', 'source_file'] + header)
        writer.writerows(all_rows)

    print(f'Wrote {len(all_rows)} rows to {OUTPUT_PATH}')


if __name__ == '__main__':
    main()

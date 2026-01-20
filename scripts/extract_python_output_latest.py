#!/usr/bin/env python3
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

BASE_DIR = Path('/home/ali/repos/porsche')
WEEKLY_DIR = BASE_DIR / 'pwc reports' / 'Weekly'
OUTPUT_PATH = BASE_DIR / 'pwc reports' / 'outputs' / 'python_output_latest.csv'

DATE_RE = re.compile(r'^(\d{8})_')
WEEK_DIR_RE = re.compile(r'^CW(\d+)$', re.IGNORECASE)


def parse_date_prefix(filename: str) -> Optional[datetime]:
    match = DATE_RE.match(filename)
    if not match:
        return None
    return datetime.strptime(match.group(1), '%d%m%Y')


def find_latest_file() -> Path:
    weekly_dirs = [
        d for d in WEEKLY_DIR.iterdir()
        if d.is_dir() and WEEK_DIR_RE.match(d.name)
    ]
    if not weekly_dirs:
        raise FileNotFoundError(f'No CW folders found in {WEEKLY_DIR}')

    latest_dir = max(weekly_dirs, key=lambda d: int(WEEK_DIR_RE.match(d.name).group(1)))
    candidates = []
    for path in latest_dir.rglob('*.xlsx'):
        if ':Zone.Identifier' in path.name or path.name.startswith('~$'):
            continue
        date = parse_date_prefix(path.name)
        candidates.append((date or datetime.min, path.name, path))
    if not candidates:
        raise RuntimeError('No weekly Excel files found in latest CW folder.')
    candidates.sort(key=lambda item: (item[0], item[1]))
    return candidates[-1][2]


def main():
    latest = find_latest_file()
    wb = load_workbook(latest, read_only=True, data_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if re.search(r'python output', name, re.IGNORECASE):
            sheet_name = name
            break
    if not sheet_name:
        raise RuntimeError('Python Output sheet not found in latest file.')

    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None and cell != '' for cell in row):
            rows.append(row)

    if not rows:
        raise RuntimeError('Python Output sheet is empty.')

    header = [str(col).strip() if col is not None else '' for col in rows[0]]
    with OUTPUT_PATH.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for row in rows[1:]:
            writer.writerow(row)

    print(f'Latest file: {latest}')
    print(f'Wrote CSV: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()

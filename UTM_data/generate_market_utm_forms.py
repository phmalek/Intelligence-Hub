from __future__ import annotations

import argparse
import datetime as dt
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


CHANNELS = [
    "Google Ads",
    "Google Inventory Ads",
    "Bing",
    "Facebook",
    "Facebook Inventory Ads",
    "Tiktok",
    "LinkedIn",
    "Pinterest",
    "DV360",
    "Amazon",
    "Direct Buys",
]

TRAFFIC_LIGHT_OPTIONS = ["Green", "Amber", "Red", "Not Live", "Unknown"]
IMPLEMENTATION_STATUS_OPTIONS = [
    "Fully implemented",
    "Partially implemented",
    "In progress",
    "Not started",
    "Not applicable",
]
DATA_QUALITY_OPTIONS = ["Good", "Needs review", "Poor", "No data"]
TRACKING_QUALITY_OPTIONS = ["Reliable", "Some issues", "Major gaps", "Unknown"]
BLOCKER_TYPE_OPTIONS = [
    "No blocker",
    "Tracking setup",
    "Tagging consistency",
    "Naming/taxonomy",
    "Platform configuration",
    "Agency process",
    "Client approval",
    "Data pipeline",
    "Other",
]
YES_NO_OPTIONS = ["Yes", "No"]
CONFIDENCE_OPTIONS = ["High", "Medium", "Low"]
PRIORITY_OPTIONS = ["Critical", "High", "Medium", "Low"]
ACTION_STATUS_OPTIONS = ["Open", "In progress", "Blocked", "Done"]
RESPONSE_STATUS_OPTIONS = ["Reviewed", "In progress", "Blocked", "Resolved", "Not applicable"]


# Theme (clean and readable)
FONT_MAIN = "Calibri"
FONT_TITLE = "Calibri"
COLOR_BG = "F7F8FA"
COLOR_HEADER_DARK = "1F2937"
COLOR_HEADER_LIGHT = "374151"
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_INPUT_BG = "EEF2F7"
COLOR_RESPONSE_BG = "FFFFFF"
COLOR_ROW_ALT = "FAFBFC"
COLOR_ACCENT = "0F766E"
COLOR_TEXT = "111827"
COLOR_MUTED = "6B7280"
COLOR_GREEN = "D1FAE5"
COLOR_AMBER = "FEF3C7"
COLOR_RED = "FEE2E2"
COLOR_BORDER = "D1D5DB"

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w\-. ]+", "_", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "market"


def _find_market_column(ws) -> tuple[int, str]:
    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    for candidate in ("Porsche Country", "Country"):
        if candidate in headers:
            return headers.index(candidate) + 1, candidate
    raise ValueError(
        f"Could not find market column in first row headers. Found: {headers}"
    )


def extract_markets(source_file: Path) -> list[str]:
    wb = load_workbook(source_file, data_only=True)
    # Prefer the review sheet if it exists.
    sheet = None
    for ws in wb.worksheets:
        if "review" in ws.title.lower():
            sheet = ws
            break
    if sheet is None:
        sheet = wb.worksheets[0]

    col_idx, col_name = _find_market_column(sheet)
    markets = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, col_idx).value
        market = str(value).strip() if value is not None else ""
        if market:
            markets.append(market)

    unique_markets = sorted(set(markets))
    if not unique_markets:
        raise ValueError(
            f"No markets found in column '{col_name}' on sheet '{sheet.title}'."
        )
    return unique_markets


def _get_review_sheet(wb):
    for ws in wb.worksheets:
        if "review" in ws.title.lower():
            return ws
    return wb.worksheets[0]


def extract_market_issue_rows(source_file: Path) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(source_file, data_only=True)
    ws = _get_review_sheet(wb)

    headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    header_map = {h: idx + 1 for idx, h in enumerate(headers) if h}

    required = ["Porsche Country", "Platform", "Traffic Light", "Status Notes", "Key Point", "Next Step"]
    # Fallback for files that use Country instead.
    if "Porsche Country" not in header_map and "Country" in header_map:
        header_map["Porsche Country"] = header_map["Country"]

    missing = [h for h in required if h not in header_map]
    if missing:
        raise ValueError(
            f"Source sheet '{ws.title}' is missing required columns: {', '.join(missing)}"
        )

    issues_by_market: dict[str, list[dict[str, str]]] = {}
    for row in range(2, ws.max_row + 1):
        market = str(ws.cell(row, header_map["Porsche Country"]).value or "").strip()
        if not market:
            continue
        platform = str(ws.cell(row, header_map["Platform"]).value or "").strip()
        if not platform:
            continue
        issue = {
            "Platform": platform,
            "Traffic Light": str(ws.cell(row, header_map["Traffic Light"]).value or "").strip(),
            "Status Notes": str(ws.cell(row, header_map["Status Notes"]).value or "").strip(),
            "Key Point": str(ws.cell(row, header_map["Key Point"]).value or "").strip(),
            "Next Step": str(ws.cell(row, header_map["Next Step"]).value or "").strip(),
        }
        issues_by_market.setdefault(market, []).append(issue)

    return issues_by_market


def _style_header(ws, row: int = 1):
    fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_DARK)
    font = Font(color=COLOR_HEADER_TEXT, bold=True, name=FONT_MAIN, size=11)
    for cell in ws[row]:
        if cell.value in (None, ""):
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_title(ws, cell_ref: str, text: str):
    ws[cell_ref] = text
    ws[cell_ref].font = Font(name=FONT_TITLE, size=15, bold=True, color=COLOR_ACCENT)


def _apply_body_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if not cell.font or cell.font.name is None:
                cell.font = Font(name=FONT_MAIN, size=10, color=COLOR_TEXT)


def _set_column_widths(ws, mapping: dict[str, int]):
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


def _add_dropdown(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _build_lists_sheet(wb: Workbook):
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Traffic Light"
    for i, val in enumerate(TRAFFIC_LIGHT_OPTIONS, start=2):
        ws[f"A{i}"] = val

    ws["B1"] = "Implementation Status"
    for i, val in enumerate(IMPLEMENTATION_STATUS_OPTIONS, start=2):
        ws[f"B{i}"] = val

    ws["C1"] = "Data Quality"
    for i, val in enumerate(DATA_QUALITY_OPTIONS, start=2):
        ws[f"C{i}"] = val

    ws["D1"] = "Tracking Quality"
    for i, val in enumerate(TRACKING_QUALITY_OPTIONS, start=2):
        ws[f"D{i}"] = val

    ws["E1"] = "Blocker Type"
    for i, val in enumerate(BLOCKER_TYPE_OPTIONS, start=2):
        ws[f"E{i}"] = val

    ws["F1"] = "Yes/No"
    for i, val in enumerate(YES_NO_OPTIONS, start=2):
        ws[f"F{i}"] = val

    ws["G1"] = "Confidence"
    for i, val in enumerate(CONFIDENCE_OPTIONS, start=2):
        ws[f"G{i}"] = val

    ws["H1"] = "Priority"
    for i, val in enumerate(PRIORITY_OPTIONS, start=2):
        ws[f"H{i}"] = val

    ws["I1"] = "Action Status"
    for i, val in enumerate(ACTION_STATUS_OPTIONS, start=2):
        ws[f"I{i}"] = val

    ws["J1"] = "Channels"
    for i, val in enumerate(CHANNELS, start=2):
        ws[f"J{i}"] = val

    ws["K1"] = "Response Status"
    for i, val in enumerate(RESPONSE_STATUS_OPTIONS, start=2):
        ws[f"K{i}"] = val

    ws.sheet_state = "hidden"


def _build_instructions_sheet(wb: Workbook, market: str, source_name: str):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws["A1"] = ""
    ws.merge_cells("A1:D1")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_DARK)
    _style_title(ws, "A2", "UTM Market Response Form")
    ws["A3"] = f"Market: {market}"
    ws["A4"] = f"Source file: {source_name}"
    ws["A3"].font = Font(name=FONT_MAIN, size=11, bold=True, color=COLOR_TEXT)
    ws["A4"].font = Font(name=FONT_MAIN, size=10, color=COLOR_MUTED)
    ws["A6"] = "How to fill this file (simple and complete):"
    ws["A6"].font = Font(name=FONT_MAIN, size=11, bold=True, color=COLOR_HEADER_LIGHT)

    guidance = [
        "1) Open 'Issue Responses': each row is an issue from the central input file for your market.",
        "2) Do NOT edit input columns (A-F). Fill only response columns (G-P).",
        "3) Use dropdown values whenever available (traffic light, status, support needed, confidence).",
        "4) If Response Traffic Light is Amber/Red, provide clear action plan and target date.",
        "5) Keep one clear response per issue row with owner, target date, and summary.",
        "6) Save and return this file with the same structure (do not rename sheets/columns).",
    ]
    for idx, line in enumerate(guidance, start=8):
        ws[f"A{idx}"] = line
        ws[f"A{idx}"].font = Font(name=FONT_MAIN, size=10, color=COLOR_TEXT)
        ws[f"A{idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A16"] = "Traffic Light meaning:"
    ws["A16"].font = Font(name=FONT_MAIN, size=11, bold=True, color=COLOR_HEADER_LIGHT)
    ws["A17"] = "Green = fully healthy / no urgent action"
    ws["A18"] = "Amber = partially healthy / action required soon"
    ws["A19"] = "Red = blocked / urgent support required"
    ws["A20"] = "Not Live = channel not active"
    ws["A21"] = "Unknown = no confirmed evidence yet"
    for row in range(17, 22):
        ws[f"A{row}"].font = Font(name=FONT_MAIN, size=10, color=COLOR_TEXT)

    ws["C8"] = "Important:"
    ws["C8"].font = Font(name=FONT_MAIN, size=11, bold=True, color=COLOR_HEADER_LIGHT)
    ws["C9"] = "Input columns in Issue Responses are prefilled from central review."
    ws["C10"] = "Please avoid editing input columns; respond in response columns."
    ws["C11"] = "Keep responses concise but complete so central review can consolidate easily."
    for row in range(9, 12):
        ws[f"C{row}"].font = Font(name=FONT_MAIN, size=10, color=COLOR_TEXT)
        ws[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    _set_column_widths(ws, {"A": 80, "B": 4, "C": 60, "D": 4})
    for row in range(6, 22):
        for col in ("A", "C"):
            ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor=COLOR_BG)


def _build_summary_sheet(wb: Workbook, market: str, market_issues: list[dict[str, str]]):
    ws = wb.create_sheet("Issue Responses")
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.showGridLines = True
    headers = [
        "Market",
        "Platform (from input)",
        "Input Traffic Light",
        "Input Status Notes",
        "Input Key Point",
        "Input Next Step",
        "Response Traffic Light",
        "Response Status",
        "Response Summary",
        "Response Action Plan",
        "Support Needed From Central",
        "Response Owner Name",
        "Response Owner Email",
        "Response Target Date",
        "Last Updated",
        "Confidence",
    ]
    ws.append(headers)
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34

    if market_issues:
        for issue in market_issues:
            ws.append([
                market,
                issue.get("Platform", ""),
                issue.get("Traffic Light", ""),
                issue.get("Status Notes", ""),
                issue.get("Key Point", ""),
                issue.get("Next Step", ""),
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ])
    else:
        for channel in CHANNELS:
            ws.append([market, channel] + [None] * (len(headers) - 2))

    last_row = ws.max_row
    _add_dropdown(ws, f"G2:G{last_row}", "'Lists'!$A$2:$A$6")
    _add_dropdown(ws, f"H2:H{last_row}", "'Lists'!$K$2:$K$6")
    _add_dropdown(ws, f"K2:K{last_row}", "'Lists'!$F$2:$F$3")
    _add_dropdown(ws, f"P2:P{last_row}", "'Lists'!$G$2:$G$4")

    for row in range(2, last_row + 1):
        ws[f"N{row}"].number_format = "yyyy-mm-dd"
        ws[f"O{row}"].number_format = "yyyy-mm-dd"
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor="F1F3F4")
            ws[f"{col}{row}"].font = Font(name=FONT_MAIN, size=10, color=COLOR_MUTED)
        for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O", "P"):
            ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor=COLOR_RESPONSE_BG)
            ws[f"{col}{row}"].font = Font(name=FONT_MAIN, size=10, color=COLOR_TEXT)
        if row % 2 == 0:
            for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O", "P"):
                ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor=COLOR_ROW_ALT)

    _apply_body_borders(ws, min_row=2, max_row=last_row, min_col=1, max_col=16)

    # Traffic light visual cues on response column.
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        FormulaRule(formula=[f'$G2="Green"'], fill=PatternFill(fill_type="solid", fgColor=COLOR_GREEN)),
    )
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        FormulaRule(formula=[f'$G2="Amber"'], fill=PatternFill(fill_type="solid", fgColor=COLOR_AMBER)),
    )
    ws.conditional_formatting.add(
        f"G2:G{last_row}",
        FormulaRule(formula=[f'$G2="Red"'], fill=PatternFill(fill_type="solid", fgColor=COLOR_RED)),
    )

    _set_column_widths(
        ws,
        {
            "A": 24,
            "B": 20,
            "C": 15,
            "D": 28,
            "E": 30,
            "F": 30,
            "G": 16,
            "H": 18,
            "I": 30,
            "J": 34,
            "K": 24,
            "L": 22,
            "M": 30,
            "N": 14,
            "O": 14,
            "P": 12,
        },
    )


def create_market_workbook(
    market: str,
    source_file: Path,
    output_dir: Path,
    issue_rows: list[dict[str, str]],
) -> Path:
    wb = Workbook()
    _build_instructions_sheet(wb, market, source_file.name)
    _build_summary_sheet(wb, market, issue_rows)
    _build_lists_sheet(wb)

    safe_market = _sanitize_filename(market)
    output_file = output_dir / f"{safe_market}_UTM_Response.xlsx"
    wb.save(output_file)
    return output_file


def generate_forms(source_file: Path, output_dir: Path, market: str = "all") -> list[Path]:
    markets = extract_markets(source_file)
    issues_by_market = extract_market_issue_rows(source_file)
    if market and market.lower() != "all":
        requested = market.strip()
        matches = [m for m in markets if m.lower() == requested.lower()]
        if not matches:
            raise ValueError(
                f"Market '{market}' not found in source file. Available: {', '.join(markets)}"
            )
        markets = matches
    output_dir.mkdir(parents=True, exist_ok=True)
    generated = []
    for market in markets:
        market_issues = issues_by_market.get(market, [])
        generated.append(create_market_workbook(market, source_file, output_dir, market_issues))
    return generated


def _iter_response_files(response_dir: Path) -> Iterable[Path]:
    for file in sorted(response_dir.glob("*_UTM_Response.xlsx")):
        if file.name.startswith("~$"):
            continue
        yield file


def consolidate_forms(response_dir: Path, output_file: Path) -> Path:
    summary_rows = [[
        "Source File",
        "Market",
        "Platform (from input)",
        "Input Traffic Light",
        "Input Status Notes",
        "Input Key Point",
        "Input Next Step",
        "Response Traffic Light",
        "Response Status",
        "Response Summary",
        "Response Action Plan",
        "Support Needed From Central",
        "Response Owner Name",
        "Response Owner Email",
        "Response Target Date",
        "Last Updated",
        "Confidence",
    ]]
    for file in _iter_response_files(response_dir):
        wb = load_workbook(file, data_only=True)
        if "Issue Responses" in wb.sheetnames:
            ws = wb["Issue Responses"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                platform = row[1]
                response_traffic = row[6]
                if platform is None and response_traffic is None:
                    continue
                summary_rows.append([file.name] + list(row[:16]))

    out_wb = Workbook()
    ws_summary = out_wb.active
    ws_summary.title = "All Issue Responses"
    for row in summary_rows:
        ws_summary.append(row)
    _style_header(ws_summary)
    _apply_body_borders(
        ws_summary,
        min_row=2,
        max_row=max(2, ws_summary.max_row),
        min_col=1,
        max_col=ws_summary.max_column,
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_file)
    return output_file


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate per-market UTM response forms and consolidate completed files."
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_generate = sub.add_parser("generate", help="Generate one workbook per market.")
    p_generate.add_argument(
        "--source",
        type=Path,
        required=True,
        help="Path to source review workbook (e.g. Porsche_UTM Adoption Feedback...xlsx).",
    )
    p_generate.add_argument(
        "--outdir",
        type=Path,
        default=Path("market_forms"),
        help="Output directory for generated market files.",
    )
    p_generate.add_argument(
        "--market",
        type=str,
        default="all",
        help="Market code from Porsche Country column, or 'all'.",
    )

    p_consolidate = sub.add_parser("consolidate", help="Combine completed market files.")
    p_consolidate.add_argument(
        "--responses",
        type=Path,
        required=True,
        help="Directory containing completed *_UTM_Response.xlsx files.",
    )
    p_consolidate.add_argument(
        "--output",
        type=Path,
        default=Path(f"utm_consolidated_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
        help="Path for consolidated output workbook.",
    )

    return parser.parse_args()


def main():
    args = parse_args()
    if args.command == "generate":
        files = generate_forms(args.source, args.outdir, market=args.market)
        print(f"Generated {len(files)} market form(s) in: {args.outdir}")
        for file in files:
            print(f" - {file.name}")
        return

    if args.command == "consolidate":
        output = consolidate_forms(args.responses, args.output)
        print(f"Consolidated workbook created: {output}")
        return

    raise ValueError(f"Unsupported command: {args.command}")


if __name__ == "__main__":
    main()
